from datetime import datetime, date
from pathlib import Path

from django.core.management.base import BaseCommand
from django.utils.text import slugify
from openpyxl import load_workbook

from football.models import (
    Competition,
    DataImportLog,
    Group,
    Match,
    MatchEvent,
    MatchReport,
    Player,
    Season,
    Team,
)


class Command(BaseCommand):
    help = 'Importa estadísticas de Benagalbón desde los archivos Excel que se colocan en data/excel/.'

    default_files = [
        Path('data/excel/FICHA_PARTIDO.xlsx'),
        Path('data/excel/Estadisticas_Partidos_20251206_213133.xlsx'),
        Path('data/excel/BDT PARTIDOS BENABALBON.xlsm'),
    ]

    def add_arguments(self, parser):
        parser.add_argument(
            '--files',
            nargs='+',
            type=Path,
            default=self.default_files,
            help='Rutas relativas a la raíz del proyecto para los Excel que quieres procesar.',
        )
        parser.add_argument('--competition', default='División de Honor Andaluza', help='Competición objetivo.')
        parser.add_argument('--season', default='2025/2026', help='Temporada a la que pertenecen las importaciones.')
        parser.add_argument('--group', default='Grupo 2', help='Grupo/liga principal de referencia.')

    def handle(self, *args, **options):
        competition_name = options['competition']
        season_name = options['season']
        group_name = options['group']

        competition, season, group = self.ensure_league_structure(
            competition_name, season_name, group_name
        )
        primary_team = self.ensure_primary_team(group)

        files = options['files']
        for path in files:
            if not path.exists():
                self.stderr.write(f'No se encuentra el archivo {path}')
                continue

            workbook = load_workbook(filename=path, read_only=True, data_only=True)
            sheets_report = self.describe_sheets(workbook)
            total_rows = sum(sheet['rows'] for sheet in sheets_report)
            log = DataImportLog.objects.create(
                file_name=path.name,
                row_count=total_rows,
                notes=f'Importado con import_benagalbon_excel desde {path}',
            )

            report_payload = {'sheets': sheets_report, 'log_id': log.id}
            self.stdout.write(self.style.NOTICE(f'Procesando {path.name} ({len(sheets_report)} hojas)...'))

            events = self.extract_bd_eventos(workbook)
            imported_events = 0
            if events:
                imported_events = self.ingest_events(
                    events,
                    season,
                    group,
                    primary_team,
                    path.name,
                )

            MatchReport.objects.create(
                match=None,
                source_file=path.name,
                raw_data={**report_payload, 'events': len(events)},
            )

            self.stdout.write(
                self.style.SUCCESS(
                    f'Procesado {path.name}: {len(sheets_report)} hojas, '
                    f'{total_rows} filas, {imported_events} eventos BD_EVENTOS'
                )
            )

    def ensure_league_structure(self, competition_name: str, season_name: str, group_name: str):
        competition, _ = Competition.objects.get_or_create(
            name=competition_name,
            defaults={
                'slug': slugify(competition_name),
                'region': 'Andalucía',
                'level': 5,
            },
        )
        season, _ = Season.objects.get_or_create(
            competition=competition,
            name=season_name,
            defaults={'is_current': True},
        )
        group, _ = Group.objects.get_or_create(
            season=season,
            slug=slugify(f'{season_name}-{group_name}'),
            defaults={'name': group_name},
        )
        return competition, season, group

    def ensure_primary_team(self, group: Group):
        team, created = Team.objects.get_or_create(
            slug='cd-benagalbon',
            defaults={
                'name': 'C.D. Benagalbón',
                'short_name': 'Benagalbón',
                'group': group,
                'is_primary': True,
            },
        )
        updated = False
        if team.group != group:
            team.group = group
            updated = True
        if not team.is_primary:
            team.is_primary = True
            updated = True
        if updated:
            team.save(update_fields=['group', 'is_primary'])
        return team

    @staticmethod
    def describe_sheets(workbook):
        sheets = []
        for name in workbook.sheetnames:
            sheet = workbook[name]
            rows = sum(1 for _ in sheet.iter_rows(values_only=True))
            sheets.append({'sheet': name, 'rows': rows})
        return sheets

    def extract_bd_eventos(self, workbook):
        sheet_name = 'BD_EVENTOS'
        if sheet_name not in workbook.sheetnames:
            return []
        sheet = workbook[sheet_name]
        rows = list(sheet.iter_rows(values_only=True))
        if len(rows) <= 1:
            return []

        headers = [self.normalize_header(value) for value in rows[0]]
        events = []
        for row in rows[1:]:
            if all(cell is None for cell in row):
                continue
            payload = {}
            for index, value in enumerate(row):
                if index >= len(headers) or not headers[index]:
                    continue
                payload[headers[index]] = value
            events.append(payload)
        return events

    def ingest_events(self, events, season, group, primary_team, source_file):
        matches = {}
        opponents = {}
        players = {}
        created_events = 0
        cleansed_matches = set()

        for row in events:
            partido_id = self.safe_text(row.get('partidoid'))
            if not partido_id:
                continue
            rival_name = self.safe_text(row.get('rival'), default='Rival desconocido')
            match_key = (partido_id, rival_name)
            match = matches.get(match_key)
            if not match:
                opponent = opponents.get(rival_name)
                if not opponent:
                    opponent_slug = slugify(rival_name) or f'rival-{partido_id}'
                    opponent, _ = Team.objects.get_or_create(
                        slug=opponent_slug,
                        defaults={'name': rival_name, 'group': group},
                    )
                    opponents[rival_name] = opponent

                round_label = self.safe_text(row.get('jornada'), default=f'Partido {partido_id}')
                match_date = self.parse_date(row.get('fecha'))
                match, _ = Match.objects.get_or_create(
                    season=season,
                    round=round_label,
                    home_team=primary_team,
                    away_team=opponent,
                    defaults={
                        'group': group,
                        'date': match_date,
                        'location': self.safe_text(row.get('campo')),
                        'source': source_file,
                    },
                )
                matches[match_key] = match

            if match.id not in cleansed_matches:
                MatchEvent.objects.filter(match=match, source_file=source_file).delete()
                cleansed_matches.add(match.id)

            player_name = self.safe_text(row.get('jugador'))
            player = None
            if player_name:
                player = players.get(player_name)
                if not player:
                    player, _ = Player.objects.get_or_create(
                        team=primary_team,
                        name=player_name,
                    )
                    players[player_name] = player

            event = MatchEvent.objects.create(
                match=match,
                player=player,
                minute=self.safe_int(row.get('minuto')),
                event_type=self.safe_text(row.get('evento')),
                result=self.safe_text(row.get('resultadoaccion')),
                zone=self.safe_text(row.get('zona')),
                tercio=self.safe_text(row.get('tercio')),
                observation=self.safe_text(row.get('observacion')),
                system=self.safe_text(row.get('sistema')),
                source_file=source_file,
                raw_data=self.clean_payload(row),
            )
            created_events += 1

        return created_events

    @staticmethod
    def normalize_header(value):
        if not value:
            return ''
        return ''.join(ch.lower() for ch in str(value).strip() if ch.isalnum())

    @staticmethod
    def safe_text(value, default=''):
        if value is None:
            return default
        text = str(value).strip()
        return text if text else default

    @staticmethod
    def safe_int(value):
        if value is None:
            return None
        try:
            return int(float(value))
        except (ValueError, TypeError):
            return None

    @staticmethod
    def parse_date(value):
        if isinstance(value, (datetime, date)):
            return value.date() if isinstance(value, datetime) else value
        if isinstance(value, str):
            for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y'):
                try:
                    return datetime.strptime(value, fmt).date()
                except ValueError:
                    continue
        return None

    @staticmethod
    def clean_payload(row):
        cleaned = {}
        for key, value in row.items():
            if isinstance(value, (datetime, date)):
                cleaned[key] = value.isoformat()
            else:
                cleaned[key] = value
        return cleaned
