import copy
import json
import logging
import os
import subprocess
import sys
import base64
import mimetypes
import io
import csv
import html
import zipfile
import uuid
import tempfile
import shutil
from collections import Counter, defaultdict
from datetime import datetime, timedelta, time, date
from html.parser import HTMLParser
from functools import wraps
from pathlib import Path
import unicodedata
import re
from types import SimpleNamespace

from django.conf import settings
from django.contrib.auth.decorators import login_required
from django.contrib.auth.models import User
from django.contrib.auth import login as auth_login
from django.contrib.auth.password_validation import validate_password
from django.contrib.auth.hashers import make_password, check_password
from django.core.cache import cache
from django.core.exceptions import ValidationError as DjangoValidationError
from django.core.files.base import ContentFile
from django.core.files.storage import default_storage
from django.db import connections
from django.db import transaction
from django.db.models import Count, Max, Q
from django.db.utils import OperationalError, ProgrammingError
from django.http import FileResponse, Http404, HttpResponse, JsonResponse
from django.shortcuts import render, redirect
from django.template.loader import render_to_string
from django.templatetags.static import static
from django.urls import reverse
from django.utils.dateparse import parse_date
from django.utils import timezone
from django.utils.text import slugify
from django.views.decorators.csrf import ensure_csrf_cookie
from django.views.decorators.http import require_POST
from django.views.decorators.csrf import csrf_exempt

from .task_backups import write_task_backup

try:
    from PIL import Image
except Exception:  # pragma: no cover
    Image = None

try:
    import pytesseract
except Exception:  # pragma: no cover
    pytesseract = None

try:
    import weasyprint
except Exception:  # pragma: no cover
    weasyprint = None

try:
    from pypdf import PdfReader
except Exception:  # pragma: no cover
    PdfReader = None

try:
    import requests
except Exception:  # pragma: no cover
    requests = None

from football.models import (
    DataSource,
    Competition,
    Group,
    Match,
    MatchEvent,
    MatchReport,
    CustomMetric,
    Player,
    PlayerInjuryRecord,
    PlayerCommunication,
    PlayerFine,
    PlayerPhysicalMetric,
    PlayerStatistic,
    SessionTask,
    ScrapeSource,
    Season,
    Team,
    TeamStatistic,
    TeamStanding,
    Workspace,
    WorkspaceMembership,
    TrainingMicrocycle,
    TrainingSession,
    ConvocationRecord,
    HomeCarouselImage,
    AnalystVideoFolder,
    RivalVideo,
    RivalAnalysisReport,
    AppUserRole,
    UserInvitation,
    ShareLink,
    AuditEvent,
    TaskBlueprint,
    TaskStudioProfile,
    TaskStudioRosterPlayer,
    TaskStudioTask,
    WorkspaceCompetitionContext,
    WorkspaceCompetitionSnapshot,
)
from football.event_taxonomy import (
    DRIBBLE_KEYWORDS,
    FIELD_ZONE_KEYS,
    FIELD_ZONES,
    PASS_KEYWORDS,
    SHOT_KEYWORDS,
    STANDARD_TERCIO_LABELS,
    build_smart_kpis,
    calculate_influence_score,
    calculate_importance_score,
    categorize_position,
    classify_duel_event,
    contains_keyword,
    extract_round_number,
    is_assist_event,
    is_goal_event,
    is_goalkeeper_save_event,
    is_key_pass_event,
    is_red_card_event,
    is_shot_attempt_event,
    is_shot_on_target_event,
    is_substitution_entry,
    is_substitution_event,
    is_substitution_exit,
    is_yellow_card_event,
    map_tercio,
    map_zone_label,
    min_or_none,
    normalize_label,
    result_is_success,
    shots_needed_per_goal,
    zone_to_tercio,
)
from football.services import (
    assign_lineup_slots,
    canonical_roster_key,
    compute_probable_eleven,
    compute_formation,
    build_rival_insights,
    update_team_standings,
    fetch_preferente_team_roster,
    find_roster_entry,
    get_roster_stats_cache,
    load_match_actions,
    load_match_results,
    normalize_player_name,
    parse_preferente_roster,
    refresh_primary_roster_cache,
    _parse_int,
)
from football.staff_briefing import build_weekly_staff_brief
from football.manual_stats import (
    get_manual_player_base_overrides,
    resolve_stats_season,
    save_manual_player_base_overrides,
    season_display_name,
)
from football.query_helpers import (
    _normalize_team_lookup_key,
    _team_match_queryset,
    confirmed_events_queryset,
    get_active_injury_player_ids,
    get_active_match,
    get_current_convocation,
    get_current_convocation_record,
    get_latest_pizarra_match,
    get_previous_match,
    get_requested_match,
    get_sanctioned_player_ids_from_previous_round,
    is_injury_record_active,
    is_manual_sanction_active,
    parse_match_date_from_ui,
)
from football.task_library import filter_task_library, prepare_task_library

logger = logging.getLogger(__name__)

def _team_standings_last_updated(group):
    if not group:
        return None
    try:
        return TeamStanding.objects.filter(group=group).aggregate(latest=Max('last_updated')).get('latest')
    except Exception:
        return None


def _latest_standings_group_for_team(primary_team):
    if not primary_team:
        return None
    try:
        standing = (
            TeamStanding.objects
            .select_related('group')
            .filter(team=primary_team)
            .order_by('-last_updated', '-played', '-id')
            .first()
        )
        return standing.group if standing else None
    except Exception:
        return None


def _refresh_rfaf_standings_inline(*, allow_fallback=False):
    """
    Refresco rápido (en-process) de la clasificación RFAF.
    Evita subprocess largos y permite devolver errores claros.
    """
    try:
        from scripts import import_from_rfef
    except Exception as exc:
        return False, f'No se pudo cargar el módulo RFAF: {exc}', None
    try:
        rows, html = import_from_rfef.parse_table(allow_fallback=bool(allow_fallback))
        next_match = import_from_rfef.extract_next_match_from_classification(html)
        if not next_match:
            next_match = import_from_rfef.fetch_next_match_from_classification(html)
        if next_match and next_match.get("status") != "next":
            next_match = None
        try:
            import_from_rfef.save_next_match_cache(next_match)
        except Exception:
            pass
        update_team_standings(
            rows,
            'RFAF',
            getattr(import_from_rfef, 'URL', '') or '',
        )
        return True, f'Clasificación actualizada (filas={len(rows)}).', next_match
    except Exception as exc:
        return False, str(exc) or 'Error desconocido actualizando RFAF.', None


@login_required
def session_keepalive(request):
    """
    Mantiene viva la sesión mientras el usuario está en el editor (pizarra) sin hacer requests.
    Evita perder trabajo por caducidad de sesión al pulsar "Guardar/Crear" tras estar tiempo editando.
    """
    request.session['__keepalive__'] = timezone.now().isoformat()
    response = JsonResponse({'ok': True})
    response['Cache-Control'] = 'no-store'
    return response


@login_required
def system_diagnostics(request):
    """Diagnóstico (solo admin) para comprobar persistencia en Render.

    No expone secretos: únicamente vendor/engine y flags.
    """
    if not _is_admin_user(request.user):
        return JsonResponse({'error': 'No autorizado'}, status=403)
    default_db = settings.DATABASES.get('default', {}) if hasattr(settings, 'DATABASES') else {}
    engine = str(default_db.get('ENGINE') or '').strip()
    name = str(default_db.get('NAME') or '').strip()
    vendor = ''
    try:
        vendor = str(getattr(connections['default'], 'vendor', '') or '')
    except Exception:
        vendor = ''
    return JsonResponse(
        {
            'ok': True,
            'database': {
                'vendor': vendor,
                'engine': engine,
                # Solo devolvemos NAME en DEBUG o sqlite (en Postgres se oculta).
                'name': name if (settings.DEBUG or vendor == 'sqlite') else '',
                'uses_database_url': bool(os.getenv('DATABASE_URL', '').strip()),
            },
            'session': {
                'engine': str(getattr(settings, 'SESSION_ENGINE', '') or '').strip(),
                'cookie_name': str(getattr(settings, 'SESSION_COOKIE_NAME', '') or '').strip(),
                'cookie_age': int(getattr(settings, 'SESSION_COOKIE_AGE', 0) or 0),
                'save_every_request': bool(getattr(settings, 'SESSION_SAVE_EVERY_REQUEST', False)),
            },
            'csrf': {
                'cookie_name': str(getattr(settings, 'CSRF_COOKIE_NAME', '') or '').strip(),
                'trusted_origins_count': len(getattr(settings, 'CSRF_TRUSTED_ORIGINS', []) or []),
            },
        }
    )

SCRIPT_PATH = Path(__file__).resolve().parents[1] / "scripts" / "import_from_rfef.py"
MANAGE_PY_DIR = SCRIPT_PATH.parents[1]
NEXT_MATCH_CACHE = Path(settings.BASE_DIR) / "data" / "input" / "rfaf-next-match.json"
UNIVERSO_SNAPSHOT_PATH = Path(settings.BASE_DIR) / "data" / "input" / "universo-rfaf-snapshot.json"
UNIVERSO_CAPTURE_PATH = Path(settings.BASE_DIR) / "data" / "input" / "universo-rfaf-capture.json"
TASK_RESOURCE_LIBRARY_PATH = Path(settings.BASE_DIR) / "data" / "input" / "task-resource-library.json"
SCRAPE_LOCK_KEY = "football:refresh_scraping_running"
SCRAPE_LOCK_TIMEOUT_SECONDS = 900
DASHBOARD_CACHE_KEY_PREFIX = "football:dashboard_payload"
DASHBOARD_CACHE_SECONDS = int(os.getenv('DASHBOARD_CACHE_SECONDS', '600'))
PLAYER_DASHBOARD_CACHE_KEY_PREFIX = "football:player_dashboard"
PLAYER_DASHBOARD_CACHE_SECONDS = int(os.getenv('PLAYER_DASHBOARD_CACHE_SECONDS', '600'))
PLAYER_PHOTO_VERSION_CACHE_KEY_PREFIX = "football:player_photo_version"
PLAYER_PHOTO_VERSION_CACHE_SECONDS = int(os.getenv('PLAYER_PHOTO_VERSION_CACHE_SECONDS', '86400'))
TEAM_METRICS_CACHE_SECONDS = int(os.getenv('TEAM_METRICS_CACHE_SECONDS', '900'))
PLAYER_METRICS_CACHE_SECONDS = int(os.getenv('PLAYER_METRICS_CACHE_SECONDS', '900'))
RFAF_LIVE_FETCH_ON_REQUEST = str(
    os.getenv('RFAF_LIVE_FETCH_ON_REQUEST', '0')
).strip().lower() in {'1', 'true', 'yes', 'on'}
UNIVERSO_API_TIMEOUT_SECONDS = max(1, int(os.getenv('UNIVERSO_API_TIMEOUT_SECONDS', '8') or 8))
UNIVERSO_EXTERNAL_IMAGES_ENABLED = str(
    os.getenv('UNIVERSO_EXTERNAL_IMAGES_ENABLED', '0')
).strip().lower() in {'1', 'true', 'yes', 'on'}

TASK_MATERIAL_LIBRARY = [
    {'label': 'CONO', 'title': 'Cono alto', 'kind': 'cone', 'category': 'delimitacion', 'icon': '△'},
    {'label': 'SETA', 'title': 'Seta baja', 'kind': 'marker', 'category': 'delimitacion', 'icon': '◉'},
    {'label': 'DISCO', 'title': 'Disco plano', 'kind': 'marker-flat', 'category': 'delimitacion', 'icon': '◍'},
    {'label': 'PICA', 'title': 'Pica', 'kind': 'pole', 'category': 'delimitacion', 'icon': '┃'},
    {'label': 'CINTA', 'title': 'Cinta delimitación', 'kind': 'tape', 'category': 'delimitacion', 'icon': '═'},
    {'label': 'ARO', 'title': 'Aro coordinación', 'kind': 'ring', 'category': 'coordinacion', 'icon': '◯'},
    {'label': 'ESC', 'title': 'Escalera coordinación', 'kind': 'ladder', 'category': 'coordinacion', 'icon': '☷'},
    {'label': 'VALLA', 'title': 'Valla coordinación', 'kind': 'hurdle', 'category': 'coordinacion', 'icon': '⊓'},
    {'label': 'SLALOM', 'title': 'Palo slalom', 'kind': 'slalom', 'category': 'coordinacion', 'icon': '╽'},
    {'label': 'MINI', 'title': 'Mini valla velocidad', 'kind': 'mini-hurdle', 'category': 'coordinacion', 'icon': '⫍'},
    {'label': 'BAL', 'title': 'Balón', 'kind': 'ball', 'category': 'juego', 'icon': '●'},
    {'label': 'MED', 'title': 'Balón medicinal', 'kind': 'medicine-ball', 'category': 'fisico', 'icon': '◒'},
    {'label': 'PORT', 'title': 'Portería móvil', 'kind': 'goal', 'category': 'porterias', 'icon': '⌷'},
    {'label': 'MPORT', 'title': 'Mini portería', 'kind': 'mini-goal', 'category': 'porterias', 'icon': '⌸'},
    {'label': 'PGK', 'title': 'Portería portero', 'kind': 'gk-goal', 'category': 'porterias', 'icon': '▭'},
    {'label': 'MANIQ', 'title': 'Maniquí', 'kind': 'mannequin', 'category': 'oposicion', 'icon': '♞'},
    {'label': 'DUMMY', 'title': 'Dummie inflable', 'kind': 'dummy', 'category': 'oposicion', 'icon': '♜'},
    {'label': 'MURO', 'title': 'Muro faltas', 'kind': 'wall', 'category': 'oposicion', 'icon': '▥'},
    {'label': 'PETO', 'title': 'Petos', 'kind': 'bib', 'category': 'organizacion', 'icon': '▣'},
    {'label': 'GPS', 'title': 'GPS chaleco', 'kind': 'gps', 'category': 'control', 'icon': '⌖'},
    {'label': 'CRONO', 'title': 'Cronómetro', 'kind': 'timer', 'category': 'control', 'icon': '◷'},
    {'label': 'SILB', 'title': 'Silbato', 'kind': 'whistle', 'category': 'control', 'icon': '⌇'},
    {'label': 'TAB', 'title': 'Pizarra coach', 'kind': 'board', 'category': 'control', 'icon': '▤'},
    {'label': 'GOMA', 'title': 'Goma resistencia', 'kind': 'band', 'category': 'fisico', 'icon': '∞'},
    {'label': 'TRX', 'title': 'Suspensión TRX', 'kind': 'trx', 'category': 'fisico', 'icon': '⟂'},
    {'label': 'CAJON', 'title': 'Cajón pliometría', 'kind': 'box', 'category': 'fisico', 'icon': '▧'},
    {'label': 'TRINEO', 'title': 'Trineo arrastre', 'kind': 'sled', 'category': 'fisico', 'icon': '⎍'},
    {'label': 'ESTACA', 'title': 'Estaca zona', 'kind': 'stake', 'category': 'delimitacion', 'icon': '⎸'},
    {'label': 'ARCO', 'title': 'Arco precisión', 'kind': 'arc', 'category': 'finalizacion', 'icon': '◠'},
    {'label': 'OBJ', 'title': 'Objetivo diana', 'kind': 'target', 'category': 'finalizacion', 'icon': '◎'},
]
TASK_MATERIAL_PPT_DIR = Path(settings.BASE_DIR) / 'static' / 'football' / 'images' / 'task-materials' / 'ppt'
TASK_SURFACE_CHOICES = [
    ('natural_grass', 'Hierba natural'),
    ('hybrid_grass', 'Hierba híbrida'),
    ('artificial_turf', 'Césped artificial'),
    ('futsal', 'Pista futsal'),
    ('sand', 'Arena'),
    ('indoor', 'Indoor'),
    ('gym', 'Gimnasio'),
    ('dirt', 'Tierra'),
    ('street', 'Asfalto'),
]
TASK_PITCH_FORMAT_CHOICES = [
    ('11v11_full', '11v11 · Campo completo'),
    ('11v11_half', '11v11 · Medio campo'),
    ('9v9', '9v9'),
    ('8v8', '8v8'),
    ('7v7', '7v7'),
    ('5v5', '5v5'),
    ('abp', 'ABP'),
    ('specific_zone', 'Zona específica'),
]
TASK_GAME_PHASE_CHOICES = [
    ('organization_attack', 'Organización ofensiva'),
    ('organization_defense', 'Organización defensiva'),
    ('offensive_transition', 'Transición ofensiva'),
    ('defensive_transition', 'Transición defensiva'),
    ('set_pieces', 'ABP'),
]
TASK_METHODOLOGY_CHOICES = [
    ('analytical', 'Analítica'),
    ('integrated', 'Integrada'),
    ('global', 'Global'),
    ('competition', 'Competitiva'),
    ('coadjuvant', 'Coadyuvante'),
]
TASK_COMPLEXITY_CHOICES = [
    ('low', 'Baja'),
    ('medium', 'Media'),
    ('high', 'Alta'),
]
TASK_STRATEGY_CHOICES = [
    ('abp', 'Acciones a Balón Parado'),
    ('combined', 'Acciones Combinadas'),
    ('circuit', 'Circuito'),
    ('conservation', 'Conservación'),
    ('adapted', 'Juego Adaptado al Fútbol'),
    ('positional', 'Juego de Posición'),
    ('positional_specific', 'Juego de Posición Específico'),
    ('waves', 'Oleadas'),
    ('matches', 'Partidos'),
    ('possession', 'Posesión'),
    ('passing_wheel', 'Rueda de Pases'),
    ('reduced_games', 'Situaciones Reducidas'),
    ('lines_work', 'Trabajo de Líneas'),
]
TASK_COORDINATION_SKILLS_CHOICES = [
    ('offball', 'Actuación por Desmarcación'),
    ('start_stop', 'Arrancar/Frenar'),
    ('direction_change', 'Cambiar de Dirección'),
    ('carry', 'Conducción'),
    ('control', 'Control'),
    ('clearances', 'Despejes'),
    ('movements', 'Desplazamientos'),
    ('shots', 'Disparos'),
    ('tackles', 'Entrada'),
    ('balance', 'Equilibrarse'),
    ('turns', 'Giros'),
    ('interceptions', 'Intercepción'),
    ('pass', 'Pase'),
    ('protection', 'Protección'),
    ('dribble', 'Regate'),
    ('jumps', 'Saltos'),
]
TASK_TACTICAL_INTENT_CHOICES = [
    ('1v1', '1 vs 1'),
    ('2v1', '2 vs 1'),
    ('2v2', '2 vs 2'),
    ('3v3', '3 vs 3'),
    ('4v4', '4 vs 4'),
    ('abp_def', 'ABP Defensiva'),
    ('abp_att', 'ABP Ofensiva'),
    ('width', 'Amplitud'),
    ('supports', 'Apoyos'),
    ('organized_attack', 'Ataque Organizado'),
    ('cover', 'Cobertura'),
    ('keep', 'Conservar'),
    ('counter', 'Contraataque'),
    ('def_build', 'Defensa Inicio de Juego'),
    ('def_direct', 'Defensa de Juego Directo'),
    ('def_organized', 'Defensa Organizada'),
    ('runs', 'Desmarques'),
    ('split', 'Dividir'),
    ('avoid_progress', 'Evitar Progresión'),
    ('phase_def', 'Fase Defensiva'),
    ('phase_att', 'Fase Ofensiva'),
    ('fix', 'Fijar'),
    ('finish', 'Finalizar'),
    ('build', 'Inicio de Juego'),
    ('direct', 'Juego Directo'),
    ('maintain', 'Mantener'),
    ('marking', 'Marcaje'),
    ('orient', 'Orientar'),
    ('swap', 'Permuta'),
    ('press', 'Presionar'),
    ('first_attacker', 'Primer Atacante'),
    ('first_defender', 'Primer Defensor'),
    ('depth', 'Profundidad'),
    ('progress', 'Progresar'),
    ('protect_goal', 'Proteger Portería'),
    ('recover', 'Recuperar'),
    ('restart', 'Reinicio de Juego'),
    ('retreat', 'Replegar'),
    ('second_attacker', 'Segundo Atacante'),
    ('second_defender', 'Segundo Defensor'),
    ('delay', 'Temporizar'),
    ('third_attacker', 'Tercer Atacante'),
    ('third_defender', 'Tercer Defensor'),
]
TASK_DYNAMICS_CHOICES = [
    ('adm', 'ADM'),
    ('extensive', 'Extensiva'),
    ('strength', 'Fuerza'),
    ('intense_action', 'Intensiva (acción)'),
    ('intense_interaction', 'Intensiva (interacción)'),
    ('recovery', 'Recuperación'),
    ('endurance', 'Resistencia'),
    ('speed', 'Velocidad'),
]
TASK_STRUCTURE_CHOICES = [
    ('complete', 'Estructura Completa'),
    ('intersectorial', 'Intersectorial'),
    ('sectorial', 'Sectorial'),
]
TASK_COORDINATION_CHOICES = [
    ('team', 'Coordinación Equipo'),
    ('player', 'Coordinación Jugador/a'),
    ('players', 'Coordinación Jugadores/as'),
]
TASK_USEFULNESS_CHOICES = [
    ('1', '1 · Baja'),
    ('2', '2'),
    ('3', '3 · Media'),
    ('4', '4'),
    ('5', '5 · Top'),
]
TASK_CONSTRAINT_CHOICES = [
    ('two_touches', '2 toques'),
    ('one_touch_zone', '1 toque zona final'),
    ('mandatory_switch', 'Cambio de orientación obligatorio'),
    ('finish_under_10', 'Finalizar < 10 segundos'),
    ('press_6_seconds', 'Presionar 6 segundos tras pérdida'),
    ('bonus_recovery_high', 'Bonus por recuperación alta'),
    ('max_3_passes_before_finish', 'Máx. 3 pases antes de finalizar'),
]
TASK_TEMPLATE_LIBRARY = [
    {
        'key': 'none',
        'label': 'Sin plantilla (manual)',
    },
    {
        'key': 'rondo_press',
        'label': 'Rondo + presión tras pérdida',
        'values': {
            'task_title': 'Rondo + presión tras pérdida',
            'task_objective': 'Mejorar orientación corporal, apoyos y reacción tras pérdida.',
            'task_space': '24x24m',
            'task_organization': '6v3 + 1 comodín',
            'task_players_distribution': '2 equipos de 5 + 1 comodín',
            'task_load_target': 'RPE 7',
            'task_work_rest': '4x3\' + 1\' pausa',
            'task_series': '4',
            'task_repetitions': '1',
            'task_coaching_points': '- Perfil corporal abierto\n- Pase tenso y línea de pase de seguridad\n- Tras pérdida: 3" de presión máxima',
            'task_confrontation_rules': '- 1 punto por 8 pases seguidos\n- Si roba defensa y sale del cuadrado: 2 puntos',
            'task_progression': '- Limitar a 2 toques\n- Reducir espacio a 20x20m',
            'task_regression': '- Aumentar espacio\n- Añadir segundo comodín',
            'task_success_criteria': 'Al menos 60% de secuencias > 6 pases y recuperación en <6".',
        },
    },
    {
        'key': 'positional_game',
        'label': 'Juego de posición',
        'values': {
            'task_title': 'Juego de posición por calles',
            'task_objective': 'Fijar por dentro y progresar por fuera con tercer hombre.',
            'task_space': '48x36m · 3 calles',
            'task_organization': '7v7 + 3 comodines',
            'task_players_distribution': 'Líneas por altura + comodines interiores',
            'task_load_target': 'RPE 6-7',
            'task_work_rest': '3x6\' + 2\' pausa',
            'task_series': '3',
            'task_repetitions': '1',
            'task_coaching_points': '- Atracción interior para liberar lado débil\n- Ocupar 5 carriles en amplitud',
            'task_confrontation_rules': '- Gol válido tras cambio de orientación\n- Máximo 3 toques en zona interior',
            'task_progression': '- Reducir toques por línea\n- Limitar comodín a 1 toque',
            'task_regression': '- Sin límite de toques\n- Añadir un comodín extra',
            'task_success_criteria': '8+ progresiones con cambio de orientación por bloque.',
        },
    },
    {
        'key': 'transition_wave',
        'label': 'Transición ataque-defensa',
        'values': {
            'task_title': 'Olas de transición 4v3 + 3v2',
            'task_objective': 'Acelerar decisión tras robo y proteger carril central en repliegue.',
            'task_space': 'Medio campo',
            'task_organization': 'Secuencias por olas',
            'task_players_distribution': '2 equipos + finalizadores + recuperadores',
            'task_load_target': 'RPE 8',
            'task_work_rest': '6x90" + 90" pausa',
            'task_series': '2',
            'task_repetitions': '6',
            'task_coaching_points': '- Primer pase tras robo hacia delante\n- Repliegue en sprint 6"',
            'task_confrontation_rules': '- Gol en <10" vale doble\n- Si recupera defensor y sale, punto',
            'task_progression': '- Reducir tiempo de finalización a 8"\n- Iniciar desde estímulo imprevisible',
            'task_regression': '- Aumentar superioridad ofensiva',
            'task_success_criteria': '50% finalizaciones en tiempo objetivo y cero goles por carril central.',
        },
    },
    {
        'key': 'abp_corner',
        'label': 'ABP córner ofensivo',
        'values': {
            'task_title': 'ABP ofensiva · córner corto/largo',
            'task_objective': 'Automatizar dos rutinas y timing de bloqueos.',
            'task_space': 'Último tercio',
            'task_organization': 'Atacantes vs defensores + portero',
            'task_players_distribution': '8 atacantes + 6 defensores',
            'task_load_target': 'RPE 5',
            'task_work_rest': '10 repeticiones por rutina',
            'task_series': '2',
            'task_repetitions': '10',
            'task_coaching_points': '- Timing de bloqueo y desmarque\n- Perfil del centrador según rutina',
            'task_confrontation_rules': '- Solo vale gol de zona objetivo\n- Si despeja defensa fuera del área: punto defensa',
            'task_progression': '- Añadir segunda jugada\n- Variar perfil de centro',
            'task_regression': '- Sin oposición en primeras repeticiones',
            'task_success_criteria': '70% de ejecuciones con remate en zona prevista.',
        },
    },
]


def build_task_material_library():
    materials = [dict(item) for item in TASK_MATERIAL_LIBRARY]
    if not TASK_MATERIAL_PPT_DIR.exists():
        return materials

    allowed_suffixes = {'.png', '.gif', '.jpg', '.jpeg', '.webp'}
    icon_files = [
        path
        for path in sorted(TASK_MATERIAL_PPT_DIR.iterdir())
        if path.is_file() and path.suffix.lower() in allowed_suffixes
    ]
    for index, icon_file in enumerate(icon_files, start=1):
        materials.append(
            {
                'label': f'P{index:02d}',
                'title': f'Recurso visual PPT {index:02d}',
                'kind': f'ppt-{index:02d}',
                'category': 'ppt',
                'icon': '',
                'asset': f"football/images/task-materials/ppt/{icon_file.name}",
            }
        )
    return materials


def _canonical_action_value(value):
    return ' '.join(str(value or '').split()).strip().lower()


def _build_pdf_response_or_html_fallback(request, html: str, filename: str):
    if not weasyprint:
        return HttpResponse(html, content_type='text/html; charset=utf-8')
    try:
        def _safe_url_fetcher(url, timeout=4, ssl_context=None):
            try:
                default_fetcher = getattr(weasyprint, 'default_url_fetcher', None)
                if callable(default_fetcher):
                    return default_fetcher(url, timeout=timeout, ssl_context=ssl_context)
            except Exception:
                pass
            return {'string': b'', 'mime_type': 'text/plain'}

        pdf_file = weasyprint.HTML(
            string=html,
            base_url=request.build_absolute_uri('/'),
            url_fetcher=_safe_url_fetcher,
        ).write_pdf()
        response = HttpResponse(pdf_file, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}.pdf"'
        return response
    except Exception:
        return HttpResponse(html, content_type='text/html; charset=utf-8')


def resolve_player_photo_static_path(player):
    if not player:
        return ''
    players_dir = Path(settings.BASE_DIR) / 'static' / 'football' / 'images' / 'players'
    if not players_dir.exists():
        return ''

    name_slug = slugify(player.name or '')
    number_value = player.number if player.number is not None else ''
    candidates = []
    candidates.extend(
        [
            f'player-{player.id}.png',
            f'player-{player.id}.jpg',
            f'player-{player.id}.jpeg',
            f'player-{player.id}.webp',
        ]
    )
    if name_slug and number_value != '':
        candidates.extend(
            [
                f'{name_slug}-n{number_value}-final.png',
                f'{name_slug}-n{number_value}.png',
                f'{name_slug}-{number_value}.png',
            ]
        )
    if name_slug:
        candidates.extend(
            [
                f'{name_slug}-final.png',
                f'{name_slug}.png',
                f'{name_slug}.jpg',
                f'{name_slug}.jpeg',
            ]
        )
    if number_value != '':
        candidates.extend(
            [
                f'n{number_value}-{name_slug}.png',
                f'{name_slug}-n{number_value}-cut.png',
                f'{name_slug}-n{number_value}-crop.png',
            ]
        )

    seen = set()
    for filename in candidates:
        if filename in seen:
            continue
        seen.add(filename)
        file_path = players_dir / filename
        if file_path.exists():
            return f'football/images/players/{filename}'
    if number_value != '':
        wildcard_patterns = [
            f'*-n{number_value}-final.*',
            f'*-n{number_value}-cut.*',
            f'*-n{number_value}-crop.*',
            f'*-n{number_value}.*',
        ]
        for pattern in wildcard_patterns:
            for file_path in players_dir.glob(pattern):
                if file_path.is_file():
                    return f'football/images/players/{file_path.name}'
    return ''


def _player_photo_storage_candidates(player):
    if not player:
        return []
    return [
        f'player-photos/player-{player.id}.png',
        f'player-photos/player-{player.id}.jpg',
        f'player-photos/player-{player.id}.jpeg',
        f'player-photos/player-{player.id}.webp',
    ]


def _build_public_media_url(request, raw_url):
    url = str(raw_url or '').strip()
    if not url:
        return ''
    if request is not None and url.startswith('/'):
        try:
            return request.build_absolute_uri(url)
        except Exception:
            return url
    return url


def product_landing_page(request):
    return render(
        request,
        'football/product_landing.html',
        {
            'brand_descriptor': 'Football Intelligence',
            'product_tracks': [
                {
                    'name': '2J Live',
                    'eyebrow': 'Partido y dato en vivo',
                    'description': 'Registro táctil en iPad, cronología de acciones, KPIs y resumen postpartido para el cuerpo técnico.',
                },
                {
                    'name': '2J Studio',
                    'eyebrow': 'Tareas y sesiones',
                    'description': 'Diseño de tareas, secuencias animadas, PDFs UEFA/Club y biblioteca reutilizable de entrenamiento.',
                },
                {
                    'name': '2J Club',
                    'eyebrow': 'Operación de cliente',
                    'description': 'Convocatoria, 11 inicial, jugadores, cuerpo técnico, planificación y seguimiento dentro del club.',
                },
            ],
            'proof_points': [
                'Registro de acciones pensado para iPad y trabajo en banquillo.',
                'Creador de tareas conectado con sesiones, microciclos y documentos.',
                'Plataforma matriz para gobernar clientes, módulos y accesos.',
            ],
        },
    )


def save_player_photo(player, uploaded_photo):
    if not player or not uploaded_photo:
        return ''
    storage_candidates = _player_photo_storage_candidates(player)
    target_name = storage_candidates[0]
    content = uploaded_photo
    try:
        if hasattr(uploaded_photo, 'seek'):
            uploaded_photo.seek(0)
        if Image is not None:
            with Image.open(uploaded_photo) as image:
                if image.mode in ('RGBA', 'LA', 'P'):
                    converted = image.convert('RGBA')
                    background = Image.new('RGBA', converted.size, (3, 7, 18, 255))
                    background.alpha_composite(converted)
                    final_image = background.convert('RGB')
                else:
                    final_image = image.convert('RGB')
                buffer = io.BytesIO()
                final_image.save(buffer, format='PNG', optimize=True)
                content = ContentFile(buffer.getvalue())
        else:
            extension = Path(str(getattr(uploaded_photo, 'name', '') or '')).suffix.lower()
            if extension in {'.jpg', '.jpeg', '.webp'}:
                target_name = f'player-photos/player-{player.id}{extension}'
            if hasattr(uploaded_photo, 'seek'):
                uploaded_photo.seek(0)
        for candidate in storage_candidates:
            try:
                if default_storage.exists(candidate):
                    default_storage.delete(candidate)
            except Exception:
                logger.exception('No se pudo limpiar una foto previa del jugador %s', player.id)
        saved_name = default_storage.save(target_name, content)
        try:
            player.photo_updated_at = timezone.now()
            player.save(update_fields=['photo_updated_at'])
        except Exception:
            pass
        try:
            cache.set(
                f'{PLAYER_PHOTO_VERSION_CACHE_KEY_PREFIX}:{player.id}',
                int(timezone.now().timestamp()),
                timeout=PLAYER_PHOTO_VERSION_CACHE_SECONDS,
            )
        except Exception:
            pass
        return saved_name
    except Exception:
        logger.exception('No se pudo guardar la foto del jugador %s', player.id)
        return ''


@login_required
def player_photo_file(request, player_id):
    primary_team = _get_player_team_for_request(request)
    if not primary_team:
        raise Http404('Equipo principal no configurado')
    player = Player.objects.filter(id=player_id, team=primary_team).first()
    if not player:
        raise Http404('Jugador no encontrado')
    forbidden = _forbid_if_no_player_access(request.user, player, primary_team=primary_team)
    if forbidden:
        return forbidden
    storage_name = ''
    for candidate in _player_photo_storage_candidates(player):
        try:
            if default_storage.exists(candidate):
                storage_name = candidate
                break
        except Exception:
            continue
    if not storage_name:
        raise Http404('Foto no disponible')
    try:
        file_field = default_storage.open(storage_name, 'rb')
    except Exception:
        return HttpResponse('No se pudo abrir la foto del jugador.', status=500)
    extension = Path(storage_name).suffix.lower()
    content_type = {
        '.png': 'image/png',
        '.jpg': 'image/jpeg',
        '.jpeg': 'image/jpeg',
        '.webp': 'image/webp',
        '.gif': 'image/gif',
    }.get(extension, 'application/octet-stream')
    response = FileResponse(file_field, content_type=content_type)
    response['Content-Disposition'] = f'inline; filename="{Path(storage_name).name}"'
    # El fichero se sobrescribe con el mismo nombre (player-{id}.png), así que evitamos caché larga.
    response['Cache-Control'] = 'private, max-age=60'
    return response


@login_required
def home_carousel_image_file(request, image_id):
    item = HomeCarouselImage.objects.filter(id=image_id).first()
    if not item or not item.image:
        raise Http404('Imagen no disponible')
    file_field = item.image
    try:
        file_field.open('rb')
    except Exception:
        return HttpResponse('No se pudo abrir la imagen.', status=500)
    extension = Path(file_field.name).suffix.lower()
    content_type = {
        '.png': 'image/png',
        '.jpg': 'image/jpeg',
        '.jpeg': 'image/jpeg',
        '.webp': 'image/webp',
        '.gif': 'image/gif',
    }.get(extension, 'application/octet-stream')
    response = FileResponse(file_field, content_type=content_type)
    response['Content-Disposition'] = f'inline; filename="{Path(file_field.name).name}"'
    response['Cache-Control'] = 'private, max-age=900'
    return response


def resolve_player_photo_url(request, player):
    resolved_storage_name = ''
    for storage_name in _player_photo_storage_candidates(player):
        try:
            if default_storage.exists(storage_name):
                resolved_storage_name = storage_name
                break
        except Exception:
            logger.exception('No se pudo resolver la foto subida del jugador %s', getattr(player, 'id', ''))
    if resolved_storage_name:
        url = reverse('player-photo-file', args=[player.id])
        version = ''
        # Fuente preferida: campo persistente (evita depender de caché en despliegues multi-instancia).
        try:
            updated_at = getattr(player, 'photo_updated_at', None)
            if updated_at:
                version = str(int(updated_at.timestamp()))
        except Exception:
            version = ''
        if not version:
            # Fallback: caché + metadatos del storage.
            cache_key = f'{PLAYER_PHOTO_VERSION_CACHE_KEY_PREFIX}:{player.id}'
            try:
                cached_version = cache.get(cache_key)
                if cached_version:
                    version = str(int(cached_version))
            except Exception:
                version = ''
            if not version:
                try:
                    modified = default_storage.get_modified_time(resolved_storage_name)
                    if modified:
                        version = str(int(modified.timestamp()))
                except Exception:
                    version = ''
            if not version:
                version = str(int(timezone.now().timestamp()))
            try:
                cache.set(cache_key, int(version), timeout=PLAYER_PHOTO_VERSION_CACHE_SECONDS)
            except Exception:
                pass
        if version:
            joiner = '&' if '?' in url else '?'
            url = f'{url}{joiner}v={version}'
        if request is None:
            return url
        return _build_public_media_url(request, url)
    static_path = resolve_player_photo_static_path(player)
    if not static_path:
        return ''
    if request is None:
        return static(static_path)
    try:
        return request.build_absolute_uri(static(static_path))
    except Exception:
        return static(static_path)


def _file_as_data_uri(file_path):
    try:
        path = Path(file_path)
        if not path.exists() or not path.is_file():
            return ''
        mime_type, _ = mimetypes.guess_type(str(path))
        if not mime_type:
            mime_type = 'image/png'
        encoded = base64.b64encode(path.read_bytes()).decode('ascii')
        return f'data:{mime_type};base64,{encoded}'
    except Exception:
        return ''


def _image_file_as_small_data_uri(file_path, max_width=1200, max_height=800, quality=72):
    """Encode image files as a lightweight JPEG data URI for faster PDF rendering."""
    if Image is None:
        return _file_as_data_uri(file_path)
    try:
        path = Path(file_path)
        if not path.exists() or not path.is_file():
            return ''
        with Image.open(path) as img:
            normalized = img.convert('RGB')
            normalized.thumbnail((int(max_width), int(max_height)))
            buffer = io.BytesIO()
            normalized.save(buffer, format='JPEG', optimize=True, quality=int(quality))
        encoded = base64.b64encode(buffer.getvalue()).decode('ascii')
        return f'data:image/jpeg;base64,{encoded}'
    except Exception:
        return _file_as_data_uri(file_path)


def resolve_team_photo_for_pdf(request):
    default_path = Path(settings.BASE_DIR) / 'static' / 'football' / 'images' / 'team-01.jpg'
    fallback_url = request.build_absolute_uri(static('football/images/team-01.jpg'))
    fallback_data_uri = _image_file_as_small_data_uri(default_path)
    source_path = default_path
    source_url = fallback_url

    carousel_cover = (
        HomeCarouselImage.objects
        .filter(is_active=True)
        .order_by('order', '-created_at', '-id')
        .first()
    )
    if not carousel_cover:
        carousel_cover = (
            HomeCarouselImage.objects
            .order_by('order', '-created_at', '-id')
            .first()
        )
    if carousel_cover and carousel_cover.image:
        try:
            source_url = request.build_absolute_uri(carousel_cover.image.url)
            if getattr(carousel_cover.image, 'path', None):
                source_path = Path(carousel_cover.image.path)
        except Exception:
            source_path = default_path
            source_url = fallback_url

    data_uri = _image_file_as_small_data_uri(source_path) or fallback_data_uri
    return {
        'url': source_url or fallback_url,
        'data_uri': data_uri or '',
    }


def get_competition_total_rounds(primary_team):
    if not primary_team:
        return 0
    group = primary_team.group
    if not group:
        return 0
    matches = Match.objects.filter(group=group)
    round_numbers = []
    for round_value in matches.values_list('round', flat=True):
        round_num = extract_round_number(round_value)
        if round_num is not None:
            round_numbers.append(round_num)
    total_by_rounds = max(round_numbers) if round_numbers else 0
    teams_count = Team.objects.filter(group=group).exclude(id__isnull=True).count()
    total_double_round_robin = max((teams_count - 1) * 2, 0) if teams_count >= 2 else 0
    return max(total_by_rounds, total_double_round_robin)


def _serialize_match_event(event, duplicate=False):
    player = event.player
    return {
        'id': event.id,
        'minute': event.minute,
        'period': event.period,
        'action': event.event_type,
        'zone': event.zone,
        'result': event.result,
        'duplicate': bool(duplicate),
        'player': {
            'id': player.id if player else None,
            'name': player.name if player else 'Equipo',
            'number': (player.number if player and player.number is not None else '--'),
        },
    }


TEAM_ONLY_ACTION_TYPES = {
    'saque de esquina a favor',
    'saque de esquina en contra',
}

TECHNICAL_ROLES = {
    AppUserRole.ROLE_COACH,
    AppUserRole.ROLE_FITNESS,
    AppUserRole.ROLE_GOALKEEPER,
    AppUserRole.ROLE_ANALYST,
    AppUserRole.ROLE_ADMIN,
}

def _has_club_workspace_access(user):
    if not user or not user.is_authenticated:
        return False
    if _get_user_role(user) == AppUserRole.ROLE_PLAYER:
        return False
    return (
        WorkspaceMembership.objects.filter(user=user, workspace__kind=Workspace.KIND_CLUB, workspace__is_active=True).exists()
        or Workspace.objects.filter(owner_user=user, kind=Workspace.KIND_CLUB, is_active=True).exists()
    )


def _get_user_role(user):
    if not user or not user.is_authenticated:
        return None
    role_obj = getattr(user, 'app_role', None)
    role = str(getattr(role_obj, 'role', '') or '').strip() or None
    legacy_map = {
        'admin': AppUserRole.ROLE_ADMIN,
        'player': AppUserRole.ROLE_PLAYER,
    }
    normalized_role = legacy_map.get(role, role)
    if normalized_role:
        return normalized_role
    if getattr(user, 'is_superuser', False) or getattr(user, 'is_staff', False):
        return AppUserRole.ROLE_ADMIN
    return None


def _is_admin_user(user):
    role = _get_user_role(user)
    return bool(user and user.is_authenticated and (user.is_superuser or user.is_staff or role == AppUserRole.ROLE_ADMIN))


def _can_edit_match_actions(user):
    if not user or not user.is_authenticated:
        return False
    if _is_admin_user(user):
        return True
    return _get_user_role(user) in TECHNICAL_ROLES or _has_club_workspace_access(user)


def _can_access_sessions_workspace(user):
    role = _get_user_role(user)
    if not user or not user.is_authenticated:
        return False
    if _is_admin_user(user):
        return True
    return role in TECHNICAL_ROLES or _has_club_workspace_access(user)


def _can_access_coach_workspace(user):
    if not user or not user.is_authenticated:
        return False
    if _is_admin_user(user):
        return True
    return _get_user_role(user) in TECHNICAL_ROLES or _has_club_workspace_access(user)


def _can_access_task_studio(user):
    if not user or not user.is_authenticated:
        return False
    if _is_admin_user(user):
        return True
    role = _get_user_role(user)
    if role not in {AppUserRole.ROLE_TASK_STUDIO, AppUserRole.ROLE_GUEST}:
        return False
    if role == AppUserRole.ROLE_GUEST:
        has_workspace = Workspace.objects.filter(kind=Workspace.KIND_TASK_STUDIO, owner_user=user, is_active=True).exists()
        has_profile = TaskStudioProfile.objects.filter(user=user).exists()
        if not has_workspace and not has_profile:
            # Invitados: si se les ha asignado el rol, deben poder entrar. Si aún no tienen workspace/perfil
            # (por ejemplo tras crearles acceso desde Platform), lo inicializamos sin crear tareas automáticamente.
            try:
                _ensure_task_studio_workspace(user)
            except Exception:
                return False
    profile = TaskStudioProfile.objects.filter(user=user).first()
    if profile and not profile.is_enabled:
        return False
    return True


def _task_studio_target_user(request):
    if not request.user.is_authenticated:
        return None
    selected_user = request.user
    selected_user_id = _parse_int(request.GET.get('user'))
    if _is_admin_user(request.user) and selected_user_id:
        candidate = User.objects.filter(id=selected_user_id).first()
        if candidate:
            selected_user = candidate
    return selected_user


def _task_studio_query_suffix(target_user, current_user):
    if not target_user or not current_user or int(getattr(target_user, 'id', 0) or 0) == int(getattr(current_user, 'id', 0) or 0):
        return ''
    if _is_admin_user(current_user):
        return f'?user={target_user.id}'
    return ''


def _task_studio_profile_for_user(user):
    if not user:
        return None
    role = _get_user_role(user)
    if role == AppUserRole.ROLE_GUEST:
        profile = TaskStudioProfile.objects.filter(user=user).first()
        workspace = Workspace.objects.filter(kind=Workspace.KIND_TASK_STUDIO, owner_user=user).first()
        if not workspace and profile and profile.workspace_id and getattr(profile.workspace, 'kind', None) == Workspace.KIND_TASK_STUDIO:
            workspace = profile.workspace
        if not profile and not workspace:
            workspace = _ensure_task_studio_workspace(user)
            if not workspace:
                return None
            profile = TaskStudioProfile.objects.create(user=user, workspace=workspace)
            return profile
        if not profile:
            profile = TaskStudioProfile.objects.create(user=user, workspace=workspace)
        elif workspace and profile.workspace_id != workspace.id:
            profile.workspace = workspace
            profile.save(update_fields=['workspace'])
        return profile
    workspace = _ensure_task_studio_workspace(user)
    profile, _ = TaskStudioProfile.objects.get_or_create(user=user, defaults={'workspace': workspace})
    if workspace and profile.workspace_id != workspace.id:
        profile.workspace = workspace
        profile.save(update_fields=['workspace'])
    return profile


def _forbid_if_no_task_studio_access(user):
    if _can_access_task_studio(user):
        return None
    return HttpResponse('No tienes permisos para acceder a Task Studio.', status=403)


def _can_access_platform(user):
    return _is_admin_user(user)


def _available_workspaces_for_user(user):
    if not user or not user.is_authenticated:
        return Workspace.objects.none()
    qs = Workspace.objects.select_related('primary_team', 'owner_user').filter(is_active=True)
    if _can_access_platform(user):
        return qs
    return qs.filter(Q(memberships__user=user) | Q(owner_user=user)).distinct()


def _get_active_workspace(request):
    if not request or not getattr(request, 'user', None) or not request.user.is_authenticated:
        return None
    available_qs = _available_workspaces_for_user(request.user)
    workspace_id = _parse_int(request.GET.get('workspace'))
    if not workspace_id:
        workspace_id = _parse_int(request.session.get('active_workspace_id'))
    if workspace_id:
        workspace = available_qs.filter(id=workspace_id).first()
        if workspace:
            request.session['active_workspace_id'] = workspace.id
            return workspace
        request.session.pop('active_workspace_id', None)
    if _can_access_platform(request.user):
        return None

    # Modo monocliente: si el sistema sólo tiene un club activo, damos acceso de lectura a roles
    # internos (incluyendo Task Studio) aunque su primer workspace sea privado. Esto permite que
    # usuarios "task" puedan navegar módulos de partido (convocatoria/11/acciones) cuando se ha
    # decidido habilitarlo para demos.
    try:
        role = _get_user_role(request.user)
        if role and role != AppUserRole.ROLE_PLAYER:
            club_ws = list(Workspace.objects.filter(kind=Workspace.KIND_CLUB, is_active=True).order_by('id')[:2])
            if len(club_ws) == 1:
                club = club_ws[0]
                if not WorkspaceMembership.objects.filter(workspace=club, user=request.user).exists() and int(getattr(club, 'owner_user_id', 0) or 0) != int(request.user.id):
                    WorkspaceMembership.objects.get_or_create(
                        workspace=club,
                        user=request.user,
                        defaults={'role': WorkspaceMembership.ROLE_VIEWER},
                    )
    except Exception:
        pass
    fallback_workspace = available_qs.order_by('kind', 'name', 'id').first()
    if fallback_workspace:
        request.session['active_workspace_id'] = fallback_workspace.id
        return fallback_workspace
    # Reparación segura: si el usuario no tiene ningún workspace asignado pero el sistema sólo tiene
    # un workspace de club activo, lo asignamos automáticamente. Esto evita 403 en entornos de un solo club.
    role = _get_user_role(request.user)
    if role in {AppUserRole.ROLE_PLAYER, AppUserRole.ROLE_COACH, AppUserRole.ROLE_FITNESS, AppUserRole.ROLE_GOALKEEPER, AppUserRole.ROLE_ANALYST}:
        club_ws = list(Workspace.objects.filter(kind=Workspace.KIND_CLUB, is_active=True).order_by('id')[:2])
        if len(club_ws) == 1:
            workspace = club_ws[0]
            try:
                WorkspaceMembership.objects.get_or_create(
                    workspace=workspace,
                    user=request.user,
                    defaults={'role': WorkspaceMembership.ROLE_MEMBER},
                )
            except Exception:
                pass
            request.session['active_workspace_id'] = workspace.id
            return workspace
    return None


def _get_primary_team_for_request(request):
    workspace = _get_active_workspace(request)
    if workspace and workspace.kind == Workspace.KIND_CLUB and workspace.primary_team_id:
        return workspace.primary_team
    if request and getattr(request, 'user', None) and request.user.is_authenticated and not _can_access_platform(request.user):
        if _get_user_role(request.user) == AppUserRole.ROLE_PLAYER:
            return Team.objects.filter(is_primary=True).first()
        return None
    return Team.objects.filter(is_primary=True).first()


def _get_player_team_for_request(request):
    workspace = _get_active_workspace(request)
    if workspace and workspace.kind == Workspace.KIND_CLUB and workspace.primary_team_id:
        return workspace.primary_team
    return Team.objects.filter(is_primary=True).first()


def _build_active_workspace_badge(request):
    workspace = _get_active_workspace(request)
    if not workspace:
        return None
    subtitle = ''
    if workspace.kind == Workspace.KIND_CLUB and workspace.primary_team_id:
        subtitle = workspace.primary_team.display_name or workspace.primary_team.name
    elif workspace.kind == Workspace.KIND_TASK_STUDIO and workspace.owner_user_id:
        subtitle = workspace.owner_user.get_username()
    return {
        'id': workspace.id,
        'name': workspace.name,
        'kind': workspace.kind,
        'kind_label': workspace.get_kind_display(),
        'subtitle': subtitle,
    }


def _unique_workspace_slug(base_text, *, exclude_id=None):
    base_slug = slugify(base_text or 'workspace') or 'workspace'
    candidate = base_slug
    suffix = 2
    qs = Workspace.objects.all()
    if exclude_id:
        qs = qs.exclude(id=exclude_id)
    while qs.filter(slug=candidate).exists():
        candidate = f'{base_slug}-{suffix}'
        suffix += 1
    return candidate


def _workspace_default_modules(kind):
    if kind == Workspace.KIND_TASK_STUDIO:
        return {
            'task_studio_home': True,
            'task_studio_profile': True,
            'task_studio_roster': True,
            'task_studio_tasks': True,
            'task_studio_pdfs': True,
        }
    return {
        'dashboard': True,
        'coach_overview': True,
        'players': True,
        'convocation': True,
        'match_actions': True,
        'sessions': True,
        'analysis': True,
        'abp_board': True,
        'manual_stats': True,
    }


def _workspace_access_module_catalog(kind):
    if kind == Workspace.KIND_TASK_STUDIO:
        return [
            {'key': 'task_studio_home', 'label': 'Inicio'},
            {'key': 'task_studio_profile', 'label': 'Perfil'},
            {'key': 'task_studio_roster', 'label': 'Plantilla'},
            {'key': 'task_studio_tasks', 'label': 'Tareas'},
            {'key': 'task_studio_pdfs', 'label': 'PDFs'},
        ]
    return [
        {'key': 'dashboard', 'label': 'Portada'},
        {'key': 'coach_overview', 'label': 'Cuerpo técnico'},
        {'key': 'players', 'label': 'Plantilla'},
        {'key': 'convocation', 'label': 'Convocatoria'},
        {'key': 'match_actions', 'label': 'Acciones'},
        {'key': 'sessions', 'label': 'Sesiones'},
        {'key': 'analysis', 'label': 'Análisis'},
        {'key': 'abp_board', 'label': 'ABP'},
        {'key': 'manual_stats', 'label': 'Est. manuales'},
    ]


def _workspace_competition_provider_choices():
    return WorkspaceCompetitionContext.PROVIDER_CHOICES


def _bootstrap_workspace_competition_context(
    workspace,
    *,
    primary_team=None,
    provider=None,
    external_competition_key=None,
    external_group_key=None,
    external_team_key=None,
    external_team_name=None,
    auto_sync_enabled=None,
):
    if not workspace or workspace.kind != Workspace.KIND_CLUB:
        return None
    primary_team = primary_team or workspace.primary_team
    group = getattr(primary_team, 'group', None)
    season = getattr(group, 'season', None)
    context, _ = WorkspaceCompetitionContext.objects.get_or_create(
        workspace=workspace,
        defaults={
            'team': primary_team,
            'group': group,
            'season': season,
            'provider': provider or WorkspaceCompetitionContext.PROVIDER_MANUAL,
            'external_competition_key': external_competition_key or '',
            'external_group_key': external_group_key or '',
            'external_team_key': external_team_key or '',
            'external_team_name': external_team_name or str(getattr(primary_team, 'name', '') or '').strip(),
            'is_auto_sync_enabled': True if auto_sync_enabled is None else bool(auto_sync_enabled),
        },
    )
    changed = False
    if context.team_id != getattr(primary_team, 'id', None):
        context.team = primary_team
        changed = True
    if context.group_id != getattr(group, 'id', None):
        context.group = group
        changed = True
    if context.season_id != getattr(season, 'id', None):
        context.season = season
        changed = True
    if provider and context.provider != provider:
        context.provider = provider
        changed = True
    if external_competition_key is not None and context.external_competition_key != (external_competition_key or ''):
        context.external_competition_key = external_competition_key or ''
        changed = True
    if external_group_key is not None and context.external_group_key != (external_group_key or ''):
        context.external_group_key = external_group_key or ''
        changed = True
    if external_team_key is not None and context.external_team_key != (external_team_key or ''):
        context.external_team_key = external_team_key or ''
        changed = True
    desired_external_team_name = context.external_team_name
    if external_team_name is not None:
        desired_external_team_name = external_team_name or str(getattr(primary_team, 'name', '') or '').strip()
    elif not desired_external_team_name:
        desired_external_team_name = str(getattr(primary_team, 'name', '') or '').strip()
    if context.external_team_name != desired_external_team_name:
        context.external_team_name = desired_external_team_name
        changed = True
    if auto_sync_enabled is not None and context.is_auto_sync_enabled != bool(auto_sync_enabled):
        context.is_auto_sync_enabled = bool(auto_sync_enabled)
        changed = True
    if changed:
        context.save()
    return context


def _build_workspace_schedule_payload(primary_team):
    if not primary_team:
        return []
    matches = (
        Match.objects
        .filter(Q(home_team=primary_team) | Q(away_team=primary_team))
        .select_related('home_team', 'away_team')
        .order_by('date', 'id')[:8]
    )
    payload = []
    for match in matches:
        match_payload = build_match_payload(
            match,
            primary_team,
            status='next' if match.date and match.date >= timezone.localdate() else 'latest',
        )
        payload.append(match_payload)
    return payload


def _expand_team_lookup_variants(raw_value):
    base_key = _normalize_team_lookup_key(raw_value)
    if not base_key:
        return set()
    variants = {base_key}
    trimmed = re.sub(r'^(cd|cf|ud|fc)+', '', base_key)
    trimmed = re.sub(r'(cd|cf|ud|fc)+$', '', trimmed)
    if trimmed:
        variants.add(trimmed)
        variants.add(f'cd{trimmed}')
        variants.add(f'{trimmed}cd')
    return {variant for variant in variants if variant}


def _context_team_lookup_keys(context, primary_team):

    keys = set()
    for raw_value in (
        getattr(primary_team, 'name', ''),
        getattr(primary_team, 'display_name', ''),
        getattr(context, 'external_team_name', ''),
    ):
        keys.update(_expand_team_lookup_variants(raw_value))
    external_team_key = str(getattr(context, 'external_team_key', '') or '').strip()
    if external_team_key:
        keys.update(_expand_team_lookup_variants(external_team_key.lower()))
    return {key for key in keys if key}


def _ensure_universo_context_binding(context, primary_team):
    if not context or str(getattr(context, 'provider', '') or '').strip() != WorkspaceCompetitionContext.PROVIDER_UNIVERSO:
        return context
    if str(getattr(context, 'external_group_key', '') or '').strip() and (
        str(getattr(context, 'external_team_key', '') or '').strip()
        or str(getattr(context, 'external_team_name', '') or '').strip()
    ):
        return context
    if not primary_team:
        return context

    competition = getattr(getattr(getattr(primary_team, 'group', None), 'season', None), 'competition', None)
    team_query = str(getattr(context, 'external_team_name', '') or getattr(primary_team, 'name', '') or '').strip()
    competition_query = str(getattr(competition, 'name', '') or '').strip()
    group_query = str(getattr(getattr(primary_team, 'group', None), 'name', '') or '').strip()

    normalized_team_query = normalize_label(team_query)
    lookup_team_query = _normalize_team_lookup_key(team_query)
    normalized_comp_query = normalize_label(competition_query)
    normalized_group_query = normalize_label(group_query)

    def _tokenize_label(value):
        return {
            token
            for token in re.split(r'[^a-z0-9]+', normalize_label(value))
            if token and token not in {'grupo', 'gr', 'senior', 's', 'cd', 'cf'}
        }

    def _query_matches_candidate(query, candidate, *, min_overlap=1):
        query_tokens = _tokenize_label(query)
        candidate_tokens = _tokenize_label(candidate)
        if not query_tokens:
            return True
        overlap = query_tokens & candidate_tokens
        return len(overlap) >= min(min_overlap, len(query_tokens))

    candidate_groups = (
        Group.objects
        .select_related('season__competition')
        .exclude(external_id__exact='')
        .filter(name__iexact=group_query)
        .order_by('-season__is_current', '-id')
    )
    best_group_match = None
    best_group_score = -1
    for candidate_group in candidate_groups:
        candidate_competition_name = str(getattr(getattr(candidate_group, 'season', None), 'competition', None).name or '').strip() if getattr(getattr(candidate_group, 'season', None), 'competition', None) else ''
        score = 0
        if group_query:
            score += 25
        if competition_query:
            score += 25 if _query_matches_candidate(competition_query, candidate_competition_name, min_overlap=2) else 0
        if score > best_group_score:
            best_group_match = candidate_group
            best_group_score = score
    if best_group_match and best_group_score >= 25:
        update_fields = []
        if best_group_match.external_id and getattr(context, 'external_group_key', '') != best_group_match.external_id:
            context.external_group_key = best_group_match.external_id
            update_fields.append('external_group_key')
        if update_fields:
            context.save(update_fields=update_fields + ['updated_at'])
        if str(getattr(context, 'external_group_key', '') or '').strip():
            return context
    catalog = _build_universo_competition_catalog()
    competitions = catalog.get('competitions') or {}
    groups = catalog.get('groups') or {}
    classifications = catalog.get('classifications') or {}
    candidates = []
    for (competition_code, group_code), classification in classifications.items():
        competition_meta = competitions.get(competition_code) or {}
        group_meta = groups.get((competition_code, group_code)) or {}
        resolved_competition_name = str(classification.get('competition_name') or competition_meta.get('name') or '').strip()
        resolved_group_name = str(classification.get('group_name') or group_meta.get('group_name') or '').strip()
        if normalized_comp_query and normalized_comp_query not in normalize_label(resolved_competition_name):
            continue
        if normalized_group_query and normalized_group_query not in normalize_label(resolved_group_name):
            continue
        for row in classification.get('rows') or []:
            if not isinstance(row, dict):
                continue
            resolved_team_name = str(row.get('nombre') or '').strip()
            if not resolved_team_name:
                continue
            normalized_team_name = normalize_label(resolved_team_name)
            lookup_team_name = _normalize_team_lookup_key(resolved_team_name)
            if lookup_team_query and not (
                lookup_team_query == lookup_team_name
                or lookup_team_query in lookup_team_name
                or lookup_team_name in lookup_team_query
            ):
                continue
            score = 0
            if lookup_team_query:
                score += 60 if lookup_team_query == lookup_team_name else 30
            if normalized_group_query:
                score += 25 if normalized_group_query == normalize_label(resolved_group_name) else 12
            if normalized_comp_query:
                score += 25 if normalized_comp_query == normalize_label(resolved_competition_name) else 12
            candidates.append(
                {
                    'team_name': resolved_team_name,
                    'competition_name': resolved_competition_name,
                    'group_name': resolved_group_name,
                    'season_name': str(competition_meta.get('season_name') or '').strip(),
                    'external_competition_key': competition_code,
                    'external_group_key': group_code,
                    'external_team_key': str(row.get('codequipo') or '').strip(),
                    'external_team_name': resolved_team_name,
                    'score': score,
                }
            )
    candidates.sort(key=lambda item: (-int(item.get('score') or 0), item['competition_name'], item['group_name'], item['team_name']))
    if not candidates:
        seasons = _fetch_universo_live_seasons()
        season_row = next((row for row in seasons if str(row.get('nombre') or '').strip() == '2025-2026'), None)
        if not season_row and seasons:
            season_row = seasons[0]
        season_id = str((season_row or {}).get('cod_temporada') or '').strip()
        season_name = str((season_row or {}).get('nombre') or '').strip()
        delegations = _fetch_universo_live_delegations()
        if season_id and delegations:
            for delegation in delegations:
                competitions = _fetch_universo_live_competitions(delegation.get('cod_delegacion'), season_id)
                for competition_row in competitions:
                    competition_code = str(competition_row.get('codigo') or '').strip()
                    resolved_competition_name = str(competition_row.get('nombre') or '').strip()
                    if not competition_code or not resolved_competition_name:
                        continue
                    if competition_query and not _query_matches_candidate(competition_query, resolved_competition_name, min_overlap=2):
                        continue
                    groups = _fetch_universo_live_groups(competition_code)
                    for group_row in groups:
                        group_code = str(group_row.get('codigo') or '').strip()
                        resolved_group_name = str(group_row.get('nombre') or '').strip()
                        if not group_code or not resolved_group_name:
                            continue
                        if group_query and not _query_matches_candidate(group_query, resolved_group_name):
                            continue
                        classification = _fetch_universo_live_classification(group_code)
                        for row in classification.get('clasificacion') or []:
                            if not isinstance(row, dict):
                                continue
                            resolved_team_name = str(row.get('nombre') or '').strip()
                            if not resolved_team_name:
                                continue
                            normalized_team_name = normalize_label(resolved_team_name)
                            lookup_team_name = _normalize_team_lookup_key(resolved_team_name)
                            if lookup_team_query and not (
                                lookup_team_query == lookup_team_name
                                or lookup_team_query in lookup_team_name
                                or lookup_team_name in lookup_team_query
                            ):
                                continue
                            score = 0
                            if lookup_team_query:
                                score += 60 if lookup_team_query == lookup_team_name else 30
                            if normalized_group_query:
                                score += 25 if _query_matches_candidate(group_query, resolved_group_name) else 12
                            if normalized_comp_query:
                                score += 25 if _query_matches_candidate(competition_query, resolved_competition_name, min_overlap=2) else 12
                            candidates.append(
                                {
                                    'team_name': resolved_team_name,
                                    'competition_name': resolved_competition_name,
                                    'group_name': resolved_group_name,
                                    'season_name': season_name,
                                    'external_competition_key': competition_code,
                                    'external_group_key': group_code,
                                    'external_team_key': str(row.get('codequipo') or '').strip(),
                                    'external_team_name': resolved_team_name,
                                    'score': score + 20,
                                }
                            )
        candidates.sort(key=lambda item: (-int(item.get('score') or 0), item['competition_name'], item['group_name'], item['team_name']))
    if not candidates:
        return context

    normalized_team = _normalize_team_lookup_key(team_query)
    exact_candidates = [
        candidate
        for candidate in candidates
        if _normalize_team_lookup_key(candidate.get('team_name')) == normalized_team
    ]
    candidate_pool = exact_candidates or candidates
    chosen = candidate_pool[0] if candidate_pool else None
    if not chosen:
        return context
    if len(candidate_pool) > 1 and int(candidate_pool[0].get('score') or 0) == int(candidate_pool[1].get('score') or 0):
        return context

    update_fields = []
    for field_name, raw_value in (
        ('external_competition_key', chosen.get('external_competition_key')),
        ('external_group_key', chosen.get('external_group_key')),
        ('external_team_key', chosen.get('external_team_key')),
        ('external_team_name', chosen.get('external_team_name') or chosen.get('team_name')),
    ):
        value = str(raw_value or '').strip()
        if value and getattr(context, field_name, '') != value:
            setattr(context, field_name, value)
            update_fields.append(field_name)
    if update_fields:
        context.save(update_fields=update_fields + ['updated_at'])
    return context


def _find_universo_next_match_for_context(context, primary_team):
    if not context or str(getattr(context, 'provider', '') or '').strip() != WorkspaceCompetitionContext.PROVIDER_UNIVERSO:
        return {}
    group_key = str(getattr(context, 'external_group_key', '') or '').strip()
    if not group_key:
        group_key = str(getattr(getattr(primary_team, 'group', None), 'external_id', '') or '').strip()
    if not group_key:
        return {}
    team_keys = _context_team_lookup_keys(context, primary_team)
    if not team_keys:
        return {}

    def _payload_from_row(row, fallback_date='', fallback_round=''):
        home_name = str(row.get('Nombre_equipo_local') or '').strip()
        away_name = str(row.get('Nombre_equipo_visitante') or '').strip()
        home_keys = _expand_team_lookup_variants(home_name)
        away_keys = _expand_team_lookup_variants(away_name)
        home_code = str(row.get('CodEquipo_local') or '').strip().lower()
        away_code = str(row.get('CodEquipo_visitante') or '').strip().lower()
        if home_code:
            home_keys.add(home_code)
        if away_code:
            away_keys.add(away_code)
        if team_keys & home_keys:
            opponent_name = away_name
            home_flag = True
            crest_url = _absolute_universo_url(row.get('url_img_visitante'))
            team_code = str(row.get('CodEquipo_visitante') or '').strip()
        elif team_keys & away_keys:
            opponent_name = home_name
            home_flag = False
            crest_url = _absolute_universo_url(row.get('url_img_local'))
            team_code = str(row.get('CodEquipo_local') or '').strip()
        else:
            return {}
        raw_date = str(row.get('fecha') or fallback_date or '').strip()
        date_iso = None
        if raw_date:
            for fmt in ('%d/%m/%Y', '%d-%m-%Y', '%Y-%m-%d'):
                try:
                    date_iso = datetime.strptime(raw_date, fmt).date().isoformat()
                    break
                except ValueError:
                    continue
        return normalize_next_match_payload(
            {
                'round': str(row.get('nombre_jornada') or row.get('jornada') or fallback_round or '').strip(),
                'date': date_iso,
                'time': str(row.get('hora') or '').strip(),
                'location': str(row.get('campojuego') or '').strip(),
                'opponent': {
                    'name': opponent_name or 'Rival por confirmar',
                    'full_name': opponent_name or 'Rival por confirmar',
                    'crest_url': crest_url,
                    'team_code': team_code,
                },
                'home': home_flag,
                'status': 'next',
                'source': 'universo-live',
            }
        )

    today = timezone.localdate()
    current_payload = _fetch_universo_live_results(group_key)
    if not current_payload:
        return {}
    rounds = []
    current_round = str(current_payload.get('jornada') or '').strip()
    if current_round:
        rounds.append(current_round)
    for bucket in current_payload.get('listado_jornadas') or []:
        if not isinstance(bucket, dict):
            continue
        for row in bucket.get('jornadas') or []:
            if not isinstance(row, dict):
                continue
            round_id = str(row.get('codjornada') or '').strip()
            if round_id and round_id not in rounds:
                rounds.append(round_id)

    # Universo suele devolver `listado_jornadas` desde la jornada 1, por lo que si iteramos "en crudo"
    # acabamos consultando sólo jornadas pasadas. Priorizamos jornada actual y siguientes.
    numeric_rounds = [rid for rid in rounds if str(rid).isdigit()]
    if numeric_rounds:
        ordered_unique = sorted({rid for rid in numeric_rounds}, key=lambda value: int(value))
        start_index = 0
        if current_round.isdigit() and current_round in ordered_unique:
            start_index = ordered_unique.index(current_round)
        rounds_to_check = ordered_unique[start_index:start_index + 6]
    else:
        rounds_to_check = rounds[:6]

    for round_id in rounds_to_check:
        payload = current_payload if round_id == current_round else _fetch_universo_live_results(group_key, round_id)
        if not payload:
            continue
        fallback_date = str(payload.get('fecha_jornada') or '').strip()
        fallback_round = str(payload.get('nombre_jornada') or payload.get('jornada') or round_id).strip()
        for row in payload.get('partidos') or []:
            if not isinstance(row, dict):
                continue
            candidate = _payload_from_row(row, fallback_date=fallback_date, fallback_round=fallback_round)
            if not candidate:
                continue
            payload_date = _parse_payload_date(candidate.get('date'))
            if payload_date and payload_date >= today:
                return candidate
    return {}


def _search_workspace_competition_candidates(provider, *, team_query='', competition_query='', group_query=''):
    team_query = str(team_query or '').strip()
    competition_query = str(competition_query or '').strip()
    group_query = str(group_query or '').strip()
    base_qs = (
        Team.objects
        .select_related('group__season__competition')
        .filter(group__isnull=False)
    )
    if team_query:
        base_qs = base_qs.filter(
            Q(name__icontains=team_query)
            | Q(short_name__icontains=team_query)
            | Q(slug__icontains=team_query)
            | Q(external_id__icontains=team_query)
        )
    if competition_query:
        base_qs = base_qs.filter(
            Q(group__season__competition__name__icontains=competition_query)
            | Q(group__season__competition__slug__icontains=competition_query)
            | Q(group__season__name__icontains=competition_query)
        )
    if group_query:
        base_qs = base_qs.filter(
            Q(group__name__icontains=group_query)
            | Q(group__slug__icontains=group_query)
            | Q(group__external_id__icontains=group_query)
        )
    candidates = []
    for team in base_qs.order_by('-group__season__is_current', 'group__season__competition__name', 'group__name', 'name')[:12]:
        group = getattr(team, 'group', None)
        season = getattr(group, 'season', None) if group else None
        competition = getattr(season, 'competition', None) if season else None
        score = 0
        normalized_team_name = normalize_label(team.name)
        normalized_short_name = normalize_label(team.short_name)
        normalized_team_query = normalize_label(team_query)
        normalized_group_name = normalize_label(getattr(group, 'name', ''))
        normalized_group_query = normalize_label(group_query)
        normalized_comp_name = normalize_label(getattr(competition, 'name', ''))
        normalized_comp_query = normalize_label(competition_query)
        if normalized_team_query:
            if normalized_team_query == normalized_team_name or normalized_team_query == normalized_short_name:
                score += 60
            elif normalized_team_query in normalized_team_name:
                score += 30
        if normalized_group_query:
            if normalized_group_query == normalized_group_name:
                score += 25
            elif normalized_group_query in normalized_group_name:
                score += 12
        if normalized_comp_query:
            if normalized_comp_query == normalized_comp_name:
                score += 25
            elif normalized_comp_query in normalized_comp_name:
                score += 12
        if season and season.is_current:
            score += 10
        candidates.append(
            {
                'team_id': team.id,
                'team_name': team.name,
                'team_slug': team.slug,
                'group_name': getattr(group, 'name', '') or 'Sin grupo',
                'group_slug': getattr(group, 'slug', '') or '',
                'group_external_id': getattr(group, 'external_id', '') or '',
                'season_name': getattr(season, 'name', '') or '',
                'competition_name': getattr(competition, 'name', '') or '',
                'competition_slug': getattr(competition, 'slug', '') or '',
                'provider': provider or WorkspaceCompetitionContext.PROVIDER_MANUAL,
                'external_competition_key': getattr(competition, 'slug', '') or getattr(competition, 'name', '') or '',
                'external_group_key': getattr(group, 'external_id', '') or getattr(group, 'slug', '') or '',
                'external_team_key': team.external_id or team.slug or '',
                'external_team_name': team.name,
                'score': score,
            }
        )
    candidates.sort(
        key=lambda item: (
            -item['score'],
            item['competition_name'],
            item['group_name'],
            item['team_name'],
        )
    )
    return candidates


def _sync_workspace_competition_context(workspace):
    if not workspace or workspace.kind != Workspace.KIND_CLUB:
        return None, 'Este cliente no admite contexto competitivo.'
    primary_team = workspace.primary_team
    context = _bootstrap_workspace_competition_context(workspace, primary_team=primary_team)
    if not context:
        return None, 'No se pudo preparar el contexto competitivo.'
    if not primary_team or not getattr(primary_team, 'group', None):
        context.sync_status = WorkspaceCompetitionContext.STATUS_ERROR
        context.sync_error = 'El cliente no tiene equipo o grupo vinculado.'
        context.last_sync_at = timezone.now()
        context.save(update_fields=['sync_status', 'sync_error', 'last_sync_at', 'updated_at'])
        return context, context.sync_error

    context = _ensure_universo_context_binding(context, primary_team)
    _sync_team_crest_from_sources(primary_team)
    if getattr(primary_team, 'group_id', None):
        for team in Team.objects.filter(group=primary_team.group).only('id', 'name', 'short_name', 'external_id', 'crest_url', 'crest_image', 'is_primary'):
            _sync_team_crest_from_sources(team)
    standings_payload = _resolve_standings_for_team(
        primary_team,
        snapshot=load_universo_snapshot(),
        provider=getattr(context, 'provider', None),
    )
    convocation_next = _build_next_match_from_convocation(primary_team)
    provider_next = _find_universo_next_match_for_context(context, primary_team)
    preferred_next = load_preferred_next_match_payload(primary_team=primary_team, competition_context=context)
    local_next = get_next_match(primary_team, primary_team.group)
    if not _next_match_payload_is_reliable(local_next):
        local_next = {}
    next_match_payload = (
        (convocation_next if _next_match_payload_is_reliable(convocation_next) else {})
        or provider_next
        or preferred_next
        or local_next
        or {}
    )
    schedule_payload = _build_workspace_schedule_payload(primary_team)
    snapshot, _ = WorkspaceCompetitionSnapshot.objects.get_or_create(
        workspace=workspace,
        defaults={'context': context},
    )
    snapshot.context = context
    snapshot.standings_payload = standings_payload or []
    snapshot.next_match_payload = next_match_payload or {}
    snapshot.schedule_payload = schedule_payload or []
    snapshot.save()
    if primary_team:
        _invalidate_team_dashboard_caches(primary_team)

    context.team = primary_team
    context.group = primary_team.group
    context.season = getattr(primary_team.group, 'season', None)
    context.last_sync_at = timezone.now()
    context.sync_status = WorkspaceCompetitionContext.STATUS_READY
    context.sync_error = ''
    if not context.external_team_name:
        context.external_team_name = str(primary_team.name or '').strip()
    context.save(update_fields=['team', 'group', 'season', 'last_sync_at', 'sync_status', 'sync_error', 'external_team_name', 'updated_at'])
    return context, ''


def _competition_payload_for_team(workspace, primary_team):
    context = None
    snapshot = None
    if workspace and workspace.kind == Workspace.KIND_CLUB:
        snapshot = WorkspaceCompetitionSnapshot.objects.filter(workspace=workspace).select_related('context').first()
        if snapshot and snapshot.context_id:
            context = snapshot.context
        elif primary_team:
            context = _bootstrap_workspace_competition_context(workspace, primary_team=primary_team)
            if context and context.is_auto_sync_enabled:
                _sync_workspace_competition_context(workspace)
                snapshot = WorkspaceCompetitionSnapshot.objects.filter(workspace=workspace).select_related('context').first()
                context = snapshot.context if snapshot and snapshot.context_id else context

    provider_key = str(getattr(context, 'provider', '') or '').strip().lower()
    standings_payload = _resolve_standings_for_team(
        primary_team,
        snapshot=load_universo_snapshot(),
        provider=provider_key,
    ) if primary_team else []
    next_match_payload = {}
    if primary_team and getattr(primary_team, 'group', None):
        next_match_payload = (
            load_preferred_next_match_payload(primary_team=primary_team, competition_context=context)
            or load_preferred_next_match_payload(primary_team=primary_team)
            or get_next_match(primary_team, primary_team.group)
            or {}
        )
    normalized_next_match_payload = normalize_next_match_payload(next_match_payload) if next_match_payload else {}
    # Si el provider es Universo y tenemos snapshot, preferimos su payload (más rico) cuando exista.
    if provider_key == WorkspaceCompetitionContext.PROVIDER_UNIVERSO and snapshot:
        if isinstance(snapshot.standings_payload, list) and snapshot.standings_payload:
            standings_payload = snapshot.standings_payload
        if isinstance(snapshot.next_match_payload, dict) and _next_match_payload_is_reliable(snapshot.next_match_payload):
            normalized_next_match_payload = snapshot.next_match_payload
        elif normalized_next_match_payload and _next_match_payload_is_reliable(normalized_next_match_payload):
            # Si el snapshot tiene un próximo partido poco fiable (sin fecha, rival placeholder, etc.),
            # lo refrescamos con el candidato fiable para que el dashboard no se quede "atascado".
            try:
                snapshot_payload = snapshot.next_match_payload if isinstance(snapshot.next_match_payload, dict) else {}
                if snapshot_payload != normalized_next_match_payload:
                    snapshot.next_match_payload = normalized_next_match_payload
                    snapshot.save(update_fields=['next_match_payload', 'updated_at'])
            except Exception:
                pass
    return {
        'standings': standings_payload or [],
        'next_match': normalized_next_match_payload or {},
    }


def _workspace_deliverable_flag(module_key, deliverable_key):
    return f'deliverable__{module_key}__{deliverable_key}'


def _workspace_module_flag(module_key):
    return f'module__{module_key}'


def _workspace_collect_route_keys(entries):
    route_keys = []
    seen = set()
    for entry in entries:
        for route_key in entry.get('route_keys', []) or []:
            if not route_key or route_key in seen:
                continue
            seen.add(route_key)
            route_keys.append(route_key)
    return route_keys


def _workspace_club_module_catalog():
    return [
        {
            'key': 'cover',
            'label': 'Portada',
            'description': 'Resumen ejecutivo del cliente: home, clasificación, contexto visual y próximo partido.',
            'deliverables': [
                {
                    'key': 'executive_home',
                    'label': 'Resumen ejecutivo',
                    'description': 'Home principal con foto, estado del equipo, alertas y contexto general.',
                    'route_keys': ['dashboard'],
                },
                {
                    'key': 'competition_context',
                    'label': 'Clasificación y próximo partido',
                    'description': 'Lectura competitiva, rival siguiente y visión rápida de la semana.',
                    'route_keys': ['dashboard'],
                },
                {
                    'key': 'visual_identity',
                    'label': 'Foto home e identidad visual',
                    'description': 'Carrusel, presencia visual y narrativa de portada del cliente.',
                    'route_keys': ['dashboard'],
                },
            ],
        },
        {
            'key': 'statistics',
            'label': 'Estadísticas',
            'description': 'KPIs de temporada, seguimiento individual y métricas manuales del cliente.',
            'deliverables': [
                {
                    'key': 'season_kpis',
                    'label': 'KPIs de temporada',
                    'description': 'Rendimiento global, disponibilidad, líderes y lectura competitiva.',
                    'route_keys': ['coach_overview'],
                },
                {
                    'key': 'player_follow_up',
                    'label': 'Seguimiento de jugadores',
                    'description': 'Panel individual, fichas y lectura técnica de la plantilla.',
                    'route_keys': ['coach_overview', 'players'],
                },
                {
                    'key': 'manual_metrics',
                    'label': 'Estadísticas manuales',
                    'description': 'Carga manual y correcciones de staff sobre métricas y estados.',
                    'route_keys': ['manual_stats'],
                },
            ],
        },
        {
            'key': 'technical_staff',
            'label': 'Cuerpo técnico',
            'description': 'Estructura de staff, plantilla operativa y accesos por área técnica.',
            'deliverables': [
                {
                    'key': 'staff_hub',
                    'label': 'Hub del staff',
                    'description': 'Portada operativa del cuerpo técnico y reparto de áreas.',
                    'route_keys': ['coach_overview'],
                },
                {
                    'key': 'staff_roster',
                    'label': 'Registro de jugadores',
                    'description': 'Alta, edición y consulta de plantilla dentro del cliente.',
                    'route_keys': ['players'],
                },
                {
                    'key': 'staff_areas',
                    'label': 'Áreas del staff',
                    'description': 'Acceso a entrenador, porteros, físico y coordinación técnica.',
                    'route_keys': ['coach_overview'],
                },
            ],
        },
        {
            'key': 'match',
            'label': 'Partido',
            'description': 'Operativa prepartido, live y postpartido del cliente.',
            'deliverables': [
                {
                    'key': 'convocation',
                    'label': 'Convocatoria',
                    'description': 'Lista oficial, disponibilidad y preparación inicial del partido.',
                    'route_keys': ['convocation'],
                },
                {
                    'key': 'starting_xi',
                    'label': '11 inicial',
                    'description': 'Once inicial, estructura táctica y banquillo del partido.',
                    'route_keys': ['convocation'],
                },
                {
                    'key': 'live_match',
                    'label': 'Registro de acciones',
                    'description': 'Operativa live, eventos de partido y lectura inmediata.',
                    'route_keys': ['match_actions'],
                },
            ],
        },
        {
            'key': 'training',
            'label': 'Entrenamiento',
            'description': 'Planificación, sesiones, tareas y áreas específicas de entrenamiento.',
            'deliverables': [
                {
                    'key': 'sessions',
                    'label': 'Sesiones',
                    'description': 'Microciclos, sesiones semanales y estructura de planificación.',
                    'route_keys': ['sessions'],
                },
                {
                    'key': 'training_areas',
                    'label': 'Áreas del entrenamiento',
                    'description': 'Entrenador, porteros y preparación física dentro del plan semanal.',
                    'route_keys': ['sessions'],
                },
                {
                    'key': 'abp',
                    'label': 'ABP',
                    'description': 'Pizarra y biblioteca específica de acciones a balón parado.',
                    'route_keys': ['sessions', 'abp_board'],
                },
            ],
        },
        {
            'key': 'analysis',
            'label': 'Análisis',
            'description': 'Scouting, informes, vídeo y lectura táctica del cliente.',
            'deliverables': [
                {
                    'key': 'rival_analysis',
                    'label': 'Análisis rival',
                    'description': 'Scouting rival y lectura táctica del oponente.',
                    'route_keys': ['analysis'],
                },
                {
                    'key': 'reports',
                    'label': 'Informes',
                    'description': 'Informes tácticos y soporte documental de análisis.',
                    'route_keys': ['analysis'],
                },
            ],
        },
    ]


def _workspace_task_studio_module_catalog():
    return [
        {
            'key': 'task_studio_access_account',
            'label': 'Acceso y cuenta',
            'description': 'Entrada al espacio, estado de acceso y base de cuenta del usuario.',
            'deliverables': [
                {
                    'key': 'home',
                    'label': 'Inicio',
                    'description': 'Home privada y punto de entrada del espacio.',
                    'route_keys': ['task_studio_home'],
                },
                {
                    'key': 'access_state',
                    'label': 'Estado de acceso',
                    'description': 'Gestión de acceso base y activación del espacio del usuario.',
                    'route_keys': ['task_studio_home'],
                },
            ],
        },
        {
            'key': 'task_studio_profile_identity',
            'label': 'Perfil e identidad',
            'description': 'Perfil profesional e identidad visual del usuario.',
            'deliverables': [
                {
                    'key': 'profile',
                    'label': 'Perfil entrenador',
                    'description': 'Datos personales y profesionales del entrenador.',
                    'route_keys': ['task_studio_profile'],
                },
                {
                    'key': 'branding',
                    'label': 'Identidad visual',
                    'description': 'Escudo, colores y marca documental para PDF e informes.',
                    'route_keys': ['task_studio_profile'],
                },
                {
                    'key': 'document_setup',
                    'label': 'Configuración documental',
                    'description': 'Firma, nombre documental y pie de documento del usuario.',
                    'route_keys': ['task_studio_profile'],
                },
            ],
        },
        {
            'key': 'task_studio_roster_area',
            'label': 'Plantilla privada',
            'description': 'Gestión de jugadores propios para pizarras, petos y documentos.',
            'deliverables': [
                {
                    'key': 'roster',
                    'label': 'Registro de plantilla',
                    'description': 'Alta, edición y consulta de jugadores privados.',
                    'route_keys': ['task_studio_roster'],
                },
                {
                    'key': 'visual_assignment',
                    'label': 'Asignación visual',
                    'description': 'Uso de plantilla en chapas, petos y recursos tácticos.',
                    'route_keys': ['task_studio_roster'],
                },
                {
                    'key': 'groups_and_bibs',
                    'label': 'Grupos y petos',
                    'description': 'Preparación de grupos y distribución visual para tareas.',
                    'route_keys': ['task_studio_roster'],
                },
            ],
        },
        {
            'key': 'task_studio_tactical_resources',
            'label': 'Recursos tácticos',
            'description': 'Superficies, recursos gráficos y presets del editor táctico.',
            'deliverables': [
                {
                    'key': 'surfaces',
                    'label': 'Superficies',
                    'description': 'Campos, variantes F7/F11 y superficies del editor.',
                    'route_keys': ['task_studio_tasks'],
                },
                {
                    'key': 'graphic_resources',
                    'label': 'Recursos gráficos',
                    'description': 'Líneas, flechas, figuras, emojis y recursos de diseño.',
                    'route_keys': ['task_studio_tasks'],
                },
                {
                    'key': 'board_presets',
                    'label': 'Presets de tablero',
                    'description': 'Presets y ayudas para construir tareas con rapidez.',
                    'route_keys': ['task_studio_tasks'],
                },
            ],
        },
        {
            'key': 'task_studio_tasks_library',
            'label': 'Tareas y biblioteca',
            'description': 'Repositorio privado, edición y reutilización de tareas.',
            'deliverables': [
                {
                    'key': 'repository',
                    'label': 'Repositorio',
                    'description': 'Listado, filtrado y gestión de tareas guardadas.',
                    'route_keys': ['task_studio_tasks'],
                },
                {
                    'key': 'editor',
                    'label': 'Editor táctico',
                    'description': 'Crear, editar y duplicar tareas dentro del workspace.',
                    'route_keys': ['task_studio_tasks'],
                },
                {
                    'key': 'library',
                    'label': 'Biblioteca',
                    'description': 'Reutilización, duplicado y trabajo sobre base de tareas.',
                    'route_keys': ['task_studio_tasks'],
                },
            ],
        },
        {
            'key': 'task_studio_documents_exports',
            'label': 'Documentos',
            'description': 'PDF UEFA, PDF Club y salidas documentales del módulo.',
            'deliverables': [
                {
                    'key': 'pdfs',
                    'label': 'PDFs',
                    'description': 'Impresión UEFA y Club desde Task Studio.',
                    'route_keys': ['task_studio_pdfs'],
                },
                {
                    'key': 'exports',
                    'label': 'Exportaciones',
                    'description': 'Salidas documentales y exportación de materiales del usuario.',
                    'route_keys': ['task_studio_pdfs'],
                },
            ],
        },
    ]


def _parse_workspace_usernames(raw_value):
    raw_chunks = re.split(r'[\n,;]+', str(raw_value or ''))
    usernames = []
    seen = set()
    for chunk in raw_chunks:
        normalized = _sanitize_task_text(chunk.strip(), multiline=False, max_len=150).lower()
        if not normalized or normalized in seen:
            continue
        usernames.append(normalized)
        seen.add(normalized)
    if not usernames:
        return [], []
    found_users = list(User.objects.filter(username__in=usernames))
    found_map = {user.username.lower(): user for user in found_users}
    missing = [username for username in usernames if username not in found_map]
    ordered_users = [found_map[username] for username in usernames if username in found_map]
    return ordered_users, missing


def _workspace_module_catalog(kind):
    if kind == Workspace.KIND_TASK_STUDIO:
        return _workspace_task_studio_module_catalog()
    return _workspace_club_module_catalog()


def _expand_workspace_module_selection(kind, selected_modules, selected_deliverables=None):
    defaults = _workspace_default_modules(kind if kind in {Workspace.KIND_TASK_STUDIO, Workspace.KIND_CLUB} else Workspace.KIND_CLUB)
    expanded = {key: False for key in defaults.keys()}
    selected_deliverables = selected_deliverables or {}
    for item in _workspace_module_catalog(kind):
        expanded[_workspace_module_flag(item['key'])] = bool(selected_modules.get(item['key']))
        if not bool(selected_modules.get(item['key'])):
            continue
        deliverables = item.get('deliverables', []) or []
        active_deliverables = []
        if deliverables:
            deliverable_states = {
                deliverable['key']: bool(selected_deliverables.get(_workspace_deliverable_flag(item['key'], deliverable['key'])))
                for deliverable in deliverables
            }
            if any(deliverable_states.values()):
                active_deliverables = [deliverable for deliverable in deliverables if deliverable_states.get(deliverable['key'])]
            else:
                active_deliverables = deliverables
            for deliverable in deliverables:
                expanded[_workspace_deliverable_flag(item['key'], deliverable['key'])] = deliverable in active_deliverables
        else:
            active_deliverables = [item]
        for route_key in _workspace_collect_route_keys(active_deliverables):
            if route_key in expanded:
                expanded[route_key] = True
    return expanded


def _workspace_selected_module_keys(kind, enabled_modules):
    selected = []
    for item in _workspace_module_catalog(kind):
        module_flag = _workspace_module_flag(item['key'])
        if module_flag in enabled_modules:
            is_enabled = bool(enabled_modules.get(module_flag))
        else:
            route_keys = _workspace_collect_route_keys(item.get('deliverables', []) or [item])
            is_enabled = any(bool(enabled_modules.get(route_key)) for route_key in route_keys)
        if is_enabled:
            selected.append(item['key'])
    return selected


def _workspace_selected_deliverable_keys(kind, enabled_modules):
    selected = []
    for item in _workspace_module_catalog(kind):
        for deliverable in item.get('deliverables', []) or []:
            flag = _workspace_deliverable_flag(item['key'], deliverable['key'])
            if flag in enabled_modules:
                is_enabled = bool(enabled_modules.get(flag))
            else:
                is_enabled = any(bool(enabled_modules.get(route_key)) for route_key in deliverable.get('route_keys', []) or [])
            if is_enabled:
                selected.append(flag)
    return selected


def _workspace_module_catalog_for_template(kind, enabled_modules=None):
    enabled_modules = enabled_modules or {}
    selected_modules = set(_workspace_selected_module_keys(kind, enabled_modules)) if enabled_modules else set()
    selected_deliverables = set(_workspace_selected_deliverable_keys(kind, enabled_modules)) if enabled_modules else set()
    rows = []
    for item in _workspace_module_catalog(kind):
        row = dict(item)
        row['enabled'] = item['key'] in selected_modules if enabled_modules else False
        deliverables = []
        for deliverable in item.get('deliverables', []) or []:
            deliverable_row = dict(deliverable)
            deliverable_row['flag_key'] = _workspace_deliverable_flag(item['key'], deliverable['key'])
            deliverable_row['enabled'] = deliverable_row['flag_key'] in selected_deliverables if enabled_modules else False
            deliverables.append(deliverable_row)
        row['deliverables'] = deliverables
        row['deliverable_count'] = len(deliverables)
        row['enabled_deliverable_count'] = sum(1 for deliverable in deliverables if deliverable.get('enabled'))
        rows.append(row)
    return rows


def _workspace_enabled_modules(workspace):
    defaults = _workspace_default_modules(workspace.kind if workspace else Workspace.KIND_CLUB)
    raw = getattr(workspace, 'enabled_modules', None)
    if not isinstance(raw, dict):
        return defaults
    normalized = dict(defaults)
    for key, value in raw.items():
        if key in defaults or str(key).startswith('deliverable__') or str(key).startswith('module__'):
            normalized[key] = bool(value)
    return normalized


def _workspace_has_module(workspace, module_key):
    return _workspace_has_module_for_user(workspace, module_key, user=None)


def _workspace_member_allows_module(workspace, user, module_key):
    if not workspace or not module_key:
        return True
    if not user or not getattr(user, 'is_authenticated', False):
        return False
    if _can_access_platform(user):
        return True
    membership = _workspace_membership_for_user(workspace, user)
    if not membership:
        # Owner explícito siempre permitido.
        if int(getattr(workspace, 'owner_user_id', 0) or 0) == int(getattr(user, 'id', 0) or 0):
            return True
        return False
    if membership.role in {WorkspaceMembership.ROLE_OWNER, WorkspaceMembership.ROLE_ADMIN}:
        return True
    raw = getattr(membership, 'module_access', None)
    if not isinstance(raw, dict) or not raw:
        return True
    # Sólo interpretamos flags explícitas a False como "denegar".
    return raw.get(module_key, True) is not False


def _workspace_has_module_for_user(workspace, module_key, *, user=None):
    if not workspace or not module_key:
        return True
    enabled_modules = _workspace_enabled_modules(workspace)
    if not bool(enabled_modules.get(module_key, False)):
        return False
    if user is None:
        return True
    return _workspace_member_allows_module(workspace, user, module_key)


def _forbid_if_workspace_module_disabled(request, module_key, label='módulo'):
    workspace = _get_active_workspace(request)
    if not workspace:
        if request and getattr(request, 'user', None) and request.user.is_authenticated and not _can_access_platform(request.user):
            # Para uso "monocliente" (sin workspaces configurados), permitimos dashboard/jugadores.
            if module_key in {'players', 'dashboard'}:
                return None
            return HttpResponse('No tienes un workspace de club asignado.', status=403)
        return None
    if workspace.kind != Workspace.KIND_CLUB:
        if request and getattr(request, 'user', None) and request.user.is_authenticated and not _can_access_platform(request.user):
            if module_key in {'players', 'dashboard'}:
                return None
            return HttpResponse('El workspace activo no es de tipo club.', status=403)
        return None
    if _workspace_has_module_for_user(workspace, module_key, user=request.user if request else None):
        return None
    return HttpResponse(f'El {label} no está activo en el workspace actual.', status=403)


def _task_studio_workspace_for_target(owner):
    if not owner:
        return None
    return _ensure_task_studio_workspace(owner)


def _forbid_if_task_studio_module_disabled(request, owner, module_key, label='módulo'):
    workspace = _task_studio_workspace_for_target(owner)
    if not workspace:
        return None
    if _workspace_has_module_for_user(workspace, module_key, user=request.user if request else None):
        return None
    return HttpResponse(f'El {label} no está activo en este Task Studio.', status=403)


def _workspace_entry_url(workspace, *, user=None):
    if not workspace:
        return reverse('platform-overview')
    if workspace.kind == Workspace.KIND_TASK_STUDIO:
        owner_id = workspace.owner_user_id
        candidates = [
            ('task_studio_home', reverse('task-studio-home')),
            ('task_studio_profile', reverse('task-studio-profile')),
            ('task_studio_roster', reverse('task-studio-roster')),
            ('task_studio_tasks', reverse('task-studio-task-create')),
        ]
        for module_key, url in candidates:
            if _workspace_has_module_for_user(workspace, module_key, user=user):
                return f'{url}?user={owner_id}' if owner_id else url
        return reverse('platform-workspace-detail', args=[workspace.id])
    candidates = [
        ('dashboard', reverse('dashboard-home')),
        ('coach_overview', reverse('coach-detail')),
        ('players', reverse('player-dashboard')),
        ('convocation', reverse('convocation')),
        ('match_actions', reverse('match-action-page')),
        ('sessions', reverse('sessions')),
        ('analysis', reverse('analysis')),
        ('abp_board', reverse('coach-abp-board')),
        ('manual_stats', reverse('manual-player-stats')),
    ]
    for module_key, url in candidates:
        if _workspace_has_module_for_user(workspace, module_key, user=user):
            return url
    return reverse('platform-workspace-detail', args=[workspace.id])


def _workspace_membership_for_user(workspace, user):
    if not workspace or not user or not user.is_authenticated:
        return None
    return WorkspaceMembership.objects.filter(workspace=workspace, user=user).first()


def _can_view_workspace(user, workspace):
    if _can_access_platform(user):
        return True
    membership = _workspace_membership_for_user(workspace, user)
    return bool(membership)


def _can_manage_workspace(user, workspace):
    if _can_access_platform(user):
        return True
    membership = _workspace_membership_for_user(workspace, user)
    return bool(membership and membership.role in {WorkspaceMembership.ROLE_OWNER, WorkspaceMembership.ROLE_ADMIN})


def _workspace_links_for_user(user):
    if not user or not user.is_authenticated:
        return []
    workspaces = list(
        Workspace.objects
        .filter(Q(memberships__user=user) | Q(owner_user=user))
        .select_related('owner_user', 'primary_team')
        .distinct()
        .order_by('kind', 'name', 'id')[:8]
    )
    links = []
    for workspace in workspaces:
        label = workspace.name
        if len(label) > 24:
            label = label[:21].rstrip() + '...'
        links.append(
            {
                'id': workspace.id,
                'label': label,
                'detail_url': reverse('platform-workspace-detail', args=[workspace.id]),
                'enter_url': reverse('platform-workspace-enter', args=[workspace.id]),
            }
        )
    return links


def _ensure_club_workspace(primary_team):
    if not primary_team:
        return None
    workspace = Workspace.objects.filter(primary_team=primary_team).first()
    if workspace:
        changed = False
        desired_name = str(primary_team.display_name or primary_team.name or 'Club').strip() or 'Club'
        if workspace.name != desired_name:
            workspace.name = desired_name
            changed = True
        if workspace.kind != Workspace.KIND_CLUB:
            workspace.kind = Workspace.KIND_CLUB
            changed = True
        if not isinstance(workspace.enabled_modules, dict) or not workspace.enabled_modules:
            workspace.enabled_modules = _workspace_default_modules(Workspace.KIND_CLUB)
            changed = True
        if changed:
            workspace.save(update_fields=['name', 'kind', 'enabled_modules', 'updated_at'])
        _bootstrap_workspace_competition_context(workspace, primary_team=primary_team)
        return workspace
    workspace = Workspace.objects.create(
        name=str(primary_team.display_name or primary_team.name or 'Club').strip() or 'Club',
        slug=_unique_workspace_slug(primary_team.display_name or primary_team.name or 'club'),
        kind=Workspace.KIND_CLUB,
        primary_team=primary_team,
        enabled_modules=_workspace_default_modules(Workspace.KIND_CLUB),
    )
    _bootstrap_workspace_competition_context(workspace, primary_team=primary_team)
    return workspace


def _ensure_task_studio_workspace(user):
    if not user:
        return None
    disabled_profile = TaskStudioProfile.objects.filter(user=user, is_enabled=False).first()
    if disabled_profile:
        return None
    workspace = Workspace.objects.filter(kind=Workspace.KIND_TASK_STUDIO, owner_user=user).first()
    default_name = (
        str(user.get_full_name() or '').strip()
        or str(getattr(user, 'first_name', '') or '').strip()
        or user.get_username()
    )
    default_name = f'Task Studio · {default_name}'
    if workspace:
        changed = False
        if workspace.name != default_name:
            workspace.name = default_name
            changed = True
        if workspace.kind != Workspace.KIND_TASK_STUDIO:
            workspace.kind = Workspace.KIND_TASK_STUDIO
            changed = True
        if not isinstance(workspace.enabled_modules, dict) or not workspace.enabled_modules:
            workspace.enabled_modules = _workspace_default_modules(Workspace.KIND_TASK_STUDIO)
            changed = True
        if changed:
            workspace.save(update_fields=['name', 'kind', 'enabled_modules', 'updated_at'])
    else:
        workspace = Workspace.objects.create(
            name=default_name,
            slug=_unique_workspace_slug(f'task-studio-{user.get_username()}'),
            kind=Workspace.KIND_TASK_STUDIO,
            owner_user=user,
            enabled_modules=_workspace_default_modules(Workspace.KIND_TASK_STUDIO),
        )
    WorkspaceMembership.objects.get_or_create(
        workspace=workspace,
        user=user,
        defaults={'role': WorkspaceMembership.ROLE_OWNER},
    )
    return workspace


def _delete_task_studio_workspace(workspace, *, disable_owner_profile=True):
    if not workspace or workspace.kind != Workspace.KIND_TASK_STUDIO:
        return
    owner_user = workspace.owner_user
    if owner_user and disable_owner_profile:
        TaskStudioProfile.objects.update_or_create(
            user=owner_user,
            defaults={'workspace': None, 'is_enabled': False},
        )
    TaskStudioTask.objects.filter(workspace=workspace).delete()
    TaskStudioRosterPlayer.objects.filter(workspace=workspace).delete()
    TaskStudioProfile.objects.filter(workspace=workspace).exclude(user=owner_user).delete()
    WorkspaceMembership.objects.filter(workspace=workspace).delete()
    workspace.delete()


def _can_access_player_resource(user, player, primary_team=None):
    if not user or not user.is_authenticated or not player:
        return False
    if _is_admin_user(user):
        return True
    role = _get_user_role(user)
    if role is None:
        # Preserve access for legacy internal users that predate AppUserRole.
        return True
    if role in TECHNICAL_ROLES:
        return True
    if role == AppUserRole.ROLE_PLAYER:
        resolved_player = _resolve_player_for_user(user, primary_team or getattr(player, 'team', None))
        return bool(resolved_player and resolved_player.id == player.id)
    return False


def _forbid_if_no_coach_access(user):
    if _can_access_coach_workspace(user):
        return None
    return HttpResponse('No tienes permisos para acceder a este espacio.', status=403)


def _forbid_if_no_player_access(user, player, primary_team=None):
    if _can_access_player_resource(user, player, primary_team=primary_team):
        return None
    return HttpResponse('No tienes permisos para acceder a este jugador.', status=403)


def _resolve_player_for_user(user, primary_team):
    if not user or not user.is_authenticated or not primary_team:
        return None
    candidates = list(Player.objects.filter(team=primary_team, is_active=True))
    if not candidates:
        return None
    if len(candidates) == 1:
        return candidates[0]
    raw_values = [
        user.get_username(),
        getattr(user, 'email', ''),
        getattr(user, 'first_name', ''),
        getattr(user, 'last_name', ''),
        user.get_full_name(),
    ]
    normalized_tokens = set()
    for raw in raw_values:
        value = str(raw or '').strip()
        if not value:
            continue
        normalized_tokens.add(normalize_player_name(value))
        if '@' in value:
            normalized_tokens.add(normalize_player_name(value.split('@', 1)[0]))
    best_player = None
    best_score = 0
    for player in candidates:
        variants = [
            player.name,
            getattr(player, 'full_name', ''),
            getattr(player, 'nickname', ''),
        ]
        score = 0
        for variant in variants:
            normalized_variant = normalize_player_name(variant)
            if not normalized_variant:
                continue
            if normalized_variant in normalized_tokens:
                score = max(score, 100)
            for token in normalized_tokens:
                if not token:
                    continue
                if token in normalized_variant or normalized_variant in token:
                    score = max(score, min(len(token), len(normalized_variant)))
        if score > best_score:
            best_score = score
            best_player = player
    return best_player if best_score >= 4 else None


def _is_team_only_action(action_type: str) -> bool:
    normalized = (action_type or '').strip().lower()
    if not normalized:
        return False
    folded = ''.join(
        ch for ch in unicodedata.normalize('NFKD', normalized) if not unicodedata.combining(ch)
    )
    team_only_aliases = {
        'saque de esquina a favor',
        'saque de esquina en contra',
        'corner a favor',
        'corner en contra',
    }
    return folded in team_only_aliases


def authenticated_write(view_func):
    @wraps(view_func)
    def _wrapped(request, *args, **kwargs):
        if not request.user.is_authenticated:
            return JsonResponse({'error': 'Autenticación requerida'}, status=401)
        return view_func(request, *args, **kwargs)

    return _wrapped


def load_cached_next_match():
    if not NEXT_MATCH_CACHE.exists():
        return None
    try:
        with NEXT_MATCH_CACHE.open(encoding="utf-8") as handle:
            payload = json.load(handle)
            if isinstance(payload, dict):
                payload = normalize_next_match_payload(payload)
                payload.setdefault("status", "next")
                status = (payload.get('status') or '').lower()
                source = str(payload.get('source') or '').strip().lower()
                date_raw = payload.get('date')
                if date_raw:
                    payload_date = _parse_payload_date(date_raw)
                    today = timezone.localdate()
                    if payload_date:
                        if status == 'next' and payload_date < today:
                            return None
                        if status == 'latest' and payload_date < (today - timedelta(days=3)):
                            return None
                    elif status == 'next':
                        # If we cannot parse the date, avoid surfacing stale "next match" payloads.
                        return None
                else:
                    # Permitimos "próximo partido" sin fecha cuando el origen es una agenda (RFAF/Universo live).
                    # Ejemplo: partidos suspendidos/aplazados en la jornada con fecha pasada.
                    if status == 'next' and source in {'', 'local-match'}:
                        return None
                return payload
    except Exception:
        return None
    return None


def normalize_next_match_payload(payload):
    if not isinstance(payload, dict):
        return payload
    opponent = payload.get('opponent')
    if isinstance(opponent, str):
        clean_name = opponent.strip() or 'Rival por confirmar'
        payload['opponent'] = {
            'name': clean_name,
            'full_name': clean_name,
            'crest_url': '',
            'team_code': '',
        }
    elif isinstance(opponent, dict):
        name = str(opponent.get('name') or opponent.get('full_name') or '').strip()
        full_name = str(opponent.get('full_name') or name).strip()
        payload['opponent'] = {
            'name': name or 'Rival por confirmar',
            'full_name': full_name or name or 'Rival por confirmar',
            'crest_url': str(opponent.get('crest_url') or '').strip(),
            'team_code': str(opponent.get('team_code') or '').strip(),
        }
    else:
        fallback = str(payload.get('rival') or '').strip()
        payload['opponent'] = {
            'name': fallback or 'Rival por confirmar',
            'full_name': fallback or 'Rival por confirmar',
            'crest_url': '',
            'team_code': '',
        }
    return payload


def _enrich_standings_rows_with_crests(rows):
    if not isinstance(rows, list):
        return []
    snapshot_lookup = _build_universo_standings_lookup(load_universo_snapshot())
    capture_lookup = _build_universo_capture_team_lookup()
    crest_lookup = _build_team_crest_lookup()
    enriched = []
    for row in rows:
        if not isinstance(row, dict):
            continue
        row_copy = dict(row)
        team_name = str(row_copy.get('full_name') or row_copy.get('team') or '').strip()
        team_key = _normalize_team_lookup_key(team_name)
        if not row_copy.get('crest_url') and team_key:
            row_copy['crest_url'] = (
                str(snapshot_lookup.get(team_key, {}).get('crest_url') or '').strip()
                or str(capture_lookup.get(team_key, {}).get('crest_url') or '').strip()
                or str(crest_lookup.get(team_key) or '').strip()
            )
        if not row_copy.get('crest_url') and team_name:
            _, rival_crest_url = _resolve_rival_identity(team_name)
            row_copy['crest_url'] = str(rival_crest_url or '').strip()
        row_copy['crest_url'] = _sanitize_universo_external_image(
            _absolute_universo_url(row_copy.get('crest_url'))
        )
        enriched.append(row_copy)
    return enriched


def _enrich_next_match_payload_with_crests(next_match, standings_rows=None):
    if not isinstance(next_match, dict):
        return {}
    payload = normalize_next_match_payload(dict(next_match))
    opponent = payload.get('opponent') if isinstance(payload.get('opponent'), dict) else {}
    opponent_name = str(opponent.get('full_name') or opponent.get('name') or payload.get('rival') or '').strip()
    if opponent_name and not str(opponent.get('crest_url') or '').strip():
        rival_full_name, rival_crest_url = _resolve_rival_identity(opponent_name, preferred_opponent=opponent)
        if rival_full_name and not opponent.get('full_name'):
            opponent['full_name'] = rival_full_name
        if rival_crest_url:
            opponent['crest_url'] = _absolute_universo_url(rival_crest_url)
    if opponent_name and not str(opponent.get('crest_url') or '').strip() and isinstance(standings_rows, list):
        normalized_opponent = _normalize_team_lookup_key(opponent_name)
        for row in standings_rows:
            if not isinstance(row, dict):
                continue
            row_name = _normalize_team_lookup_key(row.get('full_name') or row.get('team'))
            if row_name and normalized_opponent and (row_name == normalized_opponent or row_name in normalized_opponent or normalized_opponent in row_name):
                crest_url = _absolute_universo_url(row.get('crest_url'))
                if crest_url:
                    opponent['crest_url'] = crest_url
                if not opponent.get('full_name') and row.get('full_name'):
                    opponent['full_name'] = str(row.get('full_name')).strip()
                break
    payload['opponent'] = {
        'name': str(opponent.get('name') or opponent.get('full_name') or opponent_name or 'Rival por confirmar').strip() or 'Rival por confirmar',
        'full_name': str(opponent.get('full_name') or opponent.get('name') or opponent_name or 'Rival por confirmar').strip() or 'Rival por confirmar',
        'crest_url': _sanitize_universo_external_image(_absolute_universo_url(opponent.get('crest_url'))),
        'team_code': str(opponent.get('team_code') or '').strip(),
    }
    return payload


def _build_weekly_staff_brief_context(primary_team, player_cards=None):
    if not primary_team:
        return None
    player_cards = player_cards if player_cards is not None else compute_player_cards(primary_team)
    active_match = get_active_match(primary_team)
    current_convocation = get_current_convocation_record(primary_team, match=active_match)
    active_injury_ids = get_active_injury_player_ids(
        [item.get('player_id') for item in player_cards if item.get('player_id')]
    )
    sanctioned_player_ids = get_sanctioned_player_ids_from_previous_round(
        primary_team,
        reference_match=active_match,
    )
    next_match_payload = load_preferred_next_match_payload(primary_team=primary_team)
    if not _next_match_payload_is_reliable(next_match_payload) and primary_team.group:
        local_next_match = get_next_match(primary_team, primary_team.group)
        if _next_match_payload_is_reliable(local_next_match):
            next_match_payload = local_next_match
    if not _next_match_payload_is_reliable(next_match_payload):
        convocation_next_match = _build_next_match_from_convocation(primary_team)
        if _next_match_payload_is_reliable(convocation_next_match):
            next_match_payload = convocation_next_match
    return build_weekly_staff_brief(
        player_cards=player_cards,
        active_injury_ids=active_injury_ids,
        sanctioned_player_ids=sanctioned_player_ids,
        convocation_player_ids=(
            current_convocation.players.values_list('id', flat=True)
            if current_convocation
            else []
        ),
        next_match=next_match_payload,
    )


def _parse_payload_date(raw):
    if not raw:
        return None
    value = str(raw).strip()
    for fmt in ('%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y'):
        try:
            return datetime.strptime(value, fmt).date()
        except ValueError:
            continue
    return None


def _parse_payload_time(raw):
    if not raw:
        return None
    value = str(raw).strip()
    for fmt in ('%H:%M', '%H.%M', '%H,%M'):
        try:
            return datetime.strptime(value, fmt).time()
        except ValueError:
            continue
    return None


def _payload_opponent_name(payload):
    if not isinstance(payload, dict):
        return ''
    opponent = payload.get('opponent')
    if isinstance(opponent, dict):
        return str(opponent.get('full_name') or opponent.get('name') or '').strip()
    if isinstance(opponent, str):
        return opponent.strip()
    return str(payload.get('rival') or '').strip()


def _build_team_recent_activity(primary_team):
    if not primary_team:
        return []
    recent_activity = []
    recent_sessions = list(
        TrainingSession.objects
        .filter(microcycle__team=primary_team)
        .select_related('microcycle')
        .order_by('-created_at', '-id')[:3]
    )
    recent_tasks = list(
        SessionTask.objects
        .filter(session__microcycle__team=primary_team)
        .select_related('session')
        .order_by('-created_at', '-id')[:4]
    )
    for session in recent_sessions:
        recent_activity.append(
            {
                'type': 'Sesión',
                'title': str(session.focus or 'Sesión').strip() or 'Sesión',
                'meta': f"{session.session_date:%d/%m/%Y} · {session.duration_minutes or 0} min",
                'url': reverse('sessions'),
                'created_at': session.created_at,
            }
        )
    for task in recent_tasks:
        recent_activity.append(
            {
                'type': 'Tarea',
                'title': str(task.title or 'Tarea').strip() or 'Tarea',
                'meta': f"{task.duration_minutes or 0} min · {task.session.focus or 'Sesión'}",
                'url': reverse('session-task-detail', args=[task.id]),
                'created_at': task.created_at,
            }
        )
    return sorted(
        recent_activity,
        key=lambda item: item.get('created_at') or timezone.now(),
        reverse=True,
    )[:5]


def _build_team_pending_cards(primary_team, weekly_brief=None):
    if not primary_team:
        return []
    today = timezone.localdate()
    weekly_brief = weekly_brief if isinstance(weekly_brief, dict) else _build_weekly_staff_brief_context(primary_team)
    pending_cards = []

    def add_card(title, description, url, action):
        pending_cards.append(
            {
                'title': title,
                'description': description,
                'url': url,
                'action': action,
            }
        )

    if int(weekly_brief.get('convocated_count') or 0) <= 0:
        add_card(
            'Convocatoria pendiente',
            'Todavía no hay una convocatoria cerrada para el siguiente partido.',
            reverse('convocation'),
            'Abrir partido',
        )
    if int(weekly_brief.get('probable_eleven_count') or 0) <= 0:
        add_card(
            '11 inicial sin definir',
            'El partido no tiene todavía un 11 inicial o probable consolidado.',
            reverse('initial-eleven'),
            'Definir 11',
        )
    if int(weekly_brief.get('available_count') or 0) <= 0:
        add_card(
            'Disponibilidad sin consolidar',
            'La portada no tiene disponibilidad útil para leer la semana del equipo.',
            reverse('coach-role-trainer'),
            'Revisar estadísticas',
        )

    future_sessions = list(
        TrainingSession.objects
        .filter(microcycle__team=primary_team, session_date__gte=today)
        .prefetch_related('tasks')
        .order_by('session_date', 'id')[:6]
    )
    if not future_sessions:
        add_card(
            'Semana sin sesiones',
            'No hay sesiones futuras planificadas para sostener el microciclo actual.',
            reverse('sessions'),
            'Planificar semana',
        )
    elif not any(session.tasks.exists() for session in future_sessions):
        add_card(
            'Sesiones sin tareas',
            'Hay sesiones creadas, pero todavía no tienen tareas asociadas.',
            reverse('sessions') + '?tab=planning',
            'Completar sesiones',
        )

    next_match = weekly_brief.get('match') if isinstance(weekly_brief, dict) else {}
    rival_name = str(next_match.get('opponent') or '').strip() if isinstance(next_match, dict) else ''
    has_ready_rival_report = False
    if rival_name:
        has_ready_rival_report = RivalAnalysisReport.objects.filter(
            team=primary_team,
            status=RivalAnalysisReport.STATUS_READY,
        ).filter(
            Q(rival_name__icontains=rival_name) | Q(rival_team__name__icontains=rival_name)
        ).exists()
    if rival_name and not has_ready_rival_report:
        add_card(
            'Informe rival pendiente',
            f'No hay un informe rival listo para {rival_name}.',
            reverse('analysis'),
            'Abrir análisis',
        )

    return pending_cards


def _build_universo_standings_lookup(snapshot):
    lookup = {}
    if not isinstance(snapshot, dict):
        return lookup
    for row in snapshot.get('standings') or []:
        if not isinstance(row, dict):
            continue
        team_name = str(row.get('team') or row.get('full_name') or '').strip()
        key = _normalize_team_lookup_key(team_name)
        if not key:
            continue
        lookup[key] = {
            'full_name': str(row.get('full_name') or team_name).strip() or team_name,
            'crest_url': str(row.get('crest_url') or '').strip(),
            'team_code': str(row.get('team_code') or '').strip(),
        }
    return lookup


def _absolute_universo_url(path_or_url):
    value = str(path_or_url or '').strip()
    if not value:
        return ''
    if value.startswith('http://') or value.startswith('https://'):
        return value
    if value.startswith('/'):
        # Solo convertimos rutas relativas de Universo. Rutas locales (MEDIA_URL/STATIC_URL) deben quedarse tal cual.
        if value.startswith('/pnfg/') or value.startswith('/api/') or value.startswith('/_next/'):
            return f'https://www.universorfaf.es{value}'
        return value
    return f'https://www.universorfaf.es/{value.lstrip("/")}'


def _sanitize_universo_external_image(url):
    """Universo RFAF expone URLs de escudos que en la práctica pueden devolver HTML/404.

    En esos casos el navegador los bloquea (ORB) y ralentiza la home. Por defecto los anulamos
    y dejamos que el frontend muestre el equipo sin imagen.
    """
    value = str(url or '').strip()
    if not value:
        return ''
    if UNIVERSO_EXTERNAL_IMAGES_ENABLED:
        return value
    lowered = value.lower()
    if 'universorfaf.es/pnfg/pimg/' in lowered:
        return ''
    return value


def _build_universo_capture_team_lookup():
    memo = getattr(_build_universo_capture_team_lookup, '_memo', None)
    lookup = {}
    capture_path = UNIVERSO_CAPTURE_PATH
    if not capture_path.exists():
        return lookup
    try:
        mtime = capture_path.stat().st_mtime
    except Exception:
        mtime = None
    if isinstance(memo, dict) and memo.get('mtime') and mtime and float(memo.get('mtime')) == float(mtime):
        cached_lookup = memo.get('lookup')
        if isinstance(cached_lookup, dict):
            return cached_lookup
    try:
        payload = json.loads(capture_path.read_text(encoding='utf-8'))
    except Exception:
        return lookup
    items = payload.get('items') if isinstance(payload, dict) else None
    if not isinstance(items, list):
        return lookup

    def _push(team_name, crest):
        full_name = str(team_name or '').strip()
        if not full_name:
            return
        key = _normalize_team_lookup_key(full_name)
        if not key:
            return
        crest_url = _absolute_universo_url(crest)
        current = lookup.get(key, {})
        if not current:
            lookup[key] = {
                'full_name': full_name,
                'crest_url': crest_url,
            }
            return
        if len(full_name) > len(current.get('full_name') or ''):
            current['full_name'] = full_name
        if crest_url and not current.get('crest_url'):
            current['crest_url'] = crest_url
        lookup[key] = current

    for item in items:
        if not isinstance(item, dict):
            continue
        data = item.get('json')
        if isinstance(data, dict):
            _push(data.get('nombre_equipo') or data.get('equipo'), data.get('escudo_equipo'))
            competitions = data.get('competiciones_participa')
            if isinstance(competitions, list):
                for row in competitions:
                    if isinstance(row, dict):
                        _push(row.get('nombre_equipo') or row.get('nombre_club'), row.get('escudo_equipo'))

            for bucket in data.values():
                if isinstance(bucket, list):
                    for row in bucket:
                        if isinstance(row, dict):
                            _push(row.get('nombre_equipo') or row.get('equipo'), row.get('escudo_equipo'))
    try:
        _build_universo_capture_team_lookup._memo = {'mtime': mtime, 'lookup': lookup}
    except Exception:
        pass
    return lookup


def _build_team_crest_lookup():
    memo = getattr(_build_team_crest_lookup, '_memo', None)
    lookup = {}

    def _push(name='', external_id='', crest=''):
        crest_url = _absolute_universo_url(crest)
        if not crest_url:
            return
        normalized_name = _normalize_team_lookup_key(name)
        normalized_external_id = str(external_id or '').strip().lower()
        for key in (normalized_name, normalized_external_id):
            if key and key not in lookup:
                lookup[key] = crest_url

    capture_path = UNIVERSO_CAPTURE_PATH
    try:
        capture_mtime = capture_path.stat().st_mtime if capture_path.exists() else None
    except Exception:
        capture_mtime = None
    snapshot_mtime = None
    snapshot_memo = getattr(load_universo_snapshot, '_memo', None)
    if isinstance(snapshot_memo, dict):
        snapshot_mtime = snapshot_memo.get('mtime')
    if isinstance(memo, dict) and memo.get('capture_mtime') == capture_mtime and memo.get('snapshot_mtime') == snapshot_mtime:
        cached_lookup = memo.get('lookup')
        if isinstance(cached_lookup, dict):
            return cached_lookup
    if capture_path.exists():
        try:
            payload = json.loads(capture_path.read_text(encoding='utf-8'))
        except Exception:
            payload = {}
        items = payload.get('items') if isinstance(payload, dict) else []
        for item in items or []:
            if not isinstance(item, dict):
                continue
            data = item.get('json')
            if not isinstance(data, dict):
                continue
            _push(
                data.get('nombre_equipo') or data.get('equipo') or data.get('nombre_club'),
                data.get('codequipo') or data.get('cod_equipo') or data.get('CodEquipo'),
                data.get('escudo_equipo') or data.get('escudo'),
            )
            competitions = data.get('competiciones_participa')
            if isinstance(competitions, list):
                for row in competitions:
                    if not isinstance(row, dict):
                        continue
                    _push(
                        row.get('nombre_equipo') or row.get('nombre_club'),
                        row.get('codequipo') or row.get('cod_equipo'),
                        row.get('escudo_equipo') or row.get('escudo'),
                    )
            for bucket in data.values():
                if not isinstance(bucket, list):
                    continue
                for row in bucket:
                    if not isinstance(row, dict):
                        continue
                    _push(
                        row.get('nombre') or row.get('nombre_equipo') or row.get('nombre_club') or row.get('Nombre_equipo_local') or row.get('Nombre_equipo_visitante'),
                        row.get('codequipo') or row.get('CodEquipo_local') or row.get('CodEquipo_visitante'),
                        row.get('escudo_equipo') or row.get('escudo') or row.get('url_img_local') or row.get('url_img_visitante'),
                    )
    snapshot = load_universo_snapshot()
    if isinstance(snapshot, dict):
        for row in snapshot.get('standings') or []:
            if not isinstance(row, dict):
                continue
            _push(row.get('team') or row.get('full_name'), row.get('team_code'), row.get('crest_url'))
        next_match = snapshot.get('next_match')
        if isinstance(next_match, dict):
            opponent = next_match.get('opponent')
            if isinstance(opponent, dict):
                _push(opponent.get('full_name') or opponent.get('name'), opponent.get('team_code'), opponent.get('crest_url'))
    try:
        _build_team_crest_lookup._memo = {'capture_mtime': capture_mtime, 'snapshot_mtime': snapshot_mtime, 'lookup': lookup}
    except Exception:
        pass
    return lookup


def _sync_team_crest_from_sources(team):
    if not team:
        return ''
    if getattr(team, 'crest_image', None):
        return ''
    lookup = _build_team_crest_lookup()
    lookup_keys = {
        _normalize_team_lookup_key(getattr(team, 'name', '') or ''),
        _normalize_team_lookup_key(getattr(team, 'display_name', '') or ''),
        str(getattr(team, 'external_id', '') or '').strip().lower(),
    }
    resolved = ''
    for key in lookup_keys:
        if key and lookup.get(key):
            resolved = lookup[key]
            break
    if resolved and resolved != (team.crest_url or ''):
        team.crest_url = resolved
        team.save(update_fields=['crest_url'])
        _invalidate_team_dashboard_caches(team)
    return resolved


def _is_benagalbon_team(team):
    if not team:
        return False
    if bool(getattr(team, 'is_primary', False)):
        return True
    slug = str(getattr(team, 'slug', '') or '').strip().lower()
    name = str(getattr(team, 'name', '') or '').strip().lower()
    short_name = str(getattr(team, 'short_name', '') or '').strip().lower()
    return 'benagalbon' in slug or 'benagalbon' in name or 'benagalbon' in short_name


def resolve_team_crest_url(request, team, *, fallback_static='football/images/cdb-logo.png', sync=False):
    if not team:
        return request.build_absolute_uri(static(fallback_static)) if fallback_static and request else ''
    if getattr(team, 'crest_image', None):
        try:
            return request.build_absolute_uri(team.crest_image.url) if request else team.crest_image.url
        except Exception:
            pass
    crest_url = str(getattr(team, 'crest_url', '') or '').strip()
    if not crest_url and sync:
        crest_url = _sync_team_crest_from_sources(team)
    crest_url = _sanitize_universo_external_image(_absolute_universo_url(crest_url))
    if crest_url:
        return crest_url
    # Para el equipo principal, usar un escudo local estable si no hay imagen subida.
    if _is_benagalbon_team(team):
        try:
            local_primary = 'football/images/cdb-benagalbon-crest.png'
            return request.build_absolute_uri(static(local_primary)) if request else static(local_primary)
        except Exception:
            pass
    # Fallback estable: escudo generado (evita imágenes externas rotas).
    try:
        generated = reverse('team-crest-svg', args=[team.id])
        return request.build_absolute_uri(generated) if request else generated
    except Exception:
        if fallback_static:
            return request.build_absolute_uri(static(fallback_static)) if request else static(fallback_static)
        return ''


def _team_initials(label):
    text = ' '.join(str(label or '').split()).strip()
    if not text:
        return '??'
    tokens = [tok for tok in re.split(r'[^A-Za-z0-9]+', text) if tok]
    if not tokens:
        return (text[:2] if len(text) >= 2 else text).upper()
    if len(tokens) == 1:
        return (tokens[0][:2] if len(tokens[0]) >= 2 else tokens[0]).upper()
    return (tokens[0][0] + tokens[1][0]).upper()


def _team_color_seed(team):
    base = str(getattr(team, 'slug', '') or getattr(team, 'name', '') or '').strip().lower()
    if not base:
        base = str(getattr(team, 'id', '') or 'team')
    total = 0
    for ch in base:
        total = (total * 31 + ord(ch)) % 360
    return total


@login_required
def team_crest_svg(request, team_id):
    team = Team.objects.filter(id=team_id).first()
    if not team:
        raise Http404('Equipo no encontrado')
    hue = _team_color_seed(team)
    initials = _team_initials(team.display_name or team.name)
    svg = f"""<svg xmlns="http://www.w3.org/2000/svg" width="160" height="160" viewBox="0 0 160 160">
  <defs>
    <linearGradient id="g" x1="0" y1="0" x2="1" y2="1">
      <stop offset="0%" stop-color="hsl({hue}, 70%, 42%)"/>
      <stop offset="100%" stop-color="hsl({(hue + 35) % 360}, 74%, 36%)"/>
    </linearGradient>
  </defs>
  <rect x="0" y="0" width="160" height="160" rx="32" fill="url(#g)"/>
  <rect x="10" y="10" width="140" height="140" rx="28" fill="rgba(2, 6, 23, 0.25)" stroke="rgba(255,255,255,0.26)" stroke-width="2"/>
  <text x="80" y="92" text-anchor="middle" font-family="system-ui, -apple-system, Segoe UI, Roboto, Arial" font-size="56" font-weight="800" fill="rgba(255,255,255,0.92)" letter-spacing="2">{html.escape(initials)}</text>
</svg>"""
    response = HttpResponse(svg, content_type='image/svg+xml; charset=utf-8')
    response['Cache-Control'] = 'public, max-age=604800, immutable'
    return response


def load_universo_capture():
    if not UNIVERSO_CAPTURE_PATH.exists():
        return {}
    try:
        payload = json.loads(UNIVERSO_CAPTURE_PATH.read_text(encoding='utf-8'))
    except Exception:
        return {}
    return payload if isinstance(payload, dict) else {}


def _load_universo_access_token():
    storage_path = Path(settings.BASE_DIR) / 'data' / 'input' / 'rfaf_storage_state.json'
    if not storage_path.exists():
        return ''
    try:
        payload = json.loads(storage_path.read_text(encoding='utf-8'))
    except Exception:
        return ''
    for cookie in payload.get('cookies') or []:
        if not isinstance(cookie, dict):
            continue
        if str(cookie.get('name') or '').strip() == 'access_token':
            return str(cookie.get('value') or '').strip()
    return ''


def _universo_api_post(endpoint, data=None):
    if requests is None:
        return {}
    token = _load_universo_access_token()
    if not token:
        return {}
    url = f'https://www.universorfaf.es/api/novanet/{endpoint.lstrip("/")}'
    headers = {
        'Authorization': f'Bearer {token}',
        'Accept': 'application/json',
        'User-Agent': '2j-football-intelligence/1.0',
    }
    try:
        response = requests.post(url, headers=headers, data=data or {}, timeout=UNIVERSO_API_TIMEOUT_SECONDS)
        response.raise_for_status()
        payload = response.json()
    except Exception:
        return {}
    return payload if isinstance(payload, dict) else {}


def _fetch_universo_live_seasons():
    payload = _universo_api_post('competition/get-seassons')
    return [row for row in (payload.get('temporadas') or []) if isinstance(row, dict)]


def _fetch_universo_live_delegations():
    payload = _universo_api_post('competition/get-delegations')
    return [row for row in (payload.get('delegaciones') or []) if isinstance(row, dict)]


def _fetch_universo_live_competitions(delegation_id, season_id):
    payload = _universo_api_post(
        'competition/get-competitions',
        {'id_delegacion': str(delegation_id or '').strip(), 'id_season': str(season_id or '').strip()},
    )
    return [row for row in (payload.get('competiciones') or []) if isinstance(row, dict)]


def _fetch_universo_live_groups(competition_id):
    payload = _universo_api_post(
        'competition/get-groups',
        {'id_competition': str(competition_id or '').strip()},
    )
    return [row for row in (payload.get('grupos') or []) if isinstance(row, dict)]


def _fetch_universo_live_classification(group_id):
    payload = _universo_api_post(
        'competition/get-classification',
        {'id_group': str(group_id or '').strip()},
    )
    return payload if isinstance(payload, dict) else {}


def _fetch_universo_live_results(group_id, round_id=''):
    payload = _universo_api_post(
        'match/get-results',
        {'id_group': str(group_id or '').strip(), 'id_round': str(round_id or '').strip()},
    )
    return payload if isinstance(payload, dict) else {}


def _parse_capture_form_payload(raw_payload):
    parsed = {}
    raw_text = str(raw_payload or '')
    if not raw_text:
        return parsed
    for match in re.finditer(r'name="([^"]+)"\r\n\r\n(.*?)\r\n(?:--|Content-Disposition: form-data; name=)', raw_text, re.S):
        parsed[str(match.group(1) or '').strip()] = str(match.group(2) or '').strip()
    return parsed


def _derive_season_label_from_dates(start_value, end_value):
    start_date = parse_date(str(start_value or '').strip())
    end_date = parse_date(str(end_value or '').strip())
    if start_date and end_date:
        return f'{start_date.year}/{end_date.year}'
    if start_date:
        return f'{start_date.year}/{start_date.year + 1}'
    if end_date:
        return f'{end_date.year - 1}/{end_date.year}'
    today = timezone.localdate()
    if today.month >= 7:
        return f'{today.year}/{today.year + 1}'
    return f'{today.year - 1}/{today.year}'


def _extract_region_from_competition_name(name):
    text = str(name or '').strip()
    match = re.search(r'\(([^)]+)\)\s*$', text)
    return str(match.group(1) or '').strip() if match else ''


def _build_universo_competition_catalog():
    payload = load_universo_capture()
    items = payload.get('items') if isinstance(payload, dict) else None
    if not isinstance(items, list):
        return {
            'competitions': {},
            'groups': {},
            'classifications': {},
        }
    competitions = {}
    groups = {}
    classifications = {}
    for item in items:
        if not isinstance(item, dict):
            continue
        url = str(item.get('url') or '').strip()
        data = item.get('json')
        if not isinstance(data, dict):
            continue
        post_data = _parse_capture_form_payload(item.get('request_post_data'))
        if 'competition/get-competitions' in url:
            for row in data.get('competiciones') or []:
                if not isinstance(row, dict):
                    continue
                code = str(row.get('codigo') or '').strip()
                if not code:
                    continue
                competitions[code] = {
                    'code': code,
                    'name': str(row.get('nombre') or '').strip(),
                    'category_name': str(row.get('NombreCategoria') or '').strip(),
                    'game_type': str(row.get('TipoJuego') or '').strip(),
                    'season_id': str(post_data.get('id_season') or '').strip(),
                    'start_date': str(row.get('FechaInicio') or '').strip(),
                    'end_date': str(row.get('FechaFin') or '').strip(),
                    'season_name': _derive_season_label_from_dates(row.get('FechaInicio'), row.get('FechaFin')),
                    'region': _extract_region_from_competition_name(row.get('nombre')),
                }
        elif 'competition/get-groups' in url:
            competition_code = str(post_data.get('id_competition') or '').strip()
            if not competition_code:
                continue
            for row in data.get('grupos') or []:
                if not isinstance(row, dict):
                    continue
                group_code = str(row.get('codigo') or '').strip()
                if not group_code:
                    continue
                groups[(competition_code, group_code)] = {
                    'competition_code': competition_code,
                    'group_code': group_code,
                    'group_name': str(row.get('nombre') or '').strip(),
                    'total_rounds': str(row.get('total_jornadas') or '').strip(),
                    'total_teams': str(row.get('total_equipos') or '').strip(),
                }
        elif 'competition/get-classification' in url:
            competition_code = str(data.get('codigo_competicion') or '').strip()
            group_code = str(data.get('codigo_grupo') or '').strip()
            if not competition_code or not group_code:
                continue
            classifications[(competition_code, group_code)] = {
                'competition_code': competition_code,
                'competition_name': str(data.get('competicion') or '').strip(),
                'group_code': group_code,
                'group_name': str(data.get('grupo') or '').strip(),
                'round': str(data.get('jornada') or '').strip(),
                'round_date': str(data.get('fecha_jornada') or '').strip(),
                'rows': [row for row in (data.get('clasificacion') or []) if isinstance(row, dict)],
            }
    return {
        'competitions': competitions,
        'groups': groups,
        'classifications': classifications,
    }


def _search_universo_competition_candidates(*, team_query='', competition_query='', group_query=''):
    team_query = str(team_query or '').strip()
    competition_query = str(competition_query or '').strip()
    group_query = str(group_query or '').strip()
    normalized_team_query = normalize_label(team_query)
    normalized_comp_query = normalize_label(competition_query)
    normalized_group_query = normalize_label(group_query)
    live_candidates = []
    seasons = _fetch_universo_live_seasons()
    season_row = next((row for row in seasons if str(row.get('nombre') or '').strip() == '2025-2026'), None)
    if not season_row and seasons:
        season_row = seasons[0]
    season_id = str((season_row or {}).get('cod_temporada') or '').strip()
    season_name = str((season_row or {}).get('nombre') or '').strip()
    delegations = _fetch_universo_live_delegations()
    if season_id and delegations:
        for delegation in delegations:
            competitions = _fetch_universo_live_competitions(delegation.get('cod_delegacion'), season_id)
            for competition in competitions:
                competition_code = str(competition.get('codigo') or '').strip()
                competition_name = str(competition.get('nombre') or '').strip()
                if not competition_code or not competition_name:
                    continue
                if normalized_comp_query and normalized_comp_query not in normalize_label(competition_name):
                    continue
                groups = _fetch_universo_live_groups(competition_code)
                for group in groups:
                    group_code = str(group.get('codigo') or '').strip()
                    group_name = str(group.get('nombre') or '').strip()
                    if not group_code or not group_name:
                        continue
                    if normalized_group_query and normalized_group_query not in normalize_label(group_name):
                        continue
                    classification = _fetch_universo_live_classification(group_code)
                    for row in classification.get('clasificacion') or []:
                        if not isinstance(row, dict):
                            continue
                        team_name = str(row.get('nombre') or '').strip()
                        if not team_name:
                            continue
                        normalized_team_name = normalize_label(team_name)
                        if normalized_team_query and normalized_team_query not in normalized_team_name:
                            continue
                        score = 0
                        if normalized_team_query:
                            score += 60 if normalized_team_query == normalized_team_name else 30
                        if normalized_group_query:
                            score += 25 if normalized_group_query == normalize_label(group_name) else 12
                        if normalized_comp_query:
                            score += 25 if normalized_comp_query == normalize_label(competition_name) else 12
                        external_team_key = str(row.get('codequipo') or '').strip()
                        existing_team = Team.objects.filter(external_id=external_team_key).first() if external_team_key else None
                        live_candidates.append(
                            {
                                'source': 'universo_live',
                                'source_label': 'Universo RFAF · live',
                                'team_id': existing_team.id if existing_team else None,
                                'team_name': team_name,
                                'group_name': group_name or 'Sin grupo',
                                'season_name': season_name,
                                'competition_name': competition_name or 'Competición sin nombre',
                                'provider': WorkspaceCompetitionContext.PROVIDER_UNIVERSO,
                                'external_competition_key': competition_code,
                                'external_group_key': group_code,
                                'external_team_key': external_team_key,
                                'external_team_name': team_name,
                                'score': score + 20,
                                'is_import_candidate': existing_team is None,
                            }
                        )
    if live_candidates:
        live_candidates.sort(key=lambda item: (-item['score'], item['competition_name'], item['group_name'], item['team_name']))
        return live_candidates[:12]

    catalog = _build_universo_competition_catalog()
    competitions = catalog.get('competitions') or {}
    groups = catalog.get('groups') or {}
    classifications = catalog.get('classifications') or {}
    candidates = []
    for (competition_code, group_code), classification in classifications.items():
        competition_meta = competitions.get(competition_code) or {}
        group_meta = groups.get((competition_code, group_code)) or {}
        competition_name = str(classification.get('competition_name') or competition_meta.get('name') or '').strip()
        group_name = str(classification.get('group_name') or group_meta.get('group_name') or '').strip()
        if normalized_comp_query and normalized_comp_query not in normalize_label(competition_name):
            continue
        if normalized_group_query and normalized_group_query not in normalize_label(group_name):
            continue
        for row in classification.get('rows') or []:
            team_name = str(row.get('nombre') or '').strip()
            if not team_name:
                continue
            normalized_team_name = normalize_label(team_name)
            if normalized_team_query and normalized_team_query not in normalized_team_name:
                continue
            score = 0
            if normalized_team_query:
                score += 60 if normalized_team_query == normalized_team_name else 30
            if normalized_group_query:
                score += 25 if normalized_group_query == normalize_label(group_name) else 12
            if normalized_comp_query:
                score += 25 if normalized_comp_query == normalize_label(competition_name) else 12
            existing_team = Team.objects.filter(external_id=str(row.get('codequipo') or '').strip()).first()
            candidates.append(
                {
                    'source': 'universo_capture',
                    'source_label': 'Universo RFAF · captura local',
                    'team_id': existing_team.id if existing_team else None,
                    'team_name': team_name,
                    'group_name': group_name or 'Sin grupo',
                    'season_name': str(competition_meta.get('season_name') or '').strip(),
                    'competition_name': competition_name or 'Competición sin nombre',
                    'provider': WorkspaceCompetitionContext.PROVIDER_UNIVERSO,
                    'external_competition_key': competition_code,
                    'external_group_key': group_code,
                    'external_team_key': str(row.get('codequipo') or '').strip(),
                    'external_team_name': team_name,
                    'score': score,
                    'is_import_candidate': existing_team is None,
                }
            )
    candidates.sort(key=lambda item: (-item['score'], item['competition_name'], item['group_name'], item['team_name']))
    return candidates[:12]


def _import_universo_competition_candidate(*, competition_key, group_key, team_key, team_name):
    catalog = _build_universo_competition_catalog()
    competitions = catalog.get('competitions') or {}
    classifications = catalog.get('classifications') or {}
    classification = classifications.get((competition_key, group_key))
    competition_meta = competitions.get(competition_key) or {}
    if not classification:
        live_classification = _fetch_universo_live_classification(group_key)
        if isinstance(live_classification, dict) and live_classification.get('clasificacion'):
            live_seasons = _fetch_universo_live_seasons()
            season_row = next((row for row in live_seasons if str(row.get('nombre') or '').strip() == '2025-2026'), None)
            if not season_row and live_seasons:
                season_row = live_seasons[0]
            competition_meta = {
                'code': competition_key,
                'name': str(live_classification.get('competicion') or '').strip(),
                'season_name': str((season_row or {}).get('nombre') or '').strip() or _derive_season_label_from_dates('', ''),
                'start_date': str((season_row or {}).get('fecha_inicio') or '').strip(),
                'end_date': str((season_row or {}).get('fecha_fin') or '').strip(),
                'region': _extract_region_from_competition_name(live_classification.get('competicion')),
            }
            classification = {
                'competition_code': competition_key,
                'competition_name': str(live_classification.get('competicion') or '').strip(),
                'group_code': group_key,
                'group_name': str(live_classification.get('grupo') or '').strip(),
                'round': str(live_classification.get('jornada') or '').strip(),
                'round_date': str(live_classification.get('fecha_jornada') or '').strip(),
                'rows': [row for row in (live_classification.get('clasificacion') or []) if isinstance(row, dict)],
            }
    if not classification:
        return None, 'La captura local no contiene la clasificación de esa competición/grupo.'
    competition_name = str(classification.get('competition_name') or competition_meta.get('name') or '').strip()
    if not competition_name:
        return None, 'No se ha podido resolver el nombre de la competición.'
    region = str(competition_meta.get('region') or _extract_region_from_competition_name(competition_name) or '').strip()
    competition_slug = slugify(f'universo-{competition_key}-{competition_name}')[:150] or f'universo-{competition_key}'
    competition, _ = Competition.objects.get_or_create(
        slug=competition_slug,
        defaults={
            'name': competition_name,
            'region': region,
            'description': 'Importada desde captura local de Universo RFAF.',
        },
    )
    if competition.name != competition_name or competition.region != region:
        competition.name = competition_name
        competition.region = region
        competition.save(update_fields=['name', 'region'])
    season_name = str(competition_meta.get('season_name') or _derive_season_label_from_dates(competition_meta.get('start_date'), competition_meta.get('end_date'))).strip()
    season_defaults = {'is_current': True}
    start_date = parse_date(str(competition_meta.get('start_date') or '').strip())
    end_date = parse_date(str(competition_meta.get('end_date') or '').strip())
    if start_date:
        season_defaults['start_date'] = start_date
    if end_date:
        season_defaults['end_date'] = end_date
    season, _ = Season.objects.get_or_create(
        competition=competition,
        name=season_name,
        defaults=season_defaults,
    )
    season_changed = False
    if start_date and season.start_date != start_date:
        season.start_date = start_date
        season_changed = True
    if end_date and season.end_date != end_date:
        season.end_date = end_date
        season_changed = True
    if not season.is_current:
        season.is_current = True
        season_changed = True
    if season_changed:
        season.save(update_fields=['start_date', 'end_date', 'is_current'])
    group_name = str(classification.get('group_name') or '').strip() or 'Grupo importado'
    group_slug = slugify(f'{group_name}-{group_key}')[:80] or f'grupo-{group_key}'
    group, _ = Group.objects.get_or_create(
        season=season,
        slug=group_slug,
        defaults={
            'name': group_name,
            'external_id': group_key,
        },
    )
    if group.name != group_name or group.external_id != group_key:
        group.name = group_name
        group.external_id = group_key
        group.save(update_fields=['name', 'external_id'])
    selected_team = None
    for row in classification.get('rows') or []:
        row_team_name = str(row.get('nombre') or '').strip()
        row_team_key = str(row.get('codequipo') or '').strip()
        row_team_crest = _absolute_universo_url(row.get('escudo_equipo') or row.get('escudo'))
        if not row_team_name:
            continue
        team = Team.objects.filter(external_id=row_team_key).first() if row_team_key else None
        if not team:
            team = Team.objects.filter(group=group, name__iexact=row_team_name).first()
        if not team:
            team = Team.objects.create(
                name=row_team_name,
                slug=_unique_team_slug(row_team_name),
                short_name=row_team_name[:60],
                group=group,
                external_id=row_team_key,
                is_primary=False,
            )
        changed_fields = []
        if team.group_id != group.id:
            team.group = group
            changed_fields.append('group')
        if row_team_key and team.external_id != row_team_key:
            team.external_id = row_team_key
            changed_fields.append('external_id')
        short_name = row_team_name[:60]
        if team.short_name != short_name:
            team.short_name = short_name
            changed_fields.append('short_name')
        if row_team_crest and team.crest_url != row_team_crest and not getattr(team, 'crest_image', None):
            team.crest_url = row_team_crest
            changed_fields.append('crest_url')
        if changed_fields:
            team.save(update_fields=changed_fields)
        TeamStanding.objects.update_or_create(
            season=season,
            group=group,
            team=team,
            defaults={
                'position': _parse_int(str(row.get('posicion') or '0')) or 0,
                'played': _parse_int(str(row.get('jugados') or '0')),
                'wins': _parse_int(str(row.get('ganados') or '0')),
                'draws': _parse_int(str(row.get('empatados') or '0')),
                'losses': _parse_int(str(row.get('perdidos') or '0')),
                'goals_for': _parse_int(str(row.get('goles_a_favor') or '0')),
                'goals_against': _parse_int(str(row.get('goles_en_contra') or '0')),
                'goal_difference': _parse_int(str(row.get('goles_a_favor') or '0')) - _parse_int(str(row.get('goles_en_contra') or '0')),
                'points': _parse_int(str(row.get('puntos') or '0')),
                'last_updated': timezone.now(),
            },
        )
        if row_team_key and row_team_key == str(team_key or '').strip():
            selected_team = team
        elif not selected_team and normalize_label(row_team_name) == normalize_label(team_name):
            selected_team = team
    if not selected_team:
        return None, 'No se ha podido identificar el equipo seleccionado en la clasificación importada.'

    snapshot = load_universo_snapshot()
    supports_snapshot = _universo_snapshot_supports_team(snapshot, selected_team)
    next_payload = normalize_next_match_payload(snapshot.get('next_match') or {}) if isinstance(snapshot, dict) and supports_snapshot and snapshot.get('next_match') else {}
    if next_payload:
        opponent_name = _payload_opponent_name(next_payload) or 'Rival por confirmar'
        opponent_key = _normalize_team_lookup_key(opponent_name)
        opponent = Team.objects.filter(group=group, external_id__iexact=opponent_key).first()
        if not opponent:
            opponent = Team.objects.filter(group=group, name__iexact=opponent_name).first()
        if not opponent:
            opponent = Team.objects.create(
                name=opponent_name,
                slug=_unique_team_slug(opponent_name),
                short_name=opponent_name[:60],
                group=group,
                is_primary=False,
            )
        opponent_crest = ''
        opponent_payload = next_payload.get('opponent')
        if isinstance(opponent_payload, dict):
            opponent_crest = _absolute_universo_url(opponent_payload.get('crest_url'))
        if opponent_crest and opponent.crest_url != opponent_crest and not getattr(opponent, 'crest_image', None):
            opponent.crest_url = opponent_crest
            opponent.save(update_fields=['crest_url'])
        match_date = parse_date(str(next_payload.get('date') or '').strip())
        round_label = str(next_payload.get('round') or '').strip() or str(next_payload.get('round_label') or '').strip()
        location = str(next_payload.get('location') or '').strip()
        home = bool(next_payload.get('home'))
        home_team = selected_team if home else opponent
        away_team = opponent if home else selected_team
        match, _ = Match.objects.get_or_create(
            season=season,
            group=group,
            round=round_label,
            home_team=home_team,
            away_team=away_team,
            defaults={
                'date': match_date,
                'location': location,
                'notes': 'Partido importado desde snapshot local de Universo RFAF.',
            },
        )
        update_fields = []
        if match_date and match.date != match_date:
            match.date = match_date
            update_fields.append('date')
        if location and match.location != location:
            match.location = location
            update_fields.append('location')
        if update_fields:
            match.save(update_fields=update_fields)
    return selected_team, ''


def _universo_snapshot_supports_team(snapshot, primary_team):
    if not primary_team:
        return True
    if not isinstance(snapshot, dict):
        return False
    rows = snapshot.get('standings')
    candidate_keys = {
        _normalize_team_lookup_key(primary_team.name),
        _normalize_team_lookup_key(primary_team.display_name),
    }
    candidate_keys = {key for key in candidate_keys if key}
    if isinstance(rows, list) and rows:
        for row in rows:
            if not isinstance(row, dict):
                continue
            row_keys = {
                _normalize_team_lookup_key(row.get('team')),
                _normalize_team_lookup_key(row.get('full_name')),
            }
            row_keys = {key for key in row_keys if key}
            if candidate_keys & row_keys:
                return True
        return False
    return bool(getattr(primary_team, 'is_primary', False))


def _resolve_standings_for_team(primary_team, snapshot=None, provider=None):
    if not primary_team or not getattr(primary_team, 'group', None):
        return []
    snapshot = snapshot if snapshot is not None else load_universo_snapshot()
    provider_key = str(provider or '').strip().lower()
    if provider_key == WorkspaceCompetitionContext.PROVIDER_UNIVERSO:
        if _universo_snapshot_supports_team(snapshot, primary_team):
            universo_rows = _serialize_universo_standings(snapshot)
            if universo_rows:
                return universo_rows
        return serialize_standings(primary_team.group)
    # Manual/RFAF: prioriza siempre la BD (lo que importa el script de federación).
    group_for_db = _latest_standings_group_for_team(primary_team) or primary_team.group
    db_rows = serialize_standings(group_for_db)
    if db_rows:
        return db_rows
    if _universo_snapshot_supports_team(snapshot, primary_team):
        universo_rows = _serialize_universo_standings(snapshot)
        if universo_rows:
            return universo_rows
    return serialize_standings(primary_team.group)


def _resolve_rival_identity(rival_name, preferred_opponent=None):
    rival_name = str(rival_name or '').strip() or 'Rival por confirmar'
    rival_full_name = rival_name
    rival_crest_url = ''
    rival_key = _normalize_team_lookup_key(rival_name)
    known_team = (
        Team.objects
        .filter(Q(name__iexact=rival_name) | Q(short_name__iexact=rival_name) | Q(external_id__iexact=rival_name))
        .order_by('-is_primary', 'name')
        .first()
    )
    if known_team:
        rival_full_name = known_team.name
        rival_crest_url = resolve_team_crest_url(None, known_team, fallback_static='', sync=False)

    if isinstance(preferred_opponent, dict):
        preferred_name = str(preferred_opponent.get('name') or '').strip()
        preferred_full_name = str(preferred_opponent.get('full_name') or '').strip()
        preferred_key = _normalize_team_lookup_key(preferred_name or preferred_full_name)
        if preferred_key and rival_key and (preferred_key == rival_key or preferred_key in rival_key or rival_key in preferred_key):
            rival_full_name = preferred_full_name or preferred_name or rival_full_name
            rival_crest_url = str(preferred_opponent.get('crest_url') or '').strip()

    standings_lookup = _build_universo_standings_lookup(load_universo_snapshot())
    capture_lookup = _build_universo_capture_team_lookup()

    best_meta = {}
    candidates = [standings_lookup.get(rival_key, {}), capture_lookup.get(rival_key, {})]
    for source in (standings_lookup, capture_lookup):
        if best_meta.get('full_name') and best_meta.get('crest_url'):
            break
        for key, meta in source.items():
            if not key or not rival_key:
                continue
            if rival_key in key or key in rival_key:
                if len(str(meta.get('full_name') or '')) > len(str(best_meta.get('full_name') or '')):
                    best_meta['full_name'] = str(meta.get('full_name') or '').strip()
                if meta.get('crest_url') and not best_meta.get('crest_url'):
                    best_meta['crest_url'] = str(meta.get('crest_url') or '').strip()
    for meta in candidates:
        if not isinstance(meta, dict):
            continue
        if len(str(meta.get('full_name') or '')) > len(str(best_meta.get('full_name') or '')):
            best_meta['full_name'] = str(meta.get('full_name') or '').strip()
        if meta.get('crest_url') and not best_meta.get('crest_url'):
            best_meta['crest_url'] = str(meta.get('crest_url') or '').strip()

    rival_full_name = best_meta.get('full_name') or rival_full_name
    rival_crest_url = best_meta.get('crest_url') or rival_crest_url
    rival_crest_url = _sanitize_universo_external_image(_absolute_universo_url(rival_crest_url))
    return rival_full_name, rival_crest_url


def load_preferred_next_match_payload(primary_team=None, competition_context=None):
    competition_context = competition_context or (
        WorkspaceCompetitionContext.objects
        .filter(Q(team=primary_team) | Q(workspace__primary_team=primary_team))
        .select_related('workspace', 'team', 'group')
        .first()
        if primary_team else None
    )
    competition_context = _ensure_universo_context_binding(competition_context, primary_team)
    provider_next = _find_universo_next_match_for_context(competition_context, primary_team)
    if _next_match_payload_is_reliable(provider_next):
        return provider_next
    try:
        ws = getattr(competition_context, 'workspace', None)
        if ws and getattr(ws, 'id', None):
            snapshot = WorkspaceCompetitionSnapshot.objects.filter(workspace=ws).first()
            if snapshot and isinstance(snapshot.next_match_payload, dict):
                snapshot_next = normalize_next_match_payload(dict(snapshot.next_match_payload))
                if _next_match_payload_is_reliable(snapshot_next):
                    return snapshot_next
    except Exception:
        pass
    snapshot = load_universo_snapshot()
    can_use_external = _universo_snapshot_supports_team(snapshot, primary_team) if primary_team else True
    if can_use_external and isinstance(snapshot, dict) and isinstance(snapshot.get('next_match'), dict):
        snapshot_next = normalize_next_match_payload(snapshot.get('next_match'))
        if _next_match_payload_is_reliable(snapshot_next):
            return snapshot_next

    cached_next = load_cached_next_match() if can_use_external else None
    if isinstance(cached_next, dict):
        normalized_cached_next = normalize_next_match_payload(cached_next)
        if _next_match_payload_is_reliable(normalized_cached_next):
            return normalized_cached_next
    return None


def _build_next_match_from_convocation(primary_team):
    record = get_current_convocation_record(primary_team)
    if not record:
        return None
    today = timezone.localdate()
    match_date = record.match_date
    if not match_date and record.match and getattr(record.match, 'date', None):
        match_date = record.match.date
    if not match_date:
        return None
    if match_date < today:
        return None

    opponent_name = (record.opponent_name or '').strip()
    round_label = (record.round or '').strip()
    location_label = (record.location or '').strip()
    date_iso = match_date.isoformat() if match_date else None
    time_label = record.match_time.strftime('%H:%M') if record.match_time else ''

    if not any([opponent_name, round_label, date_iso, time_label, location_label]):
        return None

    home_flag = None
    match = record.match
    if match and primary_team:
        if match.home_team_id == primary_team.id:
            home_flag = True
        elif match.away_team_id == primary_team.id:
            home_flag = False

    rival_full_name, rival_crest_url = _resolve_rival_identity(opponent_name or 'Rival por confirmar')

    payload = {
        'round': round_label or 'Jornada por confirmar',
        'date': date_iso,
        'time': time_label,
        'location': location_label or 'Campo por confirmar',
        'opponent': {
            'name': opponent_name or rival_full_name or 'Rival por confirmar',
            'full_name': rival_full_name or opponent_name or 'Rival por confirmar',
            'crest_url': _absolute_universo_url(rival_crest_url),
            'team_code': '',
        },
        'home': home_flag if home_flag is not None else True,
        'status': 'next',
        'source': 'convocation-manual',
    }
    return normalize_next_match_payload(payload)


def _build_coach_rival_summary(primary_team):
    next_match_payload = load_preferred_next_match_payload(primary_team=primary_team)
    if not next_match_payload and primary_team and primary_team.group:
        next_match_payload = get_next_match(primary_team, primary_team.group)
    if not _next_match_payload_is_reliable(next_match_payload):
        convocation_next = _build_next_match_from_convocation(primary_team)
        if _next_match_payload_is_reliable(convocation_next):
            next_match_payload = convocation_next
    opponent_name = _payload_opponent_name(next_match_payload).strip() if isinstance(next_match_payload, dict) else ''

    reports_qs = RivalAnalysisReport.objects.filter(team=primary_team).select_related('rival_team').order_by('-updated_at', '-id')
    report = None
    if opponent_name:
        folded = opponent_name.lower()
        report = reports_qs.filter(
            Q(rival_name__icontains=opponent_name)
            | Q(report_title__icontains=opponent_name)
            | Q(rival_team__name__icontains=opponent_name)
        ).first()
        if not report:
            for candidate in reports_qs[:8]:
                name_pool = ' '.join(
                    [
                        str(getattr(candidate, 'rival_name', '') or ''),
                        str(getattr(getattr(candidate, 'rival_team', None), 'name', '') or ''),
                        str(getattr(candidate, 'report_title', '') or ''),
                    ]
                ).lower()
                if folded and folded in name_pool:
                    report = candidate
                    break
    if not report:
        report = reports_qs.first()

    lines = []
    if opponent_name:
        lines.append({'label': 'Próximo rival', 'value': opponent_name})
    elif report and report.rival_name:
        lines.append({'label': 'Rival analizado', 'value': report.rival_name})
    else:
        lines.append({'label': 'Rival', 'value': 'Pendiente de confirmar'})

    if report:
        if report.report_title:
            lines.append({'label': 'Informe', 'value': report.report_title})
        if report.weaknesses:
            lines.append({'label': 'Debilidad detectada', 'value': report.weaknesses[:140]})
        elif report.opportunities:
            lines.append({'label': 'Oportunidad', 'value': report.opportunities[:140]})
        elif report.key_players:
            lines.append({'label': 'Foco individual', 'value': report.key_players[:140]})
        lines.append({'label': 'Estado', 'value': report.get_status_display()})
    else:
        lines.append({'label': 'Informe', 'value': 'Sin informe manual cargado'})
        lines.append({'label': 'Acción', 'value': 'Conviene preparar el análisis rival antes de la sesión táctica'})

    return {
        'opponent': opponent_name,
        'report': report,
        'lines': lines[:4],
    }


def _next_match_payload_is_reliable(payload):
    if not isinstance(payload, dict):
        return False
    status = str(payload.get('status') or '').strip().lower()
    if status != 'next':
        return False
    source = str(payload.get('source') or '').strip().lower()
    opponent_name = _payload_opponent_name(payload).strip().lower()
    if not opponent_name or opponent_name in {'rival por confirmar', 'rival desconocido'}:
        return False
    payload_date = _parse_payload_date(payload.get('date'))
    if payload_date and payload_date < timezone.localdate():
        return False
    if not payload_date and source in {'', 'local-match'}:
        return False
    return True


def _dashboard_cache_key(team_id):
    return f'{DASHBOARD_CACHE_KEY_PREFIX}:{team_id}'

def _team_metrics_cache_key(team_id):
    return f'football:team_metrics:{team_id}'

def _player_metrics_cache_key(team_id):
    return f'football:player_metrics:{team_id}'


def _player_dashboard_cache_key(team_id):
    return f'{PLAYER_DASHBOARD_CACHE_KEY_PREFIX}:{team_id}'


def _invalidate_team_dashboard_caches(primary_team):
    if not primary_team or not getattr(primary_team, 'id', None):
        return
    cache.delete_many(
        [
            _dashboard_cache_key(primary_team.id),
            _player_dashboard_cache_key(primary_team.id),
            _team_metrics_cache_key(primary_team.id),
            _player_metrics_cache_key(primary_team.id),
        ]
    )


def _upsert_match_from_next_match_payload(primary_team, payload):
    if not primary_team or not _next_match_payload_is_reliable(payload):
        return None
    group = getattr(primary_team, 'group', None)
    season = getattr(group, 'season', None) if group else None
    if not group or not season:
        return None
    match_date = _parse_payload_date(payload.get('date'))
    if not match_date:
        return None
    opponent_name = _payload_opponent_name(payload).strip()
    if not opponent_name or opponent_name.lower() in {'rival por confirmar', 'rival desconocido'}:
        return None
    opponent = (
        Team.objects
        .filter(group=group)
        .filter(Q(name__iexact=opponent_name) | Q(short_name__iexact=opponent_name) | Q(slug__iexact=slugify(opponent_name)))
        .order_by('-is_primary', 'name')
        .first()
    ) or (
        Team.objects
        .filter(Q(name__iexact=opponent_name) | Q(short_name__iexact=opponent_name) | Q(slug__iexact=slugify(opponent_name)))
        .order_by('-is_primary', 'name')
        .first()
    )
    if not opponent:
        opponent = Team.objects.create(
            name=opponent_name,
            slug=_unique_team_slug(opponent_name),
            short_name=opponent_name[:60],
            group=group,
        )
    is_home = payload.get('home')
    if is_home is None:
        is_home = True
    is_home = bool(is_home)
    home_team = primary_team if is_home else opponent
    away_team = opponent if is_home else primary_team
    kickoff = _parse_payload_time(payload.get('time'))
    round_label = str(payload.get('round') or '').strip()
    location = str(payload.get('location') or '').strip()

    match_qs = (
        Match.objects
        .filter(season=season)
        .filter(Q(home_team=primary_team) | Q(away_team=primary_team))
        .filter(date=match_date)
    )
    match = match_qs.order_by('-id').first()
    if not match:
        return Match.objects.create(
            season=season,
            group=group,
            round=round_label,
            date=match_date,
            kickoff_time=kickoff,
            location=location,
            home_team=home_team,
            away_team=away_team,
            source='',
        )
    update_fields = []
    if match.group_id != getattr(group, 'id', None):
        match.group = group
        update_fields.append('group')
    if round_label and match.round != round_label:
        match.round = round_label
        update_fields.append('round')
    if location and match.location != location:
        match.location = location
        update_fields.append('location')
    if kickoff and match.kickoff_time != kickoff:
        match.kickoff_time = kickoff
        update_fields.append('kickoff_time')
    if match.home_team_id != getattr(home_team, 'id', None):
        match.home_team = home_team
        update_fields.append('home_team')
    if match.away_team_id != getattr(away_team, 'id', None):
        match.away_team = away_team
        update_fields.append('away_team')
    if update_fields:
        match.save(update_fields=update_fields)
    return match


def load_universo_snapshot():
    if not UNIVERSO_SNAPSHOT_PATH.exists():
        return None
    try:
        mtime = UNIVERSO_SNAPSHOT_PATH.stat().st_mtime
    except Exception:
        mtime = None
    memo = getattr(load_universo_snapshot, '_memo', None)
    if isinstance(memo, dict) and memo.get('mtime') and mtime and float(memo.get('mtime')) == float(mtime):
        cached_payload = memo.get('payload')
        if isinstance(cached_payload, dict):
            return cached_payload
    try:
        with UNIVERSO_SNAPSHOT_PATH.open(encoding='utf-8') as handle:
            payload = json.load(handle)
            if not isinstance(payload, dict):
                return None
            try:
                load_universo_snapshot._memo = {'mtime': mtime, 'payload': payload}
            except Exception:
                pass
            return payload
    except Exception:
        return None


def _safe_int(value, default=0):
    try:
        return int(str(value).strip())
    except Exception:
        return default


def _serialize_universo_standings(snapshot):
    if not isinstance(snapshot, dict):
        return []
    rows = snapshot.get('standings')
    if not isinstance(rows, list):
        return []
    crest_lookup = _build_team_crest_lookup()
    normalized = []
    for row in rows:
        if not isinstance(row, dict):
            continue
        team = str(row.get('team') or '').strip()
        if not team:
            continue
        gf = _safe_int(row.get('goals_for'))
        ga = _safe_int(row.get('goals_against'))
        gd = row.get('goal_difference')
        if gd in (None, ''):
            gd = gf - ga
        normalized.append(
            {
                'rank': _safe_int(row.get('position'), default=0),
                'team': team,
                'full_name': str(row.get('full_name') or team).strip() or team,
                'crest_url': str(row.get('crest_url') or crest_lookup.get(_normalize_team_lookup_key(team)) or '').strip(),
                'team_code': str(row.get('team_code') or '').strip(),
                'played': _safe_int(row.get('played')),
                'wins': _safe_int(row.get('wins')),
                'draws': _safe_int(row.get('draws')),
                'losses': _safe_int(row.get('losses')),
                'goals_for': gf,
                'goals_against': ga,
                'goal_difference': _safe_int(gd),
                'points': _safe_int(row.get('points')),
            }
        )
    return sorted(normalized, key=lambda x: (x['rank'] <= 0, x['rank'], -x['points'], x['full_name']))


@login_required
def dashboard_data(request):
    """Devuelve los datos principales que alimentarán la home cuerpo técnico/jugador."""
    forbidden = _forbid_if_workspace_module_disabled(request, 'dashboard', label='dashboard')
    if forbidden:
        return forbidden
    primary_team = _get_primary_team_for_request(request)
    if not primary_team:
        return JsonResponse({'error': 'No hay equipo principal configurado'}, status=400)

    group = primary_team.group
    if not group:
        return JsonResponse({'error': 'El equipo principal no está asignado a ningún grupo'}, status=400)

    force_fresh = str(request.GET.get('fresh') or '').strip().lower() in {'1', 'true', 'yes', 'on'}
    cache_key = _dashboard_cache_key(primary_team.id)
    if not force_fresh:
        cached_payload = cache.get(cache_key)
        if isinstance(cached_payload, dict):
            # Autorreparación: si en caché se guardó un próximo rival nulo (p.ej. porque todavía no
            # existía el match en BD o falló una extracción externa), intentamos reconstruirlo sin
            # coste (solo DB) y actualizar la caché.
            try:
                cached_next = cached_payload.get('next_match')
                if not _next_match_payload_is_reliable(cached_next):
                    repaired_next = load_preferred_next_match_payload(primary_team=primary_team) or (
                        get_next_match(primary_team, group) if group else None
                    )
                    if _next_match_payload_is_reliable(repaired_next):
                        cached_payload = dict(cached_payload)
                        cached_payload['next_match'] = repaired_next
                        cache.set(cache_key, cached_payload, DASHBOARD_CACHE_SECONDS)
            except Exception:
                pass
            return JsonResponse(cached_payload)

    refresh_roster_on_load = str(
        os.getenv('PREFERENTE_ROSTER_REFRESH_ON_LOAD', '0')
    ).strip().lower() in {'1', 'true', 'yes', 'on'}
    if refresh_roster_on_load:
        try:
            refresh_primary_roster_cache(primary_team, force=False)
        except Exception:
            pass

    workspace = _get_active_workspace(request)
    competition_payload = _competition_payload_for_team(workspace, primary_team)
    provider_key = str(getattr(getattr(workspace, 'competition_context', None), 'provider', '') or '').strip().lower()
    standings_group = _latest_standings_group_for_team(primary_team) or primary_team.group
    standings_last_updated = _team_standings_last_updated(standings_group)
    standings = _enrich_standings_rows_with_crests(competition_payload.get('standings') or [])
    next_match = _enrich_next_match_payload_with_crests(competition_payload.get('next_match') or {}, standings) or {}
    if not next_match:
        next_match = _enrich_next_match_payload_with_crests(
            load_preferred_next_match_payload(primary_team=primary_team) or get_next_match(primary_team, group),
            standings,
        )
    # Si detectamos un próximo partido fiable, lo persistimos como Match para que nunca dependa de
    # ficheros locales ni de providers externos (Render / múltiples instancias).
    if _next_match_payload_is_reliable(next_match):
        try:
            _upsert_match_from_next_match_payload(primary_team, next_match)
        except Exception:
            pass
    convocation_next_match = _build_next_match_from_convocation(primary_team)
    # Product rule: Home prioritizes Convocatoria only when it provides a reliable scheduled match.
    if _next_match_payload_is_reliable(convocation_next_match):
        next_match = _enrich_next_match_payload_with_crests(convocation_next_match, standings)
    # Evitar mostrar el "último partido" como si fuera el próximo rival.
    if not _next_match_payload_is_reliable(next_match):
        next_match = None
    team_metrics = compute_team_metrics(primary_team)
    player_metrics = compute_player_metrics(primary_team)
    player_cards = compute_player_cards(primary_team)
    player_cards_scope = {'type': 'global', 'label': 'Jugador · datos La Preferente'}
    competition_name = ''
    season_name = ''
    try:
        competition_name = str(getattr(getattr(getattr(group, 'season', None), 'competition', None), 'name', '') or '').strip()
        season_name = str(getattr(getattr(group, 'season', None), 'name', '') or '').strip()
    except Exception:
        competition_name = ''
        season_name = ''
    group_label = group.name
    if season_name and season_name.lower() not in str(group_label or '').lower():
        group_label = f'{season_name} · {group_label}'
    elif competition_name and competition_name.lower() not in str(group_label or '').lower():
        group_label = f'{competition_name} · {group_label}'
    corporate_parts = [competition_name, season_name, group.name]
    corporate_line = ' · '.join([part for part in corporate_parts if str(part or '').strip()])

    payload = {
        'team': {
            'name': primary_team.name,
            'group': group_label,
            'competition': competition_name,
            'season': season_name,
            'corporate_line': corporate_line,
            'crest_url': resolve_team_crest_url(request, primary_team, sync=True),
        },
        'standings': standings,
        'next_match': next_match,
        'standings_meta': {
            'provider': provider_key,
            'last_updated': standings_last_updated.isoformat() if standings_last_updated else '',
            'group': str(getattr(standings_group, 'name', '') or ''),
            'season': str(getattr(getattr(standings_group, 'season', None), 'name', '') or ''),
        },
        'team_metrics': team_metrics,
        'player_metrics': player_metrics,
        'player_cards': player_cards,
        'player_cards_scope': player_cards_scope,
    }
    cache.set(cache_key, payload, DASHBOARD_CACHE_SECONDS)
    response = JsonResponse(payload)
    if force_fresh:
        response['Cache-Control'] = 'no-store'
    return response


@login_required
@ensure_csrf_cookie
def dashboard_page(request):
    current_role = _get_user_role(request.user) or AppUserRole.ROLE_PLAYER
    if current_role == AppUserRole.ROLE_PLAYER:
        primary_team = _get_primary_team_for_request(request)
        current_player = _resolve_player_for_user(request.user, primary_team)
        if current_player:
            return redirect('player-detail', player_id=current_player.id)
        return redirect('player-dashboard')
    active_workspace_obj = _get_active_workspace(request)
    if current_role in {AppUserRole.ROLE_TASK_STUDIO, AppUserRole.ROLE_GUEST} and _can_access_task_studio(request.user):
        if active_workspace_obj and active_workspace_obj.kind == Workspace.KIND_CLUB and _has_club_workspace_access(request.user):
            pass
        else:
            return redirect('task-studio-home')
    if active_workspace_obj and not _can_access_platform(request.user):
        target_url = _workspace_entry_url(active_workspace_obj, user=request.user)
        target_path = str(target_url or '').split('?', 1)[0]
        if target_path != request.path:
            return redirect(target_url)

    sources = list(ScrapeSource.objects.filter(is_active=True))
    active_items = list(HomeCarouselImage.objects.filter(is_active=True).order_by('order', '-created_at', '-id'))
    all_items = list(HomeCarouselImage.objects.order_by('order', '-created_at', '-id'))
    candidates = active_items if active_items else all_items
    hero_image_candidates = [reverse('home-carousel-image-file', args=[item.id]) for item in candidates if item.image]
    role_labels = dict(AppUserRole.ROLE_CHOICES)
    can_access_admin = _is_admin_user(request.user)
    can_access_sessions = _can_access_sessions_workspace(request.user)
    can_access_platform = _can_access_platform(request.user)
    workspace_links = _workspace_links_for_user(request.user)
    active_workspace = _build_active_workspace_badge(request)
    dashboard_focus_items = []
    dashboard_pending_items = []
    dashboard_pending_cards = []
    dashboard_recent_activity = []
    primary_team = _get_primary_team_for_request(request)
    if primary_team and current_role not in {AppUserRole.ROLE_TASK_STUDIO, AppUserRole.ROLE_GUEST}:
        focus_by_role = {
            AppUserRole.ROLE_COACH: [
                {
                    'title': 'Partido',
                    'description': 'Cierra convocatoria, 11 inicial y lectura prepartido del rival.',
                    'url': reverse('convocation'),
                },
                {
                    'title': 'Entrenamiento',
                    'description': 'Revisa sesiones, tareas y el microciclo activo de la semana.',
                    'url': reverse('sessions'),
                },
                {
                    'title': 'Estadísticas',
                    'description': 'Controla KPIs de plantilla y seguimiento individual.',
                    'url': reverse('coach-role-trainer'),
                },
            ],
            AppUserRole.ROLE_GOALKEEPER: [
                {
                    'title': 'Porteros',
                    'description': 'Planifica tareas y sesiones específicas de portería.',
                    'url': reverse('sessions-goalkeeper'),
                },
                {
                    'title': 'Partido',
                    'description': 'Revisa 11, disponibilidad y situaciones de gol encajado.',
                    'url': reverse('convocation'),
                },
                {
                    'title': 'Análisis',
                    'description': 'Cruza rival y rendimiento de los porteros en contexto competitivo.',
                    'url': reverse('analysis'),
                },
            ],
            AppUserRole.ROLE_FITNESS: [
                {
                    'title': 'Preparación física',
                    'description': 'Organiza la carga semanal y las tareas específicas del área.',
                    'url': reverse('sessions-fitness'),
                },
                {
                    'title': 'Estadísticas',
                    'description': 'Supervisa minutos, disponibilidad y carga competitiva.',
                    'url': reverse('coach-role-trainer'),
                },
                {
                    'title': 'Partido',
                    'description': 'Confirma disponibilidad y estado de convocatoria.',
                    'url': reverse('convocation'),
                },
            ],
            AppUserRole.ROLE_ANALYST: [
                {
                    'title': 'Análisis',
                    'description': 'Mantén rival, informes y support documental al día.',
                    'url': reverse('analysis'),
                },
                {
                    'title': 'Partido',
                    'description': 'Cruza scouting con convocatoria y 11 inicial.',
                    'url': reverse('convocation'),
                },
                {
                    'title': 'Estadísticas',
                    'description': 'Valida KPIs y soporte de lectura táctica.',
                    'url': reverse('coach-role-trainer'),
                },
            ],
        }
        dashboard_focus_items = focus_by_role.get(
            current_role,
            [
                {
                    'title': 'Portada',
                    'description': 'Mantén el cliente con contexto competitivo y estado semanal.',
                    'url': reverse('dashboard-home'),
                },
                {
                    'title': 'Entrenamiento',
                    'description': 'Revisa sesiones, tareas y bibliotecas activas.',
                    'url': reverse('sessions'),
                },
                {
                    'title': 'Análisis',
                    'description': 'Comprueba rival, informes y soporte táctico.',
                    'url': reverse('analysis'),
                },
            ],
        )
        weekly_brief = _build_weekly_staff_brief_context(primary_team)
        dashboard_pending_cards = _build_team_pending_cards(primary_team, weekly_brief)
        if int(weekly_brief.get('convocated_count') or 0) <= 0:
            dashboard_pending_items.append('Falta cerrar la convocatoria actual.')
        if int(weekly_brief.get('probable_eleven_count') or 0) <= 0:
            dashboard_pending_items.append('No hay 11 inicial o 11 probable definido.')
        if int(weekly_brief.get('available_count') or 0) <= 0:
            dashboard_pending_items.append('No hay disponibilidad consolidada del equipo.')
        if not TrainingSession.objects.filter(microcycle__team=primary_team, session_date__gte=timezone.localdate()).exists():
            dashboard_pending_items.append('No hay sesiones futuras planificadas.')
        dashboard_recent_activity = _build_team_recent_activity(primary_team)
    forbidden = _forbid_if_workspace_module_disabled(request, 'dashboard', label='dashboard')
    if forbidden:
        return forbidden
    return render(
        request,
        'football/dashboard.html',
        {
            'scrape_sources': sources,
            'hero_image_candidates': hero_image_candidates,
            'current_role': current_role,
            'current_role_label': role_labels.get(current_role, 'Jugador'),
            'can_access_admin': can_access_admin,
            'can_access_sessions': can_access_sessions,
            'can_access_platform': can_access_platform,
            'workspace_links': workspace_links,
            'active_workspace': active_workspace,
            'dashboard_focus_items': dashboard_focus_items,
            'dashboard_pending_items': dashboard_pending_items,
            'dashboard_pending_cards': dashboard_pending_cards,
            'dashboard_recent_activity': dashboard_recent_activity,
        },
    )


def _split_full_name(value):
    text = str(value or '').strip()
    if not text:
        return '', ''
    parts = [part for part in text.split() if part]
    if len(parts) == 1:
        return parts[0], ''
    return parts[0], ' '.join(parts[1:])


@login_required
def platform_overview_page(request):
    if not _can_access_platform(request.user):
        return HttpResponse('No tienes permisos para acceder a la plataforma.', status=403)
    valid_tabs = {'clients', 'users', 'task-studio', 'documents', 'workspace-create', 'home-global'}
    active_tab = str(request.GET.get('tab') or 'clients').strip().lower()
    if active_tab not in valid_tabs:
        active_tab = 'clients'
    users_subtab = str(request.GET.get('subtab') or 'list').strip().lower()
    if users_subtab not in {'list', 'create'}:
        users_subtab = 'list'

    feedback = str(request.session.pop('platform_feedback', '') or '')
    error = ''
    user_message = ''
    user_error = ''
    invitation_links = []
    carousel_message = ''
    active_workspace = _build_active_workspace_badge(request)
    primary_team = Team.objects.filter(is_primary=True).first()
    workspace_form = {
        'workspace_name': '',
        'workspace_kind': Workspace.KIND_CLUB,
        'owner_username': '',
        'team_id': '',
        'team_new_name': '',
        'workspace_notes': '',
        'competition_provider': WorkspaceCompetitionContext.PROVIDER_MANUAL,
        'external_competition_key': '',
        'external_group_key': '',
        'external_team_key': '',
        'external_team_name': '',
        'competition_auto_sync': True,
        'seed_demo_data': False,
        'initial_admin_usernames': '',
        'initial_member_usernames': '',
        'modules': _workspace_default_modules(Workspace.KIND_CLUB),
        'module_keys': [item['key'] for item in _workspace_module_catalog(Workspace.KIND_CLUB)],
        'deliverable_keys': _workspace_selected_deliverable_keys(
            Workspace.KIND_CLUB,
            _expand_workspace_module_selection(
                Workspace.KIND_CLUB,
                {item['key']: True for item in _workspace_module_catalog(Workspace.KIND_CLUB)},
            ),
        ),
    }
    user_form = {
        'full_name': '',
        'username': '',
        'email': '',
        'role': AppUserRole.ROLE_PLAYER,
    }
    if primary_team:
        _ensure_club_workspace(primary_team)
    studio_users = User.objects.filter(app_role__role__in=[AppUserRole.ROLE_TASK_STUDIO, AppUserRole.ROLE_GUEST]).distinct()
    for studio_user in studio_users:
        _ensure_task_studio_workspace(studio_user)

    if request.method == 'POST':
        form_action = (request.POST.get('form_action') or 'workspace_create').strip().lower()
        if form_action == 'workspace_create':
            active_tab = 'workspace-create'
            workspace_name = _sanitize_task_text((request.POST.get('workspace_name') or '').strip(), multiline=False, max_len=160)
            workspace_kind = str(request.POST.get('workspace_kind') or Workspace.KIND_CLUB).strip()
            owner_username = _sanitize_task_text((request.POST.get('owner_username') or '').strip(), multiline=False, max_len=150).lower()
            team_id = _parse_int(request.POST.get('team_id'))
            team_new_name = _sanitize_task_text((request.POST.get('team_new_name') or '').strip(), multiline=False, max_len=150)
            workspace_notes = _sanitize_task_text((request.POST.get('workspace_notes') or '').strip(), multiline=True, max_len=1200)
            competition_provider = str(request.POST.get('competition_provider') or WorkspaceCompetitionContext.PROVIDER_MANUAL).strip()
            external_competition_key = str(request.POST.get('external_competition_key') or '').strip()[:140]
            external_group_key = str(request.POST.get('external_group_key') or '').strip()[:140]
            external_team_key = str(request.POST.get('external_team_key') or '').strip()[:140]
            external_team_name = _sanitize_task_text((request.POST.get('external_team_name') or '').strip(), multiline=False, max_len=160)
            competition_auto_sync = str(request.POST.get('competition_auto_sync') or '').lower() in {'1', 'true', 'on', 'yes'}
            seed_demo_data = str(request.POST.get('seed_demo_data') or '').lower() in {'1', 'true', 'on', 'yes'}
            initial_admin_usernames = str(request.POST.get('initial_admin_usernames') or '')
            initial_member_usernames = str(request.POST.get('initial_member_usernames') or '')
            if workspace_kind not in {Workspace.KIND_CLUB, Workspace.KIND_TASK_STUDIO}:
                workspace_kind = Workspace.KIND_CLUB
            valid_providers = {choice[0] for choice in WorkspaceCompetitionContext.PROVIDER_CHOICES}
            if competition_provider not in valid_providers:
                competition_provider = WorkspaceCompetitionContext.PROVIDER_MANUAL
            module_catalog = _workspace_module_catalog(workspace_kind)
            selected_modules = {
                item['key']: str(request.POST.get(f"module_{item['key']}") or '').lower() in {'1', 'true', 'on', 'yes'}
                for item in module_catalog
            }
            selected_deliverables = {
                _workspace_deliverable_flag(item['key'], deliverable['key']): (
                    str(request.POST.get(f"deliverable_{item['key']}__{deliverable['key']}") or '').lower() in {'1', 'true', 'on', 'yes'}
                )
                for item in module_catalog
                for deliverable in item.get('deliverables', []) or []
            }
            if not any(selected_modules.values()):
                selected_modules = {item['key']: True for item in module_catalog}
            expanded_modules = _expand_workspace_module_selection(workspace_kind, selected_modules, selected_deliverables)
            workspace_form = {
                'workspace_name': workspace_name,
                'workspace_kind': workspace_kind,
                'owner_username': owner_username,
                'team_id': str(team_id or ''),
                'team_new_name': team_new_name,
                'workspace_notes': workspace_notes,
                'competition_provider': competition_provider,
                'external_competition_key': external_competition_key,
                'external_group_key': external_group_key,
                'external_team_key': external_team_key,
                'external_team_name': external_team_name,
                'competition_auto_sync': competition_auto_sync,
                'seed_demo_data': seed_demo_data,
                'initial_admin_usernames': initial_admin_usernames,
                'initial_member_usernames': initial_member_usernames,
                'modules': expanded_modules,
                'module_keys': [key for key, enabled in selected_modules.items() if enabled],
                'deliverable_keys': [key for key, enabled in selected_deliverables.items() if enabled],
            }
            try:
                if not workspace_name:
                    raise ValueError('Indica un nombre para el workspace.')
                owner_user = User.objects.filter(username__iexact=owner_username).first() if owner_username else None
                if owner_username and not owner_user:
                    raise ValueError(f'No existe el usuario propietario "{owner_username}".')
                if workspace_kind == Workspace.KIND_TASK_STUDIO and not owner_user:
                    raise ValueError('Task Studio requiere un usuario propietario.')
                admin_users, missing_admin_users = _parse_workspace_usernames(initial_admin_usernames)
                member_users, missing_member_users = _parse_workspace_usernames(initial_member_usernames)
                if missing_admin_users:
                    raise ValueError(f'No existen estos administradores iniciales: {", ".join(missing_admin_users)}.')
                if missing_member_users:
                    raise ValueError(f'No existen estos miembros iniciales: {", ".join(missing_member_users)}.')
                primary_workspace_team = None
                if workspace_kind == Workspace.KIND_CLUB and team_new_name:
                    primary_workspace_team = _ensure_platform_team(team_new_name)
                if not primary_workspace_team and team_id:
                    primary_workspace_team = Team.objects.filter(id=team_id).first()
                workspace = Workspace.objects.create(
                    name=workspace_name,
                    slug=_unique_workspace_slug(workspace_name),
                    kind=workspace_kind,
                    owner_user=owner_user,
                    primary_team=primary_workspace_team if workspace_kind == Workspace.KIND_CLUB else None,
                    enabled_modules=expanded_modules,
                    notes=workspace_notes,
                )
                if workspace.kind == Workspace.KIND_CLUB:
                    _bootstrap_workspace_competition_context(
                        workspace,
                        primary_team=primary_workspace_team,
                        provider=competition_provider,
                        external_competition_key=external_competition_key,
                        external_group_key=external_group_key,
                        external_team_key=external_team_key,
                        external_team_name=external_team_name,
                        auto_sync_enabled=competition_auto_sync,
                    )
                    if competition_auto_sync:
                        _sync_workspace_competition_context(workspace)
                    if seed_demo_data and workspace.primary_team_id:
                        _bootstrap_demo_club_workspace(workspace)
                if owner_user:
                    WorkspaceMembership.objects.get_or_create(
                        workspace=workspace,
                        user=owner_user,
                        defaults={'role': WorkspaceMembership.ROLE_OWNER},
                    )
                for admin_user in admin_users:
                    if owner_user and admin_user.id == owner_user.id:
                        continue
                    WorkspaceMembership.objects.update_or_create(
                        workspace=workspace,
                        user=admin_user,
                        defaults={'role': WorkspaceMembership.ROLE_ADMIN},
                    )
                for member_user in member_users:
                    if owner_user and member_user.id == owner_user.id:
                        continue
                    if any(admin_user.id == member_user.id for admin_user in admin_users):
                        continue
                    WorkspaceMembership.objects.update_or_create(
                        workspace=workspace,
                        user=member_user,
                        defaults={'role': WorkspaceMembership.ROLE_MEMBER},
                    )
                feedback = f'Workspace creado: {workspace.name}.'
                workspace_form = {
                    'workspace_name': '',
                    'workspace_kind': Workspace.KIND_CLUB,
                    'owner_username': '',
                    'team_id': '',
                    'team_new_name': '',
                    'workspace_notes': '',
                    'competition_provider': WorkspaceCompetitionContext.PROVIDER_MANUAL,
                    'external_competition_key': '',
                    'external_group_key': '',
                    'external_team_key': '',
                    'external_team_name': '',
                    'competition_auto_sync': True,
                    'seed_demo_data': False,
                    'initial_admin_usernames': '',
                    'initial_member_usernames': '',
                    'modules': _workspace_default_modules(Workspace.KIND_CLUB),
                    'module_keys': [item['key'] for item in _workspace_module_catalog(Workspace.KIND_CLUB)],
                    'deliverable_keys': _workspace_selected_deliverable_keys(
                        Workspace.KIND_CLUB,
                        _expand_workspace_module_selection(
                            Workspace.KIND_CLUB,
                            {item['key']: True for item in _workspace_module_catalog(Workspace.KIND_CLUB)},
                        ),
                    ),
                }
            except ValueError as exc:
                error = str(exc)
            except Exception:
                error = 'No se pudo crear el workspace.'
        elif form_action == 'platform_user_create':
            active_tab = 'users'
            users_subtab = 'create'
            full_name = _sanitize_task_text((request.POST.get('full_name') or '').strip(), multiline=False, max_len=150)
            username = re.sub(r'\s+', '', str(request.POST.get('username') or '').strip()).lower()[:150]
            email = re.sub(r'\s+', '', str(request.POST.get('email') or '').strip()).lower()[:190]
            password = (request.POST.get('password') or '').strip()
            role_value = str(request.POST.get('role') or AppUserRole.ROLE_PLAYER).strip()
            role_choices = {choice[0] for choice in AppUserRole.ROLE_CHOICES}
            if role_value not in role_choices:
                role_value = AppUserRole.ROLE_PLAYER
            user_form = {
                'full_name': full_name,
                'username': username,
                'email': email,
                'role': role_value,
            }
            try:
                if not username:
                    raise ValueError('El usuario es obligatorio.')
                if User.objects.filter(username__iexact=username).exists():
                    raise ValueError('Ese usuario ya existe.')
                if len(password) < 6:
                    raise ValueError('La contraseña debe tener al menos 6 caracteres.')
                first_name, last_name = _split_full_name(full_name)
                user = User.objects.create_user(
                    username=username,
                    email=email,
                    password=password,
                    first_name=first_name,
                    last_name=last_name,
                )
                if role_value == AppUserRole.ROLE_ADMIN:
                    user.is_staff = True
                    user.save(update_fields=['is_staff'])
                AppUserRole.objects.update_or_create(user=user, defaults={'role': role_value})
                if role_value == AppUserRole.ROLE_TASK_STUDIO:
                    _ensure_task_studio_workspace(user)
                user_message = f'Usuario creado en Plataforma: {username}.'
                user_form = {
                    'full_name': '',
                    'username': '',
                    'email': '',
                    'role': AppUserRole.ROLE_PLAYER,
                }
            except ValueError as exc:
                user_error = str(exc)
            except Exception:
                user_error = 'No se pudo crear el usuario global.'
        elif form_action == 'platform_user_invite_create':
            active_tab = 'users'
            users_subtab = 'list'
            user_id = _parse_int(request.POST.get('user_id'))
            validity_days = _parse_int(request.POST.get('valid_days')) or 7
            validity_days = max(1, min(validity_days, 30))
            user_obj = User.objects.filter(id=user_id).first() if user_id else None
            try:
                if not user_obj:
                    raise ValueError('Usuario no encontrado.')
                UserInvitation.objects.filter(user=user_obj, is_active=True, accepted_at__isnull=True).update(is_active=False)
                invitation = UserInvitation.objects.create(
                    user=user_obj,
                    token=UserInvitation.generate_token(),
                    email=(user_obj.email or '').strip(),
                    expires_at=timezone.now() + timedelta(days=validity_days),
                    created_by=request.user.get_username() if request.user.is_authenticated else '',
                    is_active=True,
                )
                invite_url = request.build_absolute_uri(
                    reverse('user-invite-accept', args=[invitation.token])
                )
                invitation_links.append(
                    {
                        'username': user_obj.username,
                        'url': invite_url,
                        'expires_at': invitation.expires_at,
                    }
                )
                user_message = f'Invitación generada en Plataforma para {user_obj.username}.'
            except ValueError as exc:
                user_error = str(exc)
            except Exception:
                user_error = 'No se pudo generar la invitación global.'
        elif form_action == 'platform_user_update':
            active_tab = 'users'
            users_subtab = 'list'
            user_id = _parse_int(request.POST.get('user_id'))
            full_name = _sanitize_task_text((request.POST.get('full_name') or '').strip(), multiline=False, max_len=150)
            email = re.sub(r'\s+', '', str(request.POST.get('email') or '').strip()).lower()[:190]
            password = (request.POST.get('password') or '').strip()
            role_value = str(request.POST.get('role') or AppUserRole.ROLE_PLAYER).strip()
            is_active = str(request.POST.get('is_active') or '').lower() in {'1', 'true', 'on', 'yes'}
            role_choices = {choice[0] for choice in AppUserRole.ROLE_CHOICES}
            if role_value not in role_choices:
                role_value = AppUserRole.ROLE_PLAYER
            user_obj = User.objects.filter(id=user_id).first() if user_id else None
            try:
                if not user_obj:
                    raise ValueError('Usuario no encontrado.')
                if password and len(password) < 6:
                    raise ValueError('La nueva contraseña debe tener al menos 6 caracteres.')
                first_name, last_name = _split_full_name(full_name)
                user_obj.first_name = first_name
                user_obj.last_name = last_name
                user_obj.email = email
                user_obj.is_active = is_active
                user_obj.is_staff = bool(is_active and role_value == AppUserRole.ROLE_ADMIN)
                if password:
                    user_obj.set_password(password)
                update_fields = ['first_name', 'last_name', 'email', 'is_active', 'is_staff']
                if password:
                    update_fields.append('password')
                user_obj.save(update_fields=update_fields)
                AppUserRole.objects.update_or_create(user=user_obj, defaults={'role': role_value})
                if role_value == AppUserRole.ROLE_TASK_STUDIO:
                    _ensure_task_studio_workspace(user_obj)
                user_message = f'Usuario actualizado: {user_obj.username}.'
            except ValueError as exc:
                user_error = str(exc)
            except Exception:
                user_error = 'No se pudo actualizar el usuario.'
        elif form_action == 'platform_user_toggle_active':
            active_tab = 'users'
            users_subtab = 'list'
            user_id = _parse_int(request.POST.get('user_id'))
            user_obj = User.objects.filter(id=user_id).first() if user_id else None
            try:
                if not user_obj:
                    raise ValueError('Usuario no encontrado.')
                user_obj.is_active = not bool(user_obj.is_active)
                role_value = _get_user_role(user_obj)
                if not user_obj.is_active and role_value == AppUserRole.ROLE_ADMIN:
                    user_obj.is_staff = False
                elif user_obj.is_active and role_value == AppUserRole.ROLE_ADMIN:
                    user_obj.is_staff = True
                user_obj.save(update_fields=['is_active', 'is_staff'])
                user_message = f'Usuario {"activado" if user_obj.is_active else "desactivado"}: {user_obj.username}.'
            except ValueError as exc:
                user_error = str(exc)
            except Exception:
                user_error = 'No se pudo cambiar el estado del usuario.'
        elif form_action == 'platform_user_delete':
            active_tab = 'users'
            users_subtab = 'list'
            user_id = _parse_int(request.POST.get('user_id'))
            user_obj = User.objects.filter(id=user_id).first() if user_id else None
            try:
                if not user_obj:
                    raise ValueError('Usuario no encontrado.')
                owned_club_workspace = Workspace.objects.filter(kind=Workspace.KIND_CLUB, owner_user=user_obj).first()
                if owned_club_workspace:
                    raise ValueError(f'No puedes borrar {user_obj.username} mientras sea propietario de {owned_club_workspace.name}.')
                task_studio_workspace = Workspace.objects.filter(kind=Workspace.KIND_TASK_STUDIO, owner_user=user_obj).first()
                if task_studio_workspace:
                    _delete_task_studio_workspace(task_studio_workspace, disable_owner_profile=False)
                WorkspaceMembership.objects.filter(user=user_obj).delete()
                UserInvitation.objects.filter(user=user_obj).update(is_active=False)
                username = user_obj.username
                user_obj.delete()
                user_message = f'Usuario eliminado: {username}.'
            except ValueError as exc:
                user_error = str(exc)
            except Exception:
                user_error = 'No se pudo eliminar el usuario.'
        elif form_action in {'carousel_upload', 'carousel_update', 'carousel_delete'}:
            active_tab = 'home-global'
            if _handle_home_carousel_post(request):
                carousel_message = 'Cambios guardados en Home global.'

    workspaces = list(
        Workspace.objects
        .select_related('owner_user', 'primary_team')
        .annotate(member_count=Count('memberships', distinct=True))
        .order_by('kind', 'name', 'id')
    )
    for workspace in workspaces:
        workspace.task_count = TaskStudioTask.objects.filter(workspace=workspace).count()
        workspace.profile_count = TaskStudioProfile.objects.filter(workspace=workspace).count()
        workspace.active_module_count = len(_workspace_selected_module_keys(workspace.kind, _workspace_enabled_modules(workspace)))
    primary_workspace = next((workspace for workspace in workspaces if workspace.kind == Workspace.KIND_CLUB and workspace.primary_team_id), None)
    club_workspaces = [workspace for workspace in workspaces if workspace.kind == Workspace.KIND_CLUB]
    studio_workspaces = [workspace for workspace in workspaces if workspace.kind == Workspace.KIND_TASK_STUDIO and workspace.owner_user_id]
    workspace_users = list(
        WorkspaceMembership.objects
        .select_related('workspace', 'user')
        .filter(workspace__kind=Workspace.KIND_CLUB)
        .order_by('workspace__name', 'role', 'user__username')[:200]
    )
    platform_users = list(User.objects.order_by('username')[:220])
    carousel_images = list(HomeCarouselImage.objects.order_by('order', '-created_at', '-id')[:24])
    role_map = {item.user_id: item.role for item in AppUserRole.objects.select_related('user')}
    role_labels = dict(AppUserRole.ROLE_CHOICES)
    membership_counts = {
        row['user_id']: row['count']
        for row in WorkspaceMembership.objects.values('user_id').annotate(count=Count('id'))
    }
    linked_club_user_count = (
        WorkspaceMembership.objects
        .filter(workspace__kind=Workspace.KIND_CLUB)
        .values('user_id')
        .distinct()
        .count()
    )
    for item in platform_users:
        role_value = role_map.get(item.id, AppUserRole.ROLE_PLAYER)
        item.role_value = role_value
        item.role_label = role_labels.get(role_value, 'Jugador')
        item.full_name_display = item.get_full_name().strip() or item.username
        item.workspace_count = int(membership_counts.get(item.id, 0) or 0)
    club_workspaces_without_team = sum(1 for workspace in club_workspaces if not workspace.primary_team_id)
    club_workspaces_without_members = sum(1 for workspace in club_workspaces if int(workspace.member_count or 0) <= 0)
    studio_workspaces_without_tasks = sum(1 for workspace in studio_workspaces if int(workspace.task_count or 0) <= 0)
    platform_attention_items = []
    if club_workspaces_without_team:
        platform_attention_items.append(f'{club_workspaces_without_team} cliente(s) club sin equipo principal vinculado.')
    if club_workspaces_without_members:
        platform_attention_items.append(f'{club_workspaces_without_members} cliente(s) club sin miembros asignados.')
    if studio_workspaces_without_tasks:
        platform_attention_items.append(f'{studio_workspaces_without_tasks} Task Studio sin tareas creadas todavía.')
    if not platform_attention_items:
        platform_attention_items.append('La matriz no tiene alertas críticas de configuración.')

    recent_documents = []
    search_query = _sanitize_task_text((request.GET.get('q') or '').strip(), multiline=False, max_len=120)
    search_results = []
    if search_query:
        q = search_query
        session_hits = list(
            SessionTask.objects
            .select_related('session__microcycle__team')
            .filter(title__icontains=q)
            .order_by('-created_at', '-id')[:20]
        )
        for item in session_hits:
            search_results.append(
                {
                    'type': 'Tarea club',
                    'title': str(item.title or 'Tarea').strip() or 'Tarea',
                    'meta': f"{item.session.session_date:%d/%m/%Y} · {item.get_block_display()} · {item.duration_minutes or 0} min",
                    'url': reverse('sessions-task-edit', args=[item.id]),
                }
            )
        studio_hits = list(
            TaskStudioTask.objects
            .select_related('owner')
            .filter(title__icontains=q)
            .order_by('-updated_at', '-id')[:20]
        )
        for item in studio_hits:
            search_results.append(
                {
                    'type': 'Task Studio',
                    'title': str(item.title or 'Tarea').strip() or 'Tarea',
                    'meta': f"{item.owner.get_username()} · {item.updated_at:%d/%m/%Y}",
                    'url': f"{reverse('task-studio-task-edit', args=[item.id])}?user={item.owner_id}",
                }
            )
        search_results = search_results[:30]
    recent_session_tasks = list(
        SessionTask.objects
        .select_related('session__microcycle__team')
        .order_by('-created_at', '-id')[:6]
    )
    recent_sessions = list(
        TrainingSession.objects
        .select_related('microcycle__team')
        .order_by('-created_at', '-id')[:5]
    )
    recent_studio_tasks = list(
        TaskStudioTask.objects
        .select_related('owner', 'workspace')
        .order_by('-updated_at', '-id')[:6]
    )
    for item in recent_session_tasks:
        recent_documents.append(
            {
                'type': 'Tarea club',
                'title': str(item.title or 'Tarea').strip() or 'Tarea',
                'source': str(getattr(item.session.microcycle.team, 'display_name', '') or getattr(item.session.microcycle.team, 'name', '')).strip() or 'Club',
                'meta': f"{item.session.session_date:%d/%m/%Y} · {item.get_block_display()} · {item.duration_minutes or 0} min",
                'uefa_url': reverse('session-task-pdf', args=[item.id]),
                'club_url': f"{reverse('session-task-pdf', args=[item.id])}?style=club",
                'created_at': item.created_at,
            }
        )
    for item in recent_sessions:
        recent_documents.append(
            {
                'type': 'Sesión',
                'title': str(item.focus or 'Sesión').strip() or 'Sesión',
                'source': str(getattr(item.microcycle.team, 'display_name', '') or getattr(item.microcycle.team, 'name', '')).strip() or 'Club',
                'meta': f"{item.session_date:%d/%m/%Y} · {item.duration_minutes or 0} min",
                'uefa_url': reverse('session-plan-pdf', args=[item.id]),
                'club_url': f"{reverse('session-plan-pdf', args=[item.id])}?style=club",
                'created_at': item.created_at,
            }
        )
    for item in recent_studio_tasks:
        recent_documents.append(
            {
                'type': 'Task Studio',
                'title': str(item.title or 'Tarea').strip() or 'Tarea',
                'source': str(item.owner.get_full_name() or item.owner.get_username()).strip() or item.owner.get_username(),
                'meta': f"{item.updated_at:%d/%m/%Y} · {item.get_block_display()} · {item.duration_minutes or 0} min",
                'uefa_url': reverse('task-studio-task-pdf', args=[item.id]),
                'club_url': f"{reverse('task-studio-task-pdf', args=[item.id])}?style=club",
                'created_at': item.updated_at,
            }
        )
    recent_documents = sorted(
        recent_documents,
        key=lambda item: item.get('created_at') or timezone.now(),
        reverse=True,
    )[:14]

    recent_audit_events = list(
        AuditEvent.objects
        .select_related('workspace', 'actor_user')
        .order_by('-created_at', '-id')[:24]
    )

    return render(
        request,
        'football/platform_overview.html',
        {
            'feedback': feedback,
            'error': error,
            'user_message': user_message,
            'user_error': user_error,
            'invitation_links': invitation_links,
            'carousel_message': carousel_message,
            'workspaces': workspaces,
            'primary_workspace': primary_workspace,
            'club_workspaces': club_workspaces,
            'studio_workspaces': studio_workspaces,
            'workspace_users': workspace_users,
            'platform_users': platform_users,
            'carousel_images': carousel_images,
            'teams': list(Team.objects.order_by('name')[:200]),
            'workspace_kind_choices': Workspace.KIND_CHOICES,
            'workspace_module_catalog_club': _workspace_module_catalog_for_template(Workspace.KIND_CLUB),
            'workspace_module_catalog_task_studio': _workspace_module_catalog_for_template(Workspace.KIND_TASK_STUDIO),
            'workspace_form': workspace_form,
            'competition_provider_choices': WorkspaceCompetitionContext.PROVIDER_CHOICES,
            'user_form': user_form,
            'role_choices': AppUserRole.ROLE_CHOICES,
            'active_workspace': active_workspace,
            'active_tab': active_tab,
            'users_subtab': users_subtab,
            'linked_club_user_count': linked_club_user_count,
            'platform_user_count': len(platform_users),
            'club_workspaces_without_team': club_workspaces_without_team,
            'club_workspaces_without_members': club_workspaces_without_members,
            'studio_workspaces_without_tasks': studio_workspaces_without_tasks,
            'platform_attention_items': platform_attention_items,
            'recent_documents': recent_documents,
            'recent_audit_events': recent_audit_events,
            'search_query': search_query,
            'search_results': search_results,
        },
    )


@login_required
def platform_workspace_detail_page(request, workspace_id):
    workspace = (
        Workspace.objects
        .select_related('owner_user', 'primary_team', 'competition_context', 'competition_snapshot')
        .annotate(member_count=Count('memberships', distinct=True))
        .filter(id=workspace_id)
        .first()
    )
    if not workspace:
        raise Http404('Workspace no encontrado')
    if not _can_view_workspace(request.user, workspace):
        return HttpResponse('No tienes permisos para acceder a este workspace.', status=403)
    feedback = ''
    error = ''
    can_manage_workspace = _can_manage_workspace(request.user, workspace)
    invite_link = ''
    competition_search_inputs = {
        'provider': WorkspaceCompetitionContext.PROVIDER_MANUAL,
        'team_query': str(getattr(workspace.primary_team, 'name', '') or '').strip(),
        'competition_query': str(getattr(getattr(getattr(getattr(workspace.primary_team, 'group', None), 'season', None), 'competition', None), 'name', '') or '').strip(),
        'group_query': str(getattr(getattr(workspace.primary_team, 'group', None), 'name', '') or '').strip(),
    }
    competition_search_results = []
    if request.method == 'POST':
        form_action = (request.POST.get('form_action') or '').strip().lower()
        if not can_manage_workspace:
            error = 'No tienes permisos para modificar este workspace.'
        elif form_action == 'update_workspace_identity':
            workspace_name = _sanitize_task_text((request.POST.get('workspace_name') or '').strip(), multiline=False, max_len=160)
            owner_username = _sanitize_task_text((request.POST.get('owner_username') or '').strip(), multiline=False, max_len=150).lower()
            workspace_notes = _sanitize_task_text((request.POST.get('workspace_notes') or '').strip(), multiline=True, max_len=1200)
            team_id = _parse_int(request.POST.get('team_id'))
            is_active = str(request.POST.get('workspace_is_active') or '').lower() in {'1', 'true', 'on', 'yes'}
            try:
                if not workspace_name:
                    raise ValueError('Indica un nombre para el cliente.')
                owner_user = User.objects.filter(username__iexact=owner_username).first() if owner_username else None
                if owner_username and not owner_user:
                    raise ValueError(f'No existe el usuario propietario "{owner_username}".')
                if workspace.kind == Workspace.KIND_TASK_STUDIO and not owner_user:
                    raise ValueError('Task Studio requiere un usuario propietario.')
                if workspace.kind == Workspace.KIND_TASK_STUDIO and owner_user:
                    conflict = Workspace.objects.filter(kind=Workspace.KIND_TASK_STUDIO, owner_user=owner_user).exclude(id=workspace.id).first()
                    if conflict:
                        raise ValueError(f'El usuario {owner_user.username} ya tiene otro Task Studio.')
                old_owner = workspace.owner_user
                workspace.name = workspace_name
                workspace.owner_user = owner_user
                workspace.notes = workspace_notes
                workspace.is_active = is_active
                if workspace.kind == Workspace.KIND_CLUB:
                    workspace.primary_team = Team.objects.filter(id=team_id).first() if team_id else None
                workspace.save(update_fields=['name', 'owner_user', 'notes', 'is_active', 'primary_team', 'updated_at'])
                if owner_user:
                    WorkspaceMembership.objects.update_or_create(
                        workspace=workspace,
                        user=owner_user,
                        defaults={'role': WorkspaceMembership.ROLE_OWNER},
                    )
                if old_owner and (not owner_user or old_owner.id != owner_user.id):
                    old_membership = WorkspaceMembership.objects.filter(
                        workspace=workspace,
                        user=old_owner,
                        role=WorkspaceMembership.ROLE_OWNER,
                    ).first()
                    if old_membership:
                        old_membership.role = WorkspaceMembership.ROLE_ADMIN
                        old_membership.save(update_fields=['role'])
                if workspace.kind == Workspace.KIND_TASK_STUDIO:
                    if owner_user:
                        TaskStudioProfile.objects.update_or_create(
                            user=owner_user,
                            defaults={'workspace': workspace, 'is_enabled': True},
                        )
                    if old_owner and (not owner_user or old_owner.id != owner_user.id):
                        old_profile = TaskStudioProfile.objects.filter(user=old_owner, workspace=workspace).first()
                        if old_profile:
                            old_profile.workspace = None
                            old_profile.is_enabled = False
                            old_profile.save(update_fields=['workspace', 'is_enabled'])
                feedback = f'Cliente actualizado: {workspace.name}.'
            except ValueError as exc:
                error = str(exc)
            except Exception:
                error = 'No se pudo actualizar la configuración del cliente.'
        elif form_action == 'update_modules':
            module_catalog = _workspace_module_catalog(workspace.kind)
            selected_modules = {
                item['key']: str(request.POST.get(f"module_{item['key']}") or '').lower() in {'1', 'true', 'on', 'yes'}
                for item in module_catalog
            }
            selected_deliverables = {
                _workspace_deliverable_flag(item['key'], deliverable['key']): (
                    str(request.POST.get(f"deliverable_{item['key']}__{deliverable['key']}") or '').lower() in {'1', 'true', 'on', 'yes'}
                )
                for item in module_catalog
                for deliverable in item.get('deliverables', []) or []
            }
            enabled_modules = _expand_workspace_module_selection(workspace.kind, selected_modules, selected_deliverables)
            workspace.enabled_modules = enabled_modules
            workspace.save(update_fields=['enabled_modules', 'updated_at'])
            feedback = 'Módulos del workspace actualizados.'
        elif form_action == 'bootstrap_demo_data':
            if workspace.kind != Workspace.KIND_CLUB:
                error = 'Solo los clientes club pueden generar datos demo.'
            elif not workspace.primary_team_id:
                error = 'Asocia un equipo al workspace antes de generar datos demo.'
            else:
                try:
                    created = _bootstrap_demo_club_workspace(workspace)
                    feedback = (
                        'Datos demo generados. '
                        f"Jugadores: {created.get('players', 0)}, "
                        f"partidos: {created.get('matches', 0)}, "
                        f"eventos: {created.get('events', 0)}, "
                        f"sesiones: {created.get('sessions', 0)}, "
                        f"tareas: {created.get('tasks', 0)}, "
                        f"estadísticas: {created.get('stats', 0)}."
                    )
                except Exception:
                    error = 'No se pudieron generar los datos demo.'
        elif form_action == 'update_competition_context':
            if workspace.kind != Workspace.KIND_CLUB:
                error = 'Solo los clientes club tienen contexto competitivo.'
            else:
                provider = str(request.POST.get('competition_provider') or WorkspaceCompetitionContext.PROVIDER_MANUAL).strip()
                external_competition_key = str(request.POST.get('external_competition_key') or '').strip()[:140]
                external_group_key = str(request.POST.get('external_group_key') or '').strip()[:140]
                external_team_key = str(request.POST.get('external_team_key') or '').strip()[:140]
                external_team_name = _sanitize_task_text((request.POST.get('external_team_name') or '').strip(), multiline=False, max_len=160)
                auto_sync_enabled = str(request.POST.get('competition_auto_sync') or '').lower() in {'1', 'true', 'on', 'yes'}
                valid_providers = {choice[0] for choice in WorkspaceCompetitionContext.PROVIDER_CHOICES}
                if provider not in valid_providers:
                    provider = WorkspaceCompetitionContext.PROVIDER_MANUAL
                _bootstrap_workspace_competition_context(
                    workspace,
                    primary_team=workspace.primary_team,
                    provider=provider,
                    external_competition_key=external_competition_key,
                    external_group_key=external_group_key,
                    external_team_key=external_team_key,
                    external_team_name=external_team_name,
                    auto_sync_enabled=auto_sync_enabled,
                )
                if workspace.primary_team_id:
                    _invalidate_team_dashboard_caches(workspace.primary_team)
                feedback = 'Contexto competitivo actualizado.'
        elif form_action == 'sync_competition_context':
            if workspace.kind != Workspace.KIND_CLUB:
                error = 'Solo los clientes club se pueden sincronizar.'
            else:
                _, sync_error = _sync_workspace_competition_context(workspace)
                if sync_error:
                    error = sync_error
                else:
                    if workspace.primary_team_id:
                        _invalidate_team_dashboard_caches(workspace.primary_team)
                    feedback = 'Competición sincronizada para este cliente.'
        elif form_action == 'search_competition_context':
            if workspace.kind != Workspace.KIND_CLUB:
                error = 'Solo los clientes club tienen búsqueda competitiva.'
            else:
                competition_search_inputs = {
                    'provider': str(request.POST.get('competition_provider_search') or WorkspaceCompetitionContext.PROVIDER_MANUAL).strip(),
                    'team_query': _sanitize_task_text((request.POST.get('competition_team_query') or '').strip(), multiline=False, max_len=160),
                    'competition_query': _sanitize_task_text((request.POST.get('competition_competition_query') or '').strip(), multiline=False, max_len=160),
                    'group_query': _sanitize_task_text((request.POST.get('competition_group_query') or '').strip(), multiline=False, max_len=160),
                }
                valid_providers = {choice[0] for choice in WorkspaceCompetitionContext.PROVIDER_CHOICES}
                if competition_search_inputs['provider'] not in valid_providers:
                    competition_search_inputs['provider'] = WorkspaceCompetitionContext.PROVIDER_MANUAL
                if competition_search_inputs['provider'] == WorkspaceCompetitionContext.PROVIDER_UNIVERSO:
                    competition_search_results = _search_universo_competition_candidates(
                        team_query=competition_search_inputs['team_query'],
                        competition_query=competition_search_inputs['competition_query'],
                        group_query=competition_search_inputs['group_query'],
                    )
                else:
                    competition_search_results = _search_workspace_competition_candidates(
                        competition_search_inputs['provider'],
                        team_query=competition_search_inputs['team_query'],
                        competition_query=competition_search_inputs['competition_query'],
                        group_query=competition_search_inputs['group_query'],
                    )
                if competition_search_results:
                    feedback = f'Se encontraron {len(competition_search_results)} coincidencias para este cliente.'
                else:
                    error = 'No se encontraron coincidencias con los criterios indicados.'
        elif form_action == 'apply_competition_candidate':
            if workspace.kind != Workspace.KIND_CLUB:
                error = 'Solo los clientes club pueden vincular contexto competitivo.'
            else:
                candidate_team_id = _parse_int(request.POST.get('candidate_team_id'))
                candidate_team = (
                    Team.objects
                    .select_related('group__season__competition')
                    .filter(id=candidate_team_id)
                    .first()
                )
                provider = str(request.POST.get('candidate_provider') or WorkspaceCompetitionContext.PROVIDER_MANUAL).strip()
                valid_providers = {choice[0] for choice in WorkspaceCompetitionContext.PROVIDER_CHOICES}
                if provider not in valid_providers:
                    provider = WorkspaceCompetitionContext.PROVIDER_MANUAL
                auto_sync_enabled = str(request.POST.get('candidate_auto_sync') or '').lower() in {'1', 'true', 'on', 'yes'}
                external_competition_key = str(request.POST.get('candidate_external_competition_key') or '').strip()[:140]
                external_group_key = str(request.POST.get('candidate_external_group_key') or '').strip()[:140]
                external_team_key = str(request.POST.get('candidate_external_team_key') or '').strip()[:140]
                external_team_name = _sanitize_task_text((request.POST.get('candidate_external_team_name') or '').strip(), multiline=False, max_len=160)
                if not candidate_team and provider == WorkspaceCompetitionContext.PROVIDER_UNIVERSO:
                    candidate_team, import_error = _import_universo_competition_candidate(
                        competition_key=external_competition_key,
                        group_key=external_group_key,
                        team_key=external_team_key,
                        team_name=external_team_name,
                    )
                    if import_error:
                        error = import_error
                if error:
                    pass
                elif not candidate_team or not getattr(candidate_team, 'group', None):
                    error = 'La coincidencia seleccionada no es válida.'
                else:
                    workspace.primary_team = candidate_team
                    workspace.save(update_fields=['primary_team', 'updated_at'])
                    _bootstrap_workspace_competition_context(
                        workspace,
                        primary_team=candidate_team,
                        provider=provider,
                        external_competition_key=external_competition_key,
                        external_group_key=external_group_key,
                        external_team_key=external_team_key,
                        external_team_name=external_team_name,
                        auto_sync_enabled=auto_sync_enabled,
                    )
                    _, sync_error = _sync_workspace_competition_context(workspace)
                    if sync_error:
                        error = sync_error
                    else:
                        feedback = f'Cliente vinculado a {candidate_team.name} y sincronizado.'
        elif form_action == 'add_member':
            username = _sanitize_task_text((request.POST.get('member_username') or '').strip(), multiline=False, max_len=150)
            member_role = str(request.POST.get('member_role') or WorkspaceMembership.ROLE_MEMBER).strip()
            target_user = User.objects.filter(username__iexact=username).first() if username else None
            if not target_user:
                error = 'Usuario no encontrado para añadir al workspace.'
            elif member_role not in {choice[0] for choice in WorkspaceMembership.ROLE_CHOICES}:
                error = 'Rol de workspace no válido.'
            else:
                WorkspaceMembership.objects.update_or_create(
                    workspace=workspace,
                    user=target_user,
                    defaults={'role': member_role},
                )
                feedback = f'Usuario {target_user.username} vinculado al workspace.'
                _audit(
                    request,
                    'workspace_member_add',
                    workspace=workspace,
                    message='Miembro añadido al workspace',
                    payload={'user': target_user.username, 'role': member_role},
                )
        elif form_action == 'invite_member':
            # Invitación + alta opcional de usuario (para que pueda poner su contraseña).
            username = _sanitize_task_text((request.POST.get('invite_username') or '').strip(), multiline=False, max_len=150)
            full_name = _sanitize_task_text((request.POST.get('invite_full_name') or '').strip(), multiline=False, max_len=150)
            email = re.sub(r'\s+', '', str(request.POST.get('invite_email') or '').strip()).lower()[:190]
            app_role = str(request.POST.get('invite_app_role') or AppUserRole.ROLE_GUEST).strip()
            member_role = str(request.POST.get('invite_member_role') or WorkspaceMembership.ROLE_VIEWER).strip()
            validity_days = _parse_int(request.POST.get('invite_valid_days')) or 7
            validity_days = max(1, min(validity_days, 30))
            if member_role not in {choice[0] for choice in WorkspaceMembership.ROLE_CHOICES}:
                member_role = WorkspaceMembership.ROLE_VIEWER
            allowed_roles = {
                AppUserRole.ROLE_PLAYER,
                AppUserRole.ROLE_GUEST,
                AppUserRole.ROLE_COACH,
                AppUserRole.ROLE_FITNESS,
                AppUserRole.ROLE_GOALKEEPER,
                AppUserRole.ROLE_ANALYST,
                AppUserRole.ROLE_TASK_STUDIO,
            }
            if app_role not in allowed_roles:
                app_role = AppUserRole.ROLE_GUEST
            try:
                if not username:
                    raise ValueError('Indica un username para invitar.')
                username = re.sub(r'\s+', '', username).lower()[:150]
                first_name, last_name = _split_full_name(full_name)
                user_obj = User.objects.filter(username__iexact=username).first()
                if not user_obj:
                    user_obj = User.objects.create_user(
                        username=username,
                        email=email,
                        password=None,
                        first_name=first_name,
                        last_name=last_name,
                        is_active=False,
                    )
                else:
                    # Actualiza datos básicos si se proporcionan.
                    changed = False
                    if email and (user_obj.email or '').strip().lower() != email:
                        user_obj.email = email
                        changed = True
                    if full_name:
                        user_obj.first_name = first_name
                        user_obj.last_name = last_name
                        changed = True
                    if changed:
                        user_obj.save(update_fields=['email', 'first_name', 'last_name'])
                AppUserRole.objects.update_or_create(user=user_obj, defaults={'role': app_role})
                if app_role == AppUserRole.ROLE_TASK_STUDIO:
                    _ensure_task_studio_workspace(user_obj)
                WorkspaceMembership.objects.update_or_create(
                    workspace=workspace,
                    user=user_obj,
                    defaults={'role': member_role},
                )
                # Reemplaza invitaciones previas no usadas.
                UserInvitation.objects.filter(user=user_obj, is_active=True, accepted_at__isnull=True).update(is_active=False)
                invitation = UserInvitation.objects.create(
                    user=user_obj,
                    token=UserInvitation.generate_token(),
                    email=(user_obj.email or '').strip(),
                    expires_at=timezone.now() + timedelta(days=validity_days),
                    created_by=request.user.get_username() if request.user.is_authenticated else '',
                    is_active=True,
                )
                invite_link = request.build_absolute_uri(reverse('user-invite-accept', args=[invitation.token]))
                feedback = f'Invitación generada para {user_obj.username}.'
                _audit(
                    request,
                    'workspace_invite',
                    workspace=workspace,
                    message='Invitación generada',
                    payload={'user': user_obj.username, 'app_role': app_role, 'member_role': member_role, 'days': validity_days},
                )
            except ValueError as exc:
                error = str(exc)
            except Exception:
                error = 'No se pudo generar la invitación.'
        elif form_action == 'update_member_role':
            membership_id = _parse_int(request.POST.get('membership_id'))
            member_role = str(request.POST.get('member_role') or WorkspaceMembership.ROLE_MEMBER).strip()
            membership = WorkspaceMembership.objects.filter(id=membership_id, workspace=workspace).select_related('user').first()
            if not membership:
                error = 'Miembro no encontrado.'
            elif member_role not in {choice[0] for choice in WorkspaceMembership.ROLE_CHOICES}:
                error = 'Rol de workspace no válido.'
            else:
                membership.role = member_role
                membership.save(update_fields=['role'])
                feedback = f'Rol actualizado para {membership.user.username}.'
                _audit(
                    request,
                    'workspace_member_role',
                    workspace=workspace,
                    message='Rol de miembro actualizado',
                    payload={'user': membership.user.username, 'role': member_role},
                )
        elif form_action == 'update_member_modules':
            membership_id = _parse_int(request.POST.get('membership_id'))
            membership = WorkspaceMembership.objects.filter(id=membership_id, workspace=workspace).select_related('user').first()
            if not membership:
                error = 'Miembro no encontrado.'
            else:
                catalog = _workspace_access_module_catalog(workspace.kind)
                module_access = {}
                for entry in catalog:
                    key = entry.get('key')
                    if not key:
                        continue
                    # Si el módulo no está activo en el workspace, no lo guardamos.
                    if not bool(_workspace_enabled_modules(workspace).get(key, False)):
                        continue
                    checked = str(request.POST.get(f"member_module_{key}") or '').lower() in {'1', 'true', 'on', 'yes'}
                    if not checked:
                        module_access[key] = False
                membership.module_access = module_access
                membership.save(update_fields=['module_access'])
                feedback = f'Permisos por módulo actualizados para {membership.user.username}.'
                _audit(
                    request,
                    'workspace_member_modules',
                    workspace=workspace,
                    message='Permisos por módulo actualizados',
                    payload={'user': membership.user.username, 'denied': sorted([k for k, v in module_access.items() if v is False])[:40]},
                )
        elif form_action == 'reset_member_modules':
            membership_id = _parse_int(request.POST.get('membership_id'))
            membership = WorkspaceMembership.objects.filter(id=membership_id, workspace=workspace).select_related('user').first()
            if not membership:
                error = 'Miembro no encontrado.'
            else:
                membership.module_access = {}
                membership.save(update_fields=['module_access'])
                feedback = f'Permisos por módulo restablecidos para {membership.user.username}.'
                _audit(
                    request,
                    'workspace_member_modules_reset',
                    workspace=workspace,
                    message='Permisos por módulo restablecidos',
                    payload={'user': membership.user.username},
                )
        elif form_action == 'remove_member':
            membership_id = _parse_int(request.POST.get('membership_id'))
            membership = WorkspaceMembership.objects.filter(id=membership_id, workspace=workspace).select_related('user').first()
            if not membership:
                error = 'Miembro no encontrado.'
            else:
                removed_username = membership.user.username
                membership.delete()
                feedback = f'Usuario {removed_username} eliminado del workspace.'
                _audit(
                    request,
                    'workspace_member_remove',
                    workspace=workspace,
                    message='Miembro eliminado del workspace',
                    payload={'user': removed_username},
                )
    workspace.task_count = TaskStudioTask.objects.filter(workspace=workspace).count()
    workspace.profile_count = TaskStudioProfile.objects.filter(workspace=workspace).count()
    roster_count = TaskStudioRosterPlayer.objects.filter(workspace=workspace).count()
    club_player_count = Player.objects.filter(team=workspace.primary_team).count() if workspace.primary_team_id else 0
    memberships = list(
        WorkspaceMembership.objects
        .select_related('user')
        .filter(workspace=workspace)
        .order_by('role', 'user__username')
    )
    enabled_modules = _workspace_enabled_modules(workspace)
    module_catalog = _workspace_module_catalog_for_template(workspace.kind, enabled_modules)
    access_module_catalog = _workspace_access_module_catalog(workspace.kind)
    # Prepara filas de permisos por miembro para la plantilla (evita lookups dinámicos en template).
    for membership in memberships:
        raw_access = getattr(membership, 'module_access', None)
        if not isinstance(raw_access, dict):
            raw_access = {}
        membership_role = str(getattr(membership, 'role', '') or '')
        is_privileged = membership_role in {WorkspaceMembership.ROLE_OWNER, WorkspaceMembership.ROLE_ADMIN}
        rows = []
        for entry in access_module_catalog:
            key = entry.get('key')
            label = entry.get('label') or key
            ws_enabled = bool(enabled_modules.get(key, False)) if key else False
            allowed = ws_enabled and (is_privileged or (raw_access.get(key, True) is not False))
            rows.append({'key': key, 'label': label, 'workspace_enabled': ws_enabled, 'allowed': allowed})
        membership.access_module_rows = rows
    competition_context = getattr(workspace, 'competition_context', None)
    competition_snapshot = getattr(workspace, 'competition_snapshot', None)
    competition_summary = None
    if workspace.kind == Workspace.KIND_CLUB:
        if not competition_search_results:
            competition_search_inputs = {
                'provider': getattr(competition_context, 'provider', WorkspaceCompetitionContext.PROVIDER_MANUAL) if competition_context else WorkspaceCompetitionContext.PROVIDER_MANUAL,
                'team_query': str(getattr(workspace.primary_team, 'name', '') or '').strip(),
                'competition_query': str(getattr(getattr(getattr(getattr(workspace.primary_team, 'group', None), 'season', None), 'competition', None), 'name', '') or '').strip(),
                'group_query': str(getattr(getattr(workspace.primary_team, 'group', None), 'name', '') or '').strip(),
            }
        competition_context = competition_context or _bootstrap_workspace_competition_context(workspace, primary_team=workspace.primary_team)
        if competition_context and competition_context.is_auto_sync_enabled and not competition_snapshot:
            _sync_workspace_competition_context(workspace)
            workspace = (
                Workspace.objects
                .select_related('owner_user', 'primary_team', 'competition_context', 'competition_snapshot')
                .annotate(member_count=Count('memberships', distinct=True))
                .filter(id=workspace_id)
                .first()
            )
            competition_context = getattr(workspace, 'competition_context', None)
            competition_snapshot = getattr(workspace, 'competition_snapshot', None)
        snapshot_standings = competition_snapshot.standings_payload if competition_snapshot and isinstance(competition_snapshot.standings_payload, list) else []
        snapshot_next = competition_snapshot.next_match_payload if competition_snapshot and isinstance(competition_snapshot.next_match_payload, dict) else {}
        competition_summary = {
            'standings_count': len(snapshot_standings),
            'next_match_opponent': _payload_opponent_name(snapshot_next) or '',
            'next_match_round': str(snapshot_next.get('round') or '').strip() if isinstance(snapshot_next, dict) else '',
            'last_sync_at': competition_context.last_sync_at if competition_context else None,
        }
    module_cards = []
    if workspace.kind == Workspace.KIND_CLUB:
        module_cards = [
            {'title': 'Portada', 'description': 'Home, clasificación, próximo partido y contexto visual del cliente.', 'url': reverse('dashboard-home')},
            {'title': 'Estadísticas', 'description': 'KPIs de temporada, seguimiento individual y métricas manuales.', 'url': reverse('coach-role-trainer')},
            {'title': 'Cuerpo técnico', 'description': 'Hub del staff, plantilla técnica y áreas operativas del cliente.', 'url': reverse('coach-cards')},
            {'title': 'Partido', 'description': 'Convocatoria, 11 inicial y registro de acciones de matchday.', 'url': reverse('convocation')},
            {'title': 'Entrenamiento', 'description': 'Microciclos, sesiones, tareas, porteros, físico y ABP.', 'url': reverse('sessions')},
            {'title': 'Análisis', 'description': 'Rival, informes, scouting y lectura táctica del cliente seleccionado.', 'url': reverse('analysis')},
        ]
    else:
        task_studio_url = reverse('task-studio-home')
        if workspace.owner_user_id:
            task_studio_url = f'{task_studio_url}?user={workspace.owner_user_id}'
        module_cards = [
            {'title': 'Task Studio', 'description': 'Repositorio y editor privado del usuario.', 'url': task_studio_url},
            {'title': 'Perfil e identidad', 'description': 'Datos documentales, escudo y colores.', 'url': f"{reverse('task-studio-profile')}?user={workspace.owner_user_id}" if workspace.owner_user_id else reverse('task-studio-profile')},
            {'title': 'Plantilla privada', 'description': 'Jugadores propios para la pizarra y documentos.', 'url': f"{reverse('task-studio-roster')}?user={workspace.owner_user_id}" if workspace.owner_user_id else reverse('task-studio-roster')},
            {'title': 'Nueva tarea', 'description': 'Entrada directa al creador táctico del workspace.', 'url': f"{reverse('task-studio-task-create')}?user={workspace.owner_user_id}" if workspace.owner_user_id else reverse('task-studio-task-create')},
        ]
    return render(
        request,
        'football/platform_workspace_detail.html',
        {
            'workspace': workspace,
            'active_workspace': _build_active_workspace_badge(request),
            'feedback': feedback,
            'error': error,
            'can_manage_workspace': can_manage_workspace,
            'roster_count': roster_count,
            'club_player_count': club_player_count,
            'module_cards': module_cards,
            'memberships': memberships,
            'module_catalog': module_catalog,
            'enabled_modules': enabled_modules,
            'access_module_catalog': access_module_catalog,
            'workspace_role_choices': WorkspaceMembership.ROLE_CHOICES,
            'teams': list(Team.objects.order_by('name')[:200]),
            'competition_context': competition_context,
            'competition_snapshot': competition_snapshot,
            'competition_summary': competition_summary,
            'competition_provider_choices': WorkspaceCompetitionContext.PROVIDER_CHOICES,
            'competition_search_inputs': competition_search_inputs,
            'competition_search_results': competition_search_results,
            'invite_link': invite_link,
        },
    )


@login_required
@require_POST
def platform_workspace_delete_page(request, workspace_id):
    workspace = Workspace.objects.select_related('owner_user', 'primary_team').filter(id=workspace_id).first()
    if not workspace:
        raise Http404('Workspace no encontrado')
    if not _can_manage_workspace(request.user, workspace):
        return HttpResponse('No tienes permisos para eliminar este workspace.', status=403)
    if workspace.kind != Workspace.KIND_TASK_STUDIO:
        return HttpResponse('Solo se puede eliminar Task Studio desde esta acción.', status=403)
    owner_user = workspace.owner_user
    if owner_user:
        TaskStudioProfile.objects.update_or_create(
            user=owner_user,
            defaults={'workspace': None, 'is_enabled': False},
        )
    TaskStudioTask.objects.filter(workspace=workspace).delete()
    TaskStudioRosterPlayer.objects.filter(workspace=workspace).delete()
    TaskStudioProfile.objects.filter(workspace=workspace).exclude(user=owner_user).delete()
    WorkspaceMembership.objects.filter(workspace=workspace).delete()
    workspace_name = workspace.name
    workspace_id_value = workspace.id
    workspace.delete()
    if int(request.session.get('active_workspace_id') or 0) == int(workspace_id_value):
        request.session.pop('active_workspace_id', None)
    request.session['platform_feedback'] = f'Task Studio eliminado: {workspace_name}.'
    return redirect('platform-overview')


@login_required
def platform_workspace_enter_page(request, workspace_id):
    workspace = Workspace.objects.select_related('owner_user', 'primary_team').filter(id=workspace_id, is_active=True).first()
    if not workspace:
        raise Http404('Workspace no encontrado')
    if not _can_view_workspace(request.user, workspace):
        return HttpResponse('No tienes permisos para acceder a este workspace.', status=403)
    request.session['active_workspace_id'] = workspace.id
    return redirect(_workspace_entry_url(workspace, user=request.user))


@login_required
def platform_workspace_clear_page(request):
    if not _can_access_platform(request.user):
        return HttpResponse('No tienes permisos para acceder a la plataforma.', status=403)
    request.session.pop('active_workspace_id', None)
    return redirect('platform-overview')


def _handle_home_carousel_post(request):
    form_action = (request.POST.get('form_action') or '').strip().lower()

    if form_action == 'carousel_upload':
        uploaded = request.FILES.get('carousel_image')
        if uploaded:
            title = (request.POST.get('carousel_title') or '').strip()
            order = _parse_int(request.POST.get('carousel_order')) or 0
            is_active = str(request.POST.get('carousel_is_active') or '').lower() in {'1', 'true', 'on', 'yes'}
            HomeCarouselImage.objects.create(
                title=title,
                image=uploaded,
                order=order,
                is_active=is_active,
            )
        return True

    if form_action == 'carousel_update':
        image_id = _parse_int(request.POST.get('carousel_id'))
        item = HomeCarouselImage.objects.filter(id=image_id).first()
        if item:
            item.title = (request.POST.get('carousel_title') or '').strip()
            item.order = _parse_int(request.POST.get('carousel_order')) or 0
            item.is_active = str(request.POST.get('carousel_is_active') or '').lower() in {'1', 'true', 'on', 'yes'}
            item.save(update_fields=['title', 'order', 'is_active'])
        return True

    if form_action == 'carousel_delete':
        image_id = _parse_int(request.POST.get('carousel_id'))
        item = HomeCarouselImage.objects.filter(id=image_id).first()
        if item:
            if item.image:
                try:
                    item.image.delete(save=False)
                except Exception:
                    pass
            item.delete()
        return True
    return False


@login_required
@ensure_csrf_cookie
def admin_page(request):
    primary_team = _get_primary_team_for_request(request)
    current_role = AppUserRole.objects.filter(user=request.user).values_list('role', flat=True).first()
    is_admin_user = bool(request.user.is_staff or current_role == AppUserRole.ROLE_ADMIN)
    roster_message = ''
    roster_error = ''
    actions_message = ''
    actions_error = ''
    active_tab = (request.GET.get('tab') or request.POST.get('active_tab') or 'roster').strip().lower()
    if active_tab in {'carousel', 'users'}:
        target_anchor = '#home-global' if active_tab == 'carousel' else '#usuarios-club'
        return redirect(f"{reverse('platform-overview')}{target_anchor}")
    if active_tab not in {'roster', 'actions'}:
        active_tab = 'roster'

    if request.method == 'POST':
        form_action = (request.POST.get('form_action') or '').strip()
        if form_action in {
            'carousel_upload',
            'carousel_update',
            'carousel_delete',
            'user_create',
            'user_update',
            'user_invite_create',
            'user_update_role',
            'user_toggle_active',
        }:
            target_anchor = '#home-global' if form_action.startswith('carousel_') else '#usuarios-club'
            return redirect(f"{reverse('platform-overview')}{target_anchor}")
        if form_action in {'roster_add_or_update', 'roster_deactivate', 'roster_reactivate', 'crest_import_zip'} and primary_team:
            active_tab = 'roster'
            if form_action == 'crest_import_zip':
                if not is_admin_user:
                    roster_error = 'Solo administradores pueden subir escudos.'
                else:
                    uploaded_zip = request.FILES.get('crest_zip')
                    if not uploaded_zip:
                        roster_error = 'Sube un ZIP con los escudos (png/jpg/webp).'
                    else:
                        allowed_ext = {'.png', '.jpg', '.jpeg', '.webp'}
                        group_teams = list(
                            Team.objects
                            .filter(group=primary_team.group)
                            .only('id', 'name', 'slug', 'short_name', 'external_id', 'crest_url', 'crest_image')
                        )

                        def _team_keys(team_obj):
                            candidates = set()
                            for raw in (
                                getattr(team_obj, 'slug', ''),
                                getattr(team_obj, 'external_id', ''),
                                getattr(team_obj, 'name', ''),
                                getattr(team_obj, 'short_name', ''),
                                getattr(team_obj, 'display_name', ''),
                            ):
                                val = str(raw or '').strip()
                                if not val:
                                    continue
                                candidates.add(val.lower())
                                candidates.add(normalize_label(val))
                                candidates.add(_normalize_team_lookup_key(val))
                            return {key for key in candidates if key}

                        lookup = {}
                        for team in group_teams:
                            for key in _team_keys(team):
                                lookup.setdefault(key, team)

                        imported = 0
                        unknown = []
                        skipped = 0

                        def _match_team(filename):
                            stem = Path(filename).stem
                            base = str(stem or '').strip()
                            if not base:
                                return None
                            candidates = [
                                base.lower(),
                                normalize_label(base),
                                _normalize_team_lookup_key(base),
                            ]
                            for key in candidates:
                                if key and lookup.get(key):
                                    return lookup[key]
                            # Fuzzy: por inclusión (evita casos como "C.D. CANTORIA 2017 F.C.")
                            base_key = _normalize_team_lookup_key(base)
                            if base_key:
                                for key, team_obj in lookup.items():
                                    if key and (base_key in key or key in base_key):
                                        return team_obj
                            return None

                        try:
                            uploaded_zip.seek(0)
                            with zipfile.ZipFile(uploaded_zip) as zf:
                                for info in zf.infolist():
                                    if info.is_dir():
                                        continue
                                    inner_name = Path(info.filename).name
                                    ext = Path(inner_name).suffix.lower()
                                    if ext not in allowed_ext:
                                        skipped += 1
                                        continue
                                    if info.file_size and info.file_size > 5 * 1024 * 1024:
                                        skipped += 1
                                        continue
                                    team_obj = _match_team(inner_name)
                                    if not team_obj:
                                        unknown.append(inner_name)
                                        continue
                                    try:
                                        blob = zf.read(info)
                                    except Exception:
                                        skipped += 1
                                        continue
                                    if not blob:
                                        skipped += 1
                                        continue
                                    content = ContentFile(blob)
                                    content.name = inner_name
                                    # Guardar en MEDIA (S3 si está activo).
                                    try:
                                        team_obj.crest_image.save(f'{team_obj.slug}{ext}', content, save=True)
                                        if team_obj.crest_url:
                                            team_obj.crest_url = ''
                                            team_obj.save(update_fields=['crest_url'])
                                        _invalidate_team_dashboard_caches(team_obj)
                                        imported += 1
                                    except Exception:
                                        skipped += 1
                        except zipfile.BadZipFile:
                            roster_error = 'El archivo no es un ZIP válido.'
                        except Exception:
                            roster_error = 'No se pudieron importar los escudos.'

                        if not roster_error:
                            roster_message = f'Escudos importados: {imported}.'
                            if skipped:
                                roster_message += f' Saltados: {skipped}.'
                            if unknown:
                                unknown_preview = ', '.join(unknown[:6])
                                roster_message += f' Sin equipo asociado: {unknown_preview}'
                                if len(unknown) > 6:
                                    roster_message += f' (+{len(unknown) - 6}).'
                roster_players = list(Player.objects.filter(team=primary_team).order_by('name'))
                return render(
                    request,
                    'football/admin.html',
                    {
                        'primary_team': primary_team,
                        'team_name': getattr(primary_team, 'display_name', None) or getattr(primary_team, 'name', ''),
                        'is_admin_user': is_admin_user,
                        'active_tab': active_tab,
                        'roster_message': roster_message,
                        'roster_error': roster_error,
                        'roster_players': roster_players,
                        'actions_message': actions_message,
                        'actions_error': actions_error,
                    },
                )
            player_id = _parse_int(request.POST.get('player_id'))
            name = (request.POST.get('name') or '').strip()
            number_raw = (request.POST.get('number') or '').strip()
            position = (request.POST.get('position') or '').strip()
            is_active = str(request.POST.get('is_active') or '1').strip().lower() in {'1', 'true', 'on', 'yes'}
            try:
                if form_action in {'roster_deactivate', 'roster_reactivate'}:
                    if not player_id:
                        raise ValueError('Jugador no válido.')
                    player = Player.objects.filter(id=player_id, team=primary_team).first()
                    if not player:
                        raise ValueError('Jugador no encontrado.')
                    player.is_active = form_action == 'roster_reactivate'
                    player.save(update_fields=['is_active'])
                    roster_message = (
                        f'{player.name} reactivado correctamente.'
                        if player.is_active
                        else f'{player.name} marcado como inactivo.'
                    )
                else:
                    if not name:
                        raise ValueError('El nombre es obligatorio.')
                    number = _parse_int(number_raw) if number_raw else None
                    player = (
                        Player.objects.filter(team=primary_team, name__iexact=name)
                        .order_by('id')
                        .first()
                    )
                    if player:
                        player.number = number
                        player.position = position
                        player.is_active = is_active
                        player.save(update_fields=['number', 'position', 'is_active'])
                        roster_message = f'Jugador actualizado: {player.name}.'
                    else:
                        Player.objects.create(
                            team=primary_team,
                            name=name,
                            number=number,
                            position=position,
                            is_active=is_active,
                        )
                        roster_message = f'Jugador añadido: {name}.'
            except ValueError as exc:
                roster_error = str(exc)
            except Exception:
                roster_error = 'No se pudo guardar la plantilla.'
        elif form_action in {'admin_match_save', 'admin_action_bulk_add'}:
            active_tab = 'actions'
            if not is_admin_user:
                actions_error = 'Solo administradores pueden editar partidos y acciones.'
            elif not primary_team:
                actions_error = 'No hay equipo principal configurado.'
            elif form_action == 'admin_match_save':
                match_id = _parse_int(request.POST.get('match_id'))
                opponent_name = (request.POST.get('opponent_name') or '').strip()
                round_value = (request.POST.get('round') or '').strip()
                location = (request.POST.get('location') or '').strip()
                date_raw = (request.POST.get('match_date') or '').strip()
                time_raw = (request.POST.get('match_time') or '').strip()
                if not opponent_name:
                    actions_error = 'El rival es obligatorio.'
                else:
                    match_date = None
                    for fmt in ('%Y-%m-%d', '%d/%m/%Y'):
                        try:
                            match_date = datetime.strptime(date_raw, fmt).date() if date_raw else None
                            break
                        except ValueError:
                            continue
                    match_time = None
                    if time_raw:
                        try:
                            match_time = datetime.strptime(time_raw, '%H:%M').time()
                        except ValueError:
                            actions_error = 'Hora inválida. Usa HH:MM.'
                    if not actions_error:
                        rival_team = Team.objects.filter(name__iexact=opponent_name).first()
                        if not rival_team:
                            rival_team = Team.objects.create(
                                slug=_unique_team_slug(opponent_name),
                                name=opponent_name,
                                short_name=opponent_name[:24],
                                group=primary_team.group,
                            )
                        match_dt = None
                        if match_date:
                            match_dt = timezone.make_aware(
                                datetime.combine(match_date, match_time or time(hour=0, minute=0)),
                                timezone.get_current_timezone(),
                            )
                        match_obj = Match.objects.filter(id=match_id).first() if match_id else None
                        if not match_obj:
                            match_obj = Match.objects.create(
                                home_team=primary_team,
                                away_team=rival_team,
                                date=match_date,
                                kickoff_time=match_time,
                                round=round_value,
                                location=location,
                            )
                        else:
                            match_obj.away_team = rival_team
                            match_obj.date = match_date
                            match_obj.kickoff_time = match_time
                            match_obj.round = round_value
                            match_obj.location = location
                            match_obj.save(update_fields=['away_team', 'date', 'kickoff_time', 'round', 'location'])
                        actions_message = f'Partido guardado (ID {match_obj.id}).'
            elif form_action == 'admin_action_bulk_add':
                match_id = _parse_int(request.POST.get('match_id'))
                player_id = _parse_int(request.POST.get('player_id'))
                action_type = (request.POST.get('action_type') or '').strip()
                result = (request.POST.get('result') or '').strip()
                zone = (request.POST.get('zone') or '').strip()
                quantity = _parse_int(request.POST.get('quantity')) or 1
                if quantity < 1:
                    quantity = 1
                if quantity > 500:
                    quantity = 500
                match_obj = Match.objects.filter(id=match_id).first() if match_id else None
                player_obj = Player.objects.filter(id=player_id, team=primary_team).first() if player_id else None
                if not match_obj:
                    actions_error = 'Selecciona un partido válido.'
                elif not player_obj:
                    actions_error = 'Selecciona un jugador válido.'
                elif not action_type:
                    actions_error = 'Selecciona una acción.'
                elif not result:
                    actions_error = 'Selecciona un resultado.'
                else:
                    tercio = zone_to_tercio(zone)
                    base_minute = 1
                    try:
                        last_min = (
                            MatchEvent.objects.filter(match=match_obj)
                            .exclude(minute__isnull=True)
                            .order_by('-minute')
                            .values_list('minute', flat=True)
                            .first()
                        )
                        if isinstance(last_min, int):
                            base_minute = max(1, min(120, last_min))
                    except Exception:
                        pass
                    events = []
                    for idx in range(quantity):
                        minute_value = min(120, base_minute + (idx % 5))
                        events.append(
                            MatchEvent(
                                match=match_obj,
                                player=player_obj,
                                minute=minute_value,
                                period=1 if minute_value <= 45 else 2,
                                event_type=action_type,
                                result=result,
                                zone=zone,
                                tercio=tercio,
                                observation='Carga manual admin',
                                source_file='admin-manual',
                                system='touch-field-final',
                            )
                        )
                    MatchEvent.objects.bulk_create(events, batch_size=200)
                    actions_message = (
                        f'Se añadieron {quantity} acciones a {player_obj.name} '
                        f'en partido ID {match_obj.id}.'
                    )
    roster_players = []
    admin_matches = []
    selected_admin_match = None
    admin_players = []
    admin_action_choices = []
    admin_result_choices = []
    admin_zone_choices = []

    if active_tab == 'roster':
        roster_players = (
            list(Player.objects.filter(team=primary_team).order_by('-is_active', 'number', 'name'))
            if primary_team
            else []
        )

    if active_tab == 'actions' and is_admin_user:
        admin_match_qs = _team_match_queryset(primary_team) if primary_team else Match.objects.none()
        # En Admin necesitamos ver también partidos importados aunque lleguen con
        # vínculos incompletos al equipo principal. Por eso unimos:
        # 1) partidos asociados al primer equipo (query normal)
        # 2) partidos del grupo/temporada del primer equipo (fallback importador)
        # y deduplicamos por id.
        admin_map = {}
        for match in admin_match_qs.select_related('home_team', 'away_team').order_by('-date', '-id')[:500]:
            admin_map[int(match.id)] = match
        if primary_team:
            fallback_q = Q()
            if getattr(primary_team, 'group_id', None):
                fallback_q |= Q(group_id=primary_team.group_id)
                season_id = getattr(getattr(primary_team, 'group', None), 'season_id', None)
                if season_id:
                    fallback_q |= Q(season_id=season_id)
            if fallback_q:
                fallback_rows = (
                    Match.objects
                    .filter(fallback_q)
                    .select_related('home_team', 'away_team')
                    .order_by('-date', '-id')[:700]
                )
                for match in fallback_rows:
                    admin_map[int(match.id)] = match
        admin_matches = sorted(
            list(admin_map.values()),
            key=lambda item: (item.date or date.min, int(item.id or 0)),
            reverse=True,
        )
        selected_admin_match_id = _parse_int(request.GET.get('match_id') or request.POST.get('match_id'))
        selected_admin_match = (
            next((m for m in admin_matches if m.id == selected_admin_match_id), None)
            if selected_admin_match_id
            else (admin_matches[0] if admin_matches else None)
        )
        admin_players = (
            list(Player.objects.filter(team=primary_team).order_by('number', 'name'))
            if primary_team
            else []
        )
        admin_action_choices = load_match_actions()
        admin_result_choices = load_match_results()
        seen_zone_labels = set()
        for zone_def in FIELD_ZONES:
            label = str(zone_def.get('label') or '').strip()
            if not label or label in seen_zone_labels:
                continue
            seen_zone_labels.add(label)
            admin_zone_choices.append(label)
    return render(
        request,
        'football/admin.html',
        {
            'roster_players': roster_players,
            'roster_message': roster_message,
            'roster_error': roster_error,
            'active_tab': active_tab,
            'team_name': primary_team.display_name if primary_team else '',
            'primary_team_id': primary_team.id if primary_team else None,
            'is_admin_user': is_admin_user,
            'admin_matches': admin_matches,
            'selected_admin_match': selected_admin_match,
            'admin_players': admin_players,
            'admin_action_choices': admin_action_choices,
            'admin_result_choices': admin_result_choices,
            'admin_zone_choices': admin_zone_choices,
            'actions_message': actions_message,
            'actions_error': actions_error,
        },
    )


def invitation_accept_page(request, token):
    invitation = (
        UserInvitation.objects
        .select_related('user')
        .filter(token=token, is_active=True)
        .order_by('-created_at')
        .first()
    )
    now = timezone.now()
    if not invitation:
        return render(
            request,
            'football/invitation_accept.html',
            {'status': 'invalid'},
            status=404,
        )
    if invitation.accepted_at:
        return render(
            request,
            'football/invitation_accept.html',
            {'status': 'used', 'username': invitation.user.username},
            status=410,
        )
    if invitation.is_expired(now=now):
        invitation.is_active = False
        invitation.save(update_fields=['is_active'])
        return render(
            request,
            'football/invitation_accept.html',
            {'status': 'expired', 'username': invitation.user.username},
            status=410,
        )
    error = ''
    success = ''
    if request.method == 'POST':
        password = (request.POST.get('password') or '').strip()
        password_confirm = (request.POST.get('password_confirm') or '').strip()
        if not password:
            error = 'La contraseña es obligatoria.'
        elif password != password_confirm:
            error = 'Las contraseñas no coinciden.'
        else:
            try:
                validate_password(password, user=invitation.user)
                invitation.user.set_password(password)
                invitation.user.is_active = True
                invitation.user.save(update_fields=['password', 'is_active'])
                invitation.accepted_at = timezone.now()
                invitation.is_active = False
                invitation.save(update_fields=['accepted_at', 'is_active'])
                UserInvitation.objects.filter(
                    user=invitation.user,
                    is_active=True,
                    accepted_at__isnull=True,
                ).exclude(id=invitation.id).update(is_active=False)
                try:
                    auth_login(request, invitation.user)
                    return redirect('dashboard-home')
                except Exception:
                    success = 'Invitación aceptada. Ya puedes iniciar sesión.'
            except DjangoValidationError as exc:
                error = ' '.join(exc.messages) or 'Contraseña inválida.'
            except Exception:
                error = 'No se pudo completar la invitación.'
    return render(
        request,
        'football/invitation_accept.html',
        {
            'status': 'ok',
            'username': invitation.user.username,
            'expires_at': invitation.expires_at,
            'error': error,
            'success': success,
        },
    )


def _file_field_as_data_url(file_field):
    if not file_field:
        return ''
    try:
        file_field.open('rb')
        raw = file_field.read()
        file_field.close()
    except Exception:
        return ''
    if not raw:
        return ''
    ext = Path(getattr(file_field, 'name', '') or '').suffix.lower()
    mime = {
        '.png': 'image/png',
        '.jpg': 'image/jpeg',
        '.jpeg': 'image/jpeg',
        '.webp': 'image/webp',
        '.gif': 'image/gif',
    }.get(ext, 'image/jpeg')
    return f"data:{mime};base64,{base64.b64encode(raw).decode('utf-8')}"


def _client_ip(request):
    if not request:
        return ''
    forwarded = str(request.META.get('HTTP_X_FORWARDED_FOR') or '').split(',')[0].strip()
    return forwarded or str(request.META.get('REMOTE_ADDR') or '').strip()


def _audit(request, action, *, workspace=None, message='', payload=None):
    try:
        if payload is None:
            payload = {}
        actor_user = request.user if request and getattr(request, 'user', None) and request.user.is_authenticated else None
        actor = ''
        if actor_user:
            actor = actor_user.get_username()
            full_name = actor_user.get_full_name().strip() if hasattr(actor_user, 'get_full_name') else ''
            if full_name:
                actor = f'{actor} · {full_name}'
        AuditEvent.objects.create(
            workspace=workspace,
            actor_user=actor_user,
            actor=actor[:150],
            action=str(action or '')[:80],
            message=str(message or '')[:220],
            payload=payload if isinstance(payload, dict) else {},
            ip=_client_ip(request)[:60],
        )
    except Exception:
        pass


@login_required
@require_POST
def share_task_pdf_create(request):
    """
    Crea un enlace público (token) para imprimir/abrir el PDF de una tarea.
    """
    task_kind = str(request.POST.get('task_kind') or '').strip().lower()
    task_id = _parse_int(request.POST.get('task_id'))
    style = (request.POST.get('style') or 'uefa').strip().lower()
    if style not in {'uefa', 'club'}:
        style = 'uefa'
    validity_days = _parse_int(request.POST.get('valid_days')) or 14
    validity_days = max(1, min(validity_days, 60))
    password = (request.POST.get('password') or '').strip()
    try:
        if not task_kind or task_kind not in {'session', 'task_studio'}:
            raise ValueError('Tipo de tarea no válido.')
        if not task_id:
            raise ValueError('Tarea no válida.')
        if task_kind == 'session':
            if not _can_access_sessions_workspace(request.user):
                return JsonResponse({'error': 'No tienes permisos para acceder a sesiones.'}, status=403)
            forbidden = _forbid_if_workspace_module_disabled(request, 'sessions', label='sesiones')
            if forbidden:
                return JsonResponse({'error': 'El módulo sesiones no está disponible.'}, status=403)
            task = (
                SessionTask.objects
                .select_related('session__microcycle__team')
                .filter(id=task_id, deleted_at__isnull=True)
                .first()
            )
            if not task:
                raise ValueError('Tarea no encontrada.')
            ws = _get_active_workspace(request)
            if not _can_access_platform(request.user) and ws and ws.kind == Workspace.KIND_CLUB and ws.primary_team_id:
                if int(ws.primary_team_id) != int(task.session.microcycle.team_id):
                    return JsonResponse({'error': 'No tienes permisos para compartir esta tarea.'}, status=403)
        else:
            forbidden = _forbid_if_no_task_studio_access(request.user)
            if forbidden:
                return JsonResponse({'error': 'No tienes permisos para acceder a Task Studio.'}, status=403)
            task = _task_studio_task_for_request(request, task_id)
            if not task:
                raise ValueError('Tarea no encontrada.')

        ShareLink.objects.filter(
            is_active=True,
            kind=ShareLink.KIND_TASK_PDF,
            payload__task_kind=task_kind,
            payload__task_id=task_id,
            payload__style=style,
        ).update(is_active=False)
        link = ShareLink.objects.create(
            token=ShareLink.generate_token(),
            kind=ShareLink.KIND_TASK_PDF,
            payload={'task_kind': task_kind, 'task_id': task_id, 'style': style},
            password_hash=make_password(password) if password else '',
            expires_at=timezone.now() + timedelta(days=validity_days),
            created_by=request.user.get_username() if request.user.is_authenticated else '',
            created_by_user=request.user if request.user.is_authenticated else None,
            is_active=True,
        )
        url = request.build_absolute_uri(reverse('share-task-pdf', args=[link.token]))
        _audit(
            request,
            'share_link_create',
            workspace=_get_active_workspace(request),
            message='Enlace compartido creado',
            payload={'kind': 'task_pdf', 'task_kind': task_kind, 'task_id': task_id, 'style': style, 'expires_days': validity_days, 'password': bool(password)},
        )
        return JsonResponse({'ok': True, 'url': url, 'expires_at': link.expires_at})
    except ValueError as exc:
        return JsonResponse({'error': str(exc)}, status=400)
    except Exception:
        return JsonResponse({'error': 'No se pudo crear el enlace.'}, status=500)


def share_task_pdf_page(request, token):
    """
    Endpoint público para el PDF de una tarea compartida por token.
    """
    token = str(token or '').strip()
    link = (
        ShareLink.objects
        .filter(token=token, is_active=True, kind=ShareLink.KIND_TASK_PDF)
        .order_by('-created_at')
        .first()
    )
    now = timezone.now()
    if not link or not link.can_be_used(now=now):
        raise Http404('Enlace no disponible')
    if (link.password_hash or '').strip():
        if request.method != 'POST':
            return render(
                request,
                'football/share_link_gate.html',
                {'error': '', 'expires_at': link.expires_at},
                status=200,
            )
        supplied = (request.POST.get('password') or '').strip()
        if not supplied or not check_password(supplied, link.password_hash):
            return render(
                request,
                'football/share_link_gate.html',
                {'error': 'Contraseña incorrecta.', 'expires_at': link.expires_at},
                status=403,
            )
    payload = link.payload if isinstance(link.payload, dict) else {}
    task_kind = str(payload.get('task_kind') or '').strip().lower()
    task_id = _parse_int(payload.get('task_id'))
    pdf_style = str(payload.get('style') or 'uefa').strip().lower()
    if pdf_style not in {'uefa', 'club'}:
        pdf_style = 'uefa'
    if task_kind == 'session':
        task = (
            SessionTask.objects
            .select_related('session__microcycle__team')
            .filter(id=task_id, deleted_at__isnull=True)
            .first()
        )
        if not task:
            raise Http404('Tarea no encontrada')
        team = task.session.microcycle.team
        preview_url = _file_field_as_data_url(task.task_preview_image) if task.task_preview_image else ''
        context = _build_task_pdf_context(
            request,
            team=team,
            session=task.session,
            microcycle=task.session.microcycle,
            task=task,
            tactical_layout=task.tactical_layout if isinstance(task.tactical_layout, dict) else {},
            pdf_style=pdf_style,
            preview_url=preview_url,
        )
        html = render_to_string('football/session_task_pdf.html', context)
        filename = slugify(f'tarea-compartida-{task.id}-{task.title}') or f'tarea-{task.id}'
        try:
            ShareLink.objects.filter(id=link.id).update(access_count=(link.access_count or 0) + 1, last_accessed_at=timezone.now())
        except Exception:
            pass
        return _build_pdf_response_or_html_fallback(request, html, filename)
    if task_kind == 'task_studio':
        task = TaskStudioTask.objects.select_related('owner').filter(id=task_id, deleted_at__isnull=True).first()
        if not task:
            raise Http404('Tarea no encontrada')
        owner = task.owner
        preview_url = _file_field_as_data_url(task.task_preview_image) if task.task_preview_image else ''
        context = _build_task_studio_pdf_context(
            request,
            owner=owner,
            task=task,
            tactical_layout=task.tactical_layout if isinstance(task.tactical_layout, dict) else {},
            pdf_style=pdf_style,
            preview_url=preview_url,
        )
        html = render_to_string('football/session_task_pdf.html', context)
        filename = slugify(f'task-studio-compartida-{task.id}-{task.title}') or f'task-studio-{task.id}'
        try:
            ShareLink.objects.filter(id=link.id).update(access_count=(link.access_count or 0) + 1, last_accessed_at=timezone.now())
        except Exception:
            pass
        return _build_pdf_response_or_html_fallback(request, html, filename)
    raise Http404('Enlace no válido')


@login_required
@require_POST
def share_link_revoke(request, token):
    token = str(token or '').strip()
    link = ShareLink.objects.filter(token=token, is_active=True).first()
    if not link:
        return JsonResponse({'error': 'Enlace no encontrado.'}, status=404)
    if not _can_access_platform(request.user) and int(getattr(link, 'created_by_user_id', 0) or 0) != int(request.user.id):
        return JsonResponse({'error': 'No tienes permisos para revocar este enlace.'}, status=403)
    link.is_active = False
    link.save(update_fields=['is_active'])
    _audit(request, 'share_link_revoke', workspace=_get_active_workspace(request), message='Enlace compartido revocado', payload={'token': token, 'kind': link.kind})
    return JsonResponse({'ok': True})


@login_required
def player_dashboard_page(request):
    forbidden = _forbid_if_workspace_module_disabled(request, 'players', label='módulo de jugadores')
    if forbidden:
        return forbidden
    primary_team = _get_player_team_for_request(request)
    if not primary_team:
        return JsonResponse({'error': 'No hay equipo principal configurado'}, status=400)
    refresh_roster_default = str(os.getenv('PREFERENTE_ROSTER_REFRESH_ON_LOAD', '0')).strip().lower() in {'1', 'true', 'yes', 'on'}
    refresh_roster_override = os.getenv('PREFERENTE_ROSTER_REFRESH_ON_PLAYER_DASHBOARD')
    if refresh_roster_override is None:
        refresh_roster_on_load = refresh_roster_default
    else:
        refresh_roster_on_load = str(refresh_roster_override).strip().lower() in {'1', 'true', 'yes', 'on'}
    if refresh_roster_on_load:
        try:
            refresh_primary_roster_cache(primary_team, force=False)
        except Exception:
            pass
    player_stats = compute_player_dashboard(primary_team)
    team_matches = list(
        _team_match_queryset(primary_team)
        .select_related('home_team', 'away_team')
        .order_by('-date', '-id')
    )
    match_options = []
    for match in team_matches:
        opponent = (
            match.away_team.display_name
            if match.home_team == primary_team and match.away_team
            else match.home_team.display_name
            if match.away_team == primary_team and match.home_team
            else 'Rival desconocido'
        )
        label_parts = [match.round or f'Partido {match.id}', opponent]
        if match.date:
            label_parts.append(match.date.strftime('%d/%m/%Y'))
        match_options.append({'id': match.id, 'label': ' · '.join(label_parts)})

    selected_match_id = _parse_int(request.GET.get('match'))
    selected_match = None
    selected_match_total_actions = 0
    if selected_match_id:
        match_lookup = {match.id: match for match in team_matches}
        selected_match_obj = match_lookup.get(selected_match_id)
        if selected_match_obj:
            opponent = (
                selected_match_obj.away_team.display_name
                if selected_match_obj.home_team == primary_team and selected_match_obj.away_team
                else selected_match_obj.home_team.display_name
                if selected_match_obj.away_team == primary_team and selected_match_obj.home_team
                else 'Rival desconocido'
            )
            selected_match = {
                'id': selected_match_obj.id,
                'round': selected_match_obj.round or f'Partido {selected_match_obj.id}',
                'opponent': opponent,
                'date': selected_match_obj.date.strftime('%d/%m/%Y') if selected_match_obj.date else '',
                'home': selected_match_obj.home_team == primary_team,
            }
            for player in player_stats:
                match_entry = next(
                    (item for item in player.get('matches', []) if int(item.get('match_id') or 0) == selected_match_obj.id),
                    None,
                )
                actions = int(match_entry.get('actions', 0) or 0) if match_entry else 0
                successes = int(match_entry.get('successes', 0) or 0) if match_entry else 0
                player['match_actions'] = actions
                player['match_successes'] = successes
                player['match_success_rate'] = round((successes / actions) * 100, 1) if actions else 0
                selected_match_total_actions += actions

    current_role = _get_user_role(request.user) or AppUserRole.ROLE_PLAYER
    role_labels = dict(AppUserRole.ROLE_CHOICES)
    can_preview_player_view = current_role != AppUserRole.ROLE_PLAYER or _is_admin_user(request.user)
    active_workspace = _get_active_workspace(request)
    home_url = _workspace_entry_url(active_workspace, user=request.user) if active_workspace else reverse('dashboard-home')
    return render(
        request,
        'football/player_dashboard.html',
        {
            'player_stats': player_stats,
            'team_name': primary_team.display_name,
            'match_options': match_options,
            'selected_match': selected_match,
            'selected_match_id': selected_match_id,
            'selected_match_total_actions': selected_match_total_actions,
            'current_role': current_role,
            'current_role_label': role_labels.get(current_role, 'Jugador'),
            'can_preview_player_view': can_preview_player_view,
            'workspace_entry_url': home_url,
            'home_url': home_url,
        },
    )


@login_required
def coach_overview_page(request):
    forbidden = _forbid_if_no_coach_access(request.user)
    if forbidden:
        return forbidden
    forbidden = _forbid_if_workspace_module_disabled(request, 'coach_overview', label='portada staff')
    if forbidden:
        return forbidden
    sources = list(ScrapeSource.objects.filter(is_active=True))
    primary_team = _get_primary_team_for_request(request)
    technical_roles = {
        AppUserRole.ROLE_COACH,
        AppUserRole.ROLE_FITNESS,
        AppUserRole.ROLE_GOALKEEPER,
        AppUserRole.ROLE_ANALYST,
        AppUserRole.ROLE_ADMIN,
    }
    role_labels = dict(AppUserRole.ROLE_CHOICES)
    role_labels[AppUserRole.ROLE_GOALKEEPER] = 'Preparador de porteros'
    role_rows = list(AppUserRole.objects.select_related('user').filter(role__in=technical_roles))
    technical_members = []
    technical_members_lower = set()
    for role_row in role_rows:
        if not role_row.user.is_active:
            continue
        full_name = role_row.user.get_full_name().strip() or role_row.user.username
        label = f'{role_labels.get(role_row.role, "Técnico")} · {full_name}'
        normalized = label.lower()
        if normalized in technical_members_lower:
            continue
        technical_members.append(label)
        technical_members_lower.add(normalized)

    if not technical_members:
        technical_members = ['Sin miembros técnicos configurados en Admin']
    weekly_brief = _build_weekly_staff_brief_context(primary_team)
    rival_summary = _build_coach_rival_summary(primary_team)
    workspace = _get_active_workspace(request)
    competition_payload = _competition_payload_for_team(workspace, primary_team)
    standings = competition_payload.get('standings') or []
    convocation_next = _build_next_match_from_convocation(primary_team)
    next_match = competition_payload.get('next_match') or load_preferred_next_match_payload(primary_team=primary_team) or (get_next_match(primary_team, primary_team.group) if primary_team and primary_team.group else {}) or {}
    if _next_match_payload_is_reliable(convocation_next):
        next_match = convocation_next
    if isinstance(weekly_brief, dict):
        weekly_brief['match'] = next_match
        next_match = weekly_brief.get('match') or {}
    next_match_opponent = _payload_opponent_name(next_match) or 'Rival por confirmar'
    next_match_date = ''
    parsed_next_match_date = _parse_payload_date(next_match.get('date')) if isinstance(next_match, dict) else None
    if parsed_next_match_date:
        next_match_date = parsed_next_match_date.strftime('%d/%m/%Y')
    elif isinstance(next_match, dict):
        next_match_date = str(next_match.get('date') or '').strip()
    hero_items = list(HomeCarouselImage.objects.filter(is_active=True).order_by('order', '-created_at', '-id'))
    hero_image_url = hero_items[0].image.url if hero_items and getattr(hero_items[0], 'image', None) else ''
    team_name_folded = (primary_team.name or '').strip().lower() if primary_team else ''
    highlighted_standing = None
    for row in standings:
        row_name = str(row.get('full_name') or row.get('team') or '').strip().lower()
        if row_name == team_name_folded or team_name_folded in row_name or row_name in team_name_folded:
            highlighted_standing = row
            break
    standings_rows = []
    for row in standings:
        row_copy = dict(row)
        row_copy['is_team'] = bool(
            highlighted_standing
            and str(row_copy.get('team') or '').strip() == str(highlighted_standing.get('team') or '').strip()
        )
        standings_rows.append(row_copy)
    opponent_standing = None
    next_match_opponent_folded = next_match_opponent.strip().lower()
    for row in standings_rows:
        row_name = str(row.get('full_name') or row.get('team') or '').strip().lower()
        if not row_name or not next_match_opponent_folded:
            continue
        if (
            row_name == next_match_opponent_folded
            or next_match_opponent_folded in row_name
            or row_name in next_match_opponent_folded
        ):
            opponent_standing = row
            break
    competition_name = str(getattr(getattr(getattr(primary_team, 'group', None), 'season', None), 'competition', None).name or '').strip() if getattr(getattr(getattr(primary_team, 'group', None), 'season', None), 'competition', None) else ''
    group_name = str(getattr(getattr(primary_team, 'group', None), 'name', '') or '').strip()
    competition_label = ' · '.join([item for item in [competition_name, group_name] if item])
    team_display_name = str(getattr(primary_team, 'display_name', '') or getattr(primary_team, 'name', '') or 'Club').strip()
    team_crest_url = resolve_team_crest_url(request, primary_team, sync=True) if primary_team else ''
    pending_items = []
    if isinstance(weekly_brief, dict):
        if int(weekly_brief.get('convocated_count') or 0) <= 0:
            pending_items.append('Falta cerrar la convocatoria actual.')
        if int(weekly_brief.get('probable_eleven_count') or 0) <= 0:
            pending_items.append('No hay 11 probable definido.')
        if int(weekly_brief.get('available_count') or 0) <= 0:
            pending_items.append('No hay disponibilidad consolidada del equipo.')
    module_hub = [
        {
            'title': 'Portada',
            'description': 'Contexto competitivo, home visual y alertas rápidas del cliente.',
            'url': reverse('dashboard-home'),
        },
        {
            'title': 'Estadísticas',
            'description': 'KPIs, seguimiento individual y soporte manual de temporada.',
            'url': reverse('coach-role-trainer'),
        },
        {
            'title': 'Cuerpo técnico',
            'description': 'Hub del staff, plantilla técnica y accesos por área.',
            'url': reverse('coach-cards'),
        },
        {
            'title': 'Partido',
            'description': 'Convocatoria, 11 inicial y registro live de matchday.',
            'url': reverse('convocation'),
        },
        {
            'title': 'Entrenamiento',
            'description': 'Sesiones, tareas, porteros, físico y ABP.',
            'url': reverse('sessions'),
        },
        {
            'title': 'Análisis',
            'description': 'Rival, scouting, vídeo e informes tácticos.',
            'url': reverse('analysis'),
        },
    ]
    probable_eleven_names = []
    if isinstance(weekly_brief, dict):
        probable_preview = str(weekly_brief.get('probable_eleven_preview') or '').strip()
        probable_eleven_names = [item.strip() for item in probable_preview.split(',') if item.strip()][:5]
    staff_preview = technical_members[:4]
    staff_extra_count = max(0, len(technical_members) - len(staff_preview))
    pending_cards = _build_team_pending_cards(primary_team, weekly_brief)
    recent_activity = _build_team_recent_activity(primary_team)
    return render(
        request,
        'football/coach_overview.html',
        {
            'sources': sources,
            'weekly_brief': weekly_brief,
            'technical_members': technical_members,
            'staff_preview': staff_preview,
            'staff_extra_count': staff_extra_count,
            'rival_summary': rival_summary,
            'hero_image_url': hero_image_url,
            'team_display_name': team_display_name,
            'team_crest_url': team_crest_url,
            'competition_label': competition_label,
            'next_match': next_match,
            'next_match_opponent': next_match_opponent,
            'next_match_date': next_match_date,
            'standings': standings_rows,
            'highlighted_standing': highlighted_standing,
            'opponent_standing': opponent_standing,
            'probable_eleven_names': probable_eleven_names,
            'module_hub': module_hub,
            'pending_items': pending_items,
            'pending_cards': pending_cards,
            'recent_activity': recent_activity,
        },
    )


@login_required
def incident_page(request):
    primary_team = _get_primary_team_for_request(request)
    team_display_name = str(getattr(primary_team, 'display_name', '') or getattr(primary_team, 'name', '') or 'Club').strip()
    team_crest_url = resolve_team_crest_url(request, primary_team, sync=True) if primary_team else ''
    return render(
        request,
        'football/incidents.html',
        {
            'title': 'Registro de incidencias',
            'team_display_name': team_display_name,
            'team_crest_url': team_crest_url,
        },
    )


def _normalize_lineup_payload(payload, allowed_players):
    allowed = {str(player.id): player for player in (allowed_players or [])}
    base = {'starters': [], 'bench': []}
    if not isinstance(payload, dict):
        return base
    for section in ('starters', 'bench'):
        rows = payload.get(section)
        if not isinstance(rows, list):
            continue
        for row in rows:
            if not isinstance(row, dict):
                continue
            pid = str(row.get('id') or '').strip()
            if not pid or pid not in allowed:
                continue
            player = allowed[pid]
            base[section].append(
                {
                    'id': str(player.id),
                    'name': (player.name or '').upper(),
                    'number': player.number if player.number is not None else '--',
                    'position': player.position or '',
                    'photo': resolve_player_photo_url(None, player),
                }
            )
    # dedupe: a player must appear only once across sections
    seen = set()
    for section in ('starters', 'bench'):
        deduped = []
        for row in base[section]:
            pid = str(row.get('id'))
            if pid in seen:
                continue
            seen.add(pid)
            deduped.append(row)
        base[section] = deduped
    if len(base['starters']) > 11:
        overflow = base['starters'][11:]
        base['starters'] = base['starters'][:11]
        base['bench'].extend(overflow)
    return base


def _build_default_lineup_payload(convocation_players):
    if not convocation_players:
        return {'starters': [], 'bench': []}
    sorted_players = sorted(
        convocation_players,
        key=lambda p: ((p.number if p.number is not None else 999), (p.name or '').lower()),
    )
    starters = sorted_players[:11]
    bench = sorted_players[11:]
    return _normalize_lineup_payload(
        {
            'starters': [{'id': str(p.id)} for p in starters],
            'bench': [{'id': str(p.id)} for p in bench],
        },
        convocation_players,
    )


@login_required
def match_action_page(request):
    if not _can_edit_match_actions(request.user):
        return HttpResponse('Solo el cuerpo técnico puede editar estadísticas de partido.', status=403)
    forbidden = _forbid_if_workspace_module_disabled(request, 'match_actions', label='registro de acciones')
    if forbidden:
        return forbidden
    primary_team = _get_primary_team_for_request(request)
    if not primary_team:
        raise Http404('Equipo principal no configurado')
    requested_match = get_requested_match(request, primary_team)
    active_match = requested_match or get_active_match(primary_team)
    convocation_record = get_current_convocation_record(
        primary_team,
        match=active_match,
        fallback_to_latest=True,
    )
    convocation_players = convocation_record.players.order_by('name') if convocation_record else Player.objects.none()
    convocation_players = list(convocation_players)
    for player in convocation_players:
        player.photo_url = resolve_player_photo_url(request, player)
    initial_lineup_payload = {'starters': [], 'bench': []}
    if convocation_record:
        stored_lineup = convocation_record.lineup_data if isinstance(convocation_record.lineup_data, dict) else {}
        normalized_stored = _normalize_lineup_payload(stored_lineup, convocation_players)
        if normalized_stored['starters'] or normalized_stored['bench']:
            initial_lineup_payload = normalized_stored
        else:
            initial_lineup_payload = _build_default_lineup_payload(convocation_players)
            convocation_record.lineup_data = initial_lineup_payload
            convocation_record.save(update_fields=['lineup_data'])
    message = None
    if request.method == 'POST':
        action = request.POST.get('action_type', '').strip()
        player_id = request.POST.get('player')
        player = next((item for item in convocation_players if str(item.id) == str(player_id)), None)
        if action and player:
            message = f"Acción de partido para {player.name} registrada ({action})."
        else:
            message = "Completa el jugador y el tipo de acción."
    # Persistencia de registro en vivo:
    # prioriza siempre las acciones pendientes del partido activo (touch-field)
    # para que al recargar no "desaparezcan" hasta guardar/finalizar.
    if active_match:
        recent_events = (
            MatchEvent.objects.filter(
                match=active_match,
                source_file='registro-acciones',
                system='touch-field',
            )
            .select_related('player')
            .order_by('-created_at', '-id')[:120]
        )
    else:
        recent_match_ids = list(
            Match.objects.filter(Q(home_team=primary_team) | Q(away_team=primary_team))
            .order_by('-date', '-id')
            .values_list('id', flat=True)[:12]
        )
        recent_events = (
            MatchEvent.objects.filter(match_id__in=recent_match_ids)
            .select_related('player')
            .order_by('-created_at', '-id')[:20]
        )
    official_next = load_preferred_next_match_payload(primary_team=primary_team)

    def _payload_date_label(payload):
        raw_date = (payload or {}).get('date')
        if not raw_date:
            return None
        value = str(raw_date).strip()
        for fmt in ('%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y'):
            try:
                return datetime.strptime(value, fmt).strftime('%d/%m/%Y')
            except ValueError:
                continue
        return None

    substitution_history = []
    match_info = None
    if active_match:
        opponent = active_match.away_team if active_match.home_team == primary_team else active_match.home_team
        match_info = {
            'match_id': active_match.id,
            'opponent': opponent.name if opponent else '',
            'location': active_match.location or '',
            'round': active_match.round or '',
            'date': active_match.date.strftime('%d/%m/%Y') if active_match.date else None,
            'time': active_match.date.strftime('%H:%M') if active_match.date else '00:00',
        }
        if official_next:
            official_opponent = _payload_opponent_name(official_next)
            official_round = str(official_next.get('round') or '').strip()
            official_location = str(official_next.get('location') or '').strip()
            official_date_label = _payload_date_label(official_next)
            if official_opponent:
                match_info['opponent'] = official_opponent
            if official_round:
                match_info['round'] = official_round
            if official_location:
                match_info['location'] = official_location
            if official_date_label:
                match_info['date'] = official_date_label
        if convocation_record:
            if convocation_record.opponent_name:
                match_info['opponent'] = convocation_record.opponent_name
            if convocation_record.round:
                match_info['round'] = convocation_record.round
            if convocation_record.location:
                match_info['location'] = convocation_record.location
            if convocation_record.match_date:
                match_info['date'] = convocation_record.match_date.strftime('%d/%m/%Y')
            if convocation_record.match_time:
                match_info['time'] = convocation_record.match_time.strftime('%H:%M')
        substitution_events = (
            MatchEvent.objects.filter(match=active_match, player__team=primary_team)
            .select_related('player')
            .order_by('created_at', 'id')
        )
        for event in substitution_events:
            if not (
                is_substitution_event(event.event_type, event.zone)
                or is_substitution_entry(event.event_type, event.result, event.zone)
                or is_substitution_exit(event.event_type, event.result, event.zone)
            ):
                continue
            minute_label = f"{event.minute}'" if event.minute is not None else "--'"
            result_label = (event.result or '').strip() or (event.zone or '').strip() or 'Sustitución'
            player_name = event.player.name if event.player else 'Jugador'
            substitution_history.append(f"{player_name} · {minute_label} · {result_label}".upper())
    universo_lookup = _build_universo_standings_lookup(load_universo_snapshot())
    category_rivals = []
    team_fields = []
    group = primary_team.group
    if group:
        team_fields = gather_team_fields_for_group(group)
        field_map = {item['team_slug']: item['location'] for item in team_fields if item.get('team_slug')}
        standings = (
            TeamStanding.objects.filter(group=group)
            .select_related('team')
            .order_by('position')
        )
        for standing in standings:
            team = standing.team
            team_name = team.name if team else ''
            team_meta = universo_lookup.get(_normalize_team_lookup_key(team_name), {})
            full_name = team_meta.get('full_name') or team_name
            category_rivals.append(
                {
                    'position': standing.position,
                    'name': full_name,
                    'short_name': team.short_name or full_name,
                    'crest_url': team_meta.get('crest_url') or '',
                    'team_code': team_meta.get('team_code') or '',
                    'points': standing.points,
                    'played': standing.played,
                    'slug': team.slug,
                    'is_primary': team == primary_team,
                    'field_location': field_map.get(team.slug, ''),
                }
            )
    match_selector_options = list(
        _team_match_queryset(primary_team).order_by('-date', '-id')
    )
    selected_match_id = active_match.id if active_match else None
    return render(
        request,
        'football/match_actions.html',
        {
            'players': convocation_players,
            'convocation_players': convocation_players,
            'avatar_url': request.build_absolute_uri(static('football/images/player-avatar.svg')),
            'message': message,
            'team_name': primary_team.name,
            'team_crest_url': resolve_team_crest_url(request, primary_team, sync=True) or '',
            'quick_actions': load_match_actions(),
            'field_zone_defs': FIELD_ZONES,
            'result_options': load_match_results(),
            'tercio_options': STANDARD_TERCIO_LABELS,
            'recent_events': recent_events,
            'match_info': match_info,
            'category_rivals': category_rivals,
            'team_fields': team_fields,
            'substitution_history': substitution_history,
            'initial_lineup_json': json.dumps(initial_lineup_payload, ensure_ascii=False),
            'match_selector_options': match_selector_options,
            'selected_match_id': selected_match_id,
        },
    )


@authenticated_write
@require_POST
def register_match_action(request):
    if not _can_edit_match_actions(request.user):
        return JsonResponse({'error': 'Solo el cuerpo técnico puede editar estadísticas de partido.'}, status=403)
    forbidden = _forbid_if_workspace_module_disabled(request, 'match_actions', label='registro de acciones')
    if forbidden:
        return JsonResponse({'error': 'El registro de acciones no está activo en el workspace actual.'}, status=403)
    primary_team = _get_primary_team_for_request(request)
    if not primary_team:
        return JsonResponse({'error': 'Equipo principal no configurado'}, status=400)
    player_id = request.POST.get('player')
    action_type = (request.POST.get('action_type') or '').strip()
    action_type_key = action_type.lower()
    requested_match = get_requested_match(request, primary_team)
    target_match = requested_match or get_active_match(primary_team)
    convocation_record = get_current_convocation_record(
        primary_team,
        match=target_match,
        fallback_to_latest=True,
    )
    if not convocation_record:
        return JsonResponse({'error': 'No hay convocatoria activa guardada para registrar acciones'}, status=400)

    player = None
    if not _is_team_only_action(action_type_key):
        player = convocation_record.players.filter(id=player_id).first()
        if not player:
            return JsonResponse({'error': 'Selecciona un jugador convocado válido'}, status=400)

    if not action_type:
        return JsonResponse({'error': 'Especifica el tipo de acción'}, status=400)
    match = target_match
    if not match:
        return JsonResponse({'error': 'No hay partido disponible para registrar acciones'}, status=400)
    if player and (not convocation_record.players.filter(id=player.id).exists()):
        return JsonResponse({'error': 'Jugador fuera de convocatoria para este partido'}, status=400)
    minute = _parse_int(request.POST.get('minute'))
    if minute is not None:
        minute = max(0, min(minute, 120))
    period = _parse_int(request.POST.get('period'))
    result = (request.POST.get('result') or '').strip()
    zone = (request.POST.get('zone') or '').strip()
    tercio = zone_to_tercio(zone)
    observation = (request.POST.get('observation') or '').strip()
    duplicate_window = timezone.now() - timedelta(seconds=8)
    recent_duplicates = MatchEvent.objects.filter(
        match=match,
        player=player if player else None,
        minute=minute if minute is not None else None,
        period=period,
        source_file='registro-acciones',
        system='touch-field',
        created_at__gte=duplicate_window,
    ).order_by('-id')
    for existing in recent_duplicates:
        if (
            _canonical_action_value(existing.event_type) == _canonical_action_value(action_type)
            and _canonical_action_value(existing.result) == _canonical_action_value(result)
            and _canonical_action_value(existing.zone) == _canonical_action_value(zone)
            and _canonical_action_value(existing.tercio) == _canonical_action_value(tercio)
            and _canonical_action_value(existing.observation) == _canonical_action_value(observation)
        ):
            return JsonResponse(_serialize_match_event(existing, duplicate=True))

    event = MatchEvent.objects.create(
        match=match,
        player=player if player else None,
        minute=minute if minute is not None else None,
        period=period,
        event_type=action_type,
        result=result,
        zone=zone,
        tercio=tercio,
        observation=observation,
        source_file='registro-acciones',
        system='touch-field',
    )
    _invalidate_team_dashboard_caches(primary_team)
    return JsonResponse(_serialize_match_event(event, duplicate=False))


@authenticated_write
@require_POST
def save_match_lineup(request):
    if not _can_edit_match_actions(request.user):
        return JsonResponse({'error': 'Solo el cuerpo técnico puede editar estadísticas de partido.'}, status=403)
    convocation_forbidden = _forbid_if_workspace_module_disabled(request, 'convocation', label='convocatoria')
    match_actions_forbidden = _forbid_if_workspace_module_disabled(
        request,
        'match_actions',
        label='registro de acciones',
    )
    if convocation_forbidden and match_actions_forbidden:
        return JsonResponse(
            {'error': 'La convocatoria o el registro de acciones deben estar activos en el workspace actual.'},
            status=403,
        )
    primary_team = _get_primary_team_for_request(request)
    if not primary_team:
        return JsonResponse({'error': 'Equipo principal no configurado'}, status=400)
    requested_match = get_requested_match(request, primary_team)
    target_match = requested_match or get_active_match(primary_team)
    convocation_record = get_current_convocation_record(
        primary_team,
        match=target_match,
        fallback_to_latest=True,
    )
    if not convocation_record:
        return JsonResponse({'error': 'No hay convocatoria activa para guardar el 11'}, status=400)
    payload = {}
    try:
        payload = json.loads(request.body.decode('utf-8')) if request.body else {}
    except Exception:
        return JsonResponse({'error': 'Payload inválido'}, status=400)
    lineup = payload.get('lineup')
    allowed_players = list(convocation_record.players.all())
    normalized = _normalize_lineup_payload(lineup, allowed_players)
    convocation_record.lineup_data = normalized
    convocation_record.save(update_fields=['lineup_data'])
    _invalidate_team_dashboard_caches(primary_team)
    starters_count = len(normalized['starters'])
    return JsonResponse(
        {
            'saved': True,
            'starters': starters_count,
            'bench': len(normalized['bench']),
            'pending_lineup': starters_count < 11,
        }
    )


@authenticated_write
@require_POST
def delete_match_action(request):
    if not _can_edit_match_actions(request.user):
        return JsonResponse({'error': 'Solo el cuerpo técnico puede editar estadísticas de partido.'}, status=403)
    forbidden = _forbid_if_workspace_module_disabled(request, 'match_actions', label='registro de acciones')
    if forbidden:
        return JsonResponse({'error': 'El registro de acciones no está activo en el workspace actual.'}, status=403)
    primary_team = _get_primary_team_for_request(request)
    if not primary_team:
        return JsonResponse({'error': 'Equipo principal no configurado'}, status=400)
    event_id = request.POST.get('event_id')
    requested_match = get_requested_match(request, primary_team)
    candidate_events = MatchEvent.objects.filter(
        Q(match__home_team=primary_team) | Q(match__away_team=primary_team)
    ).filter(
        source_file='registro-acciones',
        system='touch-field',
    )
    if requested_match:
        candidate_events = candidate_events.filter(match=requested_match)
    try:
        event = candidate_events.get(id=event_id)
    except MatchEvent.DoesNotExist:
        return JsonResponse({'error': 'Evento no encontrado'}, status=404)
    event.delete()
    _invalidate_team_dashboard_caches(primary_team)
    return JsonResponse({'deleted': event_id})


@authenticated_write
@require_POST
def finalize_match_actions(request):
    if not _can_edit_match_actions(request.user):
        return JsonResponse({'error': 'Solo el cuerpo técnico puede editar estadísticas de partido.'}, status=403)
    forbidden = _forbid_if_workspace_module_disabled(request, 'match_actions', label='registro de acciones')
    if forbidden:
        return JsonResponse({'error': 'El registro de acciones no está activo en el workspace actual.'}, status=403)
    primary_team = _get_primary_team_for_request(request)
    if not primary_team:
        return JsonResponse({'error': 'Equipo principal no configurado'}, status=400)
    requested_match = get_requested_match(request, primary_team)
    match = requested_match or get_active_match(primary_team)
    if not match:
        return JsonResponse({'error': 'No hay partido activo para guardar'}, status=400)
    payload = {}
    try:
        if request.body:
            payload = json.loads(request.body.decode('utf-8'))
    except Exception:
        payload = {}
    _apply_match_info_overrides(match, primary_team, payload.get('match_info'))
    pending_events = list(
        MatchEvent.objects.filter(
            match=match,
            system='touch-field',
            source_file='registro-acciones',
        ).select_related('player')
    )
    if not pending_events:
        return JsonResponse({'saved': True, 'updated': 0, 'match_id': match.id})
    # Limpiar tarjetas y sustituciones anteriores
    MatchEvent.objects.filter(
        match=match,
        system='touch-field-final',
        source_file='registro-acciones',
    ).filter(
        Q(event_type__icontains='tarjeta') | Q(event_type__icontains='sustitucion') | Q(event_type__icontains='sustitución') | Q(event_type__icontains='cambio'),
    ).delete()

    # Evita consolidar duplicados por doble click/reintento de red en pocos segundos.
    # Importante: no eliminar acciones reales repetidas a lo largo del partido.
    dedupe_seconds = 12
    existing_final_by_signature = defaultdict(list)
    for event in MatchEvent.objects.filter(
        match=match,
        system='touch-field-final',
        source_file='registro-acciones',
    ).select_related('player'):
        existing_final_by_signature[_match_action_dedupe_signature(event)].append(event.created_at)
    seen_pending_by_signature = defaultdict(list)
    keep_ids = []
    drop_ids = []
    for event in sorted(pending_events, key=lambda e: e.created_at or timezone.now()):
        signature = _match_action_dedupe_signature(event)
        created_at = event.created_at or timezone.now()
        existing_times = existing_final_by_signature.get(signature, [])
        pending_times = seen_pending_by_signature.get(signature, [])
        is_near_duplicate = any(
            abs((created_at - known).total_seconds()) <= dedupe_seconds
            for known in [*existing_times, *pending_times]
        )
        if is_near_duplicate:
            drop_ids.append(event.id)
            continue
        seen_pending_by_signature[signature].append(created_at)
        keep_ids.append(event.id)

    if drop_ids:
        MatchEvent.objects.filter(id__in=drop_ids, match=match, system='touch-field').delete()

    updated = 0
    if keep_ids:
        updated = MatchEvent.objects.filter(id__in=keep_ids).update(system='touch-field-final')
    _invalidate_team_dashboard_caches(primary_team)
    return JsonResponse(
        {
            'saved': True,
            'updated': updated,
            'deduplicated': len(drop_ids),
            'match_id': match.id,
            'match_label': str(match),
        }
    )


@authenticated_write
@require_POST
def reset_match_action_register(request):
    if not _can_edit_match_actions(request.user):
        return JsonResponse({'error': 'Solo el cuerpo técnico puede editar estadísticas de partido.'}, status=403)
    forbidden = _forbid_if_workspace_module_disabled(request, 'match_actions', label='registro de acciones')
    if forbidden:
        return JsonResponse({'error': 'El registro de acciones no está activo en el workspace actual.'}, status=403)
    primary_team = _get_primary_team_for_request(request)
    if not primary_team:
        return JsonResponse({'error': 'Equipo principal no configurado'}, status=400)
    requested_match = get_requested_match(request, primary_team)
    match = requested_match or get_active_match(primary_team)
    if not match:
        return JsonResponse({'error': 'No hay partido activo para reiniciar'}, status=400)
    deleted_count, _ = MatchEvent.objects.filter(
        match=match,
        source_file='registro-acciones',
    ).filter(
        Q(system='touch-field'),
    ).delete()
    _invalidate_team_dashboard_caches(primary_team)
    return JsonResponse(
        {
            'reset': True,
            'deleted': int(deleted_count),
            'match_id': match.id,
        }
    )


@login_required
@ensure_csrf_cookie
def convocation_page(request):
    forbidden = _forbid_if_workspace_module_disabled(request, 'convocation', label='convocatoria')
    if forbidden:
        return forbidden
    primary_team = _get_primary_team_for_request(request)
    if not primary_team:
        raise Http404('Equipo principal no configurado')
    all_players = list(Player.objects.filter(team=primary_team, is_active=True).order_by('name'))
    for player in all_players:
        player.photo_url = resolve_player_photo_url(request, player)
    roster_cache = get_roster_stats_cache()
    manual_overrides = get_manual_player_base_overrides(primary_team)
    universo_snapshot = load_universo_snapshot() or {}
    universo_players = universo_snapshot.get('players') if isinstance(universo_snapshot, dict) else []
    universo_map = {}
    universo_by_number = {}
    if isinstance(universo_players, list):
        for item in universo_players:
            if not isinstance(item, dict):
                continue
            team_name = str(item.get('team') or '').strip().lower()
            if team_name and 'benagalbon' not in team_name:
                continue
            name = str(item.get('name') or '').strip()
            if not name:
                continue
            key = normalize_player_name(name)
            universo_map[key] = item
            dorsal_raw = str(item.get('dorsal') or '').strip()
            if dorsal_raw.isdigit():
                universo_by_number[int(dorsal_raw)] = item

    def _find_universo_entry(player_obj):
        key = normalize_player_name(player_obj.name)
        direct = universo_map.get(key)
        if direct:
            return direct
        if player_obj.number is not None and player_obj.number in universo_by_number:
            return universo_by_number[player_obj.number]
        compact = key.replace('-', '')
        for ukey, entry in universo_map.items():
            ucompact = ukey.replace('-', '')
            if compact in ucompact or ucompact in compact:
                return entry
        tokens = [token for token in compact.split('-') if token]
        for ukey, entry in universo_map.items():
            u_tokens = [token for token in ukey.replace('-', ' ').split() if token]
            overlap = sum(1 for token in tokens if token in u_tokens)
            if overlap >= 2:
                return entry
        return {}
    active_injury_ids = get_active_injury_player_ids([p.id for p in all_players])
    active_match = get_active_match(primary_team)
    sanctioned_player_ids = get_sanctioned_player_ids_from_previous_round(
        primary_team,
        reference_match=active_match,
    )
    filtered_players = []
    today = timezone.localdate()
    for player in all_players:
        roster_entry = find_roster_entry(player.name, roster_cache) or {}
        manual_entry = manual_overrides.get(player.id, {})
        universo_entry = _find_universo_entry(player) or {}
        yellow_cards = (
            manual_entry.get('yellow_cards')
            if manual_entry.get('yellow_cards') is not None
            else _parse_int(universo_entry.get('yellow_cards'))
            if universo_entry.get('yellow_cards') not in (None, '')
            else roster_entry.get('yellow_cards', 0)
        )
        red_cards = (
            manual_entry.get('red_cards')
            if manual_entry.get('red_cards') is not None
            else _parse_int(universo_entry.get('red_cards'))
            if universo_entry.get('red_cards') not in (None, '')
            else roster_entry.get('red_cards', 0)
        )
        player.yellow_cards = int(yellow_cards or 0)
        player.red_cards = int(red_cards or 0)
        player.is_sanctioned = (player.id in sanctioned_player_ids) or is_manual_sanction_active(player, today=today)
        player.is_apercibido = player.yellow_cards in {4, 9, 14}
        player.has_active_injury = player.id in active_injury_ids
        filtered_players.append(player)
    players = filtered_players
    convocation_record = get_current_convocation_record(primary_team)
    selected_player_ids = []
    if convocation_record:
        available_ids = {player.id for player in players}
        selected_player_ids = [
            player_id
            for player_id in convocation_record.players.values_list('id', flat=True)
            if player_id in available_ids
        ]

    next_match_payload = load_preferred_next_match_payload(primary_team=primary_team)
    if not next_match_payload and primary_team.group:
        next_match_payload = get_next_match(primary_team, primary_team.group)
    if isinstance(next_match_payload, dict) and str(next_match_payload.get('status') or '').lower() != 'next':
        next_match_payload = None
    default_match_info = normalize_next_match_payload(next_match_payload or {}) if next_match_payload else {}

    opponent_name = ''
    if isinstance(default_match_info, dict):
        opponent = default_match_info.get('opponent')
        if isinstance(opponent, dict):
            opponent_name = str(opponent.get('name') or '').strip()
        elif isinstance(opponent, str):
            opponent_name = opponent.strip()

    match_info = {
        'round': str(default_match_info.get('round') or '').strip() if isinstance(default_match_info, dict) else '',
        'date': str(default_match_info.get('date') or '').strip() if isinstance(default_match_info, dict) else '',
        'time': str(default_match_info.get('time') or '').strip() if isinstance(default_match_info, dict) else '',
        'location': str(default_match_info.get('location') or '').strip() if isinstance(default_match_info, dict) else '',
        'opponent': opponent_name,
    }
    if convocation_record:
        if convocation_record.round:
            match_info['round'] = convocation_record.round
        if convocation_record.match_date:
            match_info['date'] = convocation_record.match_date.isoformat()
        if convocation_record.match_time:
            match_info['time'] = convocation_record.match_time.strftime('%H:%M')
        if convocation_record.location:
            match_info['location'] = convocation_record.location
        if convocation_record.opponent_name:
            match_info['opponent'] = convocation_record.opponent_name
    rival_full_name, rival_crest_url = _resolve_rival_identity(
        match_info.get('opponent') or 'Rival por confirmar',
        preferred_opponent=(default_match_info.get('opponent') if isinstance(default_match_info, dict) else None),
    )
    match_info['rival_full_name'] = rival_full_name
    match_info['rival_crest_url'] = rival_crest_url

    home_location = 'ESTADIO CAÑA CHAQUETA'
    recent_home_match = (
        Match.objects.filter(
            group=primary_team.group,
            home_team=primary_team,
        )
        .exclude(location__isnull=True)
        .exclude(location__exact='')
        .order_by('-date', '-id')
        .first()
    )
    if recent_home_match and recent_home_match.location:
        home_location = recent_home_match.location.strip()

    team_fields = gather_team_fields_for_group(primary_team.group)
    field_map = {
        normalize_label(item.get('team_name') or ''): (item.get('location') or '').strip()
        for item in team_fields
        if item.get('team_name')
    }

    opponent_options = []
    seen_opponents = set()
    group_teams = Team.objects.filter(group=primary_team.group).order_by('name') if primary_team.group else Team.objects.none()
    for team in group_teams:
        if team.id == primary_team.id:
            continue
        key = normalize_label(team.name)
        if not key or key in seen_opponents:
            continue
        seen_opponents.add(key)
        opponent_options.append(
            {
                'name': team.name,
                'short_name': team.display_name,
                'location': field_map.get(key, ''),
                'crest_url': resolve_team_crest_url(request, team, fallback_static=None, sync=True),
            }
        )

    current_opponent = str(match_info.get('opponent') or '').strip()
    current_key = normalize_label(current_opponent)
    if current_opponent and current_key and current_key not in seen_opponents:
        opponent_options.append(
            {
                'name': current_opponent,
                'short_name': current_opponent,
                'location': field_map.get(current_key, ''),
                'crest_url': rival_crest_url,
            }
        )

    return render(
        request,
        'football/convocation.html',
        {
            'players': players,
            'team_name': primary_team.display_name,
            'team_crest_url': resolve_team_crest_url(request, primary_team, sync=True),
            'avatar_url': request.build_absolute_uri(static('football/images/player-avatar.svg')),
            'selected_player_ids_json': json.dumps(selected_player_ids),
            'injured_player_ids_json': json.dumps(
                [p.id for p in all_players if getattr(p, 'has_active_injury', False)]
            ),
            'match_info': match_info,
            'has_saved_convocation': bool(convocation_record and selected_player_ids),
            'has_pending_convocation': bool(convocation_record and not selected_player_ids),
            'can_generate_convocation_pdf': bool(convocation_record and selected_player_ids),
            'opponent_options_json': json.dumps(opponent_options, ensure_ascii=False),
            'home_location_label': home_location,
        },
    )


@authenticated_write
@require_POST
def save_convocation(request):
    forbidden = _forbid_if_workspace_module_disabled(request, 'convocation', label='convocatoria')
    if forbidden:
        return JsonResponse({'error': 'La convocatoria no está activa en el workspace actual.'}, status=403)
    primary_team = _get_primary_team_for_request(request)
    if not primary_team:
        return JsonResponse({'error': 'Equipo principal no configurado'}, status=400)
    try:
        payload = json.loads(request.body or '[]')
    except json.JSONDecodeError:
        return JsonResponse({'error': 'Formato JSON inválido'}, status=400)

    if isinstance(payload, dict):
        raw_players = payload.get('players') or payload.get('player_ids') or []
        match_info = payload.get('match_info') if isinstance(payload.get('match_info'), dict) else {}
    else:
        raw_players = payload
        match_info = {}

    try:
        player_ids = [int(pid) for pid in raw_players if pid]
    except (TypeError, ValueError):
        return JsonResponse({'error': 'Formato de jugadores inválido'}, status=400)
    round_value = str(match_info.get('round') or '').strip()
    location_value = str(match_info.get('location') or '').strip()
    opponent_value = str(match_info.get('opponent') or '').strip()
    date_value_raw = str(match_info.get('date') or '').strip()
    time_value_raw = str(match_info.get('time') or '').strip()

    players = Player.objects.filter(team=primary_team, is_active=True, id__in=player_ids)
    blocked_injury_ids = get_active_injury_player_ids(players.values_list('id', flat=True))
    players = players.exclude(id__in=blocked_injury_ids)
    has_match_context = any([round_value, location_value, opponent_value, date_value_raw, time_value_raw])
    if not players.exists() and not has_match_context:
        return JsonResponse({'error': 'Indica al menos los datos del próximo partido o una lista de jugadores.'}, status=400)

    parsed_match_date = None
    if date_value_raw:
        for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y'):
            try:
                parsed_match_date = datetime.strptime(date_value_raw, fmt).date()
                break
            except ValueError:
                continue
    parsed_match_time = None
    if time_value_raw:
        for fmt in ('%H:%M', '%H:%M:%S'):
            try:
                parsed_match_time = datetime.strptime(time_value_raw, fmt).time()
                break
            except ValueError:
                continue

    target_match = get_active_match(primary_team)
    if not target_match:
        season = None
        if primary_team.group and primary_team.group.season:
            season = primary_team.group.season
        if not season:
            season = Season.objects.filter(is_current=True).order_by('-start_date', '-id').first()
        if season:
            target_match = Match.objects.create(
                season=season,
                group=primary_team.group,
                round=round_value,
                date=parsed_match_date,
                location=location_value,
                home_team=primary_team,
                away_team=None,
            )

    if target_match:
        datetime_label = date_value_raw
        if time_value_raw:
            datetime_label = f'{date_value_raw} · {time_value_raw}' if date_value_raw else time_value_raw
        _apply_match_info_overrides(
            target_match,
            primary_team,
            {
                'round': round_value,
                'location': location_value,
                'datetime': datetime_label,
                'opponent': opponent_value,
            },
        )

    active_injury_ids = get_active_injury_player_ids(players.values_list('id', flat=True))
    injured_players = [
        player.name
        for player in players
        if player.id in active_injury_ids
    ]

    with transaction.atomic():
        ConvocationRecord.objects.filter(team=primary_team, is_current=True).update(is_current=False)
        record = ConvocationRecord.objects.create(
            team=primary_team,
            match=target_match,
            round=round_value,
            match_date=parsed_match_date,
            match_time=parsed_match_time,
            location=location_value,
            opponent_name=opponent_value,
        )
        record.players.set(players.distinct())
    _invalidate_team_dashboard_caches(primary_team)
    pending = players.count() == 0
    return JsonResponse(
        {
            'saved': True,
            'count': players.count(),
            'pending_convocation': pending,
            'match_id': target_match.id if target_match else None,
            'injury_warning_count': len(injured_players),
            'injury_warning_players': injured_players[:8],
        }
    )


@login_required
def convocation_pdf(request):
    forbidden = _forbid_if_workspace_module_disabled(request, 'convocation', label='convocatoria')
    if forbidden:
        return forbidden
    primary_team = _get_primary_team_for_request(request)
    if not primary_team:
        raise Http404('Equipo principal no configurado')
    convocation_record = get_current_convocation_record(primary_team)
    if not convocation_record:
        return HttpResponse('No hay convocatoria guardada.', status=400)

    players = list(convocation_record.players.order_by('number', 'name'))
    if not players:
        return HttpResponse('No hay jugadores convocados.', status=400)

    def _sort_player_key(player):
        number = player.number if player.number is not None else 999
        return (number, (player.name or '').lower())

    ordered_players = sorted(players, key=_sort_player_key)

    static_base_dir = Path(settings.BASE_DIR) / 'static'
    logo_static_path = static_base_dir / 'football' / 'images' / 'cdb-logo.png'
    avatar_static_path = static_base_dir / 'football' / 'images' / 'player-avatar.svg'
    logo_data_uri = _file_as_data_uri(logo_static_path)
    avatar_data_uri = _file_as_data_uri(avatar_static_path)

    include_player_photos = str(
        os.getenv('CONVOCATION_PDF_PLAYER_PHOTOS', '0')
    ).strip().lower() in {'1', 'true', 'yes', 'on'}

    def _player_photo_src(player_obj):
        # Default to lightweight avatar for reliability/performance on Render.
        if not include_player_photos:
            if avatar_data_uri:
                return avatar_data_uri
            return request.build_absolute_uri(static('football/images/player-avatar.svg'))

        player_static_rel = resolve_player_photo_static_path(player_obj)
        if player_static_rel:
            player_data_uri = _image_file_as_small_data_uri(
                static_base_dir / player_static_rel,
                max_width=220,
                max_height=220,
                quality=65,
            )
            if player_data_uri:
                return player_data_uri
            return request.build_absolute_uri(static(player_static_rel))
        if avatar_data_uri:
            return avatar_data_uri
        return request.build_absolute_uri(static('football/images/player-avatar.svg'))

    player_rows = [
        {
            'number': player.number,
            'name': player.name,
            'photo_src': _player_photo_src(player),
        }
        for player in ordered_players
    ]
    midpoint = (len(player_rows) + 1) // 2
    left_column_players = player_rows[:midpoint]
    right_column_players = player_rows[midpoint:]

    date_label = convocation_record.match_date.strftime('%d/%m/%Y') if convocation_record.match_date else '--'
    time_label = convocation_record.match_time.strftime('%H:%M') if convocation_record.match_time else '--:--'
    location_label = convocation_record.location or (convocation_record.match.location if convocation_record.match else '')
    rival_label = convocation_record.opponent_name
    if not rival_label and convocation_record.match:
        opponent = (
            convocation_record.match.away_team
            if convocation_record.match.home_team_id == primary_team.id
            else convocation_record.match.home_team
        )
        rival_label = opponent.name if opponent else ''

    rival_name = (rival_label or '').strip() or 'Rival por confirmar'
    preferred_next = load_preferred_next_match_payload(primary_team=primary_team)
    preferred_opponent = preferred_next.get('opponent') if isinstance(preferred_next, dict) else None
    rival_full_name, rival_crest_url = _resolve_rival_identity(
        rival_name,
        preferred_opponent=preferred_opponent,
    )
    rival_crest_url = _sanitize_universo_external_image(_absolute_universo_url(rival_crest_url))

    round_digits = ''.join(re.findall(r'\d+', convocation_record.round or ''))
    round_short = f'J{round_digits}' if round_digits else 'J'
    date_human = date_label
    if convocation_record.match_date:
        day_map = ['LUN', 'MAR', 'MIÉ', 'JUE', 'VIE', 'SÁB', 'DOM']
        month_map = ['ENE', 'FEB', 'MAR', 'ABR', 'MAY', 'JUN', 'JUL', 'AGO', 'SEP', 'OCT', 'NOV', 'DIC']
        weekday = day_map[convocation_record.match_date.weekday()]
        month = month_map[convocation_record.match_date.month - 1]
        date_human = f'{weekday} {convocation_record.match_date.day} {month}'

    rival_crest_src = str(rival_crest_url or '').strip()
    if rival_crest_src.startswith('http://') or rival_crest_src.startswith('https://'):
        # Avoid external HTTP fetches during PDF rendering to reduce timeout/502 risk on Render.
        rival_crest_src = ''

    context = {
        'team_name': primary_team.display_name,
        'team_full_name': primary_team.name,
        'round_label': convocation_record.round or 'Jornada por confirmar',
        'round_short': round_short,
        'date_label': date_label,
        'date_human': date_human,
        'time_label': time_label,
        'location_label': location_label or 'Campo por confirmar',
        'rival_label': rival_name,
        'rival_full_name': rival_full_name,
        'rival_crest_src': rival_crest_src,
        'players': player_rows,
        'left_column_players': left_column_players,
        'right_column_players': right_column_players,
        'coach_name': os.getenv('TEAM_COACH_NAME', 'Aitor Castillo'),
        'club_hashtag': os.getenv('TEAM_HASHTAG', '#VamosVerdes'),
        'logo_src': resolve_team_crest_url(request, primary_team, sync=True) or logo_data_uri or request.build_absolute_uri(static('football/images/cdb-logo.png')),
        'brand_mark_url': request.build_absolute_uri(static('football/images/2j-mark.svg')),
        'avatar_src': avatar_data_uri or request.build_absolute_uri(static('football/images/player-avatar.svg')),
        'team_photo_url': request.build_absolute_uri(static('football/images/team-01.jpg')),
        'team_photo_data_uri': '',
        'coach_photo_url': request.build_absolute_uri(static(os.getenv('TEAM_COACH_PHOTO', 'football/images/team-01.jpg'))),
    }
    team_photo = resolve_team_photo_for_pdf(request)
    context['team_photo_url'] = team_photo.get('url') or context['team_photo_url']
    context['coach_photo_url'] = context['team_photo_url']
    context['team_photo_data_uri'] = team_photo.get('data_uri') or ''

    html = render_to_string('football/convocation_pdf.html', context)
    filename = f'convocatoria-{timezone.localdate().isoformat()}'
    return _build_pdf_response_or_html_fallback(request, html, filename)


def _task_pdf_lines(value):
    text = str(value or '').replace('\r', '\n')
    lines = [line.strip() for line in text.split('\n') if line.strip()]
    return lines or ['-']


def _team_pdf_palette(team_obj, style_key='uefa'):
    primary = str(getattr(team_obj, 'primary_color', '') or '').strip() or '#0f7a35'
    secondary = str(getattr(team_obj, 'secondary_color', '') or '').strip() or '#facc15'
    accent = str(getattr(team_obj, 'accent_color', '') or '').strip() or '#102734'
    if style_key == 'club':
        return {
            'primary': primary,
            'secondary': secondary,
            'accent': accent,
            'panel': '#f5fbf6',
            'sheet': '#ffffff',
            'ink': '#102734',
            'muted': '#51606f',
        }
    return {
        'primary': '#0e7490',
        'secondary': '#dbeafe',
        'accent': '#102734',
        'panel': '#f8fafc',
        'sheet': '#ffffff',
        'ink': '#111827',
        'muted': '#64748b',
    }


def _build_task_pdf_tokens(request, tactical_layout):
    material_icon_by_kind = {
        'cone': '△',
        'marker': '◉',
        'marker-flat': '◍',
        'pole': '┃',
        'ring': '◯',
        'ball': '●',
        'hurdle': '⊓',
        'ladder': '☷',
        'slalom': '╽',
        'mini-hurdle': '⫍',
        'goal': '⌷',
        'mini-goal': '⌸',
        'gk-goal': '▭',
        'mannequin': '♞',
        'dummy': '♜',
        'wall': '▥',
        'bib': '▣',
        'gps': '⌖',
        'timer': '◷',
        'whistle': '⌇',
        'board': '▤',
        'band': '∞',
        'trx': '⟂',
        'box': '▧',
        'sled': '⎍',
        'stake': '⎸',
        'arc': '◠',
        'target': '◎',
        'medicine-ball': '◒',
        'tape': '═',
    }
    tokens = []
    raw_tokens = tactical_layout.get('tokens') if isinstance(tactical_layout, dict) else []
    if not isinstance(raw_tokens, list):
        return tokens
    for token in raw_tokens:
        if not isinstance(token, dict):
            continue
        x = _parse_int(token.get('x'))
        y = _parse_int(token.get('y'))
        x = max(2, min(x if x is not None else 50, 98))
        y = max(2, min(y if y is not None else 50, 98))
        token_type = str(token.get('type') or '').strip()
        token_kind = str(token.get('kind') or '').strip()
        token_icon = ''
        token_asset = str(token.get('asset') or '').strip()
        if token_type == 'material':
            token_icon = str(token.get('icon') or '').strip() or material_icon_by_kind.get(token_kind, '•')
        token_asset_url = ''
        if token_asset:
            token_asset_url = request.build_absolute_uri(token_asset) if token_asset.startswith('/') else request.build_absolute_uri(static(token_asset))
        tokens.append(
            {
                'label': str(token.get('label') or '?')[:16],
                'title': str(token.get('title') or token.get('label') or '').strip(),
                'type': token_type,
                'kind': token_kind,
                'icon': token_icon,
                'asset_url': token_asset_url,
                'x': x,
                'y': y,
            }
        )
    return tokens


def _build_task_pdf_tokens_from_canvas_state(request, canvas_state, canvas_width=1280, canvas_height=720):
    raw_state = canvas_state if isinstance(canvas_state, dict) else {}
    objects = raw_state.get('objects') if isinstance(raw_state.get('objects'), list) else []
    width = max(320, _parse_int(canvas_width) or 1280)
    height = max(180, _parse_int(canvas_height) or 720)
    tokens = []
    for item in objects:
        if not isinstance(item, dict):
            continue
        item_type = str(item.get('type') or '').strip()
        data = item.get('data') if isinstance(item.get('data'), dict) else {}
        kind = str(data.get('kind') or item.get('kind') or item_type).strip()
        left = float(item.get('left') or 0)
        top = float(item.get('top') or 0)
        x = max(2, min(int(round((left / width) * 100)), 98))
        y = max(2, min(int(round((top / height) * 100)), 98))
        label = ''
        token_type = 'player'
        icon = ''
        if item_type in {'group', 'circle'} and (
            kind in {'token', 'player_local', 'player_rival', 'goalkeeper_local'}
            or 'player' in kind
            or 'goalkeeper' in kind
        ):
            label = str(data.get('playerName') or data.get('label') or '?').strip()[:16] or '?'
        elif item_type in {'i-text', 'textbox', 'text'} or kind == 'text':
            label = str(item.get('text') or 'Txt').strip()[:16] or 'Txt'
        else:
            token_type = 'material'
            label = str(data.get('label') or kind or item_type or '•').replace('_', ' ').strip()[:16] or '•'
            if 'arrow' in kind:
                icon = '➜'
            elif 'line' in kind:
                icon = '━'
            elif 'shape' in kind or kind == 'zone':
                icon = '▭'
            elif 'emoji_' in kind:
                icon = str(item.get('text') or '•').strip()[:2] or '•'
            elif kind == 'ball':
                icon = '⚽'
            elif kind == 'cone':
                icon = '△'
            else:
                icon = '•'
        tokens.append(
            {
                'label': label,
                'title': label or kind or item_type,
                'type': token_type,
                'kind': kind,
                'icon': icon,
                'asset_url': '',
                'x': x,
                'y': y,
            }
        )
    return tokens


def _normalize_task_pdf_meta(meta):
    def _normalize_meta_key(raw_key):
        text = str(raw_key or '').strip()
        if not text:
            return ''
        text = re.sub(r'([a-z0-9])([A-Z])', r'\1_\2', text)
        text = text.replace(' ', '_').replace('-', '_')
        text = re.sub(r'[^a-zA-Z0-9_]', '', text)
        text = re.sub(r'__+', '_', text).strip('_').lower()
        return text

    def _normalize_meta_dict(raw_value):
        if not isinstance(raw_value, dict):
            return {}
        normalized = {}
        for key, value in raw_value.items():
            normalized_key = _normalize_meta_key(key)
            if not normalized_key:
                continue
            if isinstance(value, dict):
                normalized_value = _normalize_meta_dict(value)
            elif isinstance(value, list):
                normalized_value = [
                    _normalize_meta_dict(item) if isinstance(item, dict) else item
                    for item in value
                ]
            else:
                normalized_value = value
            existing = normalized.get(normalized_key)
            if existing in (None, '', [], {}) and normalized_value not in (None, '', [], {}):
                normalized[normalized_key] = normalized_value
            elif normalized_key not in normalized:
                normalized[normalized_key] = normalized_value
        return normalized

    def _map_choice(value, mapping):
        raw = str(value or '').strip()
        if not raw:
            return raw
        if raw in mapping:
            return mapping.get(raw, raw)
        lowered = raw.lower()
        if lowered in mapping:
            return mapping.get(lowered, raw)
        normalized = lowered.replace(' ', '_').replace('-', '_')
        if normalized in mapping:
            return mapping.get(normalized, raw)
        normalized_key = _normalize_meta_key(raw)
        if normalized_key in mapping:
            return mapping.get(normalized_key, raw)
        return raw

    surface_map = {key: label for key, label in TASK_SURFACE_CHOICES}
    pitch_map = {key: label for key, label in TASK_PITCH_FORMAT_CHOICES}
    phase_map = {key: label for key, label in TASK_GAME_PHASE_CHOICES}
    methodology_map = {key: label for key, label in TASK_METHODOLOGY_CHOICES}
    complexity_map = {key: label for key, label in TASK_COMPLEXITY_CHOICES}
    constraint_map = {key: label for key, label in TASK_CONSTRAINT_CHOICES}
    strategy_map = {key: label for key, label in TASK_STRATEGY_CHOICES}
    coordination_map = {key: label for key, label in TASK_COORDINATION_CHOICES}
    coord_skills_map = {key: label for key, label in TASK_COORDINATION_SKILLS_CHOICES}
    tactical_intent_map = {key: label for key, label in TASK_TACTICAL_INTENT_CHOICES}
    dynamics_map = {key: label for key, label in TASK_DYNAMICS_CHOICES}
    structure_map = {key: label for key, label in TASK_STRUCTURE_CHOICES}
    meta = _normalize_meta_dict(meta or {})
    # Compat: algunas importaciones/analíticas guardan claves en castellano.
    # Unificamos para que el PDF siempre muestre etiquetas correctas.
    key_aliases = {
        'estrategia': 'strategy',
        'dinamica': 'dynamics',
        'situacion_de_juego': 'structure',
        'situacion_juego': 'structure',
        'habilidades_coordinativas': 'coordination_skills',
        'habilidades_coordinativas_y': 'coordination_skills',
        'intencion_accion_tactica': 'tactical_intent',
        'intencion_accion': 'tactical_intent',
        'coordinacion': 'coordination',
        'complejidad': 'complexity',
    }
    for alias_key, canonical_key in key_aliases.items():
        if alias_key in meta and canonical_key not in meta:
            meta[canonical_key] = meta.get(alias_key)
    if not meta:
        return {}
    if not meta.get('strategy') and meta.get('training_type'):
        strategy_candidate = _map_choice(meta.get('training_type'), strategy_map)
        training_type_key = str(meta.get('training_type') or '').strip().lower().replace(' ', '_').replace('-', '_')
        if training_type_key in strategy_map or strategy_candidate in strategy_map.values():
            meta['strategy'] = strategy_candidate
    if meta.get('surface'):
        meta['surface'] = _map_choice(meta.get('surface'), surface_map) or str(meta.get('surface'))
    if meta.get('pitch_format'):
        meta['pitch_format'] = _map_choice(meta.get('pitch_format'), pitch_map) or str(meta.get('pitch_format'))
    if meta.get('game_phase'):
        meta['game_phase'] = _map_choice(meta.get('game_phase'), phase_map) or str(meta.get('game_phase'))
    if meta.get('methodology'):
        meta['methodology'] = _map_choice(meta.get('methodology'), methodology_map) or str(meta.get('methodology'))
    if meta.get('complexity'):
        meta['complexity'] = _map_choice(meta.get('complexity'), complexity_map) or str(meta.get('complexity'))
    if meta.get('strategy'):
        meta['strategy'] = _map_choice(meta.get('strategy'), strategy_map) or str(meta.get('strategy'))
    if meta.get('coordination'):
        meta['coordination'] = _map_choice(meta.get('coordination'), coordination_map) or str(meta.get('coordination'))
    if meta.get('coordination_skills'):
        meta['coordination_skills'] = _map_choice(meta.get('coordination_skills'), coord_skills_map) or str(meta.get('coordination_skills'))
    if meta.get('tactical_intent'):
        meta['tactical_intent'] = _map_choice(meta.get('tactical_intent'), tactical_intent_map) or str(meta.get('tactical_intent'))
    if meta.get('dynamics'):
        meta['dynamics'] = _map_choice(meta.get('dynamics'), dynamics_map) or str(meta.get('dynamics'))
    if meta.get('structure'):
        meta['structure'] = _map_choice(meta.get('structure'), structure_map) or str(meta.get('structure'))
    if isinstance(meta.get('constraints'), list):
        meta['constraints'] = [_map_choice(v, constraint_map) or str(v) for v in meta.get('constraints')]
    if isinstance(meta.get('category_tags'), str):
        meta['category_tags'] = [item.strip() for item in str(meta.get('category_tags') or '').split(',') if item.strip()]
    elif not isinstance(meta.get('category_tags'), list):
        meta['category_tags'] = []
    if isinstance(meta.get('assigned_player_names'), str):
        meta['assigned_player_names'] = [item.strip() for item in str(meta.get('assigned_player_names') or '').split(',') if item.strip()]
    elif not isinstance(meta.get('assigned_player_names'), list):
        meta['assigned_player_names'] = []
    return meta


def _build_task_pdf_context(request, team, session, microcycle, task, tactical_layout, pdf_style='uefa', preview_url=''):
    def _display_pdf_value(value):
        text = str(value or '').strip()
        if not text:
            return ''
        normalized = text.replace('_', ' ').replace('-', ' ').strip()
        return ' '.join(normalized.split()).title()

    meta = _normalize_task_pdf_meta(tactical_layout.get('meta') if isinstance(tactical_layout, dict) else {})
    analysis_meta = meta.get('analysis') if isinstance(meta.get('analysis'), dict) else {}
    task_sheet = analysis_meta.get('task_sheet') if isinstance(analysis_meta.get('task_sheet'), dict) else {}
    description_text = str(task_sheet.get('description') or '').strip()
    description_html = str(task_sheet.get('description_html') or '').strip()
    coaching_html = str(task_sheet.get('coaching_html') or '').strip()
    rules_html = str(task_sheet.get('rules_html') or '').strip()
    progression_html = str(meta.get('progression_html') or '').strip()
    regression_html = str(meta.get('regression_html') or '').strip()
    success_criteria_html = str(meta.get('success_criteria_html') or '').strip()
    organization_html = str(meta.get('organization_html') or '').strip()
    dimensions_text = str(task_sheet.get('dimensions') or '').strip()
    materials_text = str(task_sheet.get('materials') or meta.get('resources_summary') or '').strip()
    strategy_label = str(meta.get('strategy') or meta.get('training_type') or meta.get('methodology') or '').strip()
    if not strategy_label:
        strategy_label = ''
    space_label = ' · '.join(part for part in [dimensions_text, _display_pdf_value(meta.get('space'))] if part)
    # Compat: seguimos generando formato/superficie por si se quiere imprimir como detalle técnico.
    game_situation_label = ' · '.join(part for part in [_display_pdf_value(meta.get('pitch_format')), _display_pdf_value(meta.get('surface'))] if part)
    structure_label = str(meta.get('structure') or '').strip()
    dynamics_label = str(meta.get('dynamics') or '').strip()
    coordination_label = str(meta.get('coordination') or '').strip()
    if not coordination_label:
        coordination_label = ' · '.join(part for part in [_display_pdf_value(meta.get('organization')), _display_pdf_value(meta.get('players_distribution'))] if part)
    coordination_skills_label = str(meta.get('coordination_skills') or '').strip()
    if not coordination_skills_label:
        coordination_skills_label = ' · '.join(part for part in [_display_pdf_value(meta.get('load_target')), _display_pdf_value(meta.get('complexity'))] if part)
    tactical_intent_label = str(meta.get('tactical_intent') or '').strip()
    if not tactical_intent_label:
        tactical_intent_label = ' · '.join(
        part for part in [
            _display_pdf_value(meta.get('principle')),
            _display_pdf_value(meta.get('subprinciple')),
            _display_pdf_value(meta.get('targets')),
            str(getattr(task, 'objective', '') or '').strip(),
        ] if part
        )
    animation_frames = _normalize_animation_timeline(
        tactical_layout.get('timeline') if isinstance(tactical_layout, dict) else []
    )
    graphic_editor = meta.get('graphic_editor') if isinstance(meta.get('graphic_editor'), dict) else {}
    frame_canvas_width = _parse_int(graphic_editor.get('canvas_width')) or 1280
    frame_canvas_height = _parse_int(graphic_editor.get('canvas_height')) or 720
    animation_frame_cards = [
        {
            'title': str(frame.get('title') or f'Paso {index + 1}').strip() or f'Paso {index + 1}',
            'duration': max(1, min(_parse_int(frame.get('duration')) or 3, 20)),
            'tokens': _build_task_pdf_tokens_from_canvas_state(
                request,
                frame.get('canvas_state'),
                canvas_width=frame_canvas_width,
                canvas_height=frame_canvas_height,
            ),
        }
        for index, frame in enumerate(animation_frames)
    ]
    coach_name = (
        request.user.get_full_name().strip()
        if hasattr(request.user, 'get_full_name') and request.user.get_full_name().strip()
        else getattr(request.user, 'username', '') or 'Entrenador'
    )
    primary_club_team = _get_primary_team_for_request(request) or Team.objects.filter(is_primary=True).first()
    club_logo_url = resolve_team_crest_url(request, primary_club_team, sync=True) if primary_club_team else ''
    logo_path = 'football/images/uefa-badge.svg' if pdf_style == 'uefa' else 'football/images/cdb-logo.png'
    rich_description = _sanitize_task_rich_html(description_html) if description_html else _rich_html_from_plain_text(description_text)
    rich_coaching = _sanitize_task_rich_html(coaching_html) if coaching_html else _rich_html_from_plain_text(getattr(task, 'coaching_points', '') or '')
    rich_rules = _sanitize_task_rich_html(rules_html) if rules_html else _rich_html_from_plain_text(getattr(task, 'confrontation_rules', '') or '')
    rich_progression = _sanitize_task_rich_html(progression_html) if progression_html else _rich_html_from_plain_text(str(meta.get('progression') or ''))
    rich_regression = _sanitize_task_rich_html(regression_html) if regression_html else _rich_html_from_plain_text(str(meta.get('regression') or ''))
    rich_success = _sanitize_task_rich_html(success_criteria_html) if success_criteria_html else _rich_html_from_plain_text(str(meta.get('success_criteria') or ''))
    rich_organization = _sanitize_task_rich_html(organization_html) if organization_html else _rich_html_from_plain_text(str(meta.get('organization') or ''))
    return {
        'team_name': team.name,
        'task': task,
        'session': session,
        'microcycle': microcycle,
        'objective_lines': _task_pdf_lines(getattr(task, 'objective', '')),
        'coaching_lines': _task_pdf_lines(getattr(task, 'coaching_points', '')),
        'rules_lines': _task_pdf_lines(getattr(task, 'confrontation_rules', '')),
        'description_lines': _task_pdf_lines(description_text),
        'description_rich_html': rich_description,
        'coaching_rich_html': rich_coaching,
        'rules_rich_html': rich_rules,
        'progression_rich_html': rich_progression,
        'regression_rich_html': rich_regression,
        'success_criteria_rich_html': rich_success,
        'organization_rich_html': rich_organization,
        'tokens': _build_task_pdf_tokens(request, tactical_layout),
        'task_meta': meta,
        'strategy_label': strategy_label or '-',
        'space_label': space_label or '-',
        'dimensions_label': dimensions_text or '-',
        'materials_label': materials_text or '-',
        'game_situation_label': game_situation_label or '-',
        'structure_label': structure_label or '-',
        'dynamics_label': dynamics_label or '-',
        'coordination_label': coordination_label or '-',
        'coordination_skills_label': coordination_skills_label or '-',
        'tactical_intent_label': tactical_intent_label or '-',
        'animation_frames': animation_frames,
        'animation_frame_cards': animation_frame_cards,
        'pdf_style': pdf_style,
        'pdf_palette': _team_pdf_palette(team, pdf_style),
        'coach_name': coach_name,
        'animation_frames_count': len(animation_frames),
        'logo_url': request.build_absolute_uri(static(logo_path)),
        'brand_mark_url': request.build_absolute_uri(static('football/images/2j-mark.svg')),
        'club_logo_url': club_logo_url,
        'task_preview_url': preview_url,
        'generated_at': timezone.localtime(),
    }


def _build_task_draft_pdf_context(request, primary_team, pdf_style='uefa'):
    title = _sanitize_task_text((request.POST.get('draw_task_title') or '').strip(), multiline=False, max_len=160) or 'Tarea sin título'
    objective = _sanitize_task_text((request.POST.get('draw_task_objective') or '').strip(), multiline=False, max_len=180)
    coaching_points = _sanitize_task_text((request.POST.get('draw_task_coaching_points') or '').strip(), multiline=True)
    confrontation_rules = _sanitize_task_text((request.POST.get('draw_task_confrontation_rules') or '').strip(), multiline=True)
    block = (request.POST.get('draw_task_block') or SessionTask.BLOCK_MAIN_1).strip()
    minutes = max(5, min(_parse_int(request.POST.get('draw_task_minutes')) or 15, 90))
    target_session_id = _parse_int(request.POST.get('draw_target_session_id'))
    selected_surface = _sanitize_task_text((request.POST.get('draw_task_surface') or '').strip(), multiline=False, max_len=80)
    selected_pitch_format = _sanitize_task_text((request.POST.get('draw_task_pitch_format') or '').strip(), multiline=False, max_len=80)
    selected_phase = _sanitize_task_text((request.POST.get('draw_task_game_phase') or '').strip(), multiline=False, max_len=80)
    selected_methodology = _sanitize_task_text((request.POST.get('draw_task_methodology') or '').strip(), multiline=False, max_len=80)
    selected_complexity = _sanitize_task_text((request.POST.get('draw_task_complexity') or '').strip(), multiline=False, max_len=80)
    selected_strategy = _sanitize_task_text((request.POST.get('draw_task_strategy') or '').strip(), multiline=False, max_len=80)
    selected_dynamics = _sanitize_task_text((request.POST.get('draw_task_dynamics') or '').strip(), multiline=False, max_len=80)
    selected_structure = _sanitize_task_text((request.POST.get('draw_task_structure') or '').strip(), multiline=False, max_len=80)
    selected_coordination = _sanitize_task_text((request.POST.get('draw_task_coordination') or '').strip(), multiline=False, max_len=80)
    selected_coord_skills = _sanitize_task_text((request.POST.get('draw_task_coordination_skills') or '').strip(), multiline=False, max_len=80)
    selected_tactical_intent = _sanitize_task_text((request.POST.get('draw_task_tactical_intent') or '').strip(), multiline=False, max_len=80)
    space = _sanitize_task_text((request.POST.get('draw_task_space') or '').strip(), multiline=False, max_len=120)
    organization = _sanitize_task_text((request.POST.get('draw_task_organization') or '').strip(), multiline=True, max_len=500)
    players_distribution = _sanitize_task_text((request.POST.get('draw_task_players_distribution') or '').strip(), multiline=False, max_len=180)
    load_target = _sanitize_task_text((request.POST.get('draw_task_load_target') or '').strip(), multiline=False, max_len=180)
    work_rest = _sanitize_task_text((request.POST.get('draw_task_work_rest') or '').strip(), multiline=False, max_len=180)
    series = _sanitize_task_text((request.POST.get('draw_task_series') or '').strip(), multiline=False, max_len=100)
    repetitions = _sanitize_task_text((request.POST.get('draw_task_repetitions') or '').strip(), multiline=False, max_len=100)
    player_count = _sanitize_task_text((request.POST.get('draw_task_player_count') or '').strip(), multiline=False, max_len=100)
    age_group = _sanitize_task_text((request.POST.get('draw_task_age_group') or '').strip(), multiline=False, max_len=100)
    training_type = _sanitize_task_text((request.POST.get('draw_task_training_type') or '').strip(), multiline=False, max_len=120)
    dimensions = _sanitize_task_text((request.POST.get('draw_task_dimensions') or '').strip(), multiline=False, max_len=120)
    materials = _sanitize_task_text((request.POST.get('draw_task_materials') or '').strip(), multiline=False, max_len=180)
    progression = _sanitize_task_text((request.POST.get('draw_task_progression') or '').strip(), multiline=True, max_len=500)
    regression = _sanitize_task_text((request.POST.get('draw_task_regression') or '').strip(), multiline=True, max_len=500)
    success_criteria = _sanitize_task_text((request.POST.get('draw_task_success_criteria') or '').strip(), multiline=True, max_len=500)
    organization_html = _sanitize_task_rich_html((request.POST.get('draw_task_organization_html') or '').strip())
    progression_html = _sanitize_task_rich_html((request.POST.get('draw_task_progression_html') or '').strip())
    regression_html = _sanitize_task_rich_html((request.POST.get('draw_task_regression_html') or '').strip())
    success_criteria_html = _sanitize_task_rich_html((request.POST.get('draw_task_success_criteria_html') or '').strip())
    category_tags_raw = _sanitize_task_text((request.POST.get('draw_task_category_tags') or '').strip(), multiline=False, max_len=240)
    category_tags = [tag.strip() for tag in category_tags_raw.split(',') if tag.strip()]
    assigned_player_ids = [
        player_id
        for player_id in (_parse_int(value) for value in request.POST.getlist('assigned_player_ids'))
        if player_id
    ]
    assigned_player_names = list(
        Player.objects.filter(team=primary_team, id__in=assigned_player_ids).order_by('number', 'name').values_list('name', flat=True)
    )
    selected_session = (
        TrainingSession.objects.select_related('microcycle')
        .filter(id=target_session_id, microcycle__team=primary_team)
        .first()
    )
    if selected_session:
        session = selected_session
        microcycle = selected_session.microcycle
    else:
        today = timezone.localdate()
        session = SimpleNamespace(session_date=today, start_time=None, focus='Borrador')
        microcycle = SimpleNamespace(title='Borrador', week_start=today, week_end=today)
    canvas_state = {}
    raw_canvas_state = (request.POST.get('draw_canvas_state') or '').strip()
    if raw_canvas_state:
        try:
            parsed = json.loads(raw_canvas_state)
            if isinstance(parsed, dict):
                canvas_state = parsed
        except Exception:
            canvas_state = {}
    tactical_layout = {
        'tokens': canvas_state.get('objects') if isinstance(canvas_state.get('objects'), list) else [],
        'timeline': canvas_state.get('timeline') if isinstance(canvas_state.get('timeline'), list) else [],
	        'meta': {
	            'surface': selected_surface,
	            'pitch_format': selected_pitch_format,
	            'game_phase': selected_phase,
	            'methodology': selected_methodology,
	            'complexity': selected_complexity,
	            'strategy': selected_strategy,
	            'dynamics': selected_dynamics,
	            'structure': selected_structure,
	            'coordination': selected_coordination,
	            'coordination_skills': selected_coord_skills,
	            'tactical_intent': selected_tactical_intent,
	            'space': space,
	            'organization': organization,
	            'organization_html': organization_html,
	            'players_distribution': players_distribution,
            'load_target': load_target,
            'work_rest': work_rest,
            'series': series,
            'repetitions': repetitions,
            'player_count': player_count,
            'age_group': age_group,
            'training_type': training_type,
            'category_tags': category_tags,
            'assigned_player_names': assigned_player_names,
            'progression': progression,
            'progression_html': progression_html,
            'regression': regression,
            'regression_html': regression_html,
            'success_criteria': success_criteria,
            'success_criteria_html': success_criteria_html,
            'analysis': {
                'task_sheet': {
                    'description': _sanitize_task_text((request.POST.get('draw_task_description') or '').strip(), multiline=True),
                    'description_html': _sanitize_task_rich_html((request.POST.get('draw_task_description_html') or '').strip()),
                    'coaching_html': _sanitize_task_rich_html((request.POST.get('draw_task_coaching_points_html') or '').strip()),
                    'rules_html': _sanitize_task_rich_html((request.POST.get('draw_task_confrontation_rules_html') or '').strip()),
                    'dimensions': dimensions,
                    'materials': materials,
                }
            },
        },
    }
    draft_task = SimpleNamespace(
        id=0,
        title=title,
        duration_minutes=minutes,
        objective=objective,
        coaching_points=coaching_points,
        confrontation_rules=confrontation_rules,
        block=block,
        get_block_display=lambda: dict(SessionTask.BLOCK_CHOICES).get(block, block),
    )
    preview_data = str(request.POST.get('draw_canvas_preview_data') or '').strip()
    return _build_task_pdf_context(
        request,
        team=primary_team,
        session=session,
        microcycle=microcycle,
        task=draft_task,
        tactical_layout=tactical_layout,
        pdf_style=pdf_style,
        preview_url=preview_data,
    )


def _build_session_pdf_context(request, team, session, pdf_style='uefa'):
    tasks = list(session.tasks.filter(deleted_at__isnull=True).order_by('order', 'id'))
    total_task_minutes = sum(int(getattr(task, 'duration_minutes', 0) or 0) for task in tasks)
    coach_name = (
        request.user.get_full_name().strip()
        if hasattr(request.user, 'get_full_name') and request.user.get_full_name().strip()
        else getattr(request.user, 'username', '') or 'Entrenador'
    )
    primary_club_team = _get_primary_team_for_request(request) or Team.objects.filter(is_primary=True).first()
    club_logo_url = resolve_team_crest_url(request, primary_club_team, sync=True) if primary_club_team else ''
    logo_path = 'football/images/uefa-badge.svg' if pdf_style == 'uefa' else 'football/images/cdb-logo.png'
    return {
        'team_name': team.name,
        'session': session,
        'microcycle': session.microcycle,
        'tasks': tasks,
        'task_sheets': [_build_session_task_sheet(task) for task in tasks],
        'tasks_count': len(tasks),
        'task_minutes_total': total_task_minutes,
        'pdf_style': pdf_style,
        'pdf_palette': _team_pdf_palette(team, pdf_style),
        'coach_name': coach_name,
        'logo_url': request.build_absolute_uri(static(logo_path)),
        'brand_mark_url': request.build_absolute_uri(static('football/images/2j-mark.svg')),
        'club_logo_url': club_logo_url,
        'generated_at': timezone.localtime(),
        'intensity_label': dict(TrainingSession.INTENSITY_CHOICES).get(session.intensity, session.intensity or '-'),
        'status_label': dict(TrainingSession.STATUS_CHOICES).get(session.status, session.status or '-'),
    }


def _parse_microcycle_plan_fields(raw_notes):
    notes = str(raw_notes or '')
    defaults = {
        'attack': '',
        'defense': '',
        'set_pieces': '',
        'rival_notes': '',
        'general_notes': notes.strip(),
    }
    if not notes.strip():
        return defaults
    marker = '[2J_MICROCYCLE]'
    if marker not in notes:
        return defaults
    _, payload = notes.split(marker, 1)
    parsed = defaults.copy()
    current_key = None
    free_lines = []
    for raw_line in payload.splitlines():
        line = raw_line.rstrip()
        if not line.strip():
            if current_key:
                parsed[current_key] = (parsed.get(current_key, '') + '\n').strip('\n')
            continue
        if ':' in line:
            key, value = line.split(':', 1)
            normalized = key.strip().lower()
            if normalized in {'attack', 'defense', 'set_pieces', 'rival_notes', 'general_notes'}:
                current_key = normalized
                parsed[current_key] = value.strip()
                continue
        if current_key:
            parsed[current_key] = '\n'.join(filter(None, [parsed.get(current_key, '').strip(), line.strip()])).strip()
        else:
            free_lines.append(line.strip())
    if free_lines and not parsed.get('general_notes'):
        parsed['general_notes'] = '\n'.join(free_lines).strip()
    return parsed


def _serialize_microcycle_plan_fields(fields):
    clean = {key: str(fields.get(key) or '').strip() for key in ['attack', 'defense', 'set_pieces', 'rival_notes', 'general_notes']}
    if not any(clean.values()):
        return ''
    lines = ['[2J_MICROCYCLE]']
    for key in ['attack', 'defense', 'set_pieces', 'rival_notes', 'general_notes']:
        lines.append(f'{key}:{clean[key]}')
    return '\n'.join(lines).strip()


def _build_microcycle_week_slots(microcycle, session_rows):
    labels = ['Lun', 'Mar', 'Mié', 'Jue', 'Vie', 'Sáb', 'Dom']
    total_days = max(1, min(7, ((microcycle.week_end - microcycle.week_start).days or 0) + 1))
    rows_by_date = defaultdict(list)
    for session_row in session_rows:
        rows_by_date[session_row['obj'].session_date].append(session_row)
    slots = []
    for offset in range(total_days):
        slot_date = microcycle.week_start + timedelta(days=offset)
        slots.append(
            {
                'date': slot_date,
                'label': labels[slot_date.weekday()],
                'sessions': rows_by_date.get(slot_date, []),
            }
        )
    return slots


def _build_session_task_sheet(task):
    meta = _normalize_task_pdf_meta(task.tactical_layout.get('meta') if isinstance(task.tactical_layout, dict) else {})
    analysis_meta = meta.get('analysis') if isinstance(meta.get('analysis'), dict) else {}
    task_sheet = analysis_meta.get('task_sheet') if isinstance(analysis_meta.get('task_sheet'), dict) else {}
    contents = []
    for value in [
        meta.get('training_type'),
        meta.get('game_phase'),
        meta.get('principle'),
        meta.get('subprinciple'),
    ]:
        text = str(value or '').strip()
        if text:
            contents.append(text)
    return {
        'title': str(task.title or '').strip() or f'Tarea {task.id}',
        'block_label': task.get_block_display(),
        'minutes': int(task.duration_minutes or 0),
        'type_label': str(meta.get('training_type') or '').strip(),
        'contents_label': ' · '.join(contents) or str(task.objective or '').strip() or '-',
        'structure_label': ' · '.join(
            part for part in [
                str(meta.get('organization') or '').strip(),
                str(meta.get('players_distribution') or '').strip(),
                str(meta.get('player_count') or '').strip(),
            ] if part
        ) or '-',
        'players_label': ', '.join(meta.get('assigned_player_names') or []) or str(meta.get('player_count') or '').strip() or '-',
        'dimensions_label': str(task_sheet.get('dimensions') or meta.get('space') or '').strip() or '-',
        'materials_label': str(task_sheet.get('materials') or meta.get('resources_summary') or '').strip() or '-',
        'description': str(task_sheet.get('description') or '').strip(),
        'rules': str(task.confrontation_rules or '').strip(),
        'focus': str(task.coaching_points or '').strip(),
        'variants': str(meta.get('progression') or '').strip(),
        'success': str(meta.get('success_criteria') or '').strip(),
    }


@login_required
def session_plan_pdf(request, session_id):
    if not _can_access_sessions_workspace(request.user):
        return HttpResponse('No tienes permisos para acceder a sesiones.', status=403)
    forbidden = _forbid_if_workspace_module_disabled(request, 'sessions', label='sesiones')
    if forbidden:
        return forbidden
    session = (
        TrainingSession.objects
        .select_related('microcycle__team')
        .prefetch_related('tasks')
        .filter(id=session_id)
        .first()
    )
    if not session:
        raise Http404('Sesión no encontrada')

    pdf_style = (request.GET.get('style') or 'uefa').strip().lower()
    if pdf_style not in {'uefa', 'club'}:
        pdf_style = 'uefa'
    context = _build_session_pdf_context(request, session.microcycle.team, session, pdf_style=pdf_style)
    html = render_to_string('football/session_plan_pdf.html', context)
    filename = slugify(f'sesion-{session.session_date}-{session.focus}') or f'sesion-{session.id}'
    return _build_pdf_response_or_html_fallback(request, html, filename)


@login_required
def session_task_pdf(request, task_id):
    if not _can_access_sessions_workspace(request.user):
        return HttpResponse('No tienes permisos para acceder a sesiones.', status=403)
    task = (
        SessionTask.objects
        .select_related('session__microcycle__team')
        .filter(id=task_id)
        .first()
    )
    if not task:
        raise Http404('Tarea no encontrada')

    team = task.session.microcycle.team
    pdf_style = (request.GET.get('style') or 'uefa').strip().lower()
    if pdf_style not in {'uefa', 'club'}:
        pdf_style = 'uefa'
    context = _build_task_pdf_context(
        request,
        team=team,
        session=task.session,
        microcycle=task.session.microcycle,
        task=task,
        tactical_layout=task.tactical_layout if isinstance(task.tactical_layout, dict) else {},
        pdf_style=pdf_style,
        preview_url=request.build_absolute_uri(reverse('session-task-preview-file', args=[task.id])) if task.task_preview_image else '',
    )
    html = render_to_string('football/session_task_pdf.html', context)
    filename = slugify(f'tarea-{task.session.session_date}-{task.title}') or f'tarea-{task.id}'
    return _build_pdf_response_or_html_fallback(request, html, filename)


def _resolve_static_asset_file(asset_value):
    raw = str(asset_value or '').strip()
    if not raw:
        return None
    if raw.startswith('/static/'):
        rel = raw[len('/static/'):]
    else:
        rel = raw
    rel = rel.lstrip('/')
    roots = [
        Path(settings.BASE_DIR) / 'static',
        Path(settings.BASE_DIR) / 'football' / 'static',
    ]
    for root in roots:
        candidate = root / rel
        if candidate.exists() and candidate.is_file():
            return candidate
    return None


@login_required
def session_task_canva_export(request, task_id):
    if not _can_access_sessions_workspace(request.user):
        return HttpResponse('No tienes permisos para acceder a sesiones.', status=403)
    task = (
        SessionTask.objects
        .select_related('session__microcycle__team')
        .filter(id=task_id)
        .first()
    )
    if not task:
        raise Http404('Tarea no encontrada')

    team = task.session.microcycle.team
    session = task.session
    tactical_layout = task.tactical_layout if isinstance(task.tactical_layout, dict) else {}
    meta = tactical_layout.get('meta') if isinstance(tactical_layout, dict) else {}
    if not isinstance(meta, dict):
        meta = {}
    tokens = tactical_layout.get('tokens') if isinstance(tactical_layout, dict) else []
    if not isinstance(tokens, list):
        tokens = []
    timeline = tactical_layout.get('timeline') if isinstance(tactical_layout, dict) else []
    if not isinstance(timeline, list):
        timeline = []

    constraints_value = meta.get('constraints') if isinstance(meta.get('constraints'), list) else []
    canva_fields = {
        'equipo': team.name,
        'fecha_sesion': session.session_date.isoformat() if session.session_date else '',
        'foco_sesion': session.focus or '',
        'tarea': task.title or '',
        'bloque': task.get_block_display() if hasattr(task, 'get_block_display') else task.block,
        'duracion_min': str(task.duration_minutes or ''),
        'objetivo': task.objective or '',
        'consignas': task.coaching_points or '',
        'reglas': task.confrontation_rules or '',
        'superficie': str(meta.get('surface') or ''),
        'formato_campo': str(meta.get('pitch_format') or ''),
        'espacio': str(meta.get('space') or ''),
        'organizacion': str(meta.get('organization') or ''),
        'distribucion': str(meta.get('players_distribution') or ''),
        'carga_objetivo': str(meta.get('load_target') or ''),
        'material': str(meta.get('resources_summary') or ''),
        'progresion': str(meta.get('progression') or ''),
        'regresion': str(meta.get('regression') or ''),
        'trabajo_pausa': str(meta.get('work_rest') or ''),
        'series': str(meta.get('series') or ''),
        'repeticiones': str(meta.get('repetitions') or ''),
        'principio': str(meta.get('principle') or ''),
        'subprincipio': str(meta.get('subprinciple') or ''),
        'puntuacion': str(meta.get('scoring_model') or ''),
        'criterio_exito': str(meta.get('success_criteria') or ''),
        'fase_juego': str(meta.get('game_phase') or ''),
        'subfase': str(meta.get('game_sub_phase') or ''),
        'metodologia': str(meta.get('methodology') or ''),
        'complejidad': str(meta.get('complexity') or ''),
        'carga_cognitiva': str(meta.get('cognitive_load') or ''),
        'carga_emocional': str(meta.get('emotional_load') or ''),
        'targets': str(meta.get('targets') or ''),
        'outcomes': str(meta.get('expected_outcomes') or ''),
        'errores_frecuentes': str(meta.get('common_errors') or ''),
        'correcciones': str(meta.get('corrective_cues') or ''),
        'video_ref': str(meta.get('video_reference') or ''),
        'notas_staff': str(meta.get('staff_notes') or ''),
        'constraints': ' | '.join(str(v) for v in constraints_value),
    }

    buffer = io.BytesIO()
    with zipfile.ZipFile(buffer, mode='w', compression=zipfile.ZIP_DEFLATED) as zf:
        csv_buffer = io.StringIO()
        writer = csv.DictWriter(csv_buffer, fieldnames=list(canva_fields.keys()))
        writer.writeheader()
        writer.writerow(canva_fields)
        zf.writestr('canva/canva_task_data.csv', csv_buffer.getvalue())

        payload = {
            'task_id': task.id,
            'session_id': session.id,
            'team': team.name,
            'created_at': timezone.now().isoformat(),
            'task': canva_fields,
            'tactical_layout': tactical_layout,
        }
        zf.writestr(
            'canva/task_payload.json',
            json.dumps(payload, ensure_ascii=False, indent=2),
        )
        zf.writestr(
            'canva/tactical_layout.json',
            json.dumps(
                {'tokens': tokens, 'timeline': timeline, 'meta': meta},
                ensure_ascii=False,
                indent=2,
            ),
        )

        readme = (
            'PACK EXPORT CANVA - WEB STATS\n\n'
            '1) Canva > Crear diseño.\n'
            '2) Apps > Bulk create > Subir CSV: canva/canva_task_data.csv\n'
            '3) Vincula los campos del CSV a textos de tu plantilla.\n'
            '4) Usa canva/tactical_layout.json como referencia de pizarra.\n'
            '5) Usa assets/ para arrastrar iconos/imagenes al diseño.\n'
        )
        zf.writestr('README_CANVA.txt', readme)

        seen_names = set()
        for token in tokens:
            if not isinstance(token, dict):
                continue
            asset_value = token.get('asset')
            source_file = _resolve_static_asset_file(asset_value)
            if not source_file:
                continue
            filename = source_file.name
            if filename in seen_names:
                continue
            seen_names.add(filename)
            zf.write(source_file, arcname=f'assets/{filename}')

    buffer.seek(0)
    slug = slugify(task.title or f'tarea-{task.id}') or f'tarea-{task.id}'
    filename = f'canva-task-{slug}.zip'
    response = HttpResponse(buffer.getvalue(), content_type='application/zip')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    return response


@login_required
def coach_cards_page(request):
    forbidden = _forbid_if_no_coach_access(request.user)
    if forbidden:
        return forbidden
    forbidden = _forbid_if_workspace_module_disabled(request, 'coach_overview', label='módulos de staff')
    if forbidden:
        return forbidden
    role_labels = dict(AppUserRole.ROLE_CHOICES)
    technical_roles = {
        AppUserRole.ROLE_COACH,
        AppUserRole.ROLE_GOALKEEPER,
        AppUserRole.ROLE_FITNESS,
        AppUserRole.ROLE_ANALYST,
    }
    role_rows = list(AppUserRole.objects.select_related('user').filter(role__in=technical_roles))
    members_by_role = {key: [] for key in technical_roles}
    for role_row in role_rows:
        if not role_row.user.is_active:
            continue
        full_name = role_row.user.get_full_name().strip() or role_row.user.username
        members_by_role.setdefault(role_row.role, []).append(full_name)
    cards = [
        {
            'key': 'stats',
            'title': 'Estadísticas grupales',
            'description': 'Lectura agregada del equipo y acceso rápido al seguimiento general del rendimiento colectivo.',
            'link': 'coach-role-trainer',
            'member_name': ' · '.join(members_by_role.get(AppUserRole.ROLE_COACH) or ['Sin asignar']),
            'items': [
                {'label': 'Resumen entrenador', 'link': 'coach-role-trainer'},
                {'label': 'Registro de jugadores', 'link': 'coach-roster'},
            ],
        },
        {
            'key': 'match',
            'title': 'Partido',
            'description': 'Zona operativa de convocatoria, 11 inicial, registro en vivo y revisión de partido para el staff.',
            'link': 'convocation',
            'member_name': ' · '.join(
                list(members_by_role.get(AppUserRole.ROLE_COACH) or [])
                + list(members_by_role.get(AppUserRole.ROLE_ANALYST) or [])
                or ['Sin asignar']
            ),
            'items': [
                {'label': 'Convocatoria', 'link': 'convocation'},
                {'label': '11 inicial', 'link': 'initial-eleven'},
                {'label': 'Registro de acciones', 'link': 'match-action-page'},
            ],
        },
        {
            'key': 'training',
            'title': 'Entrenamiento',
            'description': 'Planificación semanal, sesiones y trabajo específico del staff sin salir del flujo de entrenamiento.',
            'link': 'sessions',
            'member_name': ' · '.join(
                list(members_by_role.get(AppUserRole.ROLE_COACH) or [])
                + list(members_by_role.get(AppUserRole.ROLE_GOALKEEPER) or [])
                + list(members_by_role.get(AppUserRole.ROLE_FITNESS) or [])
                or ['Sin asignar']
            ),
            'items': [
                {'label': 'Planificación general', 'link': 'sessions'},
                {'label': 'Porteros', 'link': 'sessions-goalkeeper'},
                {'label': 'Preparación física', 'link': 'sessions-fitness'},
                {'label': 'ABP', 'link': 'coach-role-abp'},
            ],
        },
        {
            'key': 'analysis',
            'title': 'Análisis',
            'description': 'Scouting, rival e informes para el trabajo técnico del staff sin duplicar módulos en otras áreas.',
            'link': 'analysis',
            'member_name': ' · '.join(members_by_role.get(AppUserRole.ROLE_ANALYST) or members_by_role.get(AppUserRole.ROLE_COACH) or ['Sin asignar']),
            'items': [
                {'label': 'Análisis rival', 'link': 'analysis'},
                {'label': 'Informes y scouting', 'link': 'analysis'},
            ],
        },
    ]
    active_area = str(request.GET.get('area') or 'stats').strip().lower()
    valid_keys = {card['key'] for card in cards}
    if active_area not in valid_keys:
        active_area = 'stats'
    active_card = next((card for card in cards if card['key'] == active_area), cards[0])
    return render(
        request,
        'football/coach_cards.html',
        {
            'cards': cards,
            'active_area': active_area,
            'active_card': active_card,
            'staff_count': sum(1 for value in members_by_role.values() for _ in value),
        },
    )


@login_required
def coach_role_trainer_page(request):
    forbidden = _forbid_if_no_coach_access(request.user)
    if forbidden:
        return forbidden
    forbidden = _forbid_if_workspace_module_disabled(request, 'coach_overview', label='módulo entrenador')
    if forbidden:
        return forbidden
    primary_team = _get_primary_team_for_request(request)
    standing = None
    if primary_team and primary_team.group and primary_team.group.season:
        standing = TeamStanding.objects.filter(
            season=primary_team.group.season,
            group=primary_team.group,
            team=primary_team,
        ).first()
    preferred_sources = preferred_event_source_by_match(primary_team) if primary_team else {}
    events = (
        _filter_stats_events(
            confirmed_events_queryset()
            .filter(player__team=primary_team)
            .select_related('player', 'match', 'match__home_team', 'match__away_team')
            .order_by('match_id', 'minute', 'id'),
            preferred_sources=preferred_sources,
        )
        if primary_team
        else []
    )
    measured_match_ids = {event.match_id for event in events if event.match_id}
    measured_matches = len(measured_match_ids)
    total_actions = len(events) if primary_team else 0
    total_matches = (
        _team_match_queryset(primary_team).count()
        if primary_team
        else 0
    )
    goals_for = standing.goals_for if standing else 0
    goals_against = standing.goals_against if standing else 0
    points = standing.points if standing else 0
    rank = standing.position if standing else 0
    duel_classifications = [
        classify_duel_event(event.event_type, event.result, event.observation, event.zone)
        for event in events
    ]
    duels = [item for item in duel_classifications if item.get('is_duel')]
    duel_won = [item for item in duels if item.get('won')]
    duel_rate = round((len(duel_won) / len(duels)) * 100, 1) if duels else 0.0
    success_actions = [event for event in events if result_is_success(event.result)]
    success_rate = round((len(success_actions) / total_actions) * 100, 1) if total_actions else 0.0
    avg_actions = round(total_actions / measured_matches, 1) if measured_matches else 0.0
    avg_duels = round(len(duels) / measured_matches, 1) if measured_matches else 0.0
    season_played_matches = standing.played if standing and standing.played else total_matches
    goals_per_match = round((goals_for / season_played_matches), 2) if season_played_matches else 0.0
    goals_conceded_per_match = round((goals_against / season_played_matches), 2) if season_played_matches else 0.0
    team_shots_attempts = sum(
        1 for event in events if is_shot_attempt_event(event.event_type, event.result, event.observation)
    )
    event_goals_for = sum(
        1 for event in events if is_goal_event(event.event_type, event.result, event.observation)
    )
    measured_goals_per_match = round((event_goals_for / measured_matches), 2) if measured_matches else 0.0
    team_shots_per_goal = shots_needed_per_goal(team_shots_attempts, event_goals_for)
    team_match_zone_profiles, team_player_zone_profiles = _build_zone_inference_profiles(events)

    def _summarize_events(event_list):
        total = len(event_list)
        successes_local = sum(1 for event in event_list if result_is_success(event.result))
        goals_local = sum(1 for event in event_list if is_goal_event(event.event_type, event.result, event.observation))
        duel_local_classifications = [
            classify_duel_event(event.event_type, event.result, event.observation, event.zone)
            for event in event_list
        ]
        duels_local = [item for item in duel_local_classifications if item.get('is_duel')]
        duels_won_local = [item for item in duels_local if item.get('won')]
        shots_attempts = 0
        shots_on_target = 0
        passes_attempts = 0
        passes_completed = 0
        yellow_local = 0
        red_local = 0
        zone_counts_local = {key: 0 for key in FIELD_ZONE_KEYS}
        for event in event_list:
            if is_yellow_card_event(event.event_type, event.result, event.zone):
                yellow_local += 1
            if is_red_card_event(event.event_type, event.result, event.zone):
                red_local += 1
            shot_event = is_shot_attempt_event(event.event_type, event.result, event.observation)
            if shot_event:
                shots_attempts += 1
                if is_shot_on_target_event(event.event_type, event.result, event.observation):
                    shots_on_target += 1
            is_pass_event = (
                contains_keyword(event.event_type, PASS_KEYWORDS)
                or contains_keyword(event.observation, PASS_KEYWORDS)
                or is_assist_event(event.event_type, event.result, event.observation)
            )
            if is_pass_event:
                passes_attempts += 1
                if result_is_success(event.result) or is_assist_event(event.event_type, event.result, event.observation):
                    passes_completed += 1
            zone_label = _resolve_zone_label(event, team_match_zone_profiles, team_player_zone_profiles)
            if zone_label:
                zone_counts_local[zone_label] += 1
        return {
            'total_actions': total,
            'success_rate': round((successes_local / total) * 100, 1) if total else 0.0,
            'duels_total': len(duels_local),
            'duel_rate': round((len(duels_won_local) / len(duels_local)) * 100, 1) if duels_local else 0.0,
            'yellow_cards': yellow_local,
            'red_cards': red_local,
            'goals': goals_local,
            'shots_attempts': shots_attempts,
            'shots_on_target': shots_on_target,
            'shots_per_goal': shots_needed_per_goal(shots_attempts, goals_local),
            'passes_attempts': passes_attempts,
            'passes_completed': passes_completed,
            'zone_counts': zone_counts_local,
            'field_zones': [],
        }

    # General overview based on player base stats (Universo/cache/manual) + actions dataset.
    player_cards = compute_player_cards(primary_team) if primary_team else []
    yellows = sum(int(item.get('yellow_cards', 0) or 0) for item in player_cards)
    reds = sum(int(item.get('red_cards', 0) or 0) for item in player_cards)
    avg_yellows = round(yellows / season_played_matches, 2) if season_played_matches else 0.0
    card_total = yellows + reds
    cards_per_match = round((card_total / season_played_matches), 2) if season_played_matches else 0.0
    weekly_staff_brief = _build_weekly_staff_brief_context(primary_team, player_cards=player_cards)
    top_minutes_player = max(player_cards, key=lambda item: item.get('minutes', 0), default={})
    top_goals_player = max(player_cards, key=lambda item: item.get('goals', 0), default={})
    top_yellow_player = max(player_cards, key=lambda item: item.get('yellow_cards', 0), default={})

    duels_won_by_player = {}
    recoveries_by_player = {}
    for event in events:
        player = event.player
        if not player:
            continue
        duel_event = classify_duel_event(event.event_type, event.result, event.observation, event.zone)
        if duel_event.get('is_duel') and duel_event.get('won'):
            duels_won_by_player[player.id] = duels_won_by_player.get(player.id, 0) + 1
        event_text = ' '.join(
            [
                str(event.event_type or ''),
                str(event.result or ''),
                str(event.observation or ''),
            ]
        )
        if contains_keyword(event_text, ['robo', 'recuper', 'intercep']):
            recoveries_by_player[player.id] = recoveries_by_player.get(player.id, 0) + 1

    def _top_player_from_counter(counter_dict):
        if not counter_dict:
            return {}
        top_id, top_value = max(counter_dict.items(), key=lambda x: x[1])
        player_obj = Player.objects.filter(id=top_id).first()
        if not player_obj:
            return {}
        return {
            'name': player_obj.name,
            'value': top_value,
        }

    top_duels_player = _top_player_from_counter(duels_won_by_player)
    top_recovery_player = _top_player_from_counter(recoveries_by_player)

    active_injuries = 0
    total_injuries = 0
    if primary_team:
        injuries_qs = PlayerInjuryRecord.objects.filter(player__team=primary_team)
        active_injuries = sum(1 for injury in injuries_qs if is_injury_record_active(injury))
        total_injuries = injuries_qs.count()

    coach_general_stats = [
        {'label': 'Partidos jugados', 'value': season_played_matches},
        {'label': 'Partidos medidos', 'value': measured_matches},
        {'label': 'Goles totales', 'value': goals_for},
        {'label': 'Goles medidos', 'value': event_goals_for},
        {'label': 'Goles por partido', 'value': goals_per_match},
        {'label': 'Goles medidos/partido', 'value': measured_goals_per_match},
        {'label': 'GC por partido', 'value': goals_conceded_per_match},
        {'label': 'Tarjetas totales', 'value': card_total},
        {'label': 'Tarjetas por partido', 'value': cards_per_match},
        {'label': 'Lesiones activas', 'value': active_injuries},
        {'label': 'Lesiones totales', 'value': total_injuries},
    ]
    coach_player_leaders = [
        {
            'label': 'Más minutos',
            'name': top_minutes_player.get('name') or '-',
            'value': top_minutes_player.get('minutes', 0),
            'suffix': 'min',
        },
        {
            'label': 'Más goles',
            'name': top_goals_player.get('name') or '-',
            'value': top_goals_player.get('goals', 0),
            'suffix': 'g',
        },
        {
            'label': 'Más robos/recuperaciones',
            'name': top_recovery_player.get('name') or '-',
            'value': top_recovery_player.get('value', 0),
            'suffix': '',
        },
        {
            'label': 'Más duelos ganados',
            'name': top_duels_player.get('name') or '-',
            'value': top_duels_player.get('value', 0),
            'suffix': '',
        },
        {
            'label': 'Más amarillas',
            'name': top_yellow_player.get('name') or '-',
            'value': top_yellow_player.get('yellow_cards', 0),
            'suffix': '',
        },
    ]

    kpis = [
        {'label': 'Clasificación', 'value': rank or '-', 'pct': min(100, max(0, 100 - (rank * 4 if rank else 0))), 'suffix': ''},
        {'label': 'Puntos', 'value': points, 'pct': min(100, round((points / 75) * 100, 1) if points else 0), 'suffix': ''},
        {'label': 'Goles a favor', 'value': goals_for, 'pct': min(100, round((goals_for / 60) * 100, 1) if goals_for else 0), 'suffix': ''},
        {'label': 'Goles en contra', 'value': goals_against, 'pct': min(100, round((goals_against / 60) * 100, 1) if goals_against else 0), 'suffix': ''},
        {'label': 'Amarillas', 'value': yellows, 'pct': min(100, round((yellows / 75) * 100, 1) if yellows else 0), 'suffix': ''},
        {'label': 'Rojas', 'value': reds, 'pct': min(100, round((reds / 20) * 100, 1) if reds else 0), 'suffix': ''},
        {'label': 'Posesión*', 'value': f'{min(80, max(35, 45 + (success_rate / 4))):.1f}%', 'pct': min(100, max(0, min(80, max(35, 45 + (success_rate / 4))))), 'suffix': ''},
        {'label': 'Duelos', 'value': len(duels), 'pct': min(100, round((len(duels) / 240) * 100, 1) if duels else 0), 'suffix': ''},
        {'label': '% Acierto', 'value': f'{success_rate:.1f}%', 'pct': min(100, max(0, success_rate)), 'suffix': ''},
        {'label': 'Disparos/Gol', 'value': '-' if team_shots_per_goal is None else team_shots_per_goal, 'pct': 0 if team_shots_per_goal is None else max(0, min(100, round(100 - (min(team_shots_per_goal, 12) / 12) * 100, 1))), 'suffix': ''},
        {'label': 'Acciones totales', 'value': total_actions, 'pct': min(100, round((total_actions / 1200) * 100, 1) if total_actions else 0), 'suffix': ''},
        {'label': 'Acciones/partido', 'value': avg_actions, 'pct': min(100, round((avg_actions / 80) * 100, 1) if avg_actions else 0), 'suffix': ''},
        {'label': 'Duelos/partido', 'value': avg_duels, 'pct': min(100, round((avg_duels / 20) * 100, 1) if avg_duels else 0), 'suffix': ''},
        {'label': 'Amarillas/partido', 'value': avg_yellows, 'pct': min(100, round((avg_yellows / 4) * 100, 1) if avg_yellows else 0), 'suffix': ''},
        {'label': '% Duelo ganado', 'value': f'{duel_rate:.1f}%', 'pct': min(100, max(0, duel_rate)), 'suffix': ''},
    ]

    team_events = list(events)
    totals_breakdown = _summarize_events(team_events)
    total_mapped_actions = sum(int(count or 0) for count in totals_breakdown['zone_counts'].values())
    totals_breakdown['field_zones'] = [
        {
            **zone,
            'count': totals_breakdown['zone_counts'].get(zone['key'], 0),
            'pct': round((totals_breakdown['zone_counts'].get(zone['key'], 0) / total_mapped_actions) * 100, 1)
            if total_mapped_actions
            else 0,
        }
        for zone in FIELD_ZONES
    ]
    coach_total_field_zones = totals_breakdown['field_zones']
    coach_total_actions_count = total_mapped_actions
    pass_accuracy = (
        round((totals_breakdown['passes_completed'] / totals_breakdown['passes_attempts']) * 100, 1)
        if totals_breakdown['passes_attempts']
        else 0.0
    )
    shot_accuracy = (
        round((totals_breakdown['shots_on_target'] / totals_breakdown['shots_attempts']) * 100, 1)
        if totals_breakdown['shots_attempts']
        else 0.0
    )
    coach_overview_stats = {
        'rings': [
            {'label': 'Tasa de éxito', 'value': f'{success_rate:.1f}%', 'pct': success_rate},
            {'label': 'Duelos ganados', 'value': f'{duel_rate:.1f}%', 'pct': duel_rate},
            {'label': 'Precisión de pase', 'value': f'{pass_accuracy:.1f}%', 'pct': pass_accuracy},
            {'label': 'Tiro a puerta', 'value': f'{shot_accuracy:.1f}%', 'pct': shot_accuracy},
        ],
        'summary': [
            {'label': 'Partidos jugados', 'value': season_played_matches},
            {'label': 'Partidos medidos', 'value': measured_matches},
            {'label': 'Goles reales', 'value': goals_for},
            {'label': 'Goles medidos', 'value': event_goals_for},
            {'label': 'Acciones', 'value': total_actions},
            {'label': 'Duelos', 'value': len(duels)},
            {'label': 'Disparos', 'value': f"{totals_breakdown['shots_on_target']}/{totals_breakdown['shots_attempts']}"},
            {'label': 'Disparos/Gol', 'value': '-' if team_shots_per_goal is None else team_shots_per_goal},
            {'label': 'Pases', 'value': f"{totals_breakdown['passes_completed']}/{totals_breakdown['passes_attempts']}"},
            {'label': 'Tarjetas', 'value': card_total},
        ],
    }
    player_dashboard_rows = compute_player_dashboard(primary_team) if primary_team else []
    coach_player_options = [
        {
            'id': item.get('player_id'),
            'name': item.get('name'),
            'number': item.get('number'),
        }
        for item in sorted(player_dashboard_rows, key=lambda row: (str(row.get('name') or '').lower(), row.get('player_id') or 0))
    ]
    coach_player_cards_all = [
        {
            'id': item.get('player_id'),
            'name': item.get('name'),
            'number': item.get('number'),
            'minutes': int(item.get('minutes', 0) or 0),
            'matches': int(item.get('pj', 0) or 0),
            'goals': int(item.get('goals', 0) or 0),
            'assists': int(item.get('assists', 0) or 0),
            'position': item.get('position') or '-',
            'photo_url': item.get('photo_url') or '',
        }
        for item in sorted(
            player_dashboard_rows,
            key=lambda row: (
                -int(row.get('minutes', 0) or 0),
                int(row.get('number') or 9999),
                str(row.get('name') or '').lower(),
            ),
        )
    ]
    selected_player_id = _parse_int(request.GET.get('player'))
    selected_player = Player.objects.filter(id=selected_player_id, team=primary_team).first() if selected_player_id else None
    selected_player_stats = next((item for item in player_dashboard_rows if item.get('player_id') == selected_player_id), None) if selected_player_id else None
    coach_player_match_options = []
    coach_player_view = None
    if selected_player and selected_player_stats:
        player_communications = list(
            selected_player.communications.select_related('match').all()[:5]
        )
        player_fines = list(selected_player.fines.all()[:5])
        player_has_manual_sanction = is_manual_sanction_active(selected_player)
        fines_total_amount = sum(int(fine.amount or 0) for fine in player_fines)
        coach_fines_summary = [
            {'label': 'Multas', 'value': len(player_fines) + (1 if player_has_manual_sanction else 0)},
            {'label': 'Importe', 'value': f"{fines_total_amount}€"},
            {'label': 'Comunicaciones', 'value': len(player_communications)},
        ]
        coach_fines_records = [
            {
                'title': dict(PlayerFine.REASON_CHOICES).get(fine.reason, fine.reason),
                'meta': f"{fine.created_at:%d/%m/%Y} · {int(fine.amount or 0)}€",
                'detail': fine.note or 'Sin detalle',
            }
            for fine in player_fines
        ]
        if player_has_manual_sanction:
            until_label = (
                selected_player.manual_sanction_until.strftime('%d/%m/%Y')
                if selected_player.manual_sanction_until
                else 'Sin fecha fin'
            )
            coach_fines_records.insert(
                0,
                {
                    'title': 'Sanción manual',
                    'meta': until_label,
                    'detail': selected_player.manual_sanction_reason or 'Sanción configurada en ficha',
                },
            )
        coach_communications = [
            {
                'title': communication.get_category_display(),
                'meta': (
                    f"{communication.created_at:%d/%m/%Y}"
                    + (
                        f" · {communication.match.away_team.display_name if communication.match and communication.match.home_team_id == primary_team.id else communication.match.home_team.display_name if communication.match and communication.match.home_team_id else 'Partido'}"
                        if communication.match_id
                        else ''
                    )
                ),
                'detail': communication.message,
            }
            for communication in player_communications
        ]
        selected_player_match_id = _parse_int(request.GET.get('player_match'))
        coach_player_match_options = [
            {
                'id': item.get('match_id'),
                'label': f"{item.get('round') or 'Partido'} · {item.get('opponent') or 'Rival'}",
            }
            for item in selected_player_stats.get('matches', [])
            if item.get('match_id')
        ]
        season_rings = [
            {'label': 'Tasa de éxito', 'value': f"{float(selected_player_stats.get('success_rate') or 0):.1f}%", 'pct': float(selected_player_stats.get('success_rate') or 0)},
            {'label': 'Duelos ganados', 'value': f"{float(selected_player_stats.get('duel_rate') or 0):.1f}%", 'pct': float(selected_player_stats.get('duel_rate') or 0)},
            {'label': 'Precisión de pase', 'value': f"{float(selected_player_stats.get('passes', {}).get('accuracy') or 0):.1f}%", 'pct': float(selected_player_stats.get('passes', {}).get('accuracy') or 0)},
            {'label': 'Tiro a puerta', 'value': f"{float(selected_player_stats.get('shots', {}).get('accuracy') or 0):.1f}%", 'pct': float(selected_player_stats.get('shots', {}).get('accuracy') or 0)},
        ]
        season_summary = [
            {'label': 'Acciones', 'value': selected_player_stats.get('total_actions', 0)},
            {'label': 'Éxitos', 'value': selected_player_stats.get('successes', 0)},
            {'label': 'Pases', 'value': f"{selected_player_stats.get('passes', {}).get('completed', 0)}/{selected_player_stats.get('passes', {}).get('attempts', 0)}"},
            {'label': 'Disparos', 'value': f"{selected_player_stats.get('shots', {}).get('on_target', 0)}/{selected_player_stats.get('shots', {}).get('attempts', 0)}"},
            {'label': 'Disparos/Gol', 'value': '-' if selected_player_stats.get('shots', {}).get('per_goal') is None else selected_player_stats.get('shots', {}).get('per_goal')},
            {'label': 'Goles', 'value': selected_player_stats.get('goals', 0)},
            {'label': 'Asistencias', 'value': selected_player_stats.get('assists', 0)},
            {'label': 'Paradas', 'value': selected_player_stats.get('goalkeeper_saves', 0)},
            {'label': 'Minutos', 'value': selected_player_stats.get('minutes', 0)},
            {'label': 'Partidos', 'value': selected_player_stats.get('pj', 0)},
            {'label': 'Importancia', 'value': round(float(selected_player_stats.get('importance_score') or 0), 1)},
            {'label': 'Influencia', 'value': round(float(selected_player_stats.get('influence_score') or 0), 1)},
            {'label': 'Posición', 'value': selected_player_stats.get('position') or '-'},
        ]
        coach_player_view = {
            'mode': 'season',
            'title': f"{selected_player.name} · Temporada",
            'meta': selected_player.position or 'Jugador',
            'photo_url': resolve_player_photo_url(request, selected_player),
            'rings': season_rings,
            'summary': season_summary,
            'field_zones': selected_player_stats.get('field_zones', []),
            'fines_summary': coach_fines_summary,
            'fines_records': coach_fines_records,
            'communications': coach_communications,
        }
        if selected_player_match_id:
            selected_match = _team_match_queryset(primary_team).filter(id=selected_player_match_id).first()
            if selected_match:
                match_stats, match_payload = _build_player_match_stats_payload(primary_team, selected_player, selected_match)
                coach_player_view = {
                    'mode': 'match',
                    'title': f"{selected_player.name} · {match_payload['opponent']}",
                    'meta': f"{match_payload['date']} · Jornada {match_payload['round']}",
                    'photo_url': resolve_player_photo_url(request, selected_player),
                    'rings': [
                        {'label': 'Tasa de éxito', 'value': f"{float(match_stats.get('success_rate') or 0):.1f}%", 'pct': float(match_stats.get('success_rate') or 0)},
                        {'label': 'Duelos ganados', 'value': f"{float(match_stats.get('duel_rate') or 0):.1f}%", 'pct': float(match_stats.get('duel_rate') or 0)},
                        {'label': 'Precisión de pase', 'value': f"{float(match_stats.get('passes', {}).get('accuracy') or 0):.1f}%", 'pct': float(match_stats.get('passes', {}).get('accuracy') or 0)},
                        {'label': 'Tiro a puerta', 'value': f"{float(match_stats.get('shots', {}).get('accuracy') or 0):.1f}%", 'pct': float(match_stats.get('shots', {}).get('accuracy') or 0)},
                    ],
                    'summary': match_stats.get('kpi_summary', []),
                    'field_zones': match_stats.get('field_zones', []),
                    'fines_summary': coach_fines_summary,
                    'fines_records': coach_fines_records,
                    'communications': coach_communications,
                }
    match_events_map = defaultdict(list)
    for event in team_events:
        if event.match_id:
            match_events_map[event.match_id].append(event)
    team_matches = list(_team_match_queryset(primary_team).select_related('home_team', 'away_team').order_by('-date', '-id'))
    coach_match_rows = []
    for match in team_matches:
        match_events = match_events_map.get(match.id, [])
        metrics = _summarize_events(match_events)
        if match.home_team == primary_team:
            opponent_name = match.away_team.display_name if match.away_team else 'Rival desconocido'
        else:
            opponent_name = match.home_team.display_name if match.home_team else 'Rival desconocido'
        coach_match_rows.append(
            {
                'match_id': match.id,
                'round': match.round or f'Partido {match.id}',
                'date': match.date.strftime('%d/%m/%Y') if match.date else '--/--/----',
                'opponent': opponent_name,
                'location': match.location or 'Campo por confirmar',
                'total_actions': metrics['total_actions'],
                'success_rate': metrics['success_rate'],
                'duel_rate': metrics['duel_rate'],
                'yellow_cards': metrics['yellow_cards'],
                'red_cards': metrics['red_cards'],
            }
        )

    selected_match_id = _parse_int(request.GET.get('match'))
    valid_match_ids = {row['match_id'] for row in coach_match_rows}
    if selected_match_id not in valid_match_ids:
        selected_match_id = None
    selected_match_row = next((row for row in coach_match_rows if row['match_id'] == selected_match_id), None)
    selected_match_metrics = _summarize_events(match_events_map.get(selected_match_id, [])) if selected_match_id else None
    coach_match_view = None
    if selected_match_id:
        selected_match = _team_match_queryset(primary_team).filter(id=selected_match_id).first()
        if selected_match and selected_match_metrics:
            selected_match_team_metrics = compute_team_metrics_for_match(selected_match, primary_team=primary_team)
            selected_match_player_cards = compute_player_cards_for_match(selected_match, primary_team)
            total_selected_zone_actions = sum(
                int(count or 0) for count in selected_match_metrics['zone_counts'].values()
            )
            selected_match_field_zones = [
                {
                    **zone,
                    'count': selected_match_metrics['zone_counts'].get(zone['key'], 0),
                    'pct': round(
                        (selected_match_metrics['zone_counts'].get(zone['key'], 0) / total_selected_zone_actions) * 100,
                        1,
                    )
                    if total_selected_zone_actions
                    else 0,
                }
                for zone in FIELD_ZONES
            ]
            selected_pass_accuracy = (
                round(
                    (selected_match_metrics['passes_completed'] / selected_match_metrics['passes_attempts']) * 100,
                    1,
                )
                if selected_match_metrics['passes_attempts']
                else 0.0
            )
            selected_shot_accuracy = (
                round(
                    (selected_match_metrics['shots_on_target'] / selected_match_metrics['shots_attempts']) * 100,
                    1,
                )
                if selected_match_metrics['shots_attempts']
                else 0.0
            )
            opponent = (
                selected_match.away_team
                if selected_match.home_team == primary_team
                else selected_match.home_team
            )
            coach_match_view = {
                'title': f"Equipo · {opponent.display_name if opponent else 'Rival desconocido'}",
                'meta': (
                    f"{selected_match.date.strftime('%d/%m/%Y') if selected_match.date else 'Fecha por definir'}"
                    f" · Jornada {selected_match.round or '--'}"
                    f" · {selected_match.location or 'Campo por confirmar'}"
                ),
                'rings': [
                    {
                        'label': 'Tasa de éxito',
                        'value': f"{selected_match_metrics['success_rate']:.1f}%",
                        'pct': selected_match_metrics['success_rate'],
                    },
                    {
                        'label': 'Duelos ganados',
                        'value': f"{selected_match_metrics['duel_rate']:.1f}%",
                        'pct': selected_match_metrics['duel_rate'],
                    },
                    {
                        'label': 'Precisión de pase',
                        'value': f"{selected_pass_accuracy:.1f}%",
                        'pct': selected_pass_accuracy,
                    },
                    {
                        'label': 'Tiro a puerta',
                        'value': f"{selected_shot_accuracy:.1f}%",
                        'pct': selected_shot_accuracy,
                    },
                ],
                'summary': [
                    {'label': 'Acciones', 'value': selected_match_metrics['total_actions']},
                    {'label': 'Goles', 'value': selected_match_metrics['goals']},
                    {
                        'label': 'Pases',
                        'value': f"{selected_match_metrics['passes_completed']}/{selected_match_metrics['passes_attempts']}",
                    },
                    {
                        'label': 'Disparos',
                        'value': f"{selected_match_metrics['shots_on_target']}/{selected_match_metrics['shots_attempts']}",
                    },
                    {
                        'label': 'Disparos/Gol',
                        'value': '-'
                        if selected_match_metrics['shots_per_goal'] is None
                        else selected_match_metrics['shots_per_goal'],
                    },
                    {
                        'label': 'Tarjetas',
                        'value': f"{selected_match_metrics['yellow_cards']}A · {selected_match_metrics['red_cards']}R",
                    },
                ],
                'field_zones': selected_match_field_zones,
                'player_cards': selected_match_player_cards,
                'top_event_types': selected_match_team_metrics['top_event_types'],
            }

    modules = [
        {'title': 'Convocatoria', 'description': 'Define lista oficial y PDF del partido.', 'link': 'convocation'},
        {'title': '11 inicial', 'description': 'Pantalla táctica visual para construir la alineación titular y banquillo.', 'link': 'initial-eleven'},
        {'title': 'Sesiones', 'description': 'Planificador semanal de sesiones y tareas.', 'link': 'sessions'},
        {'title': 'Multas', 'description': 'Control disciplinario del vestuario.', 'link': 'fines'},
    ]
    return render(
        request,
        'football/coach_role_hub.html',
        {
            'role_title': 'Entrenador',
            'role_key': 'trainer',
            'role_description': 'Área operativa principal del staff técnico.',
            'modules': modules,
            'kpis': kpis,
            'kpi_note': '* Goles, puntos y clasificación: temporada real. Métricas de acciones: solo sobre partidos medidos.',
            'coach_overview_stats': coach_overview_stats,
            'coach_general_stats': coach_general_stats,
            'coach_player_leaders': coach_player_leaders,
            'coach_total_field_zones': coach_total_field_zones,
            'coach_total_actions_count': coach_total_actions_count,
            'coach_match_rows': coach_match_rows,
            'coach_selected_match': selected_match_row,
            'coach_selected_match_metrics': selected_match_metrics,
            'coach_match_view': coach_match_view,
            'weekly_staff_brief': weekly_staff_brief,
            'coach_measured_matches': measured_matches,
            'coach_player_options': coach_player_options,
            'coach_player_cards_all': coach_player_cards_all,
            'coach_selected_player_id': selected_player_id,
            'coach_player_match_options': coach_player_match_options,
            'coach_selected_player_match_id': _parse_int(request.GET.get('player_match')),
            'coach_player_view': coach_player_view,
        },
    )


@login_required
def coach_role_goalkeeper_page(request):
    forbidden = _forbid_if_no_coach_access(request.user)
    if forbidden:
        return forbidden
    modules = [
        {
            'title': 'Tareas portero · Carga PDF',
            'description': 'Sube, analiza y crea tareas especificas de porteria.',
            'link': 'sessions-goalkeeper',
        },
        {'title': 'Pizarra táctica', 'description': 'Simula secuencias de portero en campo y guarda movimientos.', 'link': 'coach-abp-board'},
    ]
    return render(
        request,
        'football/coach_role_hub.html',
        {
            'role_title': 'Preparador de porteros',
            'role_key': 'goalkeeper',
            'role_description': 'Repositorio técnico especializado para trabajo específico de portería.',
            'modules': modules,
        },
    )


@login_required
def coach_role_fitness_page(request):
    forbidden = _forbid_if_no_coach_access(request.user)
    if forbidden:
        return forbidden
    modules = [
        {
            'title': 'Tareas fisicas · Carga PDF',
            'description': 'Sube, analiza y crea tareas de preparacion fisica.',
            'link': 'sessions-fitness',
        },
        {'title': 'Métricas físicas', 'description': 'Sección lista para cargar datos, test y seguimiento.', 'link': 'player-dashboard'},
        {'title': 'Registro individual', 'description': 'Completa datos físicos por jugador desde su ficha.', 'link': 'player-dashboard'},
    ]
    return render(
        request,
        'football/coach_role_hub.html',
        {
            'role_title': 'Preparación física',
            'role_key': 'fitness',
            'role_description': 'Base preparada para incorporar carga interna/externa y control físico.',
            'modules': modules,
        },
    )


@login_required
def coach_role_abp_page(request):
    forbidden = _forbid_if_no_coach_access(request.user)
    if forbidden:
        return forbidden
    modules = [
        {'title': 'Repositorio ABP', 'description': 'Guarda tareas y sesiones ABP en el planificador.', 'link': 'sessions'},
        {'title': 'Pizarra ABP', 'description': 'Campo interactivo con fichas, grabación y reproducción de jugadas.', 'link': 'coach-abp-board'},
    ]
    return render(
        request,
        'football/coach_role_hub.html',
        {
            'role_title': 'ABP',
            'role_key': 'abp',
            'role_description': 'Acciones a balón parado: diseño, simulación y biblioteca de jugadas.',
            'modules': modules,
        },
    )


@login_required
def coach_abp_board_page(request):
    forbidden = _forbid_if_no_coach_access(request.user)
    if forbidden:
        return forbidden
    forbidden = _forbid_if_workspace_module_disabled(request, 'abp_board', label='pizarra ABP')
    if forbidden:
        return forbidden
    primary_team = _get_primary_team_for_request(request)
    players = []
    if primary_team:
        players = list(
            Player.objects.filter(team=primary_team, is_active=True).order_by('number', 'name')[:28]
        )
    return render(
        request,
        'football/coach_abp_board.html',
        {
            'players': players,
            'team_name': primary_team.display_name if primary_team else 'Equipo principal',
        },
    )


def coach_roster_page(request):
    forbidden = _forbid_if_workspace_module_disabled(request, 'players', label='plantilla técnica')
    if forbidden:
        return forbidden
    primary_team = _get_primary_team_for_request(request)
    if not primary_team:
        raise Http404('Equipo principal no configurado')

    message = ''
    error = ''
    if request.method == 'POST':
        action = (request.POST.get('action') or 'add').strip()
        player_id = _parse_int(request.POST.get('player_id'))
        name = (request.POST.get('name') or '').strip()
        number_raw = (request.POST.get('number') or '').strip()
        position = (request.POST.get('position') or '').strip()
        is_active = (request.POST.get('is_active') or '1').strip() in {'1', 'true', 'on', 'yes'}

        try:
            if action == 'deactivate':
                if not player_id:
                    raise ValueError('Jugador no válido para desactivar.')
                player = Player.objects.filter(id=player_id, team=primary_team).first()
                if not player:
                    raise ValueError('Jugador no encontrado.')
                player.is_active = False
                player.save(update_fields=['is_active'])
                message = f'{player.name} marcado como inactivo.'
            else:
                if not name:
                    raise ValueError('El nombre es obligatorio.')
                number = _parse_int(number_raw) if number_raw else None
                player = (
                    Player.objects.filter(team=primary_team, name__iexact=name)
                    .order_by('id')
                    .first()
                )
                if player:
                    player.number = number
                    player.position = position
                    player.is_active = is_active
                    player.save(update_fields=['number', 'position', 'is_active'])
                    message = f'Jugador actualizado: {player.name}.'
                else:
                    Player.objects.create(
                        team=primary_team,
                        name=name,
                        number=number,
                        position=position,
                        is_active=is_active,
                    )
                    message = f'Jugador añadido: {name}.'
        except ValueError as exc:
            error = str(exc)
        except Exception:
            error = 'No se pudo guardar el jugador. Revisa los datos.'

    players = Player.objects.filter(team=primary_team).order_by('is_active', 'number', 'name')
    return render(
        request,
        'football/coach_roster.html',
        {
            'team_name': primary_team.display_name,
            'players': players,
            'message': message,
            'error': error,
        },
    )


def initial_eleven_page(request):
    forbidden = _forbid_if_workspace_module_disabled(request, 'convocation', label='11 inicial')
    if forbidden:
        return forbidden
    primary_team = _get_primary_team_for_request(request)
    if not primary_team:
        raise Http404('Equipo principal no configurado')
    convocation_record = get_current_convocation_record(
        primary_team,
        match=get_active_match(primary_team),
        fallback_to_latest=True,
    )
    convocation_players = list(convocation_record.players.order_by('number', 'name')) if convocation_record else []
    for player in convocation_players:
        player.photo_url = resolve_player_photo_url(request, player)

    lineup_seed = {'starters': [], 'bench': []}
    has_pending_lineup = False
    has_pending_convocation = bool(convocation_record and not convocation_players)
    if convocation_record:
        stored = convocation_record.lineup_data if isinstance(convocation_record.lineup_data, dict) else {}
        normalized = _normalize_lineup_payload(stored, convocation_players)
        if normalized['starters'] or normalized['bench']:
            lineup_seed = normalized
        else:
            has_pending_lineup = bool(convocation_players)

    return render(
        request,
        'football/coach_initial_eleven.html',
        {
            'team_name': primary_team.display_name,
            'team_crest_url': resolve_team_crest_url(request, primary_team, sync=True),
            'rival_crest_url': _absolute_universo_url(
                _resolve_rival_identity(
                    getattr(convocation_record, 'opponent_name', '') or 'Rival por confirmar'
                )[1]
            ) if convocation_record else '',
            'convocation_record': convocation_record,
            'convocation_players': convocation_players,
            'lineup_seed_json': json.dumps(lineup_seed, ensure_ascii=False),
            'has_pending_convocation': has_pending_convocation,
            'has_pending_lineup': has_pending_lineup,
        },
    )


def _extract_pdf_text(pdf_file, max_chars=12000):
    if PdfReader is None:
        raise ValueError('Falta dependencia de lectura PDF. Instala `pypdf`.')

    def _ocr_text_from_image_bytes(raw_bytes):
        if not raw_bytes or Image is None or pytesseract is None:
            return ''
        try:
            with Image.open(io.BytesIO(raw_bytes)) as img:
                rgb = img.convert('RGB')
                # OCR bilingual by default to improve mixed ES/EN task sheets.
                return (pytesseract.image_to_string(rgb, lang='spa+eng') or '').strip()
        except Exception:
            return ''

    def _needs_ocr_boost(parsed_text):
        compact_hits = re.findall(r'\b[A-ZÁÉÍÓÚÜÑ]{10,}\b', str(parsed_text or ''))
        joined_hits = 0
        for token in compact_hits[:120]:
            repaired = _split_joined_upper_token(token)
            if repaired == token:
                joined_hits += 1
        return len(str(parsed_text or '')) < 500 or joined_hits >= 2

    def _ocr_pdf_pages_with_pdftoppm(local_pdf_file, max_pages=4):
        if pytesseract is None or Image is None:
            return ''
        pdftoppm_bin = shutil.which('pdftoppm')
        if not pdftoppm_bin:
            return ''
        try:
            if hasattr(local_pdf_file, 'seek'):
                local_pdf_file.seek(0)
            pdf_bytes = local_pdf_file.read()
            if not pdf_bytes:
                return ''
            with tempfile.TemporaryDirectory(prefix='task-ocr-') as tmpdir:
                tmp_path = Path(tmpdir)
                source_pdf = tmp_path / 'source.pdf'
                source_pdf.write_bytes(pdf_bytes)
                ocr_chunks = []
                page_limit = max(1, int(max_pages or 1))
                for page_no in range(1, page_limit + 1):
                    out_base = tmp_path / f'page-{page_no}'
                    subprocess.run(
                        [
                            pdftoppm_bin,
                            '-jpeg',
                            '-r',
                            '170',
                            '-f',
                            str(page_no),
                            '-singlefile',
                            str(source_pdf),
                            str(out_base),
                        ],
                        check=True,
                        capture_output=True,
                        timeout=35,
                    )
                    page_img = tmp_path / f'page-{page_no}.jpg'
                    if not page_img.exists():
                        continue
                    raw = page_img.read_bytes()
                    ocr_text = _ocr_text_from_image_bytes(raw)
                    if ocr_text:
                        ocr_chunks.append(ocr_text)
                    if sum(len(c) for c in ocr_chunks) >= 9000:
                        break
                return '\n'.join(ocr_chunks).strip()
        except Exception:
            return ''

    try:
        if hasattr(pdf_file, 'seek'):
            pdf_file.seek(0)
        reader = PdfReader(pdf_file)
        chunks = []
        for page in reader.pages:
            chunks.append((page.extract_text() or '').strip())
        text = '\n'.join([item for item in chunks if item]).strip()
        text = re.sub(r'\n{3,}', '\n\n', text)
        if len(text) < 180:
            ocr_chunks = []
            for page in reader.pages[:5]:
                images = getattr(page, 'images', []) or []
                for image in images[:3]:
                    image_bytes = getattr(image, 'data', b'') or b''
                    ocr_text = _ocr_text_from_image_bytes(image_bytes)
                    if ocr_text:
                        ocr_chunks.append(ocr_text)
                if sum(len(c) for c in ocr_chunks) >= 5000:
                    break
            if ocr_chunks:
                text = '\n'.join([text, '\n'.join(ocr_chunks)]).strip() if text else '\n'.join(ocr_chunks)
                text = re.sub(r'\n{3,}', '\n\n', text)
        if _needs_ocr_boost(text):
            ocr_rendered = _ocr_pdf_pages_with_pdftoppm(pdf_file, max_pages=5)
            if ocr_rendered:
                merged = '\n'.join([text, ocr_rendered]).strip() if text else ocr_rendered
                text = re.sub(r'\n{3,}', '\n\n', merged)
        text = _polish_spanish_text(_repair_joined_words_text(text), multiline=True)
        return text[:max_chars]
    except Exception:
        raise ValueError('No se pudo leer el PDF. Verifica que no esté protegido o corrupto.')


def _analyze_preview_image_bytes(raw_bytes):
    if Image is None or not raw_bytes:
        return None
    try:
        with Image.open(io.BytesIO(raw_bytes)) as img:
            rgb = img.convert('RGB')
            width, height = rgb.size
            if width <= 0 or height <= 0:
                return None
            sample = rgb.copy()
            sample.thumbnail((128, 128))
            pixels = list(sample.getdata())
            total = max(1, len(pixels))
            greenish = 0
            whitish = 0
            darkish = 0
            colorful = 0
            for r, g, b in pixels:
                if g > 70 and g > (r + 14) and g > (b + 10):
                    greenish += 1
                if r > 232 and g > 232 and b > 232:
                    whitish += 1
                if r < 30 and g < 30 and b < 30:
                    darkish += 1
                if (max(r, g, b) - min(r, g, b)) >= 24:
                    colorful += 1
            green_ratio = greenish / total
            white_ratio = whitish / total
            dark_ratio = darkish / total
            color_ratio = colorful / total
            area = width * height
            aspect = width / max(1.0, float(height))

            score = 0.0
            score += min(area / float(1280 * 720), 2.6) * 28.0
            score += green_ratio * 46.0
            score += color_ratio * 16.0
            if 0.95 <= aspect <= 2.35:
                score += 8.0
            if 1.15 <= aspect <= 2.05:
                score += 6.0
            score -= white_ratio * 44.0
            if min(width, height) < 210:
                score -= 26.0
            if white_ratio > 0.78 and green_ratio < 0.06:
                score -= 48.0
            if dark_ratio > 0.88:
                score -= 20.0

            return {
                'width': int(width),
                'height': int(height),
                'area': int(area),
                'aspect': float(aspect),
                'green_ratio': float(green_ratio),
                'white_ratio': float(white_ratio),
                'dark_ratio': float(dark_ratio),
                'score': float(score),
            }
    except Exception:
        return None


def _is_preview_quality_low(raw_bytes):
    metrics = _analyze_preview_image_bytes(raw_bytes)
    if not metrics:
        return False
    score = float(metrics.get('score') or 0.0)
    area = int(metrics.get('area') or 0)
    white_ratio = float(metrics.get('white_ratio') or 0.0)
    green_ratio = float(metrics.get('green_ratio') or 0.0)
    # Typical bad previews: tiny crops, white text snippets, or non-graphic chunks.
    if area < 260 * 170:
        return True
    if score < 10.0:
        return True
    if white_ratio > 0.72 and green_ratio < 0.08:
        return True
    return False


def _extract_preview_images_from_pdf(pdf_file, max_images=8, prefer_render=False):
    if PdfReader is None or pdf_file is None:
        return []
    if prefer_render:
        rendered_payload = _render_pdf_preview_with_pdftoppm(pdf_file)
        if rendered_payload:
            return [rendered_payload]
    payloads = []
    candidates = []
    try:
        if hasattr(pdf_file, 'seek'):
            pdf_file.seek(0)
        reader = PdfReader(pdf_file)
        seq = 0
        for page_idx, page in enumerate(reader.pages):
            images = getattr(page, 'images', []) or []
            for image in images:
                seq += 1
                raw = getattr(image, 'data', b'') or b''
                if not raw:
                    continue
                ext = str(getattr(image, 'name', 'img.bin') or 'img.bin').rsplit('.', 1)[-1].lower()
                if ext not in {'png', 'jpg', 'jpeg', 'webp'}:
                    ext = 'png'
                metrics = _analyze_preview_image_bytes(raw)
                score = float(metrics.get('score') or 0.0) if metrics else 0.0
                candidates.append(
                    {
                        'raw': raw,
                        'ext': ext,
                        'score': score,
                        'page_idx': page_idx,
                        'seq': seq,
                    }
                )
    except Exception:
        candidates = []
    if candidates:
        max_count = max(1, int(max_images or 1))
        good_candidates = [item for item in candidates if float(item.get('score') or 0.0) >= 12.0]
        if good_candidates:
            # Keep original document order for multi-task PDFs once quality threshold is met.
            selected = sorted(good_candidates, key=lambda item: (int(item.get('page_idx') or 0), int(item.get('seq') or 0)))
        else:
            # If every embedded image is poor, still try top-scored ones before falling back.
            selected = sorted(candidates, key=lambda item: float(item.get('score') or 0.0), reverse=True)
        for item in selected[:max_count]:
            ext = str(item.get('ext') or 'jpg')
            filename = f'task-preview-{uuid.uuid4().hex[:10]}.{ext}'
            payloads.append((filename, ContentFile(item.get('raw') or b'')))
        if payloads:
            return payloads
    if payloads:
        return payloads

    # Fallback: render first page when PDF uses vector drawings (no embedded images).
    fallback_payload = _render_pdf_preview_with_pdftoppm(pdf_file)
    if fallback_payload:
        return [fallback_payload]

    # Last fallback: generic field image so old tasks never stay without thumbnail.
    default_payload = _default_task_preview_payload()
    if default_payload:
        return [default_payload]
    return payloads


def _extract_preview_image_from_pdf(pdf_file, prefer_render=False):
    payloads = _extract_preview_images_from_pdf(pdf_file, max_images=1, prefer_render=prefer_render)
    return payloads[0] if payloads else None


def _render_pdf_preview_with_pdftoppm(pdf_file):
    if pdf_file is None:
        return None
    pdftoppm_bin = shutil.which('pdftoppm')
    if not pdftoppm_bin:
        return None
    try:
        if hasattr(pdf_file, 'seek'):
            pdf_file.seek(0)
        pdf_bytes = pdf_file.read()
        if not pdf_bytes:
            return None
        with tempfile.TemporaryDirectory(prefix='task-preview-') as tmpdir:
            tmp_path = Path(tmpdir)
            source_pdf = tmp_path / 'source.pdf'
            output_base = tmp_path / 'preview'
            source_pdf.write_bytes(pdf_bytes)
            subprocess.run(
                [
                    pdftoppm_bin,
                    '-jpeg',
                    '-f',
                    '1',
                    '-singlefile',
                    '-scale-to',
                    '1400',
                    str(source_pdf),
                    str(output_base),
                ],
                check=True,
                capture_output=True,
                timeout=30,
            )
            output_jpg = tmp_path / 'preview.jpg'
            if not output_jpg.exists():
                return None
            raw = output_jpg.read_bytes()
            if not raw:
                return None
            if Image is not None:
                try:
                    with Image.open(io.BytesIO(raw)) as img:
                        optimized = img.convert('RGB')
                        optimized.thumbnail((1400, 1000))
                        buffer = io.BytesIO()
                        optimized.save(buffer, format='JPEG', quality=76, optimize=True)
                        raw = buffer.getvalue()
                except Exception:
                    pass
            filename = f'task-preview-{uuid.uuid4().hex[:10]}.jpg'
            return filename, ContentFile(raw)
    except Exception:
        return None


def _default_task_preview_payload():
    try:
        fallback_path = Path(settings.BASE_DIR) / 'static' / 'football' / 'campo-futbol.jpg'
        if not fallback_path.exists() or not fallback_path.is_file():
            return None
        raw = fallback_path.read_bytes()
        if Image is not None:
            try:
                with Image.open(io.BytesIO(raw)) as img:
                    normalized = img.convert('RGB')
                    normalized.thumbnail((1200, 850))
                    buffer = io.BytesIO()
                    normalized.save(buffer, format='JPEG', quality=74, optimize=True)
                    raw = buffer.getvalue()
            except Exception:
                pass
        filename = f'task-preview-{uuid.uuid4().hex[:10]}.jpg'
        return filename, ContentFile(raw)
    except Exception:
        return None


def _ensure_task_preview_image(task, prefer_render=False):
    if not task or not getattr(task, 'task_pdf', None):
        return False
    preview_payload = _extract_preview_image_from_pdf(task.task_pdf, prefer_render=prefer_render)
    if not preview_payload:
        return False
    preview_name, preview_content = preview_payload
    try:
        task.task_preview_image.save(preview_name, preview_content, save=True)
        return bool(task.task_preview_image)
    except Exception:
        return False


def _task_preview_needs_refresh(task):
    if not task:
        return False
    preview_name = str(getattr(task, 'task_preview_image', '') or '').strip()
    if not preview_name:
        return True
    try:
        if not default_storage.exists(preview_name):
            return True
    except Exception:
        return True
    cache_key = f'football:preview-quality:{preview_name}'
    cached = cache.get(cache_key)
    if cached is not None:
        return bool(cached)
    try:
        with default_storage.open(preview_name, 'rb') as handle:
            raw = handle.read()
        needs_refresh = _is_preview_quality_low(raw)
    except Exception:
        needs_refresh = True
    cache.set(cache_key, bool(needs_refresh), 60 * 60 * 6)
    return bool(needs_refresh)


def _ensure_library_task_preview(task, force=False, prefer_render=False):
    if not task:
        return False
    current_name = str(getattr(task, 'task_preview_image', '') or '').strip()
    if current_name and not force:
        try:
            if default_storage.exists(current_name):
                return True
        except Exception:
            pass
    if getattr(task, 'task_pdf', None):
        if _ensure_task_preview_image(task, prefer_render=prefer_render):
            return True
    fallback = _default_task_preview_payload()
    if not fallback:
        return False
    fallback_name, fallback_content = fallback
    try:
        task.task_preview_image.save(fallback_name, fallback_content, save=True)
        return bool(task.task_preview_image)
    except Exception:
        return False


TASK_CONTEXT_KEYWORDS = {
    'presion_alta': ['presion alta', 'presión alta', 'repliegue tras perdida', 'robo alto'],
    'salida_balon': ['salida de balon', 'salida de balón', 'inicio juego', 'construccion'],
    'finalizacion': ['finalizacion', 'finalización', 'remate', 'tiro', 'gol'],
    'transicion_ofensiva': ['transicion ofensiva', 'transición ofensiva', 'contraataque'],
    'transicion_defensiva': ['transicion defensiva', 'transición defensiva', 'balance defensivo'],
    'abp': ['abp', 'balon parado', 'balón parado', 'corner', 'falta lateral'],
    'porteros': ['portero', 'porteria', 'porteria', 'blocaje', 'despeje'],
    'fisico': ['fuerza', 'resistencia', 'velocidad', 'aceleracion', 'aceleración', 'potencia'],
}

TASK_OBJECTIVE_KEYWORDS = {
    'perfil_corporal': ['perfil corporal', 'orientacion corporal', 'orientación corporal'],
    'tercer_hombre': ['tercer hombre', 'hombre libre'],
    'amplitud': ['amplitud', 'carril exterior', 'lado debil', 'lado débil'],
    'temporizacion': ['temporizacion', 'temporización', 'esperar salto'],
    'duelo': ['duelo', '1v1', '1x1'],
    'coordinacion': ['coordinacion', 'coordinación', 'sincronizacion', 'sincronización'],
}

TASK_TYPE_KEYWORDS = {
    'rondo': ['rondo'],
    'posesion': ['posesion', 'posesión', 'conservacion', 'conservación'],
    'juego_posicional': ['juego de posicion', 'juego de posición', 'posicional'],
    'transicion': ['transicion', 'transición', 'contraataque', 'ataque rapido', 'ataque rápido'],
    'finalizacion': ['finalizacion', 'finalización', 'remate', 'tiro', 'gol'],
    'circuito': ['circuito', 'estacion', 'estación', 'posta'],
    'abp': ['abp', 'balon parado', 'balón parado', 'corner', 'falta lateral', 'saque de esquina'],
    'partido_reducido': ['partido reducido', 'ssg', 'small sided', 'juego reducido'],
}

TASK_PHASE_KEYWORDS = {
    'activacion': ['activacion', 'activación', 'calentamiento', 'entrada en calor', 'movilidad inicial'],
    'abp': ['abp', 'balon parado', 'balón parado', 'corner', 'falta lateral', 'saque de esquina'],
    'posesion': ['posesion', 'posesión', 'conservacion', 'conservación', 'rondo', 'juego de posicion', 'juego de posición'],
    'ataque': ['ataque', 'ofensiva', 'ofensivo', 'con balon', 'con balón'],
    'defensa': ['defensa', 'defensiva', 'defensivo', 'sin balon', 'sin balón'],
    'transicion': ['transicion', 'transición', 'tras perdida', 'tras pérdida', 'tras robo'],
    'mixta': ['ida y vuelta', 'ataque y defensa', 'mixto'],
}

PHASE_FOLDER_META = [
    {'key': 'activacion', 'label': 'Activación'},
    {'key': 'abp', 'label': 'ABP'},
    {'key': 'ataque', 'label': 'Ataque'},
    {'key': 'defensa', 'label': 'Defensa'},
    {'key': 'posesion', 'label': 'Posesión'},
    {'key': 'transicion', 'label': 'Transición'},
    {'key': 'mixta', 'label': 'Mixta'},
    {'key': 'sin_clasificar', 'label': 'Sin clasificar'},
]

PHASE_FOLDER_PRIORITY = [item['key'] for item in PHASE_FOLDER_META if item['key'] != 'sin_clasificar']


def _phase_folder_key_for_task(task):
    tags = getattr(task, 'phase_tags', None)
    if not isinstance(tags, list):
        tags = []
    normalized_tags = [str(tag or '').strip().lower() for tag in tags if str(tag or '').strip()]
    for key in PHASE_FOLDER_PRIORITY:
        if key in normalized_tags:
            return key
    return 'sin_clasificar'

TASK_PDF_PARSE_VERSION = 4

TASK_JOINED_WORD_VOCAB = [
    'TAREA', 'EJERCICIO', 'SESION', 'BLOQUE', 'BLOQUES', 'PARTE', 'PRINCIPAL',
    'DESCRIPCION', 'OBJETIVO', 'CONSIGNAS', 'REGLAS', 'DOSIFICACION',
    'TIEMPO', 'TOTAL', 'TRABAJO', 'PAUSA', 'SERIE', 'SERIES', 'REPETICIONES',
    'JUGADORES', 'COMODIN', 'COMODINES', 'MATERIAL', 'MATERIALES',
    'PORTERIA', 'PORTERIAS', 'PORTERO', 'PORTEROS', 'PORTERIA', 'MOVIL',
    'BALON', 'CONO', 'ARCO', 'PRECISION', 'FINALIZACION', 'DUEL', 'DUELO', 'DUELOS',
    'AEREOS', 'AEREO', 'TRANSICION', 'DEFENSA', 'ATAQUE', 'OFENSIVA', 'DEFENSIVA',
    'CAMPO', 'MEDIO', 'ZONA', 'ESPACIO', 'MEDIDAS', 'FRENTE', 'ORIENTADOS',
    'TRABAJAMOS', 'TRABAJAR', 'GENERA', 'GENERAMOS', 'HACEMOS', 'HACEMOSUN',
    'DONDE', 'CUANDO', 'CON', 'SIN', 'PARA', 'POR', 'DEL', 'AL',
    'LA', 'EL', 'LOS', 'LAS', 'UN', 'UNA', 'UNO', 'Y', 'DE', 'EN',
    'VERDE', 'ROJO', 'SALIR', 'SALE', 'SE', 'QUE', 'YA', 'VEZ',
    'MAS', 'CAIDA', 'DESPEJE', 'ORIENTADO', 'INFERIORIDAD', 'SUPERIORIDAD',
    'REACCION', 'PRESION', 'RECUPERACION', 'PASE', 'APOYO',
    'PAREDES', 'RAPIDO', 'JUEGO', 'DIRECTO', 'CAIDAS', 'DESMARQUE',
    'RUPTURA', 'DELIMITADA', 'HIDRATACION', 'ENTRENAMIENTO', 'MOVILIDAD',
    'ACTIVACION', 'ACTIVACION', 'DEFENDER', 'INCIDIMOS', 'PERDIDA',
    'TRABAJADOS', 'EQUIPOS', 'PREMISAS', 'FOMENTAR', 'COMODINES', 'FUERA',
]

TASK_JOINED_WORD_VOCAB_ES = [
    'a', 'al', 'algo', 'algunos', 'ante', 'antes', 'asi', 'aun',
    'bajo', 'bien', 'cada', 'cambio', 'casi', 'como', 'con', 'contra',
    'cuando', 'de', 'del', 'desde', 'donde', 'dos', 'durante', 'e', 'el',
    'ella', 'ellas', 'ellos', 'en', 'entre', 'era', 'es', 'esa', 'ese',
    'eso', 'esta', 'estaba', 'estado', 'estan', 'estar', 'este', 'esto',
    'final', 'frente', 'fue', 'ha', 'hacia', 'hasta', 'hay', 'hacer',
    'hacemos', 'igual', 'la', 'las', 'le', 'les', 'lo', 'los', 'mas',
    'media', 'medio', 'mientras', 'misma', 'mismo', 'muy', 'nada', 'ni',
    'no', 'nos', 'nosotros', 'o', 'otra', 'otro', 'para', 'pero', 'poca',
    'poco', 'por', 'porque', 'primera', 'primero', 'puede', 'que', 'quien',
    'rapido', 'se', 'segun', 'si', 'sin', 'sobre', 'solo', 'su', 'sus',
    'te', 'todo', 'trabajo', 'tras', 'tu', 'un', 'una', 'uno', 'unos',
    'ya', 'y',
]


def _normalize_folded_text(value):
    raw = str(value or '').strip().lower()
    if not raw:
        return ''
    normalized = unicodedata.normalize('NFKD', raw)
    return ''.join(ch for ch in normalized if not unicodedata.combining(ch))


def _normalize_upper_compact(value):
    raw = str(value or '').strip().upper()
    if not raw:
        return ''
    normalized = unicodedata.normalize('NFKD', raw)
    stripped = ''.join(ch for ch in normalized if not unicodedata.combining(ch))
    return re.sub(r'[^A-Z]+', '', stripped)


def _normalize_lower_compact(value):
    raw = str(value or '').strip().lower()
    if not raw:
        return ''
    normalized = unicodedata.normalize('NFKD', raw)
    stripped = ''.join(ch for ch in normalized if not unicodedata.combining(ch))
    return re.sub(r'[^a-z]+', '', stripped)


JOINED_WORD_MAP = {
    _normalize_upper_compact(word): str(word).upper()
    for word in TASK_JOINED_WORD_VOCAB
    if _normalize_upper_compact(word)
}

JOINED_WORD_MAP_LOWER = {}
for _word in (TASK_JOINED_WORD_VOCAB + TASK_JOINED_WORD_VOCAB_ES):
    compact = _normalize_lower_compact(_word)
    if compact:
        JOINED_WORD_MAP_LOWER[compact] = compact


def _split_joined_upper_token(token):
    compact = _normalize_upper_compact(token)
    if len(compact) < 10:
        return token
    n = len(compact)
    best = [None] * (n + 1)
    best[0] = (0, [])
    max_len = min(18, n)
    for i in range(n):
        if best[i] is None:
            continue
        base_score, base_words = best[i]
        for step in range(1, max_len + 1):
            j = i + step
            if j > n:
                break
            chunk = compact[i:j]
            mapped = JOINED_WORD_MAP.get(chunk)
            if not mapped:
                continue
            bonus = (step * step) + (4 if step >= 4 else 0)
            candidate = (base_score + bonus, base_words + [mapped])
            if best[j] is None or candidate[0] > best[j][0]:
                best[j] = candidate
    if best[n] is None:
        return token
    words = best[n][1]
    if len(words) < 2:
        return token
    joined = ' '.join(words).strip()
    # Keep guardrails: avoid over-splitting into too many tiny pieces.
    if len(words) >= 7 and (len(compact) / max(1, len(words))) < 2.2:
        return token
    return joined or token


def _split_joined_alpha_token(token):
    compact = _normalize_lower_compact(token)
    if len(compact) < 8:
        return token
    n = len(compact)
    best = [None] * (n + 1)
    best[0] = (0, [])
    max_len = min(20, n)
    for i in range(n):
        if best[i] is None:
            continue
        base_score, base_words = best[i]
        for step in range(1, max_len + 1):
            j = i + step
            if j > n:
                break
            chunk = compact[i:j]
            mapped = JOINED_WORD_MAP_LOWER.get(chunk)
            if not mapped:
                continue
            bonus = (step * step) + (6 if step >= 4 else 0) + (4 if step >= 6 else 0)
            candidate = (base_score + bonus, base_words + [mapped])
            if best[j] is None or candidate[0] > best[j][0]:
                best[j] = candidate
    if best[n] is None:
        return token
    words = best[n][1]
    if len(words) < 2:
        return token
    avg_len = len(compact) / max(1, len(words))
    if len(words) >= 8 and avg_len < 2.1:
        return token
    if avg_len < 2.5 and len(words) > 5:
        return token
    if token.isupper():
        return ' '.join(word.upper() for word in words).strip()
    if token[:1].isupper():
        joined = ' '.join(words).strip()
        return joined[:1].upper() + joined[1:]
    return ' '.join(words).strip()


def _repair_joined_words_text(value):
    text = str(value or '')
    if not text:
        return ''
    cleaned = text.replace('\r\n', '\n').replace('\r', '\n')
    cleaned = re.sub(r'(?<=\d)(?=[A-Za-zÁÉÍÓÚÜÑ])', ' ', cleaned)
    cleaned = re.sub(r'(?<=[A-Za-zÁÉÍÓÚÜÑ])(?=\d)', ' ', cleaned)
    cleaned = re.sub(r'(?<=\d)\s*[xX×]\s*(?=\d)', ' x ', cleaned)
    cleaned = re.sub(r'([,:;])(?=\S)', r'\1 ', cleaned)
    cleaned = re.sub(r'\s{2,}', ' ', cleaned)
    cleaned = re.sub(r' *\n *', '\n', cleaned)

    def _replace_upper_token(match):
        token = match.group(0)
        return _split_joined_upper_token(token)

    def _replace_alpha_token(match):
        token = match.group(0)
        return _split_joined_alpha_token(token)

    cleaned = re.sub(r'\b[A-ZÁÉÍÓÚÜÑ]{10,}\b', _replace_upper_token, cleaned)
    cleaned = re.sub(r'\b[A-Za-zÁÉÍÓÚÜÑáéíóúüñ]{8,}\b', _replace_alpha_token, cleaned)
    cleaned = re.sub(r'([\.!?])(?=[A-Za-zÁÉÍÓÚÜÑáéíóúüñ])', r'\1 ', cleaned)
    cleaned = re.sub(r'\s{2,}', ' ', cleaned)
    cleaned = re.sub(r' *\n *', '\n', cleaned)
    return cleaned.strip()


def _polish_spanish_text(value, multiline=True, max_len=None):
    text = str(value or '')
    if not text:
        return ''
    text = text.replace('\r\n', '\n').replace('\r', '\n')
    text = text.replace('’', "'").replace('`', "'")
    text = re.sub(r'[ \t]+', ' ', text)
    text = re.sub(r'\s+([,.;:!?])', r'\1', text)
    text = re.sub(r'([,.;:!?])(?=[^\s\n])', r'\1 ', text)
    text = re.sub(r'\(\s+', '(', text)
    text = re.sub(r'\s+\)', ')', text)
    text = re.sub(r'\s{2,}', ' ', text)
    text = re.sub(r' *\n *', '\n', text)
    if not multiline:
        text = text.replace('\n', ' ')
    lines = [ln.strip() for ln in text.split('\n')]
    acronym_words = {
        'ABP', 'RPE', 'GPS', 'TRX', 'MD', 'SSG', 'UEFA',
        '1V1', '2V2', '3V3', '4V4', '5V5', '6V6', '7V7', '8V8', '9V9', '11V11',
    }
    polished_lines = []
    for line in lines:
        if not line:
            continue
        bullet_prefix = ''
        bullet_match = re.match(r'^([\-•\*]\s+)', line)
        if bullet_match:
            bullet_prefix = bullet_match.group(1)
            line = line[len(bullet_prefix):].strip()
        if len(line) >= 10 and line.isupper():
            words = line.lower().split()
            fixed = []
            for idx, token in enumerate(words):
                upper_token = token.upper()
                if upper_token in acronym_words:
                    fixed.append(upper_token)
                    continue
                if idx == 0:
                    fixed.append(token.capitalize())
                else:
                    fixed.append(token)
            line = ' '.join(fixed)
        line = re.sub(r'([.!?])([A-Za-zÁÉÍÓÚÜÑáéíóúüñ])', r'\1 \2', line)
        line = line[:1].upper() + line[1:] if line else line
        polished_lines.append(f'{bullet_prefix}{line}'.strip())
    result = '\n'.join(polished_lines).strip()
    if max_len:
        result = result[:int(max_len)]
    return result


def _sanitize_task_text(value, multiline=True, max_len=None):
    return _polish_spanish_text(
        _repair_joined_words_text(value),
        multiline=multiline,
        max_len=max_len,
    )


_RICH_ALLOWED_TAGS = {
    'b',
    'strong',
    'i',
    'em',
    'u',
    's',
    'br',
    'p',
    'ul',
    'ol',
    'li',
    'div',
    'span',
    'sub',
    'sup',
}


class _RichTextSanitizer(HTMLParser):
    def __init__(self, allowed_tags):
        super().__init__()
        self.allowed = set(allowed_tags or [])
        self.out = []

    def handle_starttag(self, tag, attrs):
        name = str(tag or '').lower()
        if name not in self.allowed:
            return
        if name == 'br':
            self.out.append('<br>')
            return
        self.out.append(f'<{name}>')

    def handle_startendtag(self, tag, attrs):
        name = str(tag or '').lower()
        if name not in self.allowed:
            return
        if name == 'br':
            self.out.append('<br>')
            return
        self.out.append(f'<{name}></{name}>')

    def handle_endtag(self, tag):
        name = str(tag or '').lower()
        if name not in self.allowed or name == 'br':
            return
        self.out.append(f'</{name}>')

    def handle_data(self, data):
        if data is None:
            return
        self.out.append(html.escape(str(data)))


def _sanitize_task_rich_html(value, max_len=6000):
    raw = str(value or '').strip()
    if not raw:
        return ''
    sanitizer = _RichTextSanitizer(_RICH_ALLOWED_TAGS)
    try:
        sanitizer.feed(raw)
        sanitizer.close()
    except Exception:
        return html.escape(raw)[: int(max_len)]
    cleaned = ''.join(sanitizer.out).strip()
    if max_len:
        cleaned = cleaned[: int(max_len)]
    return cleaned


def _rich_html_from_plain_text(value, max_len=6000):
    raw = str(value or '').strip()
    if not raw:
        return ''
    if max_len:
        raw = raw[: int(max_len)]
    return html.escape(raw).replace('\n', '<br>')


def _text_has_quality_issues(value):
    text = str(value or '').strip()
    if not text:
        return False
    if re.search(r'\b[A-ZÁÉÍÓÚÜÑ]{12,}\b', text):
        return True
    if re.search(r'[,;:](\S)', text):
        return True
    if re.search(r'\s{2,}', text):
        return True
    if len(text) >= 8 and text.isupper():
        return True
    return False


MONTHS_ES = {
    'enero': 1,
    'febrero': 2,
    'marzo': 3,
    'abril': 4,
    'mayo': 5,
    'junio': 6,
    'julio': 7,
    'agosto': 8,
    'septiembre': 9,
    'setiembre': 9,
    'octubre': 10,
    'noviembre': 11,
    'diciembre': 12,
}


def _coerce_reference_date(raw_value):
    value = str(raw_value or '').strip()
    if not value:
        return None
    try:
        parsed = datetime.strptime(value, '%Y-%m-%d').date()
    except ValueError:
        return None
    if parsed.year < 2000 or parsed.year > 2100:
        return None
    return parsed


def _parse_2digit_year(year_value):
    year_int = int(year_value)
    if year_int >= 100:
        return year_int
    return 2000 + year_int if year_int <= 69 else 1900 + year_int


def _extract_dates_from_text(raw_text):
    text = str(raw_text or '')
    if not text:
        return []
    found = []

    for line in text.splitlines():
        if 'fecha' not in _normalize_folded_text(line):
            continue
        line_dates = _extract_dates_line_parser(line)
        if line_dates:
            return line_dates
    return _extract_dates_line_parser(text)


def _extract_dates_line_parser(segment):
    text = str(segment or '')
    dates_found = []
    for match in re.finditer(r'\b([0-3]?\d)[\/\-.]([01]?\d)[\/\-.](\d{2,4})\b', text):
        day_val = _parse_int(match.group(1))
        month_val = _parse_int(match.group(2))
        year_val = _parse_int(match.group(3))
        if not day_val or not month_val or year_val is None:
            continue
        year_val = _parse_2digit_year(year_val)
        try:
            parsed = date(year_val, month_val, day_val)
        except ValueError:
            continue
        if 2000 <= parsed.year <= 2100:
            dates_found.append(parsed)
    for match in re.finditer(r'\b(\d{4})[\/\-.]([01]?\d)[\/\-.]([0-3]?\d)\b', text):
        year_val = _parse_int(match.group(1))
        month_val = _parse_int(match.group(2))
        day_val = _parse_int(match.group(3))
        if not year_val or not month_val or not day_val:
            continue
        try:
            parsed = date(year_val, month_val, day_val)
        except ValueError:
            continue
        if 2000 <= parsed.year <= 2100:
            dates_found.append(parsed)
    month_regex = '|'.join(MONTHS_ES.keys())
    for match in re.finditer(
        rf'\b([0-3]?\d)\s*(?:de\s+)?({month_regex})\s*(?:de\s+)?(\d{{2,4}})\b',
        _normalize_folded_text(text),
        flags=re.IGNORECASE,
    ):
        day_val = _parse_int(match.group(1))
        month_val = MONTHS_ES.get(str(match.group(2) or '').lower())
        year_val = _parse_int(match.group(3))
        if not day_val or not month_val or year_val is None:
            continue
        year_val = _parse_2digit_year(year_val)
        try:
            parsed = date(year_val, month_val, day_val)
        except ValueError:
            continue
        if 2000 <= parsed.year <= 2100:
            dates_found.append(parsed)
    return dates_found


def _detect_reference_date_in_text(raw_text):
    candidates = _extract_dates_from_text(raw_text)
    if not candidates:
        return None
    return candidates[0]


def _task_upload_date(task):
    created = getattr(task, 'created_at', None)
    if created:
        try:
            return timezone.localtime(created).date()
        except Exception:
            try:
                return created.date()
            except Exception:
                pass
    session_obj = getattr(task, 'session', None)
    if session_obj and getattr(session_obj, 'session_date', None):
        return session_obj.session_date
    return timezone.localdate()


def _build_task_date_haystack(task, analysis_meta=None):
    sheet = analysis_meta.get('task_sheet') if isinstance(analysis_meta, dict) and isinstance(analysis_meta.get('task_sheet'), dict) else {}
    return '\n'.join(
        [
            str(getattr(task, 'title', '') or ''),
            str(getattr(task, 'objective', '') or ''),
            str(getattr(task, 'coaching_points', '') or ''),
            str(getattr(task, 'confrontation_rules', '') or ''),
            str((analysis_meta or {}).get('summary') or ''),
            str(sheet.get('description') or ''),
        ]
    ).strip()


def _extract_effective_reference_date(task, analysis_meta=None):
    upload_date = _task_upload_date(task)
    meta = analysis_meta if isinstance(analysis_meta, dict) else {}
    stored_date = _coerce_reference_date(meta.get('reference_date'))
    if stored_date and stored_date != upload_date:
        return stored_date
    detected = _detect_reference_date_in_text(_build_task_date_haystack(task, meta))
    if detected and detected != upload_date:
        return detected
    return None


def _ensure_task_reference_date(task):
    if not task:
        return False
    layout = task.tactical_layout if isinstance(task.tactical_layout, dict) else {}
    layout_copy = dict(layout)
    meta = layout_copy.get('meta') if isinstance(layout_copy.get('meta'), dict) else {}
    meta = dict(meta)
    analysis_meta = meta.get('analysis') if isinstance(meta.get('analysis'), dict) else {}
    analysis_meta = dict(analysis_meta)
    upload_date = _task_upload_date(task)
    detected = _extract_effective_reference_date(task, analysis_meta=analysis_meta)
    current = _coerce_reference_date(analysis_meta.get('reference_date'))
    changed = False
    if detected and detected != upload_date:
        if current != detected:
            analysis_meta['reference_date'] = detected.isoformat()
            analysis_meta['reference_date_source'] = 'content'
            changed = True
    elif analysis_meta.get('reference_date'):
        analysis_meta.pop('reference_date', None)
        analysis_meta.pop('reference_date_source', None)
        changed = True
    if changed:
        meta['analysis'] = analysis_meta
        layout_copy['meta'] = meta
        task.tactical_layout = layout_copy
        task.save(update_fields=['tactical_layout'])
    return changed


def _detect_keyword_tags(text, taxonomy):
    haystack = _normalize_folded_text(text)
    tags = []
    for tag, keys in (taxonomy or {}).items():
        for key in keys:
            if _normalize_folded_text(key) in haystack:
                tags.append(tag)
                break
    return tags


def _detect_materials_in_text(text):
    haystack = _normalize_folded_text(text)
    detected = []
    for item in TASK_MATERIAL_LIBRARY:
        label = _normalize_folded_text(item.get('label'))
        title = _normalize_folded_text(item.get('title'))
        kind = _normalize_folded_text(item.get('kind'))
        if any(token and token in haystack for token in {label, title, kind}):
            detected.append(
                {
                    'label': item.get('label') or '',
                    'title': item.get('title') or '',
                    'kind': item.get('kind') or '',
                    'category': item.get('category') or '',
                }
            )
    return detected


def _analysis_quality_score(analysis):
    if not isinstance(analysis, dict):
        return 0
    score = 0
    if analysis.get('title'):
        score += 20
    if analysis.get('objective'):
        score += 20
    if analysis.get('coaching_points'):
        score += 20
    if analysis.get('confrontation_rules'):
        score += 20
    if analysis.get('work_contexts'):
        score += 10
    if analysis.get('objective_tags'):
        score += 5
    if analysis.get('detected_materials'):
        score += 5
    if analysis.get('exercise_types'):
        score += 5
    if analysis.get('phase_tags'):
        score += 5
    task_sheet = analysis.get('task_sheet') if isinstance(analysis.get('task_sheet'), dict) else {}
    if task_sheet.get('description'):
        score += 5
    if task_sheet.get('players') or task_sheet.get('dimensions') or task_sheet.get('materials'):
        score += 5
    return min(100, score)


def _analysis_confidence_scores(analysis):
    data = analysis if isinstance(analysis, dict) else {}
    sheet = data.get('task_sheet') if isinstance(data.get('task_sheet'), dict) else {}
    title = str(data.get('title') or '').strip()
    objective = str(data.get('objective') or '').strip()
    coaching = str(data.get('coaching_points') or '').strip()
    rules = str(data.get('confrontation_rules') or '').strip()
    summary = str(data.get('summary') or '').strip()
    players = str(sheet.get('players') or '').strip()
    dimensions = str(sheet.get('dimensions') or '').strip()
    materials = str(sheet.get('materials') or '').strip()
    exercise_types = data.get('exercise_types') if isinstance(data.get('exercise_types'), list) else []
    phase_tags = data.get('phase_tags') if isinstance(data.get('phase_tags'), list) else []

    def _value_conf(text_value, max_len=220):
        text_value = str(text_value or '').strip()
        if not text_value:
            return 0
        score = 38
        score += min(len(text_value), max_len) / max_len * 42
        if re.search(r'\b[A-ZÁÉÍÓÚÜÑ]{12,}\b', text_value):
            score -= 32
        if re.search(r'[,;:]\S', text_value):
            score -= 8
        return max(0, min(100, int(round(score))))

    scores = {
        'title': _value_conf(title, 120),
        'objective': _value_conf(objective, 180),
        'coaching': _value_conf(coaching, 280),
        'rules': _value_conf(rules, 280),
        'summary': _value_conf(summary, 320),
        'players': _value_conf(players, 120),
        'dimensions': _value_conf(dimensions, 100),
        'materials': _value_conf(materials, 180),
    }
    semantic_bonus = 0
    if exercise_types:
        semantic_bonus += 8
    if phase_tags:
        semantic_bonus += 8
    base_values = [scores['title'], scores['objective'], scores['coaching'], scores['summary']]
    overall = int(round(sum(base_values) / max(1, len(base_values)) + semantic_bonus))
    scores['overall'] = max(0, min(100, overall))
    return scores


def _analysis_needs_review(analysis):
    if not isinstance(analysis, dict):
        return True
    conf = _analysis_confidence_scores(analysis)
    if int(conf.get('overall') or 0) < 62:
        return True
    if int(conf.get('title') or 0) < 55:
        return True
    if int(conf.get('objective') or 0) < 45:
        return True
    if re.search(r'\b[A-ZÁÉÍÓÚÜÑ]{12,}\b', str(analysis.get('summary') or '')):
        return True
    return False


def _apply_analysis_to_task(task, analysis):
    layout = task.tactical_layout if isinstance(task.tactical_layout, dict) else {}
    meta = layout.get('meta') if isinstance(layout.get('meta'), dict) else {}
    confidence = _analysis_confidence_scores(analysis)
    summary_clean = _sanitize_task_text((analysis.get('summary') or '')[:900], multiline=True, max_len=900)
    task_sheet_raw = analysis.get('task_sheet') if isinstance(analysis.get('task_sheet'), dict) else {}
    task_sheet_clean = {
        'description': _sanitize_task_text(task_sheet_raw.get('description') or '', multiline=True, max_len=1200),
        'players': _sanitize_task_text(task_sheet_raw.get('players') or '', multiline=False, max_len=120),
        'space': _sanitize_task_text(task_sheet_raw.get('space') or '', multiline=False, max_len=120),
        'dimensions': _sanitize_task_text(task_sheet_raw.get('dimensions') or '', multiline=False, max_len=120),
        'materials': _sanitize_task_text(task_sheet_raw.get('materials') or '', multiline=False, max_len=300),
    }
    upload_date = _task_upload_date(task)
    reference_date = _coerce_reference_date(analysis.get('reference_date'))
    if not reference_date:
        text_for_date = '\n'.join(
            [
                str(analysis.get('title') or ''),
                str(analysis.get('objective') or ''),
                str(analysis.get('coaching_points') or ''),
                str(analysis.get('confrontation_rules') or ''),
                str(analysis.get('summary') or ''),
                str(task_sheet_clean.get('description') or ''),
            ]
        )
        reference_date = _detect_reference_date_in_text(text_for_date)
    reference_date_iso = ''
    if reference_date and reference_date != upload_date:
        reference_date_iso = reference_date.isoformat()

    meta['analysis'] = {
        'work_contexts': analysis.get('work_contexts') or [],
        'objective_tags': analysis.get('objective_tags') or [],
        'detected_materials': analysis.get('detected_materials') or [],
        'exercise_types': analysis.get('exercise_types') or [],
        'phase_tags': analysis.get('phase_tags') or [],
        'players_count_estimate': analysis.get('players_count_estimate') or None,
        'players_band': analysis.get('players_band') or '',
        'duration_band': analysis.get('duration_band') or '',
        'quality_score': analysis.get('quality_score') or 0,
        'confidence': confidence,
        'needs_review': _analysis_needs_review(analysis),
        'pdf_template': analysis.get('pdf_template') or 'generic',
        'summary': summary_clean,
        'task_sheet': task_sheet_clean,
        'reference_date': reference_date_iso,
        'reference_date_source': 'content' if reference_date_iso else '',
        'analyzed_at': timezone.now().isoformat(),
        'parser_version': TASK_PDF_PARSE_VERSION,
    }
    layout['meta'] = meta
    task.tactical_layout = layout
    task.save(update_fields=['tactical_layout'])


def _extract_section_block(text, section_aliases):
    lines = [ln.strip() for ln in str(text or '').splitlines()]
    if not lines:
        return ''
    aliases = [a.lower() for a in section_aliases]
    start = None
    for idx, line in enumerate(lines):
        low = line.lower()
        if any(re.search(rf'^{re.escape(alias)}\s*[:\-]?\s*$', low) for alias in aliases):
            start = idx + 1
            break
        if any(low.startswith(f'{alias}:') or low.startswith(f'{alias} -') for alias in aliases):
            inline = re.split(r'[:\-]', line, maxsplit=1)
            if len(inline) == 2 and inline[1].strip():
                return inline[1].strip()
    if start is None:
        return ''

    end = len(lines)
    for idx in range(start, len(lines)):
        low = lines[idx].lower()
        if re.match(r'^[a-záéíóúüñ ]{3,30}\s*[:\-]?\s*$', low):
            if idx > start:
                end = idx
                break
    return '\n'.join([ln for ln in lines[start:end] if ln]).strip()


def _pick_bullets_or_sentences(raw_text, limit=5):
    lines = [ln.strip() for ln in str(raw_text or '').splitlines() if ln.strip()]
    items = []
    for line in lines:
        clean = re.sub(r'^[\-\•\*\d\)\.\s]+', '', line).strip()
        if not clean:
            continue
        if len(clean) < 4:
            continue
        items.append(clean)
        if len(items) >= limit:
            break
    return '\n'.join(items)


def _extract_inline_field(text, aliases, max_len=240):
    raw = str(text or '')
    if not raw:
        return ''
    alias_pattern = '|'.join(re.escape(alias) for alias in aliases)
    patterns = [
        rf'(?im)^\s*(?:{alias_pattern})\s*[:\-]\s*(.+?)\s*$',
        rf'(?i)\b(?:{alias_pattern})\s*[:\-]\s*([^\n\r]+)',
    ]
    for pattern in patterns:
        match = re.search(pattern, raw)
        if match:
            value = ' '.join(str(match.group(1) or '').split()).strip(' ,.;:-')
            if value:
                return value[:max_len]
    return ''


def _extract_first_dimension_text(text):
    raw = str(text or '')
    if not raw:
        return ''
    match = re.search(
        r'(?i)\b(\d{1,3}(?:[.,]\d{1,2})?\s*(?:m|metros?)?)\s*[xX×]\s*(\d{1,3}(?:[.,]\d{1,2})?\s*(?:m|metros?)?)\b',
        raw,
    )
    if not match:
        return ''
    left = str(match.group(1) or '').strip()
    right = str(match.group(2) or '').strip()
    return f'{left} x {right}'


def _extract_players_text(text):
    raw = str(text or '')
    if not raw:
        return ''
    explicit = _extract_inline_field(
        raw,
        ['jugadores', 'participantes', 'n jugadores', 'num jugadores', 'nº jugadores', 'numero jugadores'],
        max_len=120,
    )
    if explicit:
        return explicit
    pattern = re.search(r'(?i)\b(\d{1,2}\s*v\s*\d{1,2}(?:\s*\+\s*\d{1,2})?)\b', raw)
    if pattern:
        return str(pattern.group(1)).replace(' ', '')
    total = re.search(r'(?i)\b(\d{1,2})\s*(?:jugadores?|participantes?)\b', raw)
    if total:
        return f'{total.group(1)} jugadores'
    return ''


def _estimate_players_count(players_text, fallback_text=''):
    source = ' '.join([str(players_text or ''), str(fallback_text or '')]).strip()
    if not source:
        return None
    cleaned = source.replace('vs', 'v').replace('x', 'v')
    total = 0
    matched = False
    for group in re.findall(r'(\d{1,2}\s*v\s*\d{1,2}(?:\s*\+\s*\d{1,2})?)', cleaned, flags=re.IGNORECASE):
        nums = [_parse_int(n) for n in re.findall(r'\d{1,2}', group)]
        nums = [n for n in nums if n is not None]
        if nums:
            total += sum(nums)
            matched = True
    if matched and total > 0:
        return total
    explicit_total = re.search(r'(?i)\b(\d{1,2})\s*(?:jugadores?|participantes?)\b', cleaned)
    if explicit_total:
        parsed = _parse_int(explicit_total.group(1))
        if parsed:
            return parsed
    nums = [_parse_int(n) for n in re.findall(r'\b\d{1,2}\b', cleaned)]
    nums = [n for n in nums if n is not None]
    plausible = [n for n in nums if 4 <= n <= 30]
    if plausible:
        return max(plausible)
    return None


def _players_band_label(players_count):
    value = _parse_int(players_count)
    if not value:
        return ''
    if value <= 6:
        return 'micro (1-6)'
    if value <= 12:
        return 'grupo (7-12)'
    if value <= 18:
        return 'equipo (13-18)'
    return 'ampliado (19+)'


def _duration_band_label(minutes):
    value = _parse_int(minutes)
    if not value:
        return ''
    if value <= 12:
        return 'corta (5-12m)'
    if value <= 20:
        return 'media (13-20m)'
    if value <= 35:
        return 'larga (21-35m)'
    return 'extendida (36m+)'


def _detect_pdf_template(text):
    raw = _normalize_folded_text(text)
    if not raw:
        return 'generic'
    markers = {
        'sesion_sheet_compact': [
            'dosificacion',
            'tiempototaldetrabajo',
            'parteprincipal',
            'materialdeentrenamiento',
        ],
        'session_es_verbose': [
            'objetivo',
            'consignas',
            'reglas',
            'material',
            'jugadores',
        ],
    }
    best_key = 'generic'
    best_score = 0
    for key, keys in markers.items():
        score = sum(1 for token in keys if token in raw)
        if score > best_score:
            best_score = score
            best_key = key
    return best_key


def _extract_task_sheet_from_pdf(text, detected_materials=None, template_key='generic'):
    description_aliases = ['descripcion', 'descripción', 'desarrollo', 'organizacion', 'organización', 'estructura', 'dinamica', 'dinámica']
    if template_key == 'sesion_sheet_compact':
        description_aliases = ['consignas', 'descripcion', 'descripción', 'desarrollo', 'organizacion', 'organización']
    description = _extract_section_block(
        text,
        description_aliases,
    )
    if not description:
        description = _pick_bullets_or_sentences(text, limit=6)

    players = _extract_players_text(text)
    space = _extract_inline_field(text, ['espacio', 'superficie', 'zona', 'campo', 'area', 'área'], max_len=120)
    dimensions = _extract_inline_field(text, ['medidas', 'dimensiones', 'tamano', 'tamaño'], max_len=120)
    if not dimensions:
        dimensions = _extract_first_dimension_text(text)

    material_section = _extract_section_block(text, ['material', 'materiales', 'elementos', 'recursos'])
    material_detected = []
    for item in detected_materials or []:
        if not isinstance(item, dict):
            continue
        label = str(item.get('title') or item.get('label') or item.get('kind') or '').strip()
        if label:
            material_detected.append(label)
    if material_section:
        materials = material_section
    else:
        materials = ', '.join(sorted(set(material_detected)))[:300]

    return {
        'description': _polish_spanish_text(_repair_joined_words_text((description or '')[:1400]), multiline=True, max_len=1200),
        'players': _polish_spanish_text(_repair_joined_words_text(players), multiline=False, max_len=120),
        'space': _polish_spanish_text(_repair_joined_words_text(space), multiline=False, max_len=120),
        'dimensions': _polish_spanish_text(_repair_joined_words_text(dimensions), multiline=False, max_len=120),
        'materials': _polish_spanish_text(_repair_joined_words_text(materials), multiline=False, max_len=300),
    }


def _suggest_task_from_pdf(pdf_text):
    text = _polish_spanish_text(_repair_joined_words_text(pdf_text or ''), multiline=True)
    text = str(text or '').strip()
    template_key = _detect_pdf_template(text)
    if not text:
        return {
            'title': '',
            'objective': '',
            'minutes': 15,
            'coaching_points': '',
            'confrontation_rules': '',
            'summary': '',
            'pdf_template': template_key,
            'reference_date': '',
        }

    lines = [line.strip() for line in text.splitlines() if line.strip()]
    title = ''
    skip_title_tokens = [
        'parteprincipal',
        'materialdeentrenamiento',
        'tiempototaldesesion',
        'tiempo total de sesion',
        'fecha:',
        'hora:',
        'micro-ciclo',
        'meso-ciclo',
    ]
    if template_key == 'sesion_sheet_compact':
        skip_title_tokens.extend(['tiempototaldetrabajo', 'dosificacion', 'hidratacion'])
    for line in lines[:24]:
        if len(line) > 120 or re.search(r'^\d+$', line):
            continue
        folded = _normalize_folded_text(line)
        if not folded or folded in {',', '.', ';'}:
            continue
        if any(token in folded for token in skip_title_tokens):
            continue
        if (
            folded.startswith('2bloques')
            or folded.startswith('3bloques')
            or folded.startswith('tiempototal')
            or re.match(r'^tiempo\s*total', folded)
        ):
            continue
        title = line
        break

    objective_aliases = ['objetivo', 'objective', 'finalidad', 'meta']
    if template_key == 'sesion_sheet_compact':
        objective_aliases.extend(['consignas', 'descripcion', 'desarrollo'])
    objective = _extract_section_block(text, objective_aliases)
    if not objective:
        objective_match = re.search(r'(objetivo|objective)\s*[:\-]\s*(.+)', text, re.IGNORECASE)
        if objective_match:
            objective = objective_match.group(2).strip()
        elif len(lines) > 1:
            for candidate in lines[1:10]:
                folded = _normalize_folded_text(candidate)
                if len(candidate.strip(' ,.;:-')) < 4:
                    continue
                if re.match(r'^(dosificacion|tiempo\s*total|parteprincipal)\b', folded):
                    continue
                if re.match(r'^\d+\s*bloques', folded):
                    continue
                objective = candidate[:180]
                break

    minutes = 15
    dosage_line = re.search(r'dosificaci[oó]n\s*[:\-]?\s*([^\n]{0,120})', text, re.IGNORECASE)
    if dosage_line:
        nums = [_parse_int(num) for num in re.findall(r'\d{1,3}', dosage_line.group(1))]
        nums = [n for n in nums if n is not None]
        plausible = [n for n in nums if 5 <= n <= 90]
        if plausible:
            minutes = max(plausible)
    if minutes == 15:
        minute_match = re.search(
            r'tiempo\s*total\s*[:\-]?\s*(\d{1,3})\s*(?:min|mins|minutos|minutes|[\'’`´])?',
            text,
            re.IGNORECASE,
        )
        if not minute_match:
            minute_match = re.search(r'(\d{1,3})\s*(?:min|mins|minutos|minutes|[\'’`´])', text, re.IGNORECASE)
        if minute_match:
            parsed = _parse_int(minute_match.group(1)) or 15
            minutes = max(5, min(parsed, 90))

    coaching_points = _extract_section_block(
        text,
        ['consignas', 'coaching points', 'puntos clave', 'indicaciones', 'correcciones'],
    )
    confrontation_rules = _extract_section_block(
        text,
        ['reglas', 'reglas de confrontacion', 'puntuacion', 'normas', 'restricciones'],
    )
    if not coaching_points or not confrontation_rules:
        bullet_like = []
        for line in lines:
            if line.startswith(('-', '•', '*')) or re.match(r'^\d+[\).\s-]', line):
                bullet_like.append(re.sub(r'^[\-\•\*\d\)\.\s]+', '', line).strip())
            if len(bullet_like) >= 10:
                break
        if not coaching_points:
            coaching_points = '\n'.join(bullet_like[:5])
        if not confrontation_rules:
            confrontation_rules = '\n'.join(bullet_like[5:10])
    if not coaching_points:
        coaching_points = _pick_bullets_or_sentences(text, limit=4)
    if not confrontation_rules:
        confrontation_rules = ''

    summary = '\n'.join(lines[:12])[:900]
    work_contexts = _detect_keyword_tags(text, TASK_CONTEXT_KEYWORDS)
    objective_tags = _detect_keyword_tags(' '.join([objective, coaching_points, confrontation_rules]), TASK_OBJECTIVE_KEYWORDS)
    exercise_types = _detect_keyword_tags(text, TASK_TYPE_KEYWORDS)
    phase_tags = _detect_keyword_tags(' '.join([text, objective, coaching_points]), TASK_PHASE_KEYWORDS)
    detected_materials = _detect_materials_in_text(text)
    task_sheet = _extract_task_sheet_from_pdf(text, detected_materials=detected_materials, template_key=template_key)
    players_count_estimate = _estimate_players_count(task_sheet.get('players'), text)
    players_band = _players_band_label(players_count_estimate)
    duration_band = _duration_band_label(minutes)
    reference_date = _detect_reference_date_in_text(text)

    analysis = {
        'title': _polish_spanish_text(_repair_joined_words_text((title or 'Tarea desde PDF')[:220]), multiline=False, max_len=160),
        'objective': _polish_spanish_text(_repair_joined_words_text(objective[:240]), multiline=False, max_len=180),
        'minutes': minutes,
        'coaching_points': _polish_spanish_text(_repair_joined_words_text(coaching_points), multiline=True),
        'confrontation_rules': _polish_spanish_text(_repair_joined_words_text(confrontation_rules), multiline=True),
        'summary': _polish_spanish_text(_repair_joined_words_text(summary), multiline=True),
        'work_contexts': work_contexts,
        'objective_tags': objective_tags,
        'exercise_types': exercise_types,
        'phase_tags': phase_tags,
        'players_count_estimate': players_count_estimate,
        'players_band': players_band,
        'duration_band': duration_band,
        'detected_materials': detected_materials,
        'task_sheet': task_sheet,
        'pdf_template': template_key,
        'reference_date': reference_date.isoformat() if reference_date else '',
    }
    analysis['quality_score'] = _analysis_quality_score(analysis)
    return analysis


def _split_pdf_into_task_sections(pdf_text):
    text = str(pdf_text or '').strip()
    if not text:
        return []
    lines = [line.strip() for line in text.splitlines() if line.strip()]
    if len(lines) < 6:
        return [text]

    start_patterns = [
        re.compile(r'^(?:tarea|ejercicio|task|drill)\s*(?:n[\.\s]*[oº°])?\s*\d{1,2}\b', re.IGNORECASE),
        re.compile(r'^\d{1,2}\s*[\)\.\-]\s*(?:tarea|ejercicio|task|drill)\b', re.IGNORECASE),
        re.compile(r'^(?:task|drill)\s*\d{1,2}\b', re.IGNORECASE),
        re.compile(r'^(?:primera|segunda|tercera|cuarta)\s+tarea\b', re.IGNORECASE),
    ]

    boundaries = []
    for idx, line in enumerate(lines):
        if any(pattern.search(line) for pattern in start_patterns):
            boundaries.append(idx)

    if len(boundaries) < 2:
        objective_indices = []
        for idx, line in enumerate(lines):
            if re.match(r'^(?:objetivo|objective)\s*[:\-]', line, re.IGNORECASE):
                objective_indices.append(max(0, idx - 1))
        boundaries = sorted(set(objective_indices))

    if len(boundaries) < 2:
        return [text]

    sections = []
    seen = set()
    stop_keywords = (
        'estiramientos',
        'partefinal',
        'parte final',
        'tiempototaldesesion',
        'tiempo total de sesion',
        'observaciones',
        'ausentes',
        'lesionados',
    )
    for i, start in enumerate(boundaries):
        end = boundaries[i + 1] if i + 1 < len(boundaries) else len(lines)
        if end - start < 3:
            continue
        chunk = '\n'.join(lines[start:end]).strip()
        folded = _normalize_folded_text(chunk)
        has_load_or_task_data = (
            'dosificacion' in folded
            or 'tiempototaldetrabajo' in folded
            or 'consignas' in folded
            or 'tarea' in folded
        )
        if any(word in folded for word in stop_keywords) and not has_load_or_task_data:
            continue
        normalized = _normalize_folded_text(chunk)
        if len(chunk) < 80 or normalized in seen:
            continue
        seen.add(normalized)
        sections.append(chunk)

    # Heuristic for session sheets where "Tarea 1" is implicit and first explicit marker is Tarea 2.
    if sections:
        first_folded = _normalize_folded_text(sections[0])
        looks_like_second = any(token in first_folded for token in ('tarea2', 'tarea 2', 'segunda tarea'))
        has_first_marker = any(any(token in _normalize_folded_text(sec) for token in ('tarea1', 'tarea 1', 'primera tarea')) for sec in sections)
        if looks_like_second and not has_first_marker:
            prefix_lines = lines[:boundaries[0]]
            prefix_text = '\n'.join(prefix_lines)
            prefix_folded = _normalize_folded_text(prefix_text)
            if 'parteprincipal' in prefix_folded:
                idx = prefix_folded.rfind('parteprincipal')
                candidate = prefix_text[idx:].strip()
            else:
                candidate = '\n'.join(prefix_lines[-28:]).strip()
            candidate_folded = _normalize_folded_text(candidate)
            if (
                len(candidate) >= 140
                and ('dosificacion' in candidate_folded or 'tiempototaldetrabajo' in candidate_folded)
                and 'estiramientos' not in candidate_folded
            ):
                sections.insert(0, candidate)

    cleaned_sections = []
    for chunk in sections:
        folded = _normalize_folded_text(chunk)
        if any(word in folded for word in stop_keywords):
            if not ('dosificacion' in folded or 'tiempototaldetrabajo' in folded):
                continue
        cleaned_sections.append(chunk)
    sections = cleaned_sections

    if len(sections) < 2:
        return [text]
    return sections[:8]


def _extract_tasks_from_pdf_text(pdf_text, fallback_title='Tarea desde PDF'):
    sections = _split_pdf_into_task_sections(pdf_text)
    analyses = []
    for idx, section in enumerate(sections, start=1):
        analysis = _suggest_task_from_pdf(section)
        title = (analysis.get('title') or '').strip()
        if not title or title.lower() == 'tarea desde pdf':
            if len(sections) > 1:
                title = f'{fallback_title} · Tarea {idx}'
            else:
                title = fallback_title
            analysis['title'] = title[:160]
        analyses.append(
            {
                'analysis': analysis,
                'raw_text': section[:2500],
                'segment_index': idx,
                'segment_total': len(sections),
            }
        )
    return analyses


def _parse_bulk_tasks_text(raw_text, default_block, default_minutes):
    """
    Parse bulk task rows from a plain-text block.
    Supported row formats (separator: |, ;, or TAB):
      titulo
      titulo|minutos
      titulo|minutos|bloque
      titulo|minutos|bloque|objetivo|consignas|reglas
    """
    lines = [ln.strip() for ln in str(raw_text or '').splitlines()]
    rows = [ln for ln in lines if ln and not ln.startswith('#')]
    parsed = []
    errors = []
    allowed_blocks = {choice[0] for choice in SessionTask.BLOCK_CHOICES}

    for idx, row in enumerate(rows, start=1):
        separator = '|'
        if '|' in row:
            separator = '|'
        elif ';' in row:
            separator = ';'
        elif '\t' in row:
            separator = '\t'
        parts = [part.strip() for part in row.split(separator)]
        parts = [part for part in parts if part != '']
        if not parts:
            continue

        title = _sanitize_task_text((parts[0] or '').strip(), multiline=False, max_len=160)
        if not title:
            errors.append(f'Línea {idx}: título vacío.')
            continue

        minutes = default_minutes
        block = default_block
        objective = ''
        coaching_points = ''
        confrontation_rules = ''

        if len(parts) >= 2:
            maybe_minutes = _parse_int(parts[1])
            if maybe_minutes is None:
                errors.append(f'Línea {idx}: minutos inválidos ({parts[1]}).')
                continue
            minutes = max(5, min(maybe_minutes, 90))
        if len(parts) >= 3:
            candidate_block = str(parts[2]).strip()
            if candidate_block in allowed_blocks:
                block = candidate_block
            else:
                errors.append(f'Línea {idx}: bloque inválido ({candidate_block}).')
                continue
        if len(parts) >= 4:
            objective = _sanitize_task_text(parts[3] or '', multiline=False, max_len=180)
        if len(parts) >= 5:
            coaching_points = _sanitize_task_text(parts[4] or '', multiline=True)
        if len(parts) >= 6:
            confrontation_rules = _sanitize_task_text(parts[5] or '', multiline=True)

        parsed.append(
            {
                'title': title[:160],
                'minutes': minutes,
                'block': block,
                'objective': objective,
                'coaching_points': coaching_points,
                'confrontation_rules': confrontation_rules,
                'source_line': idx,
            }
        )
    return parsed, errors


def _persist_detected_resources_library(task_items, scope_key='coach'):
    try:
        counters = Counter()
        meta_by_key = {}
        for task in task_items or []:
            layout = task.tactical_layout if isinstance(task.tactical_layout, dict) else {}
            meta = layout.get('meta') if isinstance(layout.get('meta'), dict) else {}
            analysis_meta = meta.get('analysis') if isinstance(meta.get('analysis'), dict) else {}
            for mat in analysis_meta.get('detected_materials') or []:
                if not isinstance(mat, dict):
                    continue
                label = str(mat.get('label') or '').strip()
                title = str(mat.get('title') or '').strip()
                kind = str(mat.get('kind') or '').strip()
                category = str(mat.get('category') or '').strip()
                key = (kind or label or title or '').lower()
                if not key:
                    continue
                counters[key] += 1
                current = meta_by_key.get(key, {})
                if not current:
                    meta_by_key[key] = {
                        'label': label,
                        'title': title,
                        'kind': kind,
                        'category': category,
                    }
                else:
                    if title and len(title) > len(current.get('title') or ''):
                        current['title'] = title
                    if label and not current.get('label'):
                        current['label'] = label
                    if kind and not current.get('kind'):
                        current['kind'] = kind
                    if category and not current.get('category'):
                        current['category'] = category
                    meta_by_key[key] = current

        existing = {}
        if TASK_RESOURCE_LIBRARY_PATH.exists():
            try:
                existing = json.loads(TASK_RESOURCE_LIBRARY_PATH.read_text(encoding='utf-8'))
            except Exception:
                existing = {}
        if not isinstance(existing, dict):
            existing = {}
        resources_by_scope = existing.get('resources_by_scope')
        if not isinstance(resources_by_scope, dict):
            resources_by_scope = {}

        scope_items = []
        for key, count in counters.most_common():
            item_meta = meta_by_key.get(key, {})
            scope_items.append(
                {
                    'key': key,
                    'label': item_meta.get('label') or '',
                    'title': item_meta.get('title') or item_meta.get('label') or key,
                    'kind': item_meta.get('kind') or '',
                    'category': item_meta.get('category') or '',
                    'count': int(count),
                }
            )
        resources_by_scope[str(scope_key)] = scope_items
        payload = {
            'generated_at': timezone.now().isoformat(),
            'resources_by_scope': resources_by_scope,
        }
        TASK_RESOURCE_LIBRARY_PATH.parent.mkdir(parents=True, exist_ok=True)
        TASK_RESOURCE_LIBRARY_PATH.write_text(
            json.dumps(payload, ensure_ascii=False, indent=2),
            encoding='utf-8',
        )
    except Exception:
        # Never break the sessions workspace if library persistence fails.
        return


def _task_scope_for_item(task):
    layout = task.tactical_layout if isinstance(task.tactical_layout, dict) else {}
    meta = layout.get('meta') if isinstance(layout.get('meta'), dict) else {}
    scope = str(meta.get('scope') or '').strip()
    return scope or 'coach'


def _get_or_create_library_session(team, scope_key):
    today = timezone.localdate()
    week_start = today - timedelta(days=today.weekday())
    week_end = week_start + timedelta(days=6)
    scope_label = {
        'coach': 'Entrenador',
        'goalkeeper': 'Porteros',
        'fitness': 'Preparacion fisica',
    }.get(scope_key, 'Staff')
    microcycle, _ = TrainingMicrocycle.objects.get_or_create(
        team=team,
        week_start=week_start,
        defaults={
            'week_end': week_end,
            'title': f'Biblioteca {scope_label}',
            'objective': 'Repositorio de tareas en PDF',
            'status': TrainingMicrocycle.STATUS_DRAFT,
            'notes': 'Microciclo tecnico generado automaticamente para biblioteca.',
        },
    )
    session, _ = TrainingSession.objects.get_or_create(
        microcycle=microcycle,
        session_date=today,
        focus=f'Biblioteca PDF · {scope_label}',
        defaults={
            'duration_minutes': 90,
            'intensity': TrainingSession.INTENSITY_LOW,
            'content': 'Sesion tecnica para almacenar tareas subidas a biblioteca.',
            'order': 0,
        },
    )
    return session


def _cleanup_task_joined_text_fields(task):
    if not task:
        return False
    changed = False

    def _clean_attr(attr_name, max_len=None):
        nonlocal changed
        current = str(getattr(task, attr_name, '') or '')
        cleaned = _polish_spanish_text(_repair_joined_words_text(current), multiline=(attr_name in {'coaching_points', 'confrontation_rules'}))
        if max_len:
            cleaned = cleaned[:max_len]
        if cleaned != current:
            setattr(task, attr_name, cleaned)
            changed = True

    _clean_attr('title', 160)
    _clean_attr('objective', 180)
    _clean_attr('coaching_points')
    _clean_attr('confrontation_rules')

    layout = task.tactical_layout if isinstance(task.tactical_layout, dict) else {}
    if layout:
        layout_copy = dict(layout)
        meta = layout_copy.get('meta') if isinstance(layout_copy.get('meta'), dict) else {}
        meta = dict(meta)
        analysis_meta = meta.get('analysis') if isinstance(meta.get('analysis'), dict) else {}
        analysis_meta = dict(analysis_meta)
        summary_raw = str(analysis_meta.get('summary') or '')
        summary_clean = _polish_spanish_text(_repair_joined_words_text(summary_raw), multiline=True, max_len=900)
        if summary_clean != summary_raw:
            analysis_meta['summary'] = summary_clean
            changed = True
        task_sheet = analysis_meta.get('task_sheet') if isinstance(analysis_meta.get('task_sheet'), dict) else {}
        if task_sheet:
            task_sheet_copy = dict(task_sheet)
            for key in ('description', 'players', 'space', 'dimensions', 'materials'):
                raw_val = str(task_sheet_copy.get(key) or '')
                clean_val = _polish_spanish_text(
                    _repair_joined_words_text(raw_val),
                    multiline=(key == 'description'),
                )
                if clean_val != raw_val:
                    task_sheet_copy[key] = clean_val
                    changed = True
            analysis_meta['task_sheet'] = task_sheet_copy
        meta['analysis'] = analysis_meta
        layout_copy['meta'] = meta
        if layout_copy != layout:
            task.tactical_layout = layout_copy
            changed = True

    if changed:
        task.save(
            update_fields=[
                'title',
                'objective',
                'coaching_points',
                'confrontation_rules',
                'tactical_layout',
            ]
        )
    return changed


def _task_analysis_needs_refresh(task):
    if not task or not getattr(task, 'task_pdf', None):
        return False
    layout = task.tactical_layout if isinstance(task.tactical_layout, dict) else {}
    meta = layout.get('meta') if isinstance(layout.get('meta'), dict) else {}
    analysis_meta = meta.get('analysis') if isinstance(meta.get('analysis'), dict) else {}
    parser_version = _parse_int(analysis_meta.get('parser_version')) or 0
    if parser_version < TASK_PDF_PARSE_VERSION:
        return True
    summary = str(analysis_meta.get('summary') or '')
    if _text_has_quality_issues(summary):
        return True
    if _text_has_quality_issues(str(task.title or '')):
        return True
    if _text_has_quality_issues(str(task.objective or '')):
        return True
    return False


def _refresh_task_from_pdf_analysis(task):
    if not task or not getattr(task, 'task_pdf', None):
        return False
    try:
        extracted_text = _extract_pdf_text(task.task_pdf, max_chars=60000)
        parsed_tasks = _extract_tasks_from_pdf_text(extracted_text, fallback_title=task.title or 'Tarea desde PDF')
        selected = None
        if parsed_tasks:
            meta = task.tactical_layout.get('meta') if isinstance(task.tactical_layout, dict) else {}
            segment_index = _parse_int(meta.get('pdf_segment_index')) or 1
            segment_index = max(1, min(segment_index, len(parsed_tasks)))
            selected = parsed_tasks[segment_index - 1]
        if not selected:
            selected = {'analysis': _suggest_task_from_pdf(extracted_text), 'raw_text': extracted_text[:2500]}
        analysis = selected.get('analysis') or {}
        task.title = _sanitize_task_text(
            str(analysis.get('title') or task.title or 'Tarea desde PDF'),
            multiline=False,
            max_len=160,
        )
        task.duration_minutes = max(5, min((_parse_int(analysis.get('minutes')) or task.duration_minutes or 15), 90))
        task.objective = _sanitize_task_text(
            str(analysis.get('objective') or task.objective or ''),
            multiline=False,
            max_len=180,
        )
        task.coaching_points = _sanitize_task_text(
            str(analysis.get('coaching_points') or task.coaching_points or ''),
            multiline=True,
        )
        task.confrontation_rules = _sanitize_task_text(
            str(analysis.get('confrontation_rules') or task.confrontation_rules or ''),
            multiline=True,
        )
        layout = task.tactical_layout if isinstance(task.tactical_layout, dict) else {}
        layout = dict(layout)
        meta = layout.get('meta') if isinstance(layout.get('meta'), dict) else {}
        meta = dict(meta)
        raw_excerpt = str(selected.get('raw_text') or extracted_text or '')[:1200]
        if raw_excerpt:
            meta['pdf_segment_excerpt'] = raw_excerpt
        layout['meta'] = meta
        task.tactical_layout = layout
        task.save(
            update_fields=[
                'title',
                'duration_minutes',
                'objective',
                'coaching_points',
                'confrontation_rules',
                'tactical_layout',
            ]
        )
        _apply_analysis_to_task(task, analysis)
        if not task.task_preview_image:
            _ensure_library_task_preview(task, force=False, prefer_render=True)
        return True
    except Exception:
        return False


def _update_library_task_from_post(task, post_data, scope_key=None):
    if not task:
        raise ValueError('Tarea no encontrada.')
    if scope_key and _task_scope_for_item(task) != scope_key:
        raise ValueError('La tarea seleccionada no pertenece a este espacio.')

    _ensure_original_task_snapshot(task)

    title = (post_data.get('task_title') or '').strip()
    if not title:
        raise ValueError('El título de la tarea es obligatorio.')
    block = (post_data.get('task_block') or task.block or SessionTask.BLOCK_MAIN_1).strip()
    valid_blocks = {choice[0] for choice in SessionTask.BLOCK_CHOICES}
    if block not in valid_blocks:
        block = SessionTask.BLOCK_MAIN_1
    minutes = _parse_int(post_data.get('task_minutes'))
    if minutes is None:
        minutes = int(task.duration_minutes or 15)
    minutes = max(5, min(minutes, 90))

    task.title = _repair_joined_words_text(title[:220])[:160]
    task.block = block
    task.duration_minutes = minutes
    task.title = _polish_spanish_text(task.title, multiline=False, max_len=160)
    task.objective = _polish_spanish_text(_repair_joined_words_text((post_data.get('task_objective') or '').strip()[:260]), multiline=False, max_len=180)
    task.coaching_points = _polish_spanish_text(_repair_joined_words_text((post_data.get('task_coaching_points') or '').strip()), multiline=True)
    task.confrontation_rules = _polish_spanish_text(_repair_joined_words_text((post_data.get('task_confrontation_rules') or '').strip()), multiline=True)

    layout = task.tactical_layout if isinstance(task.tactical_layout, dict) else {}
    layout = dict(layout)
    meta = layout.get('meta') if isinstance(layout.get('meta'), dict) else {}
    meta = dict(meta)
    analysis_meta = meta.get('analysis') if isinstance(meta.get('analysis'), dict) else {}
    analysis_meta = dict(analysis_meta)
    task_sheet = analysis_meta.get('task_sheet') if isinstance(analysis_meta.get('task_sheet'), dict) else {}
    task_sheet = dict(task_sheet)

    sheet_field_map = {
        'task_sheet_description': 'description',
        'task_sheet_players': 'players',
        'task_sheet_space': 'space',
        'task_sheet_dimensions': 'dimensions',
        'task_sheet_materials': 'materials',
    }
    task_sheet_touched = False
    for post_key, sheet_key in sheet_field_map.items():
        if post_key in post_data:
            task_sheet[sheet_key] = _polish_spanish_text(
                _repair_joined_words_text(str(post_data.get(post_key) or '').strip()),
                multiline=(sheet_key == 'description'),
            )
            task_sheet_touched = True
    if task_sheet_touched:
        analysis_meta['task_sheet'] = task_sheet
        meta['analysis'] = analysis_meta
        layout['meta'] = meta
        task.tactical_layout = layout
        task.save(
            update_fields=[
                'title',
                'block',
                'duration_minutes',
                'objective',
                'coaching_points',
                'confrontation_rules',
                'tactical_layout',
            ]
        )
    else:
        layout['meta'] = meta
        task.tactical_layout = layout
        task.save(
            update_fields=[
                'title',
                'block',
                'duration_minutes',
                'objective',
                'coaching_points',
                'confrontation_rules',
                'tactical_layout',
            ]
        )


def _storage_url_or_empty(name):
    value = str(name or '').strip()
    if not value:
        return ''
    try:
        return default_storage.url(value)
    except Exception:
        return ''


def _ensure_original_task_snapshot(task):
    if not task:
        return {}
    layout = task.tactical_layout if isinstance(task.tactical_layout, dict) else {}
    layout = dict(layout)
    meta = layout.get('meta') if isinstance(layout.get('meta'), dict) else {}
    meta = dict(meta)
    if isinstance(meta.get('original_version'), dict):
        return meta.get('original_version') or {}
    analysis_meta = meta.get('analysis') if isinstance(meta.get('analysis'), dict) else {}
    analysis_meta = dict(analysis_meta)
    task_sheet = analysis_meta.get('task_sheet') if isinstance(analysis_meta.get('task_sheet'), dict) else {}
    snapshot = {
        'captured_at': timezone.now().isoformat(),
        'title': task.title or '',
        'block': task.block or '',
        'duration_minutes': int(task.duration_minutes or 0),
        'objective': task.objective or '',
        'coaching_points': task.coaching_points or '',
        'confrontation_rules': task.confrontation_rules or '',
        'task_sheet': dict(task_sheet),
        'graphic_editor': meta.get('graphic_editor') if isinstance(meta.get('graphic_editor'), dict) else {},
        'task_preview_image': task.task_preview_image.name if task.task_preview_image else '',
        'task_pdf': task.task_pdf.name if task.task_pdf else '',
    }
    meta['original_version'] = snapshot
    layout['meta'] = meta
    task.tactical_layout = layout
    task.save(update_fields=['tactical_layout'])
    return snapshot


def _is_imported_task(task):
    if not task:
        return False
    if bool(getattr(task, 'task_pdf', None)):
        return True
    layout = task.tactical_layout if isinstance(task.tactical_layout, dict) else {}
    meta = layout.get('meta') if isinstance(layout.get('meta'), dict) else {}
    source = str(meta.get('source') or '').strip().lower()
    if source in {'pdf_analysis', 'pdf_import', 'pdf'}:
        return True
    if str(meta.get('pdf_source_name') or '').strip():
        return True
    return False


def _is_task_editable(task):
    return not _is_imported_task(task)


def _restore_task_from_original_snapshot(task, scope_key=None):
    if not task:
        raise ValueError('Tarea no encontrada.')
    if scope_key and _task_scope_for_item(task) != scope_key:
        raise ValueError('La tarea seleccionada no pertenece a este espacio.')
    layout = task.tactical_layout if isinstance(task.tactical_layout, dict) else {}
    layout = dict(layout)
    meta = layout.get('meta') if isinstance(layout.get('meta'), dict) else {}
    meta = dict(meta)
    original = meta.get('original_version') if isinstance(meta.get('original_version'), dict) else {}
    if not original:
        raise ValueError('La tarea no tiene versión original guardada.')

    task.title = str(original.get('title') or task.title or '')[:160]
    block = str(original.get('block') or task.block or SessionTask.BLOCK_MAIN_1).strip()
    if block not in {choice[0] for choice in SessionTask.BLOCK_CHOICES}:
        block = SessionTask.BLOCK_MAIN_1
    task.block = block
    task.duration_minutes = max(5, min(_parse_int(original.get('duration_minutes')) or int(task.duration_minutes or 15), 90))
    task.objective = str(original.get('objective') or '')[:180]
    task.coaching_points = str(original.get('coaching_points') or '')
    task.confrontation_rules = str(original.get('confrontation_rules') or '')

    analysis_meta = meta.get('analysis') if isinstance(meta.get('analysis'), dict) else {}
    analysis_meta = dict(analysis_meta)
    original_sheet = original.get('task_sheet') if isinstance(original.get('task_sheet'), dict) else {}
    if original_sheet:
        analysis_meta['task_sheet'] = dict(original_sheet)
        meta['analysis'] = analysis_meta

    original_graphic = original.get('graphic_editor') if isinstance(original.get('graphic_editor'), dict) else {}
    if original_graphic:
        meta['graphic_editor'] = dict(original_graphic)

    layout['meta'] = meta
    task.tactical_layout = layout
    preview_name = str(original.get('task_preview_image') or '').strip()
    update_fields = [
        'title',
        'block',
        'duration_minutes',
        'objective',
        'coaching_points',
        'confrontation_rules',
        'tactical_layout',
    ]
    if preview_name:
        task.task_preview_image = preview_name
        update_fields.append('task_preview_image')
    task.save(update_fields=update_fields)


def _decode_canvas_data_url(data_url):
    value = str(data_url or '').strip()
    if not value or ';base64,' not in value:
        return None, None
    header, encoded = value.split(';base64,', 1)
    mime = header.replace('data:', '').strip().lower()
    allowed_mimes = {
        'image/png': '.png',
        'image/jpeg': '.jpg',
        'image/webp': '.webp',
    }
    extension = allowed_mimes.get(mime)
    if not extension:
        return None, None
    try:
        raw_bytes = base64.b64decode(encoded)
    except Exception:
        return None, None
    if not raw_bytes:
        return None, None
    return raw_bytes, extension


def _sessions_tab_from_action(action):
    action_key = str(action or '').strip()
    if action_key == 'library_upload_pdf':
        return 'import'
    if action_key in {'bulk_create_tasks', 'create_draw_task'}:
        return 'create'
    if action_key in {
        'create_microcycle_plan',
        'create_session_plan',
        'update_microcycle_plan',
        'clone_microcycle_plan',
        'copy_library_task_to_session',
        'update_session_plan',
        'delete_session_plan',
        'duplicate_session_plan',
        'move_session_task',
        'duplicate_session_task',
        'delete_session_task',
        'restore_session_task',
        'purge_session_task',
    }:
        return 'planning'
    if action_key in {
        'analyze_all_library_pdfs',
        'split_existing_library_pdfs',
        'analyze_library_pdf',
        'auto_fix_task_text',
        'delete_library_task',
        'restore_library_task',
        'purge_library_task',
        'update_library_task',
        'create_task_from_analysis',
    }:
        return 'library'
    return 'library'


def _next_session_task_order(session):
    return (SessionTask.objects.filter(session=session, deleted_at__isnull=True).aggregate(Max('order')).get('order__max') or 0) + 1


def _clone_session_task_to_session(source_task, target_session, note=''):
    cloned = SessionTask.objects.create(
        session=target_session,
        title=source_task.title,
        block=source_task.block,
        duration_minutes=source_task.duration_minutes,
        objective=source_task.objective,
        coaching_points=source_task.coaching_points,
        confrontation_rules=source_task.confrontation_rules,
        tactical_layout=source_task.tactical_layout if isinstance(source_task.tactical_layout, dict) else {},
        task_pdf=source_task.task_pdf.name if source_task.task_pdf else None,
        task_preview_image=source_task.task_preview_image.name if source_task.task_preview_image else None,
        status=SessionTask.STATUS_PLANNED,
        order=_next_session_task_order(target_session),
        notes=note or f'Clonada desde tarea #{source_task.id}',
    )
    return cloned


def _normalize_session_task_orders(session):
    ordered_tasks = list(SessionTask.objects.filter(session=session, deleted_at__isnull=True).order_by('order', 'id'))
    for index, item in enumerate(ordered_tasks, start=1):
        if int(item.order or 0) != index:
            item.order = index
            item.save(update_fields=['order'])


def _move_session_task(task, direction):
    siblings = list(SessionTask.objects.filter(session=task.session, deleted_at__isnull=True).order_by('order', 'id'))
    if len(siblings) < 2:
        return False
    current_index = None
    for index, item in enumerate(siblings):
        if item.id == task.id:
            current_index = index
            break
    if current_index is None:
        return False
    target_index = current_index - 1 if direction == 'up' else current_index + 1
    if target_index < 0 or target_index >= len(siblings):
        return False
    target = siblings[target_index]
    with transaction.atomic():
        current_order = int(task.order or current_index + 1)
        target_order = int(target.order or target_index + 1)
        task.order = 0
        task.save(update_fields=['order'])
        target.order = current_order
        target.save(update_fields=['order'])
        task.order = target_order
        task.save(update_fields=['order'])
    _normalize_session_task_orders(task.session)
    return True


def _clone_training_session(source_session, target_microcycle, target_date=None, target_focus=''):
    clone_date = target_date or source_session.session_date
    clone_focus = str(target_focus or source_session.focus or '').strip()[:140]
    if not clone_focus:
        raise ValueError('Indica un foco válido para la sesión duplicada.')
    if clone_date < target_microcycle.week_start or clone_date > target_microcycle.week_end:
        raise ValueError('La fecha duplicada debe quedar dentro del microciclo destino.')
    if TrainingSession.objects.filter(
        microcycle=target_microcycle,
        session_date=clone_date,
        focus__iexact=clone_focus,
    ).exists():
        raise ValueError('Ya existe una sesión con esa fecha y foco en el microciclo destino.')

    cloned_session = TrainingSession.objects.create(
        microcycle=target_microcycle,
        session_date=clone_date,
        start_time=source_session.start_time,
        duration_minutes=source_session.duration_minutes,
        intensity=source_session.intensity,
        focus=clone_focus,
        content=source_session.content,
        status=TrainingSession.STATUS_PLANNED,
        order=(TrainingSession.objects.filter(microcycle=target_microcycle).aggregate(Max('order')).get('order__max') or 0) + 1,
    )
    for source_task in source_session.tasks.filter(deleted_at__isnull=True).order_by('order', 'id'):
        _clone_session_task_to_session(
            source_task,
            cloned_session,
            note=f'Clonada desde sesión #{source_session.id} · tarea #{source_task.id}',
        )
    return cloned_session


def _clone_microcycle_plan(source_microcycle, week_start, week_end=None):
    if TrainingMicrocycle.objects.filter(team=source_microcycle.team, week_start=week_start).exists():
        raise ValueError('Ya existe un microciclo con esa fecha de inicio.')
    source_span = max(0, (source_microcycle.week_end - source_microcycle.week_start).days)
    target_week_end = week_end or (week_start + timedelta(days=source_span))
    if target_week_end < week_start:
        raise ValueError('La fecha de fin clonada no puede ser anterior al inicio.')

    cloned_microcycle = TrainingMicrocycle.objects.create(
        team=source_microcycle.team,
        title=(str(source_microcycle.title or 'Microciclo')[:110] + ' · copia')[:140],
        objective=source_microcycle.objective,
        week_start=week_start,
        week_end=target_week_end,
        status=TrainingMicrocycle.STATUS_DRAFT,
        notes=source_microcycle.notes,
    )
    source_sessions = list(source_microcycle.sessions.prefetch_related('tasks').order_by('session_date', 'start_time', 'order', 'id'))
    for source_session in source_sessions:
        day_offset = (source_session.session_date - source_microcycle.week_start).days
        clone_date = week_start + timedelta(days=day_offset)
        if clone_date > target_week_end:
            continue
        _clone_training_session(source_session, cloned_microcycle, target_date=clone_date, target_focus=source_session.focus)
    return cloned_microcycle


def _starter_canvas_state(preset):
    return {'version': '5.3.0', 'objects': []}


def _sessions_workspace_page(request, scope_key='coach', scope_title='Sesiones'):
    if not _can_access_sessions_workspace(request.user):
        return HttpResponse('No tienes permisos para acceder a sesiones.', status=403)
    forbidden = _forbid_if_workspace_module_disabled(request, 'sessions', label='sesiones')
    if forbidden:
        return forbidden
    primary_team = _get_primary_team_for_request(request)
    if not primary_team:
        raise Http404('Equipo principal no configurado')

    feedback = ''
    error = ''
    analysis = None
    analysis_task = None
    active_tab = 'library'
    allowed_library_views = {'overview', 'phase', 'type', 'players', 'duration', 'quality', 'date'}
    library_view = str(request.GET.get('library_view') or request.POST.get('library_view') or 'overview').strip().lower()
    if library_view not in allowed_library_views:
        library_view = 'overview'
    library_key = str(request.GET.get('library_key') or request.POST.get('library_key') or '').strip()

    planner_tables_ready = True
    try:
        SessionTask.objects.order_by('-id').values_list('id', flat=True).first()
    except (OperationalError, ProgrammingError):
        planner_tables_ready = False
        error = (
            'El módulo de sesiones requiere migración de base de datos. '
            'Ejecuta `python manage.py migrate` y recarga la página.'
        )

    all_sessions = []

    if request.method == 'POST' and planner_tables_ready:
        planner_action = (request.POST.get('planner_action') or '').strip()
        posted_tab = (request.POST.get('planner_tab') or '').strip()
        if posted_tab in {'import', 'create', 'library', 'planning'}:
            active_tab = posted_tab
        else:
            active_tab = _sessions_tab_from_action(planner_action)
        try:
            if planner_action == 'library_upload_pdf':
                title = _sanitize_task_text((request.POST.get('pdf_task_title') or '').strip(), multiline=False, max_len=160)
                objective = _sanitize_task_text((request.POST.get('pdf_task_objective') or '').strip(), multiline=False, max_len=180)
                block = (request.POST.get('pdf_task_block') or SessionTask.BLOCK_MAIN_1).strip()
                minutes = _parse_int(request.POST.get('pdf_task_minutes')) or 15
                pdf_files = list(request.FILES.getlist('library_task_pdf'))
                if block not in {choice[0] for choice in SessionTask.BLOCK_CHOICES}:
                    block = SessionTask.BLOCK_MAIN_1
                target_session = _get_or_create_library_session(primary_team, scope_key)
                if not pdf_files:
                    raise ValueError('Selecciona al menos un PDF.')
                for item in pdf_files:
                    if not str(getattr(item, 'name', '') or '').lower().endswith('.pdf'):
                        raise ValueError('Todos los archivos deben ser PDF.')
                base_order = SessionTask.objects.filter(session=target_session).count()
                created_count = 0
                processed_pdfs = 0
                for idx, pdf_file in enumerate(pdf_files, start=1):
                    raw_name = str(getattr(pdf_file, 'name', '') or '').rsplit('/', 1)[-1]
                    file_stem = raw_name.rsplit('.', 1)[0].strip() or f'Tarea PDF {idx}'
                    task_title = title or file_stem
                    if title and len(pdf_files) > 1:
                        task_title = f'{title} · {file_stem}'

                    extracted_text = ''
                    try:
                        extracted_text = _extract_pdf_text(pdf_file, max_chars=60000)
                    except Exception:
                        pass
                    parsed_tasks = _extract_tasks_from_pdf_text(extracted_text, fallback_title=task_title)
                    if not parsed_tasks:
                        parsed_tasks = [
                            {
                                'analysis': {
                                    'title': task_title[:160],
                                    'objective': objective[:180],
                                    'minutes': max(5, min(minutes, 90)),
                                    'coaching_points': '',
                                    'confrontation_rules': '',
                                    'summary': '',
                                    'work_contexts': [],
                                    'objective_tags': [],
                                    'exercise_types': [],
                                    'phase_tags': [],
                                    'players_count_estimate': None,
                                    'players_band': '',
                                    'duration_band': _duration_band_label(max(5, min(minutes, 90))),
                                    'detected_materials': [],
                                    'quality_score': 0,
                                },
                                'raw_text': '',
                                'segment_index': 1,
                                'segment_total': 1,
                            }
                        ]

                    if hasattr(pdf_file, 'seek'):
                        pdf_file.seek(0)
                    preview_payloads = _extract_preview_images_from_pdf(
                        pdf_file,
                        max_images=max(1, len(parsed_tasks)),
                    )

                    first = parsed_tasks[0]
                    first_analysis = first.get('analysis') or {}
                    first_title = (first_analysis.get('title') or task_title or 'Tarea desde PDF')[:160]
                    first_task = SessionTask.objects.create(
                        session=target_session,
                        title=first_title,
                        block=block,
                        duration_minutes=max(5, min((_parse_int(first_analysis.get('minutes')) or minutes), 90)),
                        objective=((first_analysis.get('objective') or objective or '')[:180]),
                        coaching_points=(first_analysis.get('coaching_points') or ''),
                        confrontation_rules=(first_analysis.get('confrontation_rules') or ''),
                        tactical_layout={
                            'meta': {
                                'scope': scope_key,
                                'pdf_source_name': raw_name,
                                'pdf_segment_index': first.get('segment_index') or 1,
                                'pdf_segments_total': first.get('segment_total') or 1,
                                'pdf_segment_excerpt': (first.get('raw_text') or '')[:1200],
                                'pdf_split_done': True,
                            }
                        },
                        task_pdf=pdf_file,
                        status=SessionTask.STATUS_PLANNED,
                        order=base_order + created_count + 1,
                        notes='Cargada desde Biblioteca PDF',
                    )
                    preview_payload = preview_payloads[0] if preview_payloads else None
                    if preview_payload:
                        preview_name, preview_content = preview_payload
                        first_task.task_preview_image.save(preview_name, preview_content, save=True)
                    try:
                        _apply_analysis_to_task(first_task, first_analysis)
                    except Exception:
                        pass
                    created_count += 1
                    processed_pdfs += 1

                    shared_pdf_name = first_task.task_pdf.name if first_task.task_pdf else ''
                    shared_preview_name = first_task.task_preview_image.name if first_task.task_preview_image else ''
                    for extra in parsed_tasks[1:]:
                        extra_analysis = extra.get('analysis') or {}
                        extra_title = (extra_analysis.get('title') or f'{task_title} · Tarea {extra.get("segment_index") or 0}')[:160]
                        extra_task = SessionTask.objects.create(
                            session=target_session,
                            title=extra_title,
                            block=block,
                            duration_minutes=max(5, min((_parse_int(extra_analysis.get('minutes')) or minutes), 90)),
                            objective=((extra_analysis.get('objective') or objective or '')[:180]),
                            coaching_points=(extra_analysis.get('coaching_points') or ''),
                            confrontation_rules=(extra_analysis.get('confrontation_rules') or ''),
                            tactical_layout={
                                'meta': {
                                    'scope': scope_key,
                                    'pdf_source_name': raw_name,
                                    'pdf_segment_index': extra.get('segment_index') or 1,
                                    'pdf_segments_total': extra.get('segment_total') or 1,
                                    'pdf_segment_excerpt': (extra.get('raw_text') or '')[:1200],
                                    'pdf_split_done': True,
                                }
                            },
                            task_pdf=shared_pdf_name or None,
                            task_preview_image=shared_preview_name or None,
                            status=SessionTask.STATUS_PLANNED,
                            order=base_order + created_count + 1,
                            notes='Extraída automáticamente desde PDF multi-tarea',
                        )
                        segment_index = max(1, int(extra.get('segment_index') or 1))
                        extra_preview_payload = None
                        if preview_payloads:
                            extra_preview_payload = preview_payloads[min(segment_index - 1, len(preview_payloads) - 1)]
                        if extra_preview_payload:
                            extra_preview_name, extra_preview_content = extra_preview_payload
                            extra_task.task_preview_image.save(extra_preview_name, extra_preview_content, save=True)
                        elif shared_preview_name:
                            extra_task.task_preview_image = shared_preview_name
                            extra_task.save(update_fields=['task_preview_image'])
                        try:
                            _apply_analysis_to_task(extra_task, extra_analysis)
                        except Exception:
                            pass
                        created_count += 1
                feedback = (
                    f'Se procesó 1 PDF y se creó 1 tarea.'
                    if created_count == 1 and processed_pdfs == 1
                    else f'Se procesaron {processed_pdfs} PDFs y se crearon {created_count} tareas.'
                )

            elif planner_action == 'create_microcycle_plan':
                title = str(request.POST.get('plan_microcycle_title') or 'Microciclo semanal').strip()[:140]
                objective = str(request.POST.get('plan_microcycle_objective') or '').strip()[:200]
                week_start_raw = str(request.POST.get('plan_week_start') or '').strip()
                week_end_raw = str(request.POST.get('plan_week_end') or '').strip()
                notes = _serialize_microcycle_plan_fields(
                    {
                        'attack': request.POST.get('plan_microcycle_attack'),
                        'defense': request.POST.get('plan_microcycle_defense'),
                        'set_pieces': request.POST.get('plan_microcycle_set_pieces'),
                        'rival_notes': request.POST.get('plan_microcycle_rival_notes'),
                        'general_notes': request.POST.get('plan_microcycle_notes'),
                    }
                )
                if not week_start_raw or not week_end_raw:
                    raise ValueError('Indica semana inicio y fin para crear el microciclo.')
                try:
                    week_start = datetime.strptime(week_start_raw, '%Y-%m-%d').date()
                    week_end = datetime.strptime(week_end_raw, '%Y-%m-%d').date()
                except ValueError:
                    raise ValueError('Formato de fechas no válido para microciclo.')
                if week_end < week_start:
                    raise ValueError('La fecha de fin del microciclo no puede ser anterior al inicio.')
                status = str(request.POST.get('plan_microcycle_status') or TrainingMicrocycle.STATUS_DRAFT).strip()
                if status not in {item[0] for item in TrainingMicrocycle.STATUS_CHOICES}:
                    status = TrainingMicrocycle.STATUS_DRAFT
                microcycle, created = TrainingMicrocycle.objects.get_or_create(
                    team=primary_team,
                    week_start=week_start,
                    defaults={
                        'week_end': week_end,
                        'title': title or 'Microciclo semanal',
                        'objective': objective,
                        'status': status,
                        'notes': notes,
                    },
                )
                if not created:
                    microcycle.week_end = week_end
                    microcycle.title = title or microcycle.title or 'Microciclo semanal'
                    microcycle.objective = objective
                    microcycle.status = status
                    microcycle.notes = notes
                    microcycle.save(update_fields=['week_end', 'title', 'objective', 'status', 'notes', 'updated_at'])
                    feedback = f'Microciclo actualizado: {microcycle.title}.'
                else:
                    feedback = f'Microciclo creado: {microcycle.title}.'

            elif planner_action == 'update_microcycle_plan':
                microcycle_id = _parse_int(request.POST.get('edit_microcycle_id'))
                if not microcycle_id:
                    raise ValueError('No se pudo identificar el microciclo.')
                microcycle = TrainingMicrocycle.objects.filter(id=microcycle_id, team=primary_team).first()
                if not microcycle:
                    raise ValueError('Microciclo no encontrado.')
                title = str(request.POST.get('edit_microcycle_title') or microcycle.title or 'Microciclo semanal').strip()[:140]
                objective = str(request.POST.get('edit_microcycle_objective') or '').strip()[:200]
                notes = _serialize_microcycle_plan_fields(
                    {
                        'attack': request.POST.get('edit_microcycle_attack'),
                        'defense': request.POST.get('edit_microcycle_defense'),
                        'set_pieces': request.POST.get('edit_microcycle_set_pieces'),
                        'rival_notes': request.POST.get('edit_microcycle_rival_notes'),
                        'general_notes': request.POST.get('edit_microcycle_notes'),
                    }
                )
                week_start_raw = str(request.POST.get('edit_week_start') or '').strip()
                week_end_raw = str(request.POST.get('edit_week_end') or '').strip()
                if not week_start_raw or not week_end_raw:
                    raise ValueError('Indica inicio y fin para actualizar el microciclo.')
                try:
                    week_start = datetime.strptime(week_start_raw, '%Y-%m-%d').date()
                    week_end = datetime.strptime(week_end_raw, '%Y-%m-%d').date()
                except ValueError:
                    raise ValueError('Formato de fechas no válido para microciclo.')
                if week_end < week_start:
                    raise ValueError('La fecha de fin del microciclo no puede ser anterior al inicio.')
                status = str(request.POST.get('edit_microcycle_status') or microcycle.status or TrainingMicrocycle.STATUS_DRAFT).strip()
                if status not in {item[0] for item in TrainingMicrocycle.STATUS_CHOICES}:
                    status = TrainingMicrocycle.STATUS_DRAFT
                if TrainingMicrocycle.objects.filter(team=primary_team, week_start=week_start).exclude(id=microcycle.id).exists():
                    raise ValueError('Ya existe otro microciclo con esa fecha de inicio.')
                sessions_outside = microcycle.sessions.exclude(session_date__range=(week_start, week_end)).exists()
                if sessions_outside:
                    raise ValueError('Ajusta o mueve primero las sesiones fuera del nuevo rango.')
                microcycle.title = title or 'Microciclo semanal'
                microcycle.objective = objective
                microcycle.week_start = week_start
                microcycle.week_end = week_end
                microcycle.status = status
                microcycle.notes = notes
                microcycle.save()
                feedback = f'Microciclo actualizado: {microcycle.title}.'

            elif planner_action == 'clone_microcycle_plan':
                source_microcycle_id = _parse_int(request.POST.get('source_microcycle_id'))
                source_microcycle = TrainingMicrocycle.objects.filter(id=source_microcycle_id, team=primary_team).first()
                if not source_microcycle:
                    raise ValueError('Microciclo origen no encontrado.')
                week_start_raw = str(request.POST.get('clone_week_start') or '').strip()
                week_end_raw = str(request.POST.get('clone_week_end') or '').strip()
                if not week_start_raw:
                    raise ValueError('Indica la semana de inicio para clonar el microciclo.')
                try:
                    clone_week_start = datetime.strptime(week_start_raw, '%Y-%m-%d').date()
                    clone_week_end = datetime.strptime(week_end_raw, '%Y-%m-%d').date() if week_end_raw else None
                except ValueError:
                    raise ValueError('Fechas no válidas para la clonación del microciclo.')
                cloned_microcycle = _clone_microcycle_plan(source_microcycle, clone_week_start, week_end=clone_week_end)
                feedback = f'Microciclo clonado: {cloned_microcycle.title}.'

            elif planner_action == 'create_session_plan':
                microcycle_id = _parse_int(request.POST.get('plan_microcycle_id'))
                if not microcycle_id:
                    raise ValueError('Selecciona microciclo para crear la sesión.')
                microcycle = (
                    TrainingMicrocycle.objects
                    .filter(id=microcycle_id, team=primary_team)
                    .first()
                )
                if not microcycle:
                    raise ValueError('Microciclo no encontrado para crear la sesión.')
                session_date_raw = str(request.POST.get('plan_session_date') or '').strip()
                focus = str(request.POST.get('plan_session_focus') or '').strip()[:140]
                if not session_date_raw or not focus:
                    raise ValueError('Completa fecha y foco para crear la sesión.')
                try:
                    session_date = datetime.strptime(session_date_raw, '%Y-%m-%d').date()
                except ValueError:
                    raise ValueError('Fecha de sesión no válida.')
                if session_date < microcycle.week_start or session_date > microcycle.week_end:
                    raise ValueError('La fecha de la sesión debe estar dentro del microciclo.')
                start_time_raw = str(request.POST.get('plan_session_start_time') or '').strip()
                start_time = None
                if start_time_raw:
                    try:
                        start_time = datetime.strptime(start_time_raw, '%H:%M').time()
                    except ValueError:
                        raise ValueError('Hora de sesión no válida.')
                duration_minutes = max(30, min(_parse_int(request.POST.get('plan_session_minutes')) or 90, 180))
                intensity = str(request.POST.get('plan_session_intensity') or TrainingSession.INTENSITY_MEDIUM).strip()
                if intensity not in {item[0] for item in TrainingSession.INTENSITY_CHOICES}:
                    intensity = TrainingSession.INTENSITY_MEDIUM
                status = str(request.POST.get('plan_session_status') or TrainingSession.STATUS_PLANNED).strip()
                if status not in {item[0] for item in TrainingSession.STATUS_CHOICES}:
                    status = TrainingSession.STATUS_PLANNED
                content = str(request.POST.get('plan_session_content') or '').strip()
                duplicate_exists = TrainingSession.objects.filter(
                    microcycle=microcycle,
                    session_date=session_date,
                    focus__iexact=focus,
                ).exists()
                if duplicate_exists:
                    raise ValueError('Ya existe una sesión con la misma fecha y foco en este microciclo.')
                next_order = (TrainingSession.objects.filter(microcycle=microcycle).aggregate(Max('order')).get('order__max') or 0) + 1
                session_obj = TrainingSession.objects.create(
                    microcycle=microcycle,
                    session_date=session_date,
                    start_time=start_time,
                    duration_minutes=duration_minutes,
                    intensity=intensity,
                    focus=focus,
                    content=content,
                    status=status,
                    order=next_order,
                )
                source_task_ids = [
                    task_id
                    for task_id in (_parse_int(value) for value in request.POST.getlist('plan_session_task_ids'))
                    if task_id
                ]
                attached_count = 0
                if source_task_ids:
                    source_tasks = list(
                        SessionTask.objects
                        .select_related('session__microcycle')
                        .filter(id__in=source_task_ids, session__microcycle__team=primary_team)
                        .order_by('order', 'id')
                    )
                    for source_task in source_tasks:
                        if _task_scope_for_item(source_task) != scope_key:
                            continue
                        _clone_session_task_to_session(
                            source_task,
                            session_obj,
                            note=f'Añadida al crear sesión desde tarea #{source_task.id}',
                        )
                        attached_count += 1
                feedback = (
                    f'Sesión creada en {microcycle.title}: {focus}.'
                    if not attached_count
                    else f'Sesión creada en {microcycle.title}: {focus}. Tareas añadidas: {attached_count}.'
                )

            elif planner_action == 'update_session_plan':
                session_id = _parse_int(request.POST.get('edit_session_id'))
                microcycle_id = _parse_int(request.POST.get('edit_microcycle_id'))
                if not session_id or not microcycle_id:
                    raise ValueError('No se pudo identificar la sesión a actualizar.')
                session_obj = (
                    TrainingSession.objects
                    .select_related('microcycle')
                    .filter(id=session_id, microcycle__team=primary_team)
                    .first()
                )
                microcycle = (
                    TrainingMicrocycle.objects
                    .filter(id=microcycle_id, team=primary_team)
                    .first()
                )
                if not session_obj or not microcycle:
                    raise ValueError('Sesión o microciclo no encontrado.')
                session_date_raw = str(request.POST.get('edit_session_date') or '').strip()
                focus = str(request.POST.get('edit_session_focus') or '').strip()[:140]
                if not session_date_raw or not focus:
                    raise ValueError('Completa fecha y foco para actualizar la sesión.')
                try:
                    session_date = datetime.strptime(session_date_raw, '%Y-%m-%d').date()
                except ValueError:
                    raise ValueError('Fecha de sesión no válida.')
                if session_date < microcycle.week_start or session_date > microcycle.week_end:
                    raise ValueError('La fecha de la sesión debe estar dentro del microciclo.')
                start_time_raw = str(request.POST.get('edit_session_start_time') or '').strip()
                start_time = None
                if start_time_raw:
                    try:
                        start_time = datetime.strptime(start_time_raw, '%H:%M').time()
                    except ValueError:
                        raise ValueError('Hora de sesión no válida.')
                duration_minutes = max(30, min(_parse_int(request.POST.get('edit_session_minutes')) or 90, 180))
                intensity = str(request.POST.get('edit_session_intensity') or TrainingSession.INTENSITY_MEDIUM).strip()
                if intensity not in {item[0] for item in TrainingSession.INTENSITY_CHOICES}:
                    intensity = TrainingSession.INTENSITY_MEDIUM
                status = str(request.POST.get('edit_session_status') or TrainingSession.STATUS_PLANNED).strip()
                if status not in {item[0] for item in TrainingSession.STATUS_CHOICES}:
                    status = TrainingSession.STATUS_PLANNED
                content = str(request.POST.get('edit_session_content') or '').strip()
                duplicate_exists = (
                    TrainingSession.objects
                    .filter(microcycle=microcycle, session_date=session_date, focus__iexact=focus)
                    .exclude(id=session_obj.id)
                    .exists()
                )
                if duplicate_exists:
                    raise ValueError('Ya existe otra sesión con la misma fecha y foco en este microciclo.')
                if session_obj.microcycle_id != microcycle.id:
                    session_obj.order = (TrainingSession.objects.filter(microcycle=microcycle).aggregate(Max('order')).get('order__max') or 0) + 1
                session_obj.microcycle = microcycle
                session_obj.session_date = session_date
                session_obj.start_time = start_time
                session_obj.duration_minutes = duration_minutes
                session_obj.intensity = intensity
                session_obj.focus = focus
                session_obj.content = content
                session_obj.status = status
                session_obj.save()
                feedback = f'Sesión actualizada: {focus}.'

            elif planner_action == 'delete_session_plan':
                session_id = _parse_int(request.POST.get('delete_session_id'))
                if not session_id:
                    raise ValueError('No se pudo identificar la sesión a eliminar.')
                session_obj = (
                    TrainingSession.objects
                    .select_related('microcycle')
                    .filter(id=session_id, microcycle__team=primary_team)
                    .first()
                )
                if not session_obj:
                    raise ValueError('Sesión no encontrada.')
                if session_obj.tasks.exists():
                    raise ValueError('No puedes borrar una sesión que ya tiene tareas asociadas.')
                deleted_focus = session_obj.focus
                session_obj.delete()
                feedback = f'Sesión eliminada: {deleted_focus}.'

            elif planner_action == 'duplicate_session_plan':
                source_session_id = _parse_int(request.POST.get('source_session_id'))
                source_session = (
                    TrainingSession.objects
                    .select_related('microcycle')
                    .prefetch_related('tasks')
                    .filter(id=source_session_id, microcycle__team=primary_team)
                    .first()
                )
                if not source_session:
                    raise ValueError('Sesión origen no encontrada.')
                target_microcycle_id = _parse_int(request.POST.get('duplicate_microcycle_id')) or source_session.microcycle_id
                target_microcycle = TrainingMicrocycle.objects.filter(id=target_microcycle_id, team=primary_team).first()
                if not target_microcycle:
                    raise ValueError('Microciclo destino no encontrado.')
                duplicate_date_raw = str(request.POST.get('duplicate_session_date') or '').strip()
                duplicate_focus = str(request.POST.get('duplicate_session_focus') or '').strip()
                if not duplicate_date_raw:
                    raise ValueError('Indica una fecha para duplicar la sesión.')
                try:
                    duplicate_date = datetime.strptime(duplicate_date_raw, '%Y-%m-%d').date()
                except ValueError:
                    raise ValueError('Fecha no válida para duplicar la sesión.')
                duplicated_session = _clone_training_session(
                    source_session,
                    target_microcycle,
                    target_date=duplicate_date,
                    target_focus=duplicate_focus or source_session.focus,
                )
                feedback = f'Sesión duplicada: {duplicated_session.focus}.'

            elif planner_action == 'copy_library_task_to_session':
                source_task_id = _parse_int(request.POST.get('source_task_id'))
                target_session_id = _parse_int(request.POST.get('target_session_id'))
                source_task = (
                    SessionTask.objects
                    .select_related('session__microcycle')
                    .filter(id=source_task_id, session__microcycle__team=primary_team, deleted_at__isnull=True)
                    .first()
                )
                target_session = (
                    TrainingSession.objects
                    .select_related('microcycle')
                    .filter(id=target_session_id, microcycle__team=primary_team)
                    .first()
                )
                if not source_task or not target_session:
                    raise ValueError('No se pudo copiar: tarea origen o sesión destino no válidas.')
                if _task_scope_for_item(source_task) != scope_key:
                    raise ValueError('La tarea origen no pertenece a este espacio.')
                copied = _clone_session_task_to_session(
                    source_task,
                    target_session,
                    note=f'Copiada desde biblioteca (tarea #{source_task.id})',
                )
                feedback = f'Tarea copiada a sesión: {copied.title}.'

            elif planner_action == 'move_session_task':
                task_id = _parse_int(request.POST.get('task_id'))
                direction = str(request.POST.get('move_direction') or '').strip().lower()
                target_task = (
                    SessionTask.objects
                    .select_related('session__microcycle')
                    .filter(id=task_id, session__microcycle__team=primary_team, deleted_at__isnull=True)
                    .first()
                )
                if not target_task:
                    raise ValueError('Tarea de sesión no encontrada.')
                if direction not in {'up', 'down'}:
                    raise ValueError('Movimiento no válido para la tarea.')
                if _move_session_task(target_task, direction):
                    feedback = f'Orden actualizado para: {target_task.title}.'
                else:
                    error = 'La tarea ya está en el límite del bloque.'

            elif planner_action == 'duplicate_session_task':
                task_id = _parse_int(request.POST.get('task_id'))
                target_task = (
                    SessionTask.objects
                    .select_related('session__microcycle')
                    .filter(id=task_id, session__microcycle__team=primary_team, deleted_at__isnull=True)
                    .first()
                )
                if not target_task:
                    raise ValueError('Tarea de sesión no encontrada.')
                duplicated_task = _clone_session_task_to_session(
                    target_task,
                    target_task.session,
                    note=f'Duplicada desde tarea #{target_task.id}',
                )
                feedback = f'Tarea duplicada en la sesión: {duplicated_task.title}.'

            elif planner_action == 'delete_session_task':
                task_id = _parse_int(request.POST.get('task_id'))
                target_task = (
                    SessionTask.objects
                    .select_related('session__microcycle')
                    .filter(id=task_id, session__microcycle__team=primary_team, deleted_at__isnull=True)
                    .first()
                )
                if not target_task:
                    raise ValueError('Tarea de sesión no encontrada.')
                session_for_order = target_task.session
                deleted_title = str(target_task.title or f'Tarea {target_task.id}')
                try:
                    write_task_backup(
                        target_task,
                        kind='session_task',
                        reason='delete',
                        actor_username=(request.user.username if getattr(request, 'user', None) and request.user.is_authenticated else ''),
                    )
                except Exception:
                    pass
                target_task.deleted_at = timezone.localtime()
                target_task.deleted_by = request.user if getattr(request, 'user', None) and request.user.is_authenticated else None
                target_task.save(update_fields=['deleted_at', 'deleted_by'])
                _normalize_session_task_orders(session_for_order)
                feedback = f'Tarea enviada a papelera: {deleted_title}.'

            elif planner_action == 'restore_session_task':
                task_id = _parse_int(request.POST.get('task_id'))
                target_task = (
                    SessionTask.objects
                    .select_related('session__microcycle')
                    .filter(id=task_id, session__microcycle__team=primary_team, deleted_at__isnull=False)
                    .first()
                )
                if not target_task:
                    raise ValueError('No se encontró la tarea en papelera.')
                session_for_order = target_task.session
                target_task.deleted_at = None
                target_task.deleted_by = None
                target_task.order = _next_session_task_order(session_for_order)
                target_task.save(update_fields=['deleted_at', 'deleted_by', 'order'])
                _normalize_session_task_orders(session_for_order)
                restored_title = str(target_task.title or f'Tarea {target_task.id}')
                feedback = f'Tarea restaurada: {restored_title}.'

            elif planner_action == 'purge_session_task':
                task_id = _parse_int(request.POST.get('task_id'))
                target_task = (
                    SessionTask.objects
                    .select_related('session__microcycle')
                    .filter(id=task_id, session__microcycle__team=primary_team, deleted_at__isnull=False)
                    .first()
                )
                if not target_task:
                    raise ValueError('No se encontró la tarea en papelera.')
                session_for_order = target_task.session
                task_title = str(target_task.title or f'Tarea {target_task.id}')
                try:
                    if getattr(target_task, 'task_pdf', None):
                        target_task.task_pdf.delete(save=False)
                except Exception:
                    pass
                try:
                    if getattr(target_task, 'task_preview_image', None):
                        target_task.task_preview_image.delete(save=False)
                except Exception:
                    pass
                target_task.delete()
                _normalize_session_task_orders(session_for_order)
                feedback = f'Tarea eliminada definitivamente: {task_title}.'

            elif planner_action == 'bulk_create_tasks':
                bulk_text = (request.POST.get('bulk_tasks_text') or '').strip()
                block_default = (request.POST.get('bulk_default_block') or SessionTask.BLOCK_MAIN_1).strip()
                minutes_default = _parse_int(request.POST.get('bulk_default_minutes')) or 15
                target_session_id = _parse_int(request.POST.get('bulk_target_session_id'))
                if block_default not in {choice[0] for choice in SessionTask.BLOCK_CHOICES}:
                    block_default = SessionTask.BLOCK_MAIN_1
                minutes_default = max(5, min(minutes_default, 90))
                if not bulk_text:
                    raise ValueError('Pega al menos una línea para crear tareas.')

                target_session = None
                if target_session_id:
                    target_session = (
                        TrainingSession.objects
                        .select_related('microcycle')
                        .filter(id=target_session_id, microcycle__team=primary_team)
                        .first()
                    )
                if not target_session:
                    target_session = _get_or_create_library_session(primary_team, scope_key)

                parsed_rows, parse_errors = _parse_bulk_tasks_text(
                    bulk_text,
                    default_block=block_default,
                    default_minutes=minutes_default,
                )
                if not parsed_rows and parse_errors:
                    raise ValueError(parse_errors[0])
                if not parsed_rows:
                    raise ValueError('No se detectaron tareas válidas en el bloque pegado.')

                base_order = SessionTask.objects.filter(session=target_session).count()
                created_count = 0
                for row in parsed_rows:
                    SessionTask.objects.create(
                        session=target_session,
                        title=row['title'],
                        block=row['block'],
                        duration_minutes=row['minutes'],
                        objective=row['objective'],
                        coaching_points=row['coaching_points'],
                        confrontation_rules=row['confrontation_rules'],
                        tactical_layout={
                            'meta': {
                                'scope': scope_key,
                                'bulk_source': 'manual_batch',
                                'bulk_source_line': row.get('source_line'),
                            }
                        },
                        status=SessionTask.STATUS_PLANNED,
                        order=base_order + created_count + 1,
                        notes='Creada mediante carga masiva',
                    )
                    created_count += 1

                if parse_errors:
                    feedback = (
                        f'Carga masiva completada: {created_count} tareas creadas. '
                        f'Incidencias: {len(parse_errors)} (primer error: {parse_errors[0]}).'
                    )
                else:
                    feedback = f'Carga masiva completada: {created_count} tareas creadas.'

            elif planner_action == 'create_draw_task':
                target_session_id = _parse_int(request.POST.get('draw_target_session_id'))
                title = _sanitize_task_text((request.POST.get('draw_task_title') or '').strip(), multiline=False, max_len=160)
                block = (request.POST.get('draw_task_block') or SessionTask.BLOCK_MAIN_1).strip()
                minutes = _parse_int(request.POST.get('draw_task_minutes')) or 15
                objective = _sanitize_task_text((request.POST.get('draw_task_objective') or '').strip(), multiline=False, max_len=180)
                coaching_points = _sanitize_task_text((request.POST.get('draw_task_coaching_points') or '').strip(), multiline=True)
                confrontation_rules = _sanitize_task_text((request.POST.get('draw_task_confrontation_rules') or '').strip(), multiline=True)
                description = _sanitize_task_text((request.POST.get('draw_task_description') or '').strip(), multiline=True, max_len=1200)
                description_html = _sanitize_task_rich_html((request.POST.get('draw_task_description_html') or '').strip())
                coaching_html = _sanitize_task_rich_html((request.POST.get('draw_task_coaching_points_html') or '').strip())
                rules_html = _sanitize_task_rich_html((request.POST.get('draw_task_confrontation_rules_html') or '').strip())
                players = _sanitize_task_text((request.POST.get('draw_task_players') or '').strip(), multiline=False, max_len=120)
                dimensions = _sanitize_task_text((request.POST.get('draw_task_dimensions') or '').strip(), multiline=False, max_len=120)
                space = _sanitize_task_text((request.POST.get('draw_task_space') or '').strip(), multiline=False, max_len=120)
                materials = _sanitize_task_text((request.POST.get('draw_task_materials') or '').strip(), multiline=False, max_len=300)
                organization = _sanitize_task_text((request.POST.get('draw_task_organization') or '').strip(), multiline=True, max_len=500)
                organization_html = _sanitize_task_rich_html((request.POST.get('draw_task_organization_html') or '').strip())
                work_rest = _sanitize_task_text((request.POST.get('draw_task_work_rest') or '').strip(), multiline=False, max_len=180)
                load_target = _sanitize_task_text((request.POST.get('draw_task_load_target') or '').strip(), multiline=False, max_len=180)
                players_distribution = _sanitize_task_text((request.POST.get('draw_task_players_distribution') or '').strip(), multiline=False, max_len=180)
                progression = _sanitize_task_text((request.POST.get('draw_task_progression') or '').strip(), multiline=True, max_len=500)
                progression_html = _sanitize_task_rich_html((request.POST.get('draw_task_progression_html') or '').strip())
                regression = _sanitize_task_text((request.POST.get('draw_task_regression') or '').strip(), multiline=True, max_len=500)
                regression_html = _sanitize_task_rich_html((request.POST.get('draw_task_regression_html') or '').strip())
                success_criteria = _sanitize_task_text((request.POST.get('draw_task_success_criteria') or '').strip(), multiline=True, max_len=500)
                success_criteria_html = _sanitize_task_rich_html((request.POST.get('draw_task_success_criteria_html') or '').strip())
                selected_surface = (request.POST.get('draw_task_surface') or '').strip()
                selected_pitch_format = (request.POST.get('draw_task_pitch_format') or '').strip()
                selected_phase = (request.POST.get('draw_task_game_phase') or '').strip()
                selected_methodology = (request.POST.get('draw_task_methodology') or '').strip()
                selected_complexity = (request.POST.get('draw_task_complexity') or '').strip()
                template_key = (request.POST.get('draw_task_template') or 'none').strip()
                pitch_preset = (request.POST.get('draw_task_pitch_preset') or 'full_pitch').strip()
                pitch_orientation = (request.POST.get('draw_task_pitch_orientation') or 'landscape').strip().lower()
                constraints = [str(v).strip() for v in request.POST.getlist('draw_constraints') if str(v).strip()]
                series = (request.POST.get('draw_task_series') or '').strip()
                repetitions = (request.POST.get('draw_task_repetitions') or '').strip()

                template_map = {
                    str(item.get('key') or ''): dict(item.get('values') or {})
                    for item in TASK_TEMPLATE_LIBRARY
                }
                template_values = template_map.get(template_key) or {}
                if not title:
                    title = _sanitize_task_text(str(template_values.get('task_title') or '').strip(), multiline=False, max_len=160)
                if not objective:
                    objective = _sanitize_task_text(str(template_values.get('task_objective') or '').strip(), multiline=False, max_len=180)
                if not coaching_points:
                    coaching_points = _sanitize_task_text(str(template_values.get('task_coaching_points') or '').strip(), multiline=True)
                if not confrontation_rules:
                    confrontation_rules = _sanitize_task_text(str(template_values.get('task_confrontation_rules') or '').strip(), multiline=True)
                if not space:
                    space = _sanitize_task_text(str(template_values.get('task_space') or '').strip(), multiline=False, max_len=120)
                if not organization:
                    organization = _sanitize_task_text(str(template_values.get('task_organization') or '').strip(), multiline=True, max_len=500)
                if not players_distribution:
                    players_distribution = _sanitize_task_text(str(template_values.get('task_players_distribution') or '').strip(), multiline=False, max_len=180)
                if not load_target:
                    load_target = _sanitize_task_text(str(template_values.get('task_load_target') or '').strip(), multiline=False, max_len=180)
                if not work_rest:
                    work_rest = _sanitize_task_text(str(template_values.get('task_work_rest') or '').strip(), multiline=False, max_len=180)
                if not series:
                    series = _sanitize_task_text(str(template_values.get('task_series') or '').strip(), multiline=False, max_len=100)
                if not repetitions:
                    repetitions = _sanitize_task_text(str(template_values.get('task_repetitions') or '').strip(), multiline=False, max_len=100)
                if not progression:
                    progression = _sanitize_task_text(str(template_values.get('task_progression') or '').strip(), multiline=True, max_len=500)
                if not regression:
                    regression = _sanitize_task_text(str(template_values.get('task_regression') or '').strip(), multiline=True, max_len=500)
                if not success_criteria:
                    success_criteria = _sanitize_task_text(str(template_values.get('task_success_criteria') or '').strip(), multiline=True, max_len=500)

                if not title:
                    raise ValueError('Indica un título para la tarea dibujada.')
                if block not in {choice[0] for choice in SessionTask.BLOCK_CHOICES}:
                    block = SessionTask.BLOCK_MAIN_1
                minutes = max(5, min(minutes, 90))
                valid_surfaces = {key for key, _ in TASK_SURFACE_CHOICES}
                valid_pitch_formats = {key for key, _ in TASK_PITCH_FORMAT_CHOICES}
                valid_phases = {key for key, _ in TASK_GAME_PHASE_CHOICES}
                valid_methodologies = {key for key, _ in TASK_METHODOLOGY_CHOICES}
                valid_complexities = {key for key, _ in TASK_COMPLEXITY_CHOICES}
                valid_constraints = {key for key, _ in TASK_CONSTRAINT_CHOICES}
                constraints = [item for item in constraints if item in valid_constraints]
                if selected_surface not in valid_surfaces:
                    selected_surface = ''
                if selected_pitch_format not in valid_pitch_formats:
                    selected_pitch_format = ''
                if selected_phase not in valid_phases:
                    selected_phase = ''
                if selected_methodology not in valid_methodologies:
                    selected_methodology = ''
                if selected_complexity not in valid_complexities:
                    selected_complexity = ''
                if pitch_preset not in {'full_pitch', 'half_pitch', 'attacking_third', 'middle_third', 'defensive_third', 'seven_side', 'seven_side_single', 'futsal', 'blank'}:
                    pitch_preset = 'full_pitch'
                if pitch_orientation not in {'landscape', 'portrait'}:
                    pitch_orientation = 'landscape'
                target_session = None
                if target_session_id:
                    target_session = (
                        TrainingSession.objects
                        .select_related('microcycle')
                        .filter(id=target_session_id, microcycle__team=primary_team)
                        .first()
                    )
                if not target_session:
                    target_session = _get_or_create_library_session(primary_team, scope_key)
                canvas_state = None
                raw_canvas_state = (request.POST.get('draw_canvas_state') or '').strip()
                if raw_canvas_state:
                    try:
                        parsed_state = json.loads(raw_canvas_state)
                        if isinstance(parsed_state, dict):
                            canvas_state = parsed_state
                    except Exception:
                        canvas_state = None
                if not isinstance(canvas_state, dict):
                    canvas_state = _starter_canvas_state(pitch_preset)

                canvas_width = _parse_int(request.POST.get('draw_canvas_width')) or 1280
                canvas_height = _parse_int(request.POST.get('draw_canvas_height')) or 720
                canvas_width = max(320, min(canvas_width, 3840))
                canvas_height = max(180, min(canvas_height, 2160))

                task = SessionTask.objects.create(
                    session=target_session,
                    title=title[:160],
                    block=block,
                    duration_minutes=minutes,
                    objective=objective[:180],
                    coaching_points=coaching_points,
                    confrontation_rules=confrontation_rules,
                    tactical_layout={
                        'meta': {
                            'scope': scope_key,
                            'source': 'manual-draw',
                            'template_key': template_key,
                            'surface': selected_surface,
                            'pitch_format': selected_pitch_format,
                            'pitch_orientation': pitch_orientation,
                            'game_phase': selected_phase,
                            'methodology': selected_methodology,
                            'complexity': selected_complexity,
                            'space': space,
                            'organization': organization,
                            'organization_html': organization_html,
                            'players_distribution': players_distribution,
                            'load_target': load_target,
                            'work_rest': work_rest,
                            'series': series,
                            'repetitions': repetitions,
                            'progression': progression,
                            'regression': regression,
                            'success_criteria': success_criteria,
                            'progression_html': progression_html,
                            'regression_html': regression_html,
                            'success_criteria_html': success_criteria_html,
                            'constraints': constraints,
                            'graphic_editor': {
                                'canvas_state': canvas_state,
                                'canvas_width': canvas_width,
                                'canvas_height': canvas_height,
                            },
                            'analysis': {
                                'task_sheet': {
                                    'description': description,
                                    'description_html': description_html,
                                    'coaching_html': coaching_html,
                                    'rules_html': rules_html,
                                    'players': players,
                                    'space': space,
                                    'dimensions': dimensions,
                                    'materials': materials,
                                }
                            },
                        }
                    },
                    status=SessionTask.STATUS_PLANNED,
                    order=SessionTask.objects.filter(session=target_session).count() + 1,
                    notes='Tarea creada para dibujo manual',
                )
                preview_data = request.POST.get('draw_canvas_preview_data')
                if preview_data:
                    raw_bytes, extension = _decode_canvas_data_url(preview_data)
                    if raw_bytes and extension:
                        filename = f'task_preview_{task.id}{extension}'
                        task.task_preview_image.save(filename, ContentFile(raw_bytes), save=False)
                        task.save(update_fields=['task_preview_image'])
                feedback = 'Tarea creada con pizarra táctica.'

            elif planner_action == 'analyze_all_library_pdfs':
                tasks_for_scope = list(
                    SessionTask.objects
                    .select_related('session__microcycle')
                    .filter(session__microcycle__team=primary_team, task_pdf__isnull=False)
                    .order_by('-id')[:400]
                )
                tasks_for_scope = [item for item in tasks_for_scope if _task_scope_for_item(item) == scope_key]
                success_count = 0
                fail_count = 0
                for item in tasks_for_scope:
                    try:
                        extracted_text = _extract_pdf_text(item.task_pdf)
                        parsed = _suggest_task_from_pdf(extracted_text)
                        _apply_analysis_to_task(item, parsed)
                        success_count += 1
                    except Exception:
                        fail_count += 1
                feedback = f'Análisis masivo completado. OK: {success_count} · Error: {fail_count}.'

            elif planner_action == 'split_existing_library_pdfs':
                tasks_for_scope = list(
                    SessionTask.objects
                    .select_related('session__microcycle')
                    .filter(session__microcycle__team=primary_team, task_pdf__isnull=False)
                    .order_by('id')[:600]
                )
                tasks_for_scope = [item for item in tasks_for_scope if _task_scope_for_item(item) == scope_key]
                split_ok = 0
                split_created = 0
                split_skipped = 0
                split_fail = 0
                for item in tasks_for_scope:
                    layout = item.tactical_layout if isinstance(item.tactical_layout, dict) else {}
                    meta = layout.get('meta') if isinstance(layout.get('meta'), dict) else {}
                    if meta.get('pdf_split_done'):
                        split_skipped += 1
                        continue
                    if int(meta.get('pdf_segments_total') or 1) > 1:
                        meta['pdf_split_done'] = True
                        layout['meta'] = meta
                        item.tactical_layout = layout
                        item.save(update_fields=['tactical_layout'])
                        split_skipped += 1
                        continue
                    try:
                        extracted_text = _extract_pdf_text(item.task_pdf, max_chars=30000)
                        parsed_tasks = _extract_tasks_from_pdf_text(extracted_text, fallback_title=item.title or 'Tarea desde PDF')
                        if not parsed_tasks:
                            meta['pdf_split_done'] = True
                            layout['meta'] = meta
                            item.tactical_layout = layout
                            item.save(update_fields=['tactical_layout'])
                            split_skipped += 1
                            continue
                        if len(parsed_tasks) <= 1:
                            single = parsed_tasks[0]
                            single_analysis = single.get('analysis') or {}
                            if single_analysis:
                                item.title = (single_analysis.get('title') or item.title or 'Tarea desde PDF')[:160]
                                item.duration_minutes = max(
                                    5,
                                    min((_parse_int(single_analysis.get('minutes')) or item.duration_minutes or 15), 90),
                                )
                                item.objective = (single_analysis.get('objective') or item.objective or '')[:180]
                                item.coaching_points = single_analysis.get('coaching_points') or item.coaching_points or ''
                                item.confrontation_rules = single_analysis.get('confrontation_rules') or item.confrontation_rules or ''
                            meta['pdf_segment_index'] = 1
                            meta['pdf_segments_total'] = 1
                            meta['pdf_segment_excerpt'] = (single.get('raw_text') or '')[:1200]
                            meta['pdf_split_done'] = True
                            layout['meta'] = meta
                            item.tactical_layout = layout
                            item.save(
                                update_fields=[
                                    'title',
                                    'duration_minutes',
                                    'objective',
                                    'coaching_points',
                                    'confrontation_rules',
                                    'tactical_layout',
                                ]
                            )
                            try:
                                _apply_analysis_to_task(item, single_analysis)
                            except Exception:
                                pass
                            split_ok += 1
                            continue

                        if hasattr(item.task_pdf, 'seek'):
                            item.task_pdf.seek(0)
                        preview_payloads = _extract_preview_images_from_pdf(
                            item.task_pdf,
                            max_images=max(1, len(parsed_tasks)),
                        )
                        original_order = int(item.order or 0)
                        first = parsed_tasks[0]
                        first_analysis = first.get('analysis') or {}
                        item.title = (first_analysis.get('title') or item.title or 'Tarea desde PDF')[:160]
                        item.duration_minutes = max(
                            5,
                            min((_parse_int(first_analysis.get('minutes')) or item.duration_minutes or 15), 90),
                        )
                        item.objective = (first_analysis.get('objective') or item.objective or '')[:180]
                        item.coaching_points = first_analysis.get('coaching_points') or item.coaching_points or ''
                        item.confrontation_rules = first_analysis.get('confrontation_rules') or item.confrontation_rules or ''
                        meta['pdf_segment_index'] = 1
                        meta['pdf_segments_total'] = len(parsed_tasks)
                        meta['pdf_segment_excerpt'] = (first.get('raw_text') or '')[:1200]
                        meta['pdf_split_done'] = True
                        layout['meta'] = meta
                        item.tactical_layout = layout
                        if preview_payloads:
                            first_preview_name, first_preview_content = preview_payloads[0]
                            item.task_preview_image.save(first_preview_name, first_preview_content, save=False)
                        item.save(
                            update_fields=[
                                'title',
                                'duration_minutes',
                                'objective',
                                'coaching_points',
                                'confrontation_rules',
                                'tactical_layout',
                                'task_preview_image',
                            ]
                        )
                        try:
                            _apply_analysis_to_task(item, first_analysis)
                        except Exception:
                            pass

                        for offset, extra in enumerate(parsed_tasks[1:], start=1):
                            extra_analysis = extra.get('analysis') or {}
                            extra_title = (extra_analysis.get('title') or f'{item.title} · Tarea {offset + 1}')[:160]
                            extra_layout = {
                                'meta': {
                                    'scope': scope_key,
                                    'pdf_source_name': str(meta.get('pdf_source_name') or Path(item.task_pdf.name).name),
                                    'pdf_segment_index': extra.get('segment_index') or (offset + 1),
                                    'pdf_segments_total': len(parsed_tasks),
                                    'pdf_segment_excerpt': (extra.get('raw_text') or '')[:1200],
                                    'pdf_split_done': True,
                                }
                            }
                            new_task = SessionTask.objects.create(
                                session=item.session,
                                title=extra_title,
                                block=item.block,
                                duration_minutes=max(
                                    5,
                                    min((_parse_int(extra_analysis.get('minutes')) or item.duration_minutes or 15), 90),
                                ),
                                objective=((extra_analysis.get('objective') or '')[:180]),
                                coaching_points=(extra_analysis.get('coaching_points') or ''),
                                confrontation_rules=(extra_analysis.get('confrontation_rules') or ''),
                                tactical_layout=extra_layout,
                                task_pdf=item.task_pdf.name,
                                status=SessionTask.STATUS_PLANNED,
                                order=original_order + offset,
                                notes='Separada automáticamente desde PDF existente',
                            )
                            preview_idx = min(offset, len(preview_payloads) - 1) if preview_payloads else -1
                            if preview_idx >= 0:
                                extra_preview_name, extra_preview_content = preview_payloads[preview_idx]
                                new_task.task_preview_image.save(extra_preview_name, extra_preview_content, save=True)
                            elif item.task_preview_image:
                                new_task.task_preview_image = item.task_preview_image.name
                                new_task.save(update_fields=['task_preview_image'])
                            try:
                                _apply_analysis_to_task(new_task, extra_analysis)
                            except Exception:
                                pass
                            split_created += 1

                        split_ok += 1
                    except Exception:
                        split_fail += 1
                feedback = (
                    f'Separación completada. PDFs revisados: {split_ok} · '
                    f'Tareas nuevas: {split_created} · Omitidos: {split_skipped} · Error: {split_fail}.'
                )

            elif planner_action == 'analyze_library_pdf':
                task_id = _parse_int(request.POST.get('task_id'))
                analysis_task = (
                    SessionTask.objects
                    .select_related('session__microcycle')
                    .filter(id=task_id, session__microcycle__team=primary_team)
                    .first()
                )
                if not analysis_task or not analysis_task.task_pdf:
                    raise ValueError('Selecciona una tarea con PDF guardado para analizar.')
                if _task_scope_for_item(analysis_task) != scope_key:
                    raise ValueError('La tarea seleccionada no pertenece a este espacio.')
                pdf_text = _extract_pdf_text(analysis_task.task_pdf)
                analysis = _suggest_task_from_pdf(pdf_text)
                analysis['raw_text'] = pdf_text[:2500]
                _apply_analysis_to_task(analysis_task, analysis)
                feedback = 'Análisis actualizado para la tarea seleccionada.'

            elif planner_action == 'auto_fix_task_text':
                task_id = _parse_int(request.POST.get('task_id'))
                target_task = (
                    SessionTask.objects
                    .select_related('session__microcycle')
                    .filter(id=task_id, session__microcycle__team=primary_team, deleted_at__isnull=True)
                    .first()
                )
                if not target_task:
                    raise ValueError('Tarea no encontrada.')
                if _task_scope_for_item(target_task) != scope_key:
                    raise ValueError('La tarea seleccionada no pertenece a este espacio.')
                fixed = False
                if target_task.task_pdf:
                    fixed = _refresh_task_from_pdf_analysis(target_task) or fixed
                if _cleanup_task_joined_text_fields(target_task):
                    fixed = True
                if _task_preview_needs_refresh(target_task):
                    _ensure_library_task_preview(target_task, force=True, prefer_render=True)
                feedback = 'Tarea autocorregida.' if fixed else 'No se detectaron cambios para autocorregir.'

            elif planner_action == 'delete_library_task':
                task_id = _parse_int(request.POST.get('task_id'))
                target_task = (
                    SessionTask.objects
                    .select_related('session__microcycle')
                    .filter(id=task_id, session__microcycle__team=primary_team, deleted_at__isnull=True)
                    .first()
                )
                if not target_task:
                    raise ValueError('Tarea no encontrada.')
                if _task_scope_for_item(target_task) != scope_key:
                    raise ValueError('La tarea seleccionada no pertenece a este espacio.')
                task_title = str(target_task.title or f'Tarea {target_task.id}')
                target_task.deleted_at = timezone.localtime()
                target_task.deleted_by = request.user if getattr(request, 'user', None) and request.user.is_authenticated else None
                target_task.save(update_fields=['deleted_at', 'deleted_by'])
                feedback = f'Tarea enviada a papelera: {task_title}. (Puedes borrarla definitivamente desde “Papelera biblioteca”)'

            elif planner_action == 'restore_library_task':
                task_id = _parse_int(request.POST.get('task_id'))
                target_task = (
                    SessionTask.objects
                    .select_related('session__microcycle')
                    .filter(id=task_id, session__microcycle__team=primary_team, deleted_at__isnull=False)
                    .first()
                )
                if not target_task:
                    raise ValueError('No se encontró la tarea en papelera.')
                if _task_scope_for_item(target_task) != scope_key:
                    raise ValueError('La tarea seleccionada no pertenece a este espacio.')
                session_for_order = target_task.session
                target_task.deleted_at = None
                target_task.deleted_by = None
                target_task.order = _next_session_task_order(session_for_order)
                target_task.save(update_fields=['deleted_at', 'deleted_by', 'order'])
                _normalize_session_task_orders(session_for_order)
                feedback = f'Tarea restaurada: {str(target_task.title or f"Tarea {target_task.id}")}.'

            elif planner_action == 'purge_library_task':
                task_id = _parse_int(request.POST.get('task_id'))
                target_task = (
                    SessionTask.objects
                    .select_related('session__microcycle')
                    .filter(id=task_id, session__microcycle__team=primary_team, deleted_at__isnull=False)
                    .first()
                )
                if not target_task:
                    raise ValueError('No se encontró la tarea en papelera.')
                if _task_scope_for_item(target_task) != scope_key:
                    raise ValueError('La tarea seleccionada no pertenece a este espacio.')
                task_title = str(target_task.title or f'Tarea {target_task.id}')
                try:
                    if getattr(target_task, 'task_pdf', None):
                        target_task.task_pdf.delete(save=False)
                except Exception:
                    pass
                try:
                    if getattr(target_task, 'task_preview_image', None):
                        target_task.task_preview_image.delete(save=False)
                except Exception:
                    pass
                target_task.delete()
                feedback = f'Tarea eliminada definitivamente: {task_title}.'

            elif planner_action == 'update_library_task':
                task_id = _parse_int(request.POST.get('task_id'))
                target_task = (
                    SessionTask.objects
                    .select_related('session__microcycle')
                    .filter(id=task_id, session__microcycle__team=primary_team, deleted_at__isnull=True)
                    .first()
                )
                _update_library_task_from_post(target_task, request.POST, scope_key=scope_key)
                feedback = 'Tarea actualizada correctamente.'

            elif planner_action == 'create_task_from_analysis':
                source_task_id = _parse_int(request.POST.get('source_task_id'))
                target_session_id = _parse_int(request.POST.get('target_session_id'))
                title = (request.POST.get('task_title') or '').strip()
                objective = (request.POST.get('task_objective') or '').strip()
                coaching_points = (request.POST.get('task_coaching_points') or '').strip()
                confrontation_rules = (request.POST.get('task_confrontation_rules') or '').strip()
                minutes = _parse_int(request.POST.get('task_minutes')) or 15
                block = (request.POST.get('task_block') or SessionTask.BLOCK_MAIN_1).strip()

                if not title:
                    raise ValueError('El título de la tarea es obligatorio.')
                if block not in {choice[0] for choice in SessionTask.BLOCK_CHOICES}:
                    block = SessionTask.BLOCK_MAIN_1

                source_task = (
                    SessionTask.objects
                    .select_related('session__microcycle')
                    .filter(id=source_task_id, session__microcycle__team=primary_team)
                    .first()
                )
                target_session = (
                    TrainingSession.objects
                    .select_related('microcycle')
                    .filter(id=target_session_id, microcycle__team=primary_team)
                    .first()
                )
                if not source_task or not source_task.task_pdf:
                    raise ValueError('No se encontró el PDF fuente para crear la tarea.')
                if _task_scope_for_item(source_task) != scope_key:
                    raise ValueError('El PDF fuente no pertenece a este espacio.')
                if not target_session:
                    raise ValueError('Selecciona una sesión destino válida.')

                extracted_text = (request.POST.get('task_raw_text') or '').strip()
                layout_meta = {
                    'source': 'pdf_analysis',
                    'source_task_id': source_task.id,
                    'extracted_text_excerpt': extracted_text[:1200],
                    'scope': scope_key,
                }
                created_task = SessionTask.objects.create(
                    session=target_session,
                    title=title[:160],
                    block=block,
                    duration_minutes=max(5, min(minutes, 90)),
                    objective=objective[:180],
                    coaching_points=coaching_points,
                    confrontation_rules=confrontation_rules,
                    tactical_layout={'meta': layout_meta},
                    task_pdf=source_task.task_pdf,
                    task_preview_image=source_task.task_preview_image,
                    status=SessionTask.STATUS_PLANNED,
                    order=SessionTask.objects.filter(session=target_session).count() + 1,
                    notes='Creada con editor basado en análisis de PDF',
                )
                if not created_task.task_preview_image and created_task.task_pdf:
                    preview_payload = _extract_preview_image_from_pdf(created_task.task_pdf)
                    if preview_payload:
                        preview_name, preview_content = preview_payload
                        created_task.task_preview_image.save(preview_name, preview_content, save=True)
                if source_task and isinstance(source_task.tactical_layout, dict):
                    source_meta = source_task.tactical_layout.get('meta') if isinstance(source_task.tactical_layout.get('meta'), dict) else {}
                    source_analysis = source_meta.get('analysis') if isinstance(source_meta.get('analysis'), dict) else {}
                    if source_analysis:
                        _apply_analysis_to_task(created_task, source_analysis)
                feedback = 'Tarea creada desde análisis de PDF.'
            else:
                error = 'Acción no reconocida.'
        except ValueError as exc:
            error = str(exc)
        except Exception:
            error = 'No se pudo completar la operación. Revisa los datos e inténtalo de nuevo.'
    elif request.method == 'GET':
        requested_tab = (request.GET.get('tab') or '').strip()
        if requested_tab in {'import', 'create', 'library', 'planning'}:
            active_tab = requested_tab

    all_sessions = list(
        TrainingSession.objects
        .select_related('microcycle')
        .filter(microcycle__team=primary_team)
        .order_by('-session_date', '-id')[:150]
    ) if planner_tables_ready else []

    analyze_task_id = _parse_int(request.GET.get('analyze'))
    if planner_tables_ready and analyze_task_id and not analysis:
        candidate = (
            SessionTask.objects
            .select_related('session__microcycle')
            .filter(id=analyze_task_id, session__microcycle__team=primary_team, deleted_at__isnull=True)
            .first()
        )
        if candidate and candidate.task_pdf:
            try:
                pdf_text = _extract_pdf_text(candidate.task_pdf)
                analysis_task = candidate
                analysis = _suggest_task_from_pdf(pdf_text)
                analysis['raw_text'] = pdf_text[:2500]
                _apply_analysis_to_task(analysis_task, analysis)
            except ValueError as exc:
                error = str(exc)

    task_library_raw = []
    task_library = []
    library_deleted_tasks = []
    if planner_tables_ready and active_tab == 'library':
        task_library_raw = list(
            SessionTask.objects
            .select_related('session__microcycle')
            .filter(session__microcycle__team=primary_team, deleted_at__isnull=True)
            .order_by('-id')[:300]
        )
        task_library = [item for item in task_library_raw if _task_scope_for_item(item) == scope_key]
        deleted_candidates = list(
            SessionTask.objects
            .select_related('session__microcycle')
            .filter(session__microcycle__team=primary_team, deleted_at__isnull=False)
            .order_by('-deleted_at', '-id')[:80]
        )
        library_deleted_tasks = [item for item in deleted_candidates if _task_scope_for_item(item) == scope_key]

    # IMPORTANTE (rendimiento): no hagas mantenimiento pesado (parseo PDF / render previews)
    # dentro de la petición de la vista. En Render esto puede bloquear varios minutos el primer acceso
    # tras un restart y degradar toda la zona de Sesiones/Tareas.
    # Para operaciones de mantenimiento usa `python manage.py reanalyze_library_tasks_pro`.

    task_library_context = prepare_task_library(
        task_library,
        parse_int=_parse_int,
        sanitize_text=_sanitize_task_text,
        analysis_confidence_scores=_analysis_confidence_scores,
        task_upload_date=_task_upload_date,
        extract_effective_reference_date=_extract_effective_reference_date,
        detect_keyword_tags=_detect_keyword_tags,
        task_type_keywords=TASK_TYPE_KEYWORDS,
        task_phase_keywords=TASK_PHASE_KEYWORDS,
        players_band_label=_players_band_label,
        estimate_players_count=_estimate_players_count,
        duration_band_label=_duration_band_label,
        phase_folder_key_for_task=_phase_folder_key_for_task,
        phase_folder_meta=PHASE_FOLDER_META,
        coerce_reference_date=_coerce_reference_date,
        is_imported_task=_is_imported_task,
    )
    task_library = task_library_context['task_library']
    if task_library:
        # Evita I/O en cada request: persistimos el catálogo detectado como máximo cada 30 minutos.
        persist_cache_key = f'sessions_resources_library:{primary_team.id}:{scope_key}'
        if cache.add(persist_cache_key, '1', timeout=60 * 30):
            _persist_detected_resources_library(task_library, scope_key=scope_key)
    context_group_rows = task_library_context['context_group_rows']
    objective_group_rows = task_library_context['objective_group_rows']
    type_group_rows = task_library_context['type_group_rows']
    phase_group_rows = task_library_context['phase_group_rows']
    players_band_group_rows = task_library_context['players_band_group_rows']
    duration_band_group_rows = task_library_context['duration_band_group_rows']
    date_group_rows = task_library_context['date_group_rows']
    quality_group_rows = task_library_context['quality_group_rows']
    task_library_filtered = filter_task_library(
        task_library,
        library_view=library_view,
        library_key=library_key,
    )

    planning_microcycles = []
    planning_sessions = []
    planning_session_items = []
    if planner_tables_ready and active_tab == 'planning':
        planning_microcycles = list(
            TrainingMicrocycle.objects
            .filter(team=primary_team)
            .order_by('-week_start', '-id')[:24]
        )
        planning_session_qs = (
            TrainingSession.objects
            .select_related('microcycle')
            .filter(microcycle__team=primary_team)
            .order_by('-microcycle__week_start', 'session_date', 'start_time', 'order', 'id')
        )
        planning_sessions = list(planning_session_qs[:200])
    planning_session_ids = [int(item.id) for item in planning_sessions if getattr(item, 'id', None)]
    planning_tasks_by_session = defaultdict(list)
    if planning_session_ids:
        planning_tasks = list(
            SessionTask.objects
            .filter(session_id__in=planning_session_ids, deleted_at__isnull=True)
            .order_by('session_id', 'order', 'id')
        )
        for item in planning_tasks:
            planning_tasks_by_session[int(item.session_id)].append(item)
    planning_deleted_tasks_by_session = defaultdict(list)
    if planning_session_ids:
        deleted_tasks = list(
            SessionTask.objects
            .filter(session_id__in=planning_session_ids, deleted_at__isnull=False)
            .order_by('session_id', '-deleted_at', '-id')
        )
        for item in deleted_tasks:
            planning_deleted_tasks_by_session[int(item.session_id)].append(item)
    sessions_count_map = defaultdict(int)
    tasks_count_by_session = defaultdict(int)
    tasks_count_map = defaultdict(int)
    session_minutes_by_microcycle = defaultdict(int)
    task_minutes_by_microcycle = defaultdict(int)
    intensity_by_microcycle = defaultdict(Counter)
    sessions_by_microcycle = defaultdict(list)
    for session_item in planning_sessions:
        sessions_count_map[int(session_item.microcycle_id)] += 1
        session_minutes_by_microcycle[int(session_item.microcycle_id)] += int(session_item.duration_minutes or 0)
        intensity_by_microcycle[int(session_item.microcycle_id)][str(session_item.intensity or '')] += 1
        sessions_by_microcycle[int(session_item.microcycle_id)].append(session_item)
    if planner_tables_ready:
        session_task_counts = (
            SessionTask.objects
            .filter(session__microcycle__team=primary_team, deleted_at__isnull=True)
            .values('session_id')
            .annotate(total=Count('id'))
        )
        for row in session_task_counts:
            key = _parse_int(row.get('session_id'))
            if key:
                tasks_count_by_session[int(key)] = int(row.get('total') or 0)
        task_counts = (
            SessionTask.objects
            .filter(session__microcycle__team=primary_team, deleted_at__isnull=True)
            .values('session__microcycle_id')
            .annotate(total=Count('id'))
        )
        for row in task_counts:
            key = _parse_int(row.get('session__microcycle_id'))
            if key:
                tasks_count_map[int(key)] = int(row.get('total') or 0)
    microcycle_rows = []
    for micro in planning_microcycles:
        plan_fields = _parse_microcycle_plan_fields(getattr(micro, 'notes', ''))
        session_rows = []
        for session_item in sessions_by_microcycle.get(int(micro.id), []):
            session_tasks = planning_tasks_by_session.get(int(session_item.id), [])
            deleted_session_tasks = planning_deleted_tasks_by_session.get(int(session_item.id), [])
            task_minutes_total = sum(int(getattr(task_obj, 'duration_minutes', 0) or 0) for task_obj in session_tasks)
            task_sheets = [_build_session_task_sheet(task_obj) for task_obj in session_tasks]
            task_rows = []
            for index, task_obj in enumerate(session_tasks):
                is_imported_task = _is_imported_task(task_obj)
                task_rows.append(
                    {
                        'obj': task_obj,
                        'sheet': task_sheets[index],
                        'is_imported_task': is_imported_task,
                        'is_editable_task': not is_imported_task,
                    }
                )
            session_rows.append(
                {
                    'obj': session_item,
                    'tasks_count': tasks_count_by_session.get(int(session_item.id), 0),
                    'task_minutes_total': task_minutes_total,
                    'tasks': session_tasks,
                    'task_sheets': task_sheets,
                    'task_rows': task_rows,
                    'deleted_tasks': deleted_session_tasks,
                }
            )
            task_minutes_by_microcycle[int(micro.id)] += task_minutes_total
        week_slots = _build_microcycle_week_slots(micro, session_rows)
        microcycle_rows.append(
            {
                'obj': micro,
                'plan_fields': plan_fields,
                'week_slots': week_slots,
                'sessions_count': sessions_count_map.get(int(micro.id), 0),
                'tasks_count': tasks_count_map.get(int(micro.id), 0),
                'session_minutes_total': session_minutes_by_microcycle.get(int(micro.id), 0),
                'task_minutes_total': task_minutes_by_microcycle.get(int(micro.id), 0),
                'intensity_summary': [
                    {
                        'key': key,
                        'label': dict(TrainingSession.INTENSITY_CHOICES).get(key, key or '-'),
                        'count': count,
                    }
                    for key, count in intensity_by_microcycle.get(int(micro.id), Counter()).items()
                    if count
                ],
                'sessions': session_rows,
            }
        )
    planning_session_items = [
        {
            'id': int(session.id),
            'label': f"{session.session_date:%d/%m/%Y}"
            + (f" · {session.start_time:%H:%M}" if session.start_time else '')
            + f" · {session.focus}",
        }
        for session in (planning_sessions or all_sessions)
    ]
    planning_task_source_options = []
    if planner_tables_ready and active_tab == 'planning':
        source_task_candidates = list(
            SessionTask.objects
            .select_related('session__microcycle')
            .filter(session__microcycle__team=primary_team, deleted_at__isnull=True)
            .order_by('-id')[:120]
        )
        for task_item in source_task_candidates:
            if _task_scope_for_item(task_item) != scope_key:
                continue
            session_focus = str(getattr(getattr(task_item, 'session', None), 'focus', '') or '').strip()
            planning_task_source_options.append(
                {
                    'id': int(task_item.id),
                    'title': str(task_item.title or '').strip() or f'Tarea {task_item.id}',
                    'block_label': task_item.get_block_display(),
                    'minutes': int(task_item.duration_minutes or 0),
                    'session_label': session_focus or 'Repositorio',
                }
            )
    tactical_player_catalog = []
    try:
        squad_players = (
            Player.objects
            .filter(team=primary_team, is_active=True)
            .order_by('number', 'name')[:60]
        )
        for player in squad_players:
            tactical_player_catalog.append(
                {
                    'id': int(player.id),
                    'name': str(player.name or '').strip(),
                    'number': _parse_int(player.number) or '',
                    'position': str(player.position or '').strip(),
                    'photo_url': str(resolve_player_photo_url(request, player) or '').strip(),
                }
            )
    except Exception:
        tactical_player_catalog = []
    planner_summary = {
        'library_tasks': len(task_library),
        'filtered_tasks': len(task_library_filtered),
        'microcycles': len(planning_microcycles),
        'sessions': len(planning_sessions),
        'task_sources': len(planning_task_source_options),
    }
    planner_focus_items = []
    if not planner_summary['microcycles']:
        planner_focus_items.append('No hay microciclos creados para planificar la semana.')
    if planner_summary['microcycles'] and not planner_summary['sessions']:
        planner_focus_items.append('Hay microciclos activos, pero aún no se han cargado sesiones.')
    if planner_summary['sessions'] and not planner_summary['task_sources']:
        planner_focus_items.append('No hay tareas reutilizables para insertar en nuevas sesiones.')
    if not planner_focus_items:
        planner_focus_items.append('La planificación tiene base suficiente para seguir construyendo sesiones y microciclos.')

    return render(
        request,
        'football/sessions_planner.html',
        {
            'team_name': primary_team.display_name,
            'feedback': feedback,
            'error': error,
            'planner_tables_ready': planner_tables_ready,
            'task_builder_route_name': _task_builder_route_name(scope_key),
            'task_builder_edit_route_name': _task_builder_edit_route_name(scope_key),
            'task_blocks': SessionTask.BLOCK_CHOICES,
            'all_sessions': all_sessions,
            'task_library': task_library,
            'task_library_filtered': task_library_filtered,
            'library_deleted_tasks': library_deleted_tasks,
            'analysis': analysis,
            'analysis_task': analysis_task,
            'scope_key': scope_key,
            'scope_title': scope_title,
            'context_group_rows': context_group_rows,
            'objective_group_rows': objective_group_rows,
            'type_group_rows': type_group_rows,
            'phase_group_rows': phase_group_rows,
            'players_band_group_rows': players_band_group_rows,
            'duration_band_group_rows': duration_band_group_rows,
            'quality_group_rows': quality_group_rows,
            'date_group_rows': date_group_rows,
            'active_tab': active_tab,
            'library_view': library_view,
            'library_key': library_key,
            'planning_microcycle_rows': microcycle_rows,
            'planning_sessions': planning_sessions,
            'planning_session_items': planning_session_items,
            'planning_task_source_options': planning_task_source_options,
            'microcycle_status_choices': TrainingMicrocycle.STATUS_CHOICES,
            'session_intensity_choices': TrainingSession.INTENSITY_CHOICES,
            'session_status_choices': TrainingSession.STATUS_CHOICES,
            'planning_microcycles': planning_microcycles,
            'planner_summary': planner_summary,
            'planner_focus_items': planner_focus_items,
            'task_surface_choices': TASK_SURFACE_CHOICES,
            'task_pitch_choices': TASK_PITCH_FORMAT_CHOICES,
            'tactical_player_catalog': tactical_player_catalog,
        },
    )


def _sessions_scope_route_name(scope_key):
    return {
        'coach': 'sessions',
        'goalkeeper': 'sessions-goalkeeper',
        'fitness': 'sessions-fitness',
    }.get(scope_key, 'sessions')


def _task_builder_route_name(scope_key):
    return {
        'coach': 'sessions-task-create',
        'goalkeeper': 'sessions-goalkeeper-task-create',
        'fitness': 'sessions-fitness-task-create',
    }.get(scope_key, 'sessions-task-create')


def _task_builder_edit_route_name(scope_key):
    return {
        'coach': 'sessions-task-edit',
        'goalkeeper': 'sessions-goalkeeper-task-edit',
        'fitness': 'sessions-fitness-task-edit',
    }.get(scope_key, 'sessions-task-edit')


def _build_tactical_player_catalog(request, primary_team):
    catalog = []
    if not primary_team:
        return catalog
    players = (
        Player.objects
        .filter(team=primary_team, is_active=True)
        .order_by('number', 'name')[:60]
    )
    for player in players:
        catalog.append(
            {
                'id': int(player.id),
                'name': str(player.name or '').strip(),
                'number': _parse_int(player.number) or '',
                'position': str(player.position or '').strip(),
                'photo_url': str(resolve_player_photo_url(request, player) or '').strip(),
            }
        )
    return catalog


def _task_studio_roster_photo_url(request, roster_player):
    if not roster_player or not getattr(roster_player, 'photo', None):
        return ''
    try:
        url = roster_player.photo.url
    except Exception:
        return ''
    if not url:
        return ''
    return request.build_absolute_uri(url) if request else url


def _build_task_studio_player_catalog(request, owner):
    if not owner:
        return []
    workspace = _ensure_task_studio_workspace(owner)
    catalog = []
    players = (
        TaskStudioRosterPlayer.objects
        .filter(owner=owner, is_active=True)
        .filter(Q(workspace=workspace) | Q(workspace__isnull=True))
        .order_by('number', 'name')[:60]
    )
    for player in players:
        catalog.append(
            {
                'id': int(player.id),
                'name': str(player.name or '').strip(),
                'number': _parse_int(player.number) or '',
                'position': str(player.position or '').strip(),
                'photo_url': str(_task_studio_roster_photo_url(request, player) or '').strip(),
            }
        )
    return catalog


def _task_builder_initial_values(task):
    meta = {}
    timeline = []
    if task and isinstance(task.tactical_layout, dict):
        meta = task.tactical_layout.get('meta') if isinstance(task.tactical_layout.get('meta'), dict) else {}
        timeline = task.tactical_layout.get('timeline') if isinstance(task.tactical_layout.get('timeline'), list) else []
    meta = meta if isinstance(meta, dict) else {}
    analysis = meta.get('analysis') if isinstance(meta.get('analysis'), dict) else {}
    task_sheet = analysis.get('task_sheet') if isinstance(analysis.get('task_sheet'), dict) else {}
    graphic_editor = meta.get('graphic_editor') if isinstance(meta.get('graphic_editor'), dict) else {}
    canvas_state = graphic_editor.get('canvas_state') or _starter_canvas_state(str(meta.get('pitch_preset') or 'full_pitch'))
    if not isinstance(canvas_state, dict):
        canvas_state = _starter_canvas_state(str(meta.get('pitch_preset') or 'full_pitch'))
    canvas_state = dict(canvas_state)
    if isinstance(timeline, list) and timeline:
        canvas_state['timeline'] = timeline
        canvas_state['active_step_index'] = 0
    return {
        'target_session_id': str(getattr(task, 'session_id', '') or ''),
        'title': str(getattr(task, 'title', '') or ''),
        'block': str(getattr(task, 'block', '') or SessionTask.BLOCK_MAIN_1),
        'minutes': int(getattr(task, 'duration_minutes', 15) or 15),
        'objective': str(getattr(task, 'objective', '') or ''),
        'coaching_points': str(getattr(task, 'coaching_points', '') or ''),
        'confrontation_rules': str(getattr(task, 'confrontation_rules', '') or ''),
        'description': str(task_sheet.get('description') or ''),
        'description_html': str(task_sheet.get('description_html') or ''),
        'coaching_points_html': str(task_sheet.get('coaching_html') or ''),
        'confrontation_rules_html': str(task_sheet.get('rules_html') or ''),
        'players': str(task_sheet.get('players') or ''),
        'materials': str(task_sheet.get('materials') or ''),
        'dimensions': str(task_sheet.get('dimensions') or ''),
        'space': str(meta.get('space') or task_sheet.get('space') or ''),
        'organization': str(meta.get('organization') or ''),
        'organization_html': str(meta.get('organization_html') or ''),
        'work_rest': str(meta.get('work_rest') or ''),
        'load_target': str(meta.get('load_target') or ''),
        'players_distribution': str(meta.get('players_distribution') or ''),
        'progression': str(meta.get('progression') or ''),
        'progression_html': str(meta.get('progression_html') or ''),
        'regression': str(meta.get('regression') or ''),
        'regression_html': str(meta.get('regression_html') or ''),
        'success_criteria': str(meta.get('success_criteria') or ''),
        'success_criteria_html': str(meta.get('success_criteria_html') or ''),
        'surface': str(meta.get('surface') or ''),
        'pitch_format': str(meta.get('pitch_format') or ''),
        'game_phase': str(meta.get('game_phase') or ''),
        'methodology': str(meta.get('methodology') or ''),
        'complexity': str(meta.get('complexity') or ''),
        'strategy': str(meta.get('strategy') or ''),
        'coordination': str(meta.get('coordination') or ''),
        'coordination_skills': str(meta.get('coordination_skills') or ''),
        'tactical_intent': str(meta.get('tactical_intent') or ''),
        'dynamics': str(meta.get('dynamics') or ''),
        'structure': str(meta.get('structure') or ''),
        'template_key': str(meta.get('template_key') or 'none'),
        'pitch_preset': str(meta.get('pitch_preset') or 'full_pitch'),
        'pitch_orientation': str(meta.get('pitch_orientation') or 'landscape'),
        'pitch_zoom': str(meta.get('pitch_zoom') or '1.00'),
        'series': str(meta.get('series') or ''),
        'repetitions': str(meta.get('repetitions') or ''),
        'player_count': str(meta.get('player_count') or ''),
        'age_group': str(meta.get('age_group') or ''),
        'training_type': str(meta.get('training_type') or ''),
        'category_tags': ', '.join(meta.get('category_tags') or []) if isinstance(meta.get('category_tags'), list) else str(meta.get('category_tags') or ''),
        'assigned_player_ids': [int(value) for value in (meta.get('assigned_player_ids') or []) if _parse_int(value)],
        'constraints': [str(value) for value in (meta.get('constraints') or []) if str(value).strip()],
        'canvas_state': json.dumps(canvas_state, ensure_ascii=False),
        'canvas_width': int(graphic_editor.get('canvas_width') or 1280),
        'canvas_height': int(graphic_editor.get('canvas_height') or 720),
    }


def _task_existing_meta(task):
    if not task or not isinstance(getattr(task, 'tactical_layout', None), dict):
        return {}
    meta = task.tactical_layout.get('meta')
    return meta if isinstance(meta, dict) else {}


def _normalize_animation_timeline(raw_timeline):
    if not isinstance(raw_timeline, list):
        return []
    normalized = []
    for index, item in enumerate(raw_timeline[:24]):
        if not isinstance(item, dict):
            continue
        canvas_state = item.get('canvas_state')
        if not isinstance(canvas_state, dict):
            continue
        title = _sanitize_task_text(str(item.get('title') or '').strip(), multiline=False, max_len=80)
        duration = max(1, min(_parse_int(item.get('duration')) or 3, 20))
        normalized.append(
            {
                'title': title or f'Paso {index + 1}',
                'duration': duration,
                'canvas_state': canvas_state,
            }
        )
    return normalized


def _save_task_builder_entry(request, primary_team, scope_key, existing_task=None):
    existing_meta = _task_existing_meta(existing_task)
    target_session_id = _parse_int(request.POST.get('draw_target_session_id'))
    title = _sanitize_task_text((request.POST.get('draw_task_title') or '').strip(), multiline=False, max_len=160)
    block = (request.POST.get('draw_task_block') or SessionTask.BLOCK_MAIN_1).strip()
    minutes = _parse_int(request.POST.get('draw_task_minutes')) or 15
    objective = _sanitize_task_text((request.POST.get('draw_task_objective') or '').strip(), multiline=False, max_len=180)
    coaching_points = _sanitize_task_text((request.POST.get('draw_task_coaching_points') or '').strip(), multiline=True)
    confrontation_rules = _sanitize_task_text((request.POST.get('draw_task_confrontation_rules') or '').strip(), multiline=True)
    description = _sanitize_task_text((request.POST.get('draw_task_description') or '').strip(), multiline=True, max_len=1200)
    description_html = _sanitize_task_rich_html((request.POST.get('draw_task_description_html') or '').strip())
    coaching_html = _sanitize_task_rich_html((request.POST.get('draw_task_coaching_points_html') or '').strip())
    rules_html = _sanitize_task_rich_html((request.POST.get('draw_task_confrontation_rules_html') or '').strip())
    players = _sanitize_task_text((request.POST.get('draw_task_players') or '').strip(), multiline=False, max_len=120)
    dimensions = _sanitize_task_text((request.POST.get('draw_task_dimensions') or '').strip(), multiline=False, max_len=120)
    space = _sanitize_task_text((request.POST.get('draw_task_space') or '').strip(), multiline=False, max_len=120)
    materials = _sanitize_task_text((request.POST.get('draw_task_materials') or '').strip(), multiline=False, max_len=300)
    organization = _sanitize_task_text((request.POST.get('draw_task_organization') or '').strip(), multiline=True, max_len=500)
    organization_html = _sanitize_task_rich_html((request.POST.get('draw_task_organization_html') or '').strip())
    raw_work_rest = request.POST.get('draw_task_work_rest')
    work_rest = _sanitize_task_text((raw_work_rest or '').strip(), multiline=False, max_len=180) if raw_work_rest is not None else str(existing_meta.get('work_rest') or '')
    load_target = _sanitize_task_text((request.POST.get('draw_task_load_target') or '').strip(), multiline=False, max_len=180)
    players_distribution = _sanitize_task_text((request.POST.get('draw_task_players_distribution') or '').strip(), multiline=False, max_len=180)
    progression = _sanitize_task_text((request.POST.get('draw_task_progression') or '').strip(), multiline=True, max_len=500)
    progression_html = _sanitize_task_rich_html((request.POST.get('draw_task_progression_html') or '').strip())
    regression = _sanitize_task_text((request.POST.get('draw_task_regression') or '').strip(), multiline=True, max_len=500)
    regression_html = _sanitize_task_rich_html((request.POST.get('draw_task_regression_html') or '').strip())
    success_criteria = _sanitize_task_text((request.POST.get('draw_task_success_criteria') or '').strip(), multiline=True, max_len=500)
    success_criteria_html = _sanitize_task_rich_html((request.POST.get('draw_task_success_criteria_html') or '').strip())
    selected_surface = (request.POST.get('draw_task_surface') or '').strip()
    selected_pitch_format = (request.POST.get('draw_task_pitch_format') or '').strip()
    raw_selected_phase = request.POST.get('draw_task_game_phase')
    selected_phase = (raw_selected_phase or '').strip() if raw_selected_phase is not None else str(existing_meta.get('game_phase') or '')
    raw_selected_methodology = request.POST.get('draw_task_methodology')
    selected_methodology = (raw_selected_methodology or '').strip() if raw_selected_methodology is not None else str(existing_meta.get('methodology') or '')
    raw_selected_complexity = request.POST.get('draw_task_complexity')
    selected_complexity = (raw_selected_complexity or '').strip() if raw_selected_complexity is not None else str(existing_meta.get('complexity') or '')
    raw_strategy = request.POST.get('draw_task_strategy')
    selected_strategy = (raw_strategy or '').strip() if raw_strategy is not None else str(existing_meta.get('strategy') or '')
    raw_dynamics = request.POST.get('draw_task_dynamics')
    selected_dynamics = (raw_dynamics or '').strip() if raw_dynamics is not None else str(existing_meta.get('dynamics') or '')
    raw_structure = request.POST.get('draw_task_structure')
    selected_structure = (raw_structure or '').strip() if raw_structure is not None else str(existing_meta.get('structure') or '')
    raw_coordination = request.POST.get('draw_task_coordination')
    selected_coordination = (raw_coordination or '').strip() if raw_coordination is not None else str(existing_meta.get('coordination') or '')
    raw_coord_skills = request.POST.get('draw_task_coordination_skills')
    selected_coord_skills = (raw_coord_skills or '').strip() if raw_coord_skills is not None else str(existing_meta.get('coordination_skills') or '')
    raw_tactical_intent = request.POST.get('draw_task_tactical_intent')
    selected_tactical_intent = (raw_tactical_intent or '').strip() if raw_tactical_intent is not None else str(existing_meta.get('tactical_intent') or '')
    raw_template_key = request.POST.get('draw_task_template')
    template_key = (raw_template_key or 'none').strip() if raw_template_key is not None else str(existing_meta.get('template_key') or 'none')
    pitch_preset = (request.POST.get('draw_task_pitch_preset') or 'full_pitch').strip()
    raw_pitch_orientation = request.POST.get('draw_task_pitch_orientation')
    pitch_orientation = (raw_pitch_orientation or '').strip().lower() if raw_pitch_orientation is not None else str(existing_meta.get('pitch_orientation') or 'landscape')
    raw_pitch_zoom = request.POST.get('draw_task_pitch_zoom')
    pitch_zoom = None
    if raw_pitch_zoom is None:
        pitch_zoom = float(existing_meta.get('pitch_zoom') or 1.0)
    else:
        try:
            pitch_zoom = float(str(raw_pitch_zoom).strip())
        except Exception:
            pitch_zoom = float(existing_meta.get('pitch_zoom') or 1.0)
    if 'draw_constraints' in request.POST:
        constraints = [str(v).strip() for v in request.POST.getlist('draw_constraints') if str(v).strip()]
    else:
        constraints = [str(v).strip() for v in (existing_meta.get('constraints') or []) if str(v).strip()]
    series = _sanitize_task_text((request.POST.get('draw_task_series') or '').strip(), multiline=False, max_len=100)
    repetitions = _sanitize_task_text((request.POST.get('draw_task_repetitions') or '').strip(), multiline=False, max_len=100)
    player_count = _sanitize_task_text((request.POST.get('draw_task_player_count') or '').strip(), multiline=False, max_len=100)
    age_group = _sanitize_task_text((request.POST.get('draw_task_age_group') or '').strip(), multiline=False, max_len=100)
    training_type = _sanitize_task_text((request.POST.get('draw_task_training_type') or '').strip(), multiline=False, max_len=120)
    raw_category_tags = request.POST.get('draw_task_category_tags')
    if raw_category_tags is None:
        existing_tags = existing_meta.get('category_tags') or []
        if isinstance(existing_tags, list):
            category_tags = [str(tag).strip() for tag in existing_tags if str(tag).strip()]
        else:
            category_tags = [tag.strip() for tag in str(existing_tags).split(',') if tag.strip()]
    else:
        category_tags_raw = _sanitize_task_text((raw_category_tags or '').strip(), multiline=False, max_len=240)
        category_tags = [tag.strip() for tag in category_tags_raw.split(',') if tag.strip()]
    assigned_player_ids = [
        player_id
        for player_id in (_parse_int(value) for value in request.POST.getlist('assigned_player_ids'))
        if player_id
    ]
    assigned_players = list(Player.objects.filter(team=primary_team, id__in=assigned_player_ids).order_by('number', 'name'))
    assigned_player_ids = [int(player.id) for player in assigned_players]

    template_map = {
        str(item.get('key') or ''): dict(item.get('values') or {})
        for item in TASK_TEMPLATE_LIBRARY
    }
    template_values = template_map.get(template_key) or {}
    if not title:
        title = _sanitize_task_text(str(template_values.get('task_title') or '').strip(), multiline=False, max_len=160)
    if not objective:
        objective = _sanitize_task_text(str(template_values.get('task_objective') or '').strip(), multiline=False, max_len=180)
    if not coaching_points:
        coaching_points = _sanitize_task_text(str(template_values.get('task_coaching_points') or '').strip(), multiline=True)
    if not confrontation_rules:
        confrontation_rules = _sanitize_task_text(str(template_values.get('task_confrontation_rules') or '').strip(), multiline=True)
    if not space:
        space = _sanitize_task_text(str(template_values.get('task_space') or '').strip(), multiline=False, max_len=120)
    if not organization:
        organization = _sanitize_task_text(str(template_values.get('task_organization') or '').strip(), multiline=True, max_len=500)
    if not players_distribution:
        players_distribution = _sanitize_task_text(str(template_values.get('task_players_distribution') or '').strip(), multiline=False, max_len=180)
    if not load_target:
        load_target = _sanitize_task_text(str(template_values.get('task_load_target') or '').strip(), multiline=False, max_len=180)
    if not work_rest:
        work_rest = _sanitize_task_text(str(template_values.get('task_work_rest') or '').strip(), multiline=False, max_len=180)
    if not progression:
        progression = _sanitize_task_text(str(template_values.get('task_progression') or '').strip(), multiline=True, max_len=500)
    if not regression:
        regression = _sanitize_task_text(str(template_values.get('task_regression') or '').strip(), multiline=True, max_len=500)
    if not success_criteria:
        success_criteria = _sanitize_task_text(str(template_values.get('task_success_criteria') or '').strip(), multiline=True, max_len=500)

    if not title:
        raise ValueError('Indica un título para la tarea.')
    if block not in {choice[0] for choice in SessionTask.BLOCK_CHOICES}:
        block = SessionTask.BLOCK_MAIN_1
    minutes = max(5, min(minutes, 90))
    valid_surfaces = {key for key, _ in TASK_SURFACE_CHOICES}
    valid_pitch_formats = {key for key, _ in TASK_PITCH_FORMAT_CHOICES}
    valid_phases = {key for key, _ in TASK_GAME_PHASE_CHOICES}
    valid_methodologies = {key for key, _ in TASK_METHODOLOGY_CHOICES}
    valid_complexities = {key for key, _ in TASK_COMPLEXITY_CHOICES}
    valid_strategies = {key for key, _ in TASK_STRATEGY_CHOICES}
    valid_dynamics = {key for key, _ in TASK_DYNAMICS_CHOICES}
    valid_structures = {key for key, _ in TASK_STRUCTURE_CHOICES}
    valid_coordination = {key for key, _ in TASK_COORDINATION_CHOICES}
    valid_coord_skills = {key for key, _ in TASK_COORDINATION_SKILLS_CHOICES}
    valid_tactical_intents = {key for key, _ in TASK_TACTICAL_INTENT_CHOICES}
    valid_constraints = {key for key, _ in TASK_CONSTRAINT_CHOICES}
    constraints = [item for item in constraints if item in valid_constraints]
    if selected_surface not in valid_surfaces:
        selected_surface = ''
    if selected_pitch_format not in valid_pitch_formats:
        selected_pitch_format = ''
    if selected_phase not in valid_phases:
        selected_phase = ''
    if selected_methodology not in valid_methodologies:
        selected_methodology = ''
    if selected_complexity not in valid_complexities:
        selected_complexity = ''
    if selected_strategy not in valid_strategies:
        selected_strategy = ''
    if selected_dynamics not in valid_dynamics:
        selected_dynamics = ''
    if selected_structure not in valid_structures:
        selected_structure = ''
    if selected_coordination not in valid_coordination:
        selected_coordination = ''
    if selected_coord_skills not in valid_coord_skills:
        selected_coord_skills = ''
    if selected_tactical_intent not in valid_tactical_intents:
        selected_tactical_intent = ''
    if pitch_preset not in {'full_pitch', 'half_pitch', 'attacking_third', 'middle_third', 'defensive_third', 'seven_side', 'seven_side_single', 'futsal', 'blank'}:
        pitch_preset = 'full_pitch'
    if pitch_orientation not in {'landscape', 'portrait'}:
        pitch_orientation = 'landscape'
    try:
        pitch_zoom = float(pitch_zoom or 1.0)
    except Exception:
        pitch_zoom = 1.0
    pitch_zoom = max(0.8, min(pitch_zoom, 1.6))

    target_session = existing_task.session if existing_task else None
    if target_session_id:
        target_session = (
            TrainingSession.objects
            .select_related('microcycle')
            .filter(id=target_session_id, microcycle__team=primary_team)
            .first()
        )
    if not target_session:
        target_session = _get_or_create_library_session(primary_team, scope_key)

    canvas_state = None
    raw_canvas_state = (request.POST.get('draw_canvas_state') or '').strip()
    if raw_canvas_state:
        try:
            parsed_state = json.loads(raw_canvas_state)
            if isinstance(parsed_state, dict):
                canvas_state = parsed_state
        except Exception:
            canvas_state = None
    if not isinstance(canvas_state, dict):
        canvas_state = _starter_canvas_state(pitch_preset)
    timeline = _normalize_animation_timeline(canvas_state.get('timeline'))

    canvas_width = max(320, min(_parse_int(request.POST.get('draw_canvas_width')) or 1280, 3840))
    canvas_height = max(180, min(_parse_int(request.POST.get('draw_canvas_height')) or 720, 2160))
    tactical_layout = {
        'tokens': canvas_state.get('objects') if isinstance(canvas_state.get('objects'), list) else [],
        'timeline': timeline,
        'meta': {
            'scope': scope_key,
            'source': 'manual-studio',
            'template_key': template_key,
            'surface': selected_surface,
            'pitch_format': selected_pitch_format,
            'pitch_preset': pitch_preset,
            'pitch_orientation': pitch_orientation,
            'pitch_zoom': pitch_zoom,
            'game_phase': selected_phase,
            'methodology': selected_methodology,
            'complexity': selected_complexity,
            'strategy': selected_strategy,
            'dynamics': selected_dynamics,
            'structure': selected_structure,
            'coordination': selected_coordination,
            'coordination_skills': selected_coord_skills,
            'tactical_intent': selected_tactical_intent,
            'space': space,
            'organization': organization,
            'organization_html': organization_html,
            'players_distribution': players_distribution,
            'load_target': load_target,
            'work_rest': work_rest,
            'series': series,
            'repetitions': repetitions,
            'progression': progression,
            'regression': regression,
            'progression_html': progression_html,
            'regression_html': regression_html,
            'success_criteria': success_criteria,
            'success_criteria_html': success_criteria_html,
            'constraints': constraints,
            'player_count': player_count,
            'age_group': age_group,
            'training_type': training_type,
            'category_tags': category_tags,
            'assigned_player_ids': assigned_player_ids,
            'assigned_player_names': [player.name for player in assigned_players],
            'graphic_editor': {
                'canvas_state': canvas_state,
                'canvas_width': canvas_width,
                'canvas_height': canvas_height,
            },
            'analysis': {
                'task_sheet': {
                    'description': description,
                    'description_html': description_html,
                    'coaching_html': coaching_html,
                    'rules_html': rules_html,
                    'players': players,
                    'space': space,
                    'dimensions': dimensions,
                    'materials': materials,
                }
            },
        }
    }
    if existing_task:
        task = existing_task
        if task.session_id != target_session.id:
            task.order = (SessionTask.objects.filter(session=target_session).aggregate(Max('order')).get('order__max') or 0) + 1
        task.session = target_session
        task.title = title[:160]
        task.block = block
        task.duration_minutes = minutes
        task.objective = objective[:180]
        task.coaching_points = coaching_points
        task.confrontation_rules = confrontation_rules
        task.tactical_layout = tactical_layout
        task.notes = 'Tarea actualizada en editor visual'
        task.save()
    else:
        task = SessionTask.objects.create(
            session=target_session,
            title=title[:160],
            block=block,
            duration_minutes=minutes,
            objective=objective[:180],
            coaching_points=coaching_points,
            confrontation_rules=confrontation_rules,
            tactical_layout=tactical_layout,
            status=SessionTask.STATUS_PLANNED,
            order=SessionTask.objects.filter(session=target_session).count() + 1,
            notes='Tarea creada en editor visual',
        )
    preview_data = request.POST.get('draw_canvas_preview_data')
    if preview_data:
        raw_bytes, extension = _decode_canvas_data_url(preview_data)
        if raw_bytes and extension:
            filename = f'task_preview_{task.id}{extension}'
            task.task_preview_image.save(filename, ContentFile(raw_bytes), save=False)
            task.save(update_fields=['task_preview_image'])
    try:
        write_task_backup(
            task,
            kind='session_task',
            reason='save',
            actor_username=(request.user.username if request and getattr(request, 'user', None) and request.user.is_authenticated else ''),
        )
    except Exception:
        pass
    return task


def _save_task_studio_entry(request, owner, existing_task=None):
    workspace = _ensure_task_studio_workspace(owner)
    existing_meta = _task_existing_meta(existing_task)
    title = _sanitize_task_text((request.POST.get('draw_task_title') or '').strip(), multiline=False, max_len=160)
    block = (request.POST.get('draw_task_block') or SessionTask.BLOCK_MAIN_1).strip()
    minutes = _parse_int(request.POST.get('draw_task_minutes')) or 15
    objective = _sanitize_task_text((request.POST.get('draw_task_objective') or '').strip(), multiline=False, max_len=180)
    coaching_points = _sanitize_task_text((request.POST.get('draw_task_coaching_points') or '').strip(), multiline=True)
    confrontation_rules = _sanitize_task_text((request.POST.get('draw_task_confrontation_rules') or '').strip(), multiline=True)
    description = _sanitize_task_text((request.POST.get('draw_task_description') or '').strip(), multiline=True, max_len=1200)
    description_html = _sanitize_task_rich_html((request.POST.get('draw_task_description_html') or '').strip())
    coaching_html = _sanitize_task_rich_html((request.POST.get('draw_task_coaching_points_html') or '').strip())
    rules_html = _sanitize_task_rich_html((request.POST.get('draw_task_confrontation_rules_html') or '').strip())
    players = _sanitize_task_text((request.POST.get('draw_task_players') or '').strip(), multiline=False, max_len=120)
    dimensions = _sanitize_task_text((request.POST.get('draw_task_dimensions') or '').strip(), multiline=False, max_len=120)
    space = _sanitize_task_text((request.POST.get('draw_task_space') or '').strip(), multiline=False, max_len=120)
    materials = _sanitize_task_text((request.POST.get('draw_task_materials') or '').strip(), multiline=False, max_len=300)
    organization = _sanitize_task_text((request.POST.get('draw_task_organization') or '').strip(), multiline=True, max_len=500)
    organization_html = _sanitize_task_rich_html((request.POST.get('draw_task_organization_html') or '').strip())
    raw_work_rest = request.POST.get('draw_task_work_rest')
    work_rest = _sanitize_task_text((raw_work_rest or '').strip(), multiline=False, max_len=180) if raw_work_rest is not None else str(existing_meta.get('work_rest') or '')
    load_target = _sanitize_task_text((request.POST.get('draw_task_load_target') or '').strip(), multiline=False, max_len=180)
    players_distribution = _sanitize_task_text((request.POST.get('draw_task_players_distribution') or '').strip(), multiline=False, max_len=180)
    progression = _sanitize_task_text((request.POST.get('draw_task_progression') or '').strip(), multiline=True, max_len=500)
    progression_html = _sanitize_task_rich_html((request.POST.get('draw_task_progression_html') or '').strip())
    regression = _sanitize_task_text((request.POST.get('draw_task_regression') or '').strip(), multiline=True, max_len=500)
    regression_html = _sanitize_task_rich_html((request.POST.get('draw_task_regression_html') or '').strip())
    success_criteria = _sanitize_task_text((request.POST.get('draw_task_success_criteria') or '').strip(), multiline=True, max_len=500)
    success_criteria_html = _sanitize_task_rich_html((request.POST.get('draw_task_success_criteria_html') or '').strip())
    selected_surface = (request.POST.get('draw_task_surface') or '').strip()
    selected_pitch_format = (request.POST.get('draw_task_pitch_format') or '').strip()
    raw_selected_phase = request.POST.get('draw_task_game_phase')
    selected_phase = (raw_selected_phase or '').strip() if raw_selected_phase is not None else str(existing_meta.get('game_phase') or '')
    raw_selected_methodology = request.POST.get('draw_task_methodology')
    selected_methodology = (raw_selected_methodology or '').strip() if raw_selected_methodology is not None else str(existing_meta.get('methodology') or '')
    raw_selected_complexity = request.POST.get('draw_task_complexity')
    selected_complexity = (raw_selected_complexity or '').strip() if raw_selected_complexity is not None else str(existing_meta.get('complexity') or '')
    raw_strategy = request.POST.get('draw_task_strategy')
    selected_strategy = (raw_strategy or '').strip() if raw_strategy is not None else str(existing_meta.get('strategy') or '')
    raw_dynamics = request.POST.get('draw_task_dynamics')
    selected_dynamics = (raw_dynamics or '').strip() if raw_dynamics is not None else str(existing_meta.get('dynamics') or '')
    raw_structure = request.POST.get('draw_task_structure')
    selected_structure = (raw_structure or '').strip() if raw_structure is not None else str(existing_meta.get('structure') or '')
    raw_coordination = request.POST.get('draw_task_coordination')
    selected_coordination = (raw_coordination or '').strip() if raw_coordination is not None else str(existing_meta.get('coordination') or '')
    raw_coord_skills = request.POST.get('draw_task_coordination_skills')
    selected_coord_skills = (raw_coord_skills or '').strip() if raw_coord_skills is not None else str(existing_meta.get('coordination_skills') or '')
    raw_tactical_intent = request.POST.get('draw_task_tactical_intent')
    selected_tactical_intent = (raw_tactical_intent or '').strip() if raw_tactical_intent is not None else str(existing_meta.get('tactical_intent') or '')
    raw_template_key = request.POST.get('draw_task_template')
    template_key = (raw_template_key or 'none').strip() if raw_template_key is not None else str(existing_meta.get('template_key') or 'none')
    pitch_preset = (request.POST.get('draw_task_pitch_preset') or 'full_pitch').strip()
    raw_pitch_orientation = request.POST.get('draw_task_pitch_orientation')
    pitch_orientation = (raw_pitch_orientation or '').strip().lower() if raw_pitch_orientation is not None else str(existing_meta.get('pitch_orientation') or 'landscape')
    if 'draw_constraints' in request.POST:
        constraints = [str(v).strip() for v in request.POST.getlist('draw_constraints') if str(v).strip()]
    else:
        constraints = [str(v).strip() for v in (existing_meta.get('constraints') or []) if str(v).strip()]
    series = _sanitize_task_text((request.POST.get('draw_task_series') or '').strip(), multiline=False, max_len=100)
    repetitions = _sanitize_task_text((request.POST.get('draw_task_repetitions') or '').strip(), multiline=False, max_len=100)
    player_count = _sanitize_task_text((request.POST.get('draw_task_player_count') or '').strip(), multiline=False, max_len=100)
    age_group = _sanitize_task_text((request.POST.get('draw_task_age_group') or '').strip(), multiline=False, max_len=100)
    training_type = _sanitize_task_text((request.POST.get('draw_task_training_type') or '').strip(), multiline=False, max_len=120)
    raw_category_tags = request.POST.get('draw_task_category_tags')
    if raw_category_tags is None:
        existing_tags = existing_meta.get('category_tags') or []
        if isinstance(existing_tags, list):
            category_tags = [str(tag).strip() for tag in existing_tags if str(tag).strip()]
        else:
            category_tags = [tag.strip() for tag in str(existing_tags).split(',') if tag.strip()]
    else:
        category_tags_raw = _sanitize_task_text((raw_category_tags or '').strip(), multiline=False, max_len=240)
        category_tags = [tag.strip() for tag in category_tags_raw.split(',') if tag.strip()]
    assigned_player_ids = [
        player_id
        for player_id in (_parse_int(value) for value in request.POST.getlist('assigned_player_ids'))
        if player_id
    ]
    assigned_players = list(
        TaskStudioRosterPlayer.objects
        .filter(owner=owner, id__in=assigned_player_ids, is_active=True)
        .filter(Q(workspace=workspace) | Q(workspace__isnull=True))
        .order_by('number', 'name')
    )
    assigned_player_ids = [int(player.id) for player in assigned_players]

    if not title:
        raise ValueError('Indica un título para la tarea.')
    if block not in {choice[0] for choice in SessionTask.BLOCK_CHOICES}:
        block = SessionTask.BLOCK_MAIN_1
    minutes = max(5, min(minutes, 90))
    if pitch_preset not in {'full_pitch', 'half_pitch', 'attacking_third', 'middle_third', 'defensive_third', 'seven_side', 'seven_side_single', 'futsal', 'blank'}:
        pitch_preset = 'full_pitch'
    if pitch_orientation not in {'landscape', 'portrait'}:
        pitch_orientation = 'landscape'

    valid_surfaces = {key for key, _ in TASK_SURFACE_CHOICES}
    valid_pitch_formats = {key for key, _ in TASK_PITCH_FORMAT_CHOICES}
    valid_phases = {key for key, _ in TASK_GAME_PHASE_CHOICES}
    valid_methodologies = {key for key, _ in TASK_METHODOLOGY_CHOICES}
    valid_complexities = {key for key, _ in TASK_COMPLEXITY_CHOICES}
    valid_strategies = {key for key, _ in TASK_STRATEGY_CHOICES}
    valid_dynamics = {key for key, _ in TASK_DYNAMICS_CHOICES}
    valid_structures = {key for key, _ in TASK_STRUCTURE_CHOICES}
    valid_coordination = {key for key, _ in TASK_COORDINATION_CHOICES}
    valid_coord_skills = {key for key, _ in TASK_COORDINATION_SKILLS_CHOICES}
    valid_tactical_intents = {key for key, _ in TASK_TACTICAL_INTENT_CHOICES}
    valid_constraints = {key for key, _ in TASK_CONSTRAINT_CHOICES}
    constraints = [item for item in constraints if item in valid_constraints]
    if selected_surface not in valid_surfaces:
        selected_surface = ''
    if selected_pitch_format not in valid_pitch_formats:
        selected_pitch_format = ''
    if selected_phase not in valid_phases:
        selected_phase = ''
    if selected_methodology not in valid_methodologies:
        selected_methodology = ''
    if selected_complexity not in valid_complexities:
        selected_complexity = ''
    if selected_strategy not in valid_strategies:
        selected_strategy = ''
    if selected_dynamics not in valid_dynamics:
        selected_dynamics = ''
    if selected_structure not in valid_structures:
        selected_structure = ''
    if selected_coordination not in valid_coordination:
        selected_coordination = ''
    if selected_coord_skills not in valid_coord_skills:
        selected_coord_skills = ''
    if selected_tactical_intent not in valid_tactical_intents:
        selected_tactical_intent = ''

    canvas_state = None
    raw_canvas_state = (request.POST.get('draw_canvas_state') or '').strip()
    if raw_canvas_state:
        try:
            parsed_state = json.loads(raw_canvas_state)
            if isinstance(parsed_state, dict):
                canvas_state = parsed_state
        except Exception:
            canvas_state = None
    if not isinstance(canvas_state, dict):
        canvas_state = _starter_canvas_state(pitch_preset)
    timeline = _normalize_animation_timeline(canvas_state.get('timeline'))

    canvas_width = max(320, min(_parse_int(request.POST.get('draw_canvas_width')) or 1280, 3840))
    canvas_height = max(180, min(_parse_int(request.POST.get('draw_canvas_height')) or 720, 2160))
    tactical_layout = {
        'tokens': canvas_state.get('objects') if isinstance(canvas_state.get('objects'), list) else [],
        'timeline': timeline,
        'meta': {
            'scope': 'task_studio',
            'source': 'task-studio',
            'template_key': template_key,
            'surface': selected_surface,
            'pitch_format': selected_pitch_format,
            'pitch_preset': pitch_preset,
            'pitch_orientation': pitch_orientation,
            'game_phase': selected_phase,
	            'methodology': selected_methodology,
	            'complexity': selected_complexity,
	            'strategy': selected_strategy,
	            'dynamics': selected_dynamics,
	            'structure': selected_structure,
	            'coordination': selected_coordination,
	            'coordination_skills': selected_coord_skills,
	            'tactical_intent': selected_tactical_intent,
	            'space': space,
	            'organization': organization,
	            'organization_html': organization_html,
	            'players_distribution': players_distribution,
            'load_target': load_target,
            'work_rest': work_rest,
            'series': series,
            'repetitions': repetitions,
            'progression': progression,
            'regression': regression,
            'success_criteria': success_criteria,
            'progression_html': progression_html,
            'regression_html': regression_html,
            'success_criteria_html': success_criteria_html,
            'constraints': constraints,
            'player_count': player_count,
            'age_group': age_group,
            'training_type': training_type,
            'category_tags': category_tags,
            'assigned_player_ids': assigned_player_ids,
            'assigned_player_names': [player.name for player in assigned_players],
            'graphic_editor': {
                'canvas_state': canvas_state,
                'canvas_width': canvas_width,
                'canvas_height': canvas_height,
            },
            'analysis': {
                'task_sheet': {
                    'description': description,
                    'description_html': description_html,
                    'coaching_html': coaching_html,
                    'rules_html': rules_html,
                    'players': players,
                    'space': space,
                    'dimensions': dimensions,
                    'materials': materials,
                }
            },
        }
    }
    if existing_task:
        task = existing_task
        task.title = title[:160]
        task.block = block
        task.duration_minutes = minutes
        task.objective = objective[:180]
        task.coaching_points = coaching_points
        task.confrontation_rules = confrontation_rules
        task.tactical_layout = tactical_layout
        task.notes = 'Tarea actualizada en Task Studio'
        task.workspace = workspace
        task.save()
    else:
        task = TaskStudioTask.objects.create(
            workspace=workspace,
            owner=owner,
            title=title[:160],
            block=block,
            duration_minutes=minutes,
            objective=objective[:180],
            coaching_points=coaching_points,
            confrontation_rules=confrontation_rules,
            tactical_layout=tactical_layout,
            notes='Tarea creada en Task Studio',
        )
    preview_data = request.POST.get('draw_canvas_preview_data')
    if preview_data:
        raw_bytes, extension = _decode_canvas_data_url(preview_data)
        if raw_bytes and extension:
            filename = f'task_studio_preview_{task.id}{extension}'
            task.task_preview_image.save(filename, ContentFile(raw_bytes), save=False)
            task.save(update_fields=['task_preview_image'])
    try:
        write_task_backup(
            task,
            kind='task_studio_task',
            reason='save',
            actor_username=(request.user.username if request and getattr(request, 'user', None) and request.user.is_authenticated else ''),
        )
    except Exception:
        pass
    return task


@csrf_exempt
@login_required
def session_task_builder_page(request, scope_key='coach', scope_title='Sesiones · Entrenador', task_id=None):
    if not _can_access_sessions_workspace(request.user):
        return HttpResponse('No tienes permisos para acceder a sesiones.', status=403)
    forbidden = _forbid_if_workspace_module_disabled(request, 'sessions', label='sesiones')
    if forbidden:
        return forbidden
    primary_team = _get_primary_team_for_request(request)
    if not primary_team:
        raise Http404('Equipo principal no configurado')
    task = None
    if task_id:
        task = (
            SessionTask.objects
            .select_related('session__microcycle')
            .filter(id=task_id, session__microcycle__team=primary_team, deleted_at__isnull=True)
            .first()
        )
        if not task:
            raise Http404('Tarea no encontrada')
        if _task_scope_for_item(task) != scope_key:
            return HttpResponse('La tarea no pertenece a este espacio.', status=403)
        if not _is_task_editable(task):
            return HttpResponse('Las tareas importadas son de solo lectura.', status=403)

    feedback = ''
    error = ''
    if request.method == 'POST':
        try:
            task = _save_task_builder_entry(request, primary_team, scope_key, existing_task=task)
            feedback = 'Tarea guardada correctamente.'
        except ValueError as exc:
            error = str(exc)
        except Exception:
            error = 'No se pudo guardar la tarea.'

    initial = _task_builder_initial_values(task)
    all_sessions = list(
        TrainingSession.objects
        .select_related('microcycle')
        .filter(microcycle__team=primary_team)
        .order_by('-session_date', '-id')[:150]
    )
    player_catalog = _build_tactical_player_catalog(request, primary_team)
    available_players = list(
        Player.objects
        .filter(team=primary_team, is_active=True)
        .order_by('number', 'name')[:60]
    )
    return render(
        request,
        'football/task_builder.html',
        {
            'scope_key': scope_key,
            'scope_title': scope_title,
            'scope_route_name': _sessions_scope_route_name(scope_key),
            'task': task,
            'feedback': feedback,
            'error': error,
            'task_blocks': SessionTask.BLOCK_CHOICES,
            'all_sessions': all_sessions,
            'task_surface_choices': TASK_SURFACE_CHOICES,
            'task_pitch_choices': TASK_PITCH_FORMAT_CHOICES,
            'task_complexity_choices': TASK_COMPLEXITY_CHOICES,
            'task_strategy_choices': TASK_STRATEGY_CHOICES,
            'task_coordination_skills_choices': TASK_COORDINATION_SKILLS_CHOICES,
            'task_tactical_intent_choices': TASK_TACTICAL_INTENT_CHOICES,
            'task_dynamics_choices': TASK_DYNAMICS_CHOICES,
            'task_structure_choices': TASK_STRUCTURE_CHOICES,
            'task_coordination_choices': TASK_COORDINATION_CHOICES,
            'tactical_player_catalog': player_catalog,
            'available_players': available_players,
            'initial': initial,
            'back_url': reverse(_sessions_scope_route_name(scope_key)),
            'back_label': 'Volver a sesiones',
            'pdf_preview_url': reverse('sessions-task-pdf-preview'),
            'task_preview_url': (reverse('session-task-preview-file', args=[task.id]) if task and task.task_preview_image else ''),
            'show_session_selector': True,
            'show_dragon_nav': True,
        },
    )


@csrf_exempt
@login_required
@require_POST
def session_task_pdf_preview(request):
    if not _can_access_sessions_workspace(request.user):
        return HttpResponse('No tienes permisos para acceder a sesiones.', status=403)
    forbidden = _forbid_if_workspace_module_disabled(request, 'sessions', label='sesiones')
    if forbidden:
        return forbidden
    primary_team = _get_primary_team_for_request(request)
    if not primary_team:
        raise Http404('Equipo principal no configurado')
    pdf_style = (request.GET.get('style') or 'uefa').strip().lower()
    if pdf_style not in {'uefa', 'club'}:
        pdf_style = 'uefa'
    context = _build_task_draft_pdf_context(request, primary_team, pdf_style=pdf_style)
    html = render_to_string('football/session_task_pdf.html', context)
    filename = slugify(f"borrador-{context['task'].title}") or 'borrador-tarea'
    return _build_pdf_response_or_html_fallback(request, html, filename)


def _task_studio_identity(request, owner):
    profile = _task_studio_profile_for_user(owner)
    team_name = str(profile.club_name or profile.document_name or profile.display_name or owner.get_full_name() or owner.get_username()).strip() or owner.get_username()
    coach_name = str(profile.document_name or profile.display_name or owner.get_full_name() or owner.get_username()).strip() or 'Entrenador'
    crest_url = ''
    if getattr(profile, 'crest_image', None):
        try:
            crest_url = request.build_absolute_uri(profile.crest_image.url)
        except Exception:
            crest_url = ''
    if not crest_url:
        crest_url = request.build_absolute_uri(static('football/images/cdb-logo.png'))
    team_stub = SimpleNamespace(
        name=team_name,
        primary_color=str(profile.primary_color or '#0f7a35').strip() or '#0f7a35',
        secondary_color=str(profile.secondary_color or '#f8fafc').strip() or '#f8fafc',
        accent_color=str(profile.accent_color or '#102734').strip() or '#102734',
    )
    return profile, team_stub, coach_name, crest_url


def _build_task_studio_pdf_context(request, owner, task, tactical_layout, pdf_style='uefa', preview_url=''):
    profile, team_stub, coach_name, crest_url = _task_studio_identity(request, owner)
    today = timezone.localdate()
    session = SimpleNamespace(session_date=today, start_time=None, focus='Task Studio')
    microcycle = SimpleNamespace(title='Repositorio privado', week_start=today, week_end=today)
    context = _build_task_pdf_context(
        request,
        team=team_stub,
        session=session,
        microcycle=microcycle,
        task=task,
        tactical_layout=tactical_layout,
        pdf_style=pdf_style,
        preview_url=preview_url,
    )
    context.update(
        {
            'team_name': team_stub.name,
            'coach_name': coach_name,
            'logo_url': request.build_absolute_uri(static('football/images/uefa-badge.svg')) if pdf_style == 'uefa' else crest_url,
            'pdf_palette': _team_pdf_palette(team_stub, pdf_style),
            'task_studio_profile': profile,
        }
    )
    return context


def _build_task_studio_draft_pdf_context(request, owner, pdf_style='uefa'):
    title = _sanitize_task_text((request.POST.get('draw_task_title') or '').strip(), multiline=False, max_len=160) or 'Tarea sin título'
    objective = _sanitize_task_text((request.POST.get('draw_task_objective') or '').strip(), multiline=False, max_len=180)
    coaching_points = _sanitize_task_text((request.POST.get('draw_task_coaching_points') or '').strip(), multiline=True)
    confrontation_rules = _sanitize_task_text((request.POST.get('draw_task_confrontation_rules') or '').strip(), multiline=True)
    block = (request.POST.get('draw_task_block') or SessionTask.BLOCK_MAIN_1).strip()
    minutes = max(5, min(_parse_int(request.POST.get('draw_task_minutes')) or 15, 90))
    selected_surface = _sanitize_task_text((request.POST.get('draw_task_surface') or '').strip(), multiline=False, max_len=80)
    selected_pitch_format = _sanitize_task_text((request.POST.get('draw_task_pitch_format') or '').strip(), multiline=False, max_len=80)
    selected_phase = _sanitize_task_text((request.POST.get('draw_task_game_phase') or '').strip(), multiline=False, max_len=80)
    selected_methodology = _sanitize_task_text((request.POST.get('draw_task_methodology') or '').strip(), multiline=False, max_len=80)
    selected_complexity = _sanitize_task_text((request.POST.get('draw_task_complexity') or '').strip(), multiline=False, max_len=80)
    selected_strategy = _sanitize_task_text((request.POST.get('draw_task_strategy') or '').strip(), multiline=False, max_len=80)
    selected_dynamics = _sanitize_task_text((request.POST.get('draw_task_dynamics') or '').strip(), multiline=False, max_len=80)
    selected_structure = _sanitize_task_text((request.POST.get('draw_task_structure') or '').strip(), multiline=False, max_len=80)
    selected_coordination = _sanitize_task_text((request.POST.get('draw_task_coordination') or '').strip(), multiline=False, max_len=80)
    selected_coord_skills = _sanitize_task_text((request.POST.get('draw_task_coordination_skills') or '').strip(), multiline=False, max_len=80)
    selected_tactical_intent = _sanitize_task_text((request.POST.get('draw_task_tactical_intent') or '').strip(), multiline=False, max_len=80)
    space = _sanitize_task_text((request.POST.get('draw_task_space') or '').strip(), multiline=False, max_len=120)
    organization = _sanitize_task_text((request.POST.get('draw_task_organization') or '').strip(), multiline=True, max_len=500)
    players_distribution = _sanitize_task_text((request.POST.get('draw_task_players_distribution') or '').strip(), multiline=False, max_len=180)
    load_target = _sanitize_task_text((request.POST.get('draw_task_load_target') or '').strip(), multiline=False, max_len=180)
    work_rest = _sanitize_task_text((request.POST.get('draw_task_work_rest') or '').strip(), multiline=False, max_len=180)
    series = _sanitize_task_text((request.POST.get('draw_task_series') or '').strip(), multiline=False, max_len=100)
    repetitions = _sanitize_task_text((request.POST.get('draw_task_repetitions') or '').strip(), multiline=False, max_len=100)
    player_count = _sanitize_task_text((request.POST.get('draw_task_player_count') or '').strip(), multiline=False, max_len=100)
    age_group = _sanitize_task_text((request.POST.get('draw_task_age_group') or '').strip(), multiline=False, max_len=100)
    training_type = _sanitize_task_text((request.POST.get('draw_task_training_type') or '').strip(), multiline=False, max_len=120)
    dimensions = _sanitize_task_text((request.POST.get('draw_task_dimensions') or '').strip(), multiline=False, max_len=120)
    materials = _sanitize_task_text((request.POST.get('draw_task_materials') or '').strip(), multiline=False, max_len=180)
    progression = _sanitize_task_text((request.POST.get('draw_task_progression') or '').strip(), multiline=True, max_len=500)
    regression = _sanitize_task_text((request.POST.get('draw_task_regression') or '').strip(), multiline=True, max_len=500)
    success_criteria = _sanitize_task_text((request.POST.get('draw_task_success_criteria') or '').strip(), multiline=True, max_len=500)
    organization_html = _sanitize_task_rich_html((request.POST.get('draw_task_organization_html') or '').strip())
    progression_html = _sanitize_task_rich_html((request.POST.get('draw_task_progression_html') or '').strip())
    regression_html = _sanitize_task_rich_html((request.POST.get('draw_task_regression_html') or '').strip())
    success_criteria_html = _sanitize_task_rich_html((request.POST.get('draw_task_success_criteria_html') or '').strip())
    category_tags_raw = _sanitize_task_text((request.POST.get('draw_task_category_tags') or '').strip(), multiline=False, max_len=240)
    category_tags = [tag.strip() for tag in category_tags_raw.split(',') if tag.strip()]
    assigned_player_ids = [
        player_id
        for player_id in (_parse_int(value) for value in request.POST.getlist('assigned_player_ids'))
        if player_id
    ]
    assigned_player_names = list(
        TaskStudioRosterPlayer.objects
        .filter(owner=owner, id__in=assigned_player_ids, is_active=True)
        .order_by('number', 'name')
        .values_list('name', flat=True)
    )
    today = timezone.localdate()
    session = SimpleNamespace(session_date=today, start_time=None, focus='Borrador')
    microcycle = SimpleNamespace(title='Task Studio', week_start=today, week_end=today)
    canvas_state = {}
    raw_canvas_state = (request.POST.get('draw_canvas_state') or '').strip()
    if raw_canvas_state:
        try:
            parsed = json.loads(raw_canvas_state)
            if isinstance(parsed, dict):
                canvas_state = parsed
        except Exception:
            canvas_state = {}
    tactical_layout = {
        'tokens': canvas_state.get('objects') if isinstance(canvas_state.get('objects'), list) else [],
        'timeline': canvas_state.get('timeline') if isinstance(canvas_state.get('timeline'), list) else [],
        'meta': {
            'surface': selected_surface,
            'pitch_format': selected_pitch_format,
            'game_phase': selected_phase,
            'methodology': selected_methodology,
            'complexity': selected_complexity,
            'strategy': selected_strategy,
            'dynamics': selected_dynamics,
            'structure': selected_structure,
            'coordination': selected_coordination,
            'coordination_skills': selected_coord_skills,
            'tactical_intent': selected_tactical_intent,
            'space': space,
            'organization': organization,
            'organization_html': organization_html,
            'players_distribution': players_distribution,
            'load_target': load_target,
            'work_rest': work_rest,
            'series': series,
            'repetitions': repetitions,
            'player_count': player_count,
            'age_group': age_group,
            'training_type': training_type,
            'category_tags': category_tags,
            'assigned_player_names': assigned_player_names,
            'progression': progression,
            'progression_html': progression_html,
            'regression': regression,
            'regression_html': regression_html,
            'success_criteria': success_criteria,
            'success_criteria_html': success_criteria_html,
            'analysis': {
                'task_sheet': {
                    'description': _sanitize_task_text((request.POST.get('draw_task_description') or '').strip(), multiline=True),
                    'description_html': _sanitize_task_rich_html((request.POST.get('draw_task_description_html') or '').strip()),
                    'coaching_html': _sanitize_task_rich_html((request.POST.get('draw_task_coaching_points_html') or '').strip()),
                    'rules_html': _sanitize_task_rich_html((request.POST.get('draw_task_confrontation_rules_html') or '').strip()),
                    'dimensions': dimensions,
                    'materials': materials,
                }
            },
        },
    }
    draft_task = SimpleNamespace(
        id=0,
        title=title,
        duration_minutes=minutes,
        objective=objective,
        coaching_points=coaching_points,
        confrontation_rules=confrontation_rules,
        block=block,
        get_block_display=lambda: dict(SessionTask.BLOCK_CHOICES).get(block, block),
    )
    preview_data = str(request.POST.get('draw_canvas_preview_data') or '').strip()
    return _build_task_studio_pdf_context(
        request,
        owner=owner,
        task=draft_task,
        tactical_layout=tactical_layout,
        pdf_style=pdf_style,
        preview_url=preview_data,
    )


def _task_studio_task_for_request(request, task_id):
    task = TaskStudioTask.objects.select_related('owner').filter(id=task_id).first()
    if not task:
        return None
    if getattr(task, 'deleted_at', None):
        return None
    if _is_admin_user(request.user):
        return task
    if int(task.owner_id) != int(request.user.id):
        return None
    return task


def _clone_task_studio_task(source_task):
    clone_title = str(source_task.title or 'Tarea').strip() or 'Tarea'
    suffix = ' (copia)'
    if len(clone_title) + len(suffix) > 160:
        clone_title = clone_title[: 160 - len(suffix)].rstrip()
    clone = TaskStudioTask.objects.create(
        workspace=source_task.workspace,
        owner=source_task.owner,
        title=f'{clone_title}{suffix}',
        block=source_task.block,
        duration_minutes=source_task.duration_minutes,
        objective=source_task.objective,
        coaching_points=source_task.coaching_points,
        confrontation_rules=source_task.confrontation_rules,
        tactical_layout=copy.deepcopy(source_task.tactical_layout) if isinstance(source_task.tactical_layout, dict) else {},
        task_pdf=(source_task.task_pdf.name if source_task.task_pdf else None),
        task_preview_image=(source_task.task_preview_image.name if source_task.task_preview_image else None),
        notes=source_task.notes,
    )
    return clone


@login_required
@ensure_csrf_cookie
def task_studio_home_page(request):
    forbidden = _forbid_if_no_task_studio_access(request.user)
    if forbidden:
        return forbidden
    target_user = _task_studio_target_user(request)
    forbidden = _forbid_if_task_studio_module_disabled(request, target_user, 'task_studio_home', label='inicio Task Studio')
    if forbidden:
        return forbidden
    feedback = str(request.session.pop('task_studio_feedback', '') or '')
    browse_all = bool(_is_admin_user(request.user) and not request.GET.get('user'))
    profile = _task_studio_profile_for_user(target_user)
    task_qs = TaskStudioTask.objects.select_related('owner').filter(deleted_at__isnull=True)
    if not browse_all:
        task_qs = task_qs.filter(owner=target_user)
    search_query = _sanitize_task_text((request.GET.get('q') or '').strip(), multiline=False, max_len=120)
    if search_query:
        task_qs = task_qs.filter(
            Q(title__icontains=search_query)
            | Q(objective__icontains=search_query)
            | Q(notes__icontains=search_query)
        )
    tasks = list(task_qs.order_by('-updated_at', '-id')[:80])
    recent_document_tasks = tasks[:6]
    deleted_task_qs = TaskStudioTask.objects.select_related('owner').filter(deleted_at__isnull=False)
    if not browse_all:
        deleted_task_qs = deleted_task_qs.filter(owner=target_user)
    deleted_tasks = list(deleted_task_qs.order_by('-deleted_at', '-id')[:24])
    roster_count = TaskStudioRosterPlayer.objects.filter(owner=target_user, is_active=True).count()
    query_suffix = _task_studio_query_suffix(target_user, request.user)
    task_count = task_qs.count()
    has_identity = bool(
        (profile.club_name or '').strip()
        or (profile.document_name or '').strip()
        or profile.crest_image
    )
    has_profile_basics = bool(
        (profile.display_name or '').strip()
        or (profile.license_name or '').strip()
        or (profile.phone or '').strip()
        or (profile.category_label or '').strip()
    )
    has_roster = roster_count > 0
    has_tasks = task_count > 0
    club_workspaces = []
    club_primary = None
    try:
        # Links de "modo club" para usuarios Task Studio: muestran módulos de partido (convocatoria/11/acciones)
        # en el contexto del club activo (monocliente) o de clubs donde tengan membresía.
        club_qs = _available_workspaces_for_user(request.user).filter(kind=Workspace.KIND_CLUB).select_related('primary_team')
        club_workspaces = list(club_qs.order_by('name', 'id')[:6])
        club_primary = club_workspaces[0] if club_workspaces else None
    except Exception:
        club_workspaces = []
        club_primary = None
    onboarding_steps = [
        {
            'title': 'Perfil e identidad',
            'description': 'Define nombre documental, escudo, colores y firma.',
            'done': has_identity or has_profile_basics,
            'url': f"{reverse('task-studio-profile')}{query_suffix}",
            'action': 'Configurar',
        },
        {
            'title': 'Plantilla privada',
            'description': 'Carga jugadores para usarlos como chapas reales en la pizarra.',
            'done': has_roster,
            'url': f"{reverse('task-studio-roster')}{query_suffix}",
            'action': 'Cargar plantilla',
        },
        {
            'title': 'Primera tarea',
            'description': 'Diseña, guarda e imprime tu primera tarea de entrenamiento.',
            'done': has_tasks,
            'url': f"{reverse('task-studio-task-create')}{query_suffix}",
            'action': 'Crear tarea',
        },
    ]
    next_actions = [step for step in onboarding_steps if not step['done']]
    return render(
        request,
        'football/task_studio_home.html',
        {
            'target_user': target_user,
            'profile': profile,
            'tasks': tasks,
            'task_count': task_count,
            'roster_count': roster_count,
            'browse_all': browse_all,
            'query_suffix': query_suffix,
            'feedback': feedback,
            'onboarding_steps': onboarding_steps,
            'next_actions': next_actions,
            'has_identity': has_identity,
            'search_query': search_query,
            'recent_document_tasks': recent_document_tasks,
            'deleted_tasks': deleted_tasks,
            'club_workspaces': club_workspaces,
            'club_primary': club_primary,
        },
    )


@login_required
def task_studio_profile_page(request):
    forbidden = _forbid_if_no_task_studio_access(request.user)
    if forbidden:
        return forbidden
    target_user = _task_studio_target_user(request)
    forbidden = _forbid_if_task_studio_module_disabled(request, target_user, 'task_studio_profile', label='perfil Task Studio')
    if forbidden:
        return forbidden
    profile = _task_studio_profile_for_user(target_user)
    feedback = ''
    error = ''

    def _clean_color(raw, fallback):
        value = str(raw or '').strip()
        return value if re.fullmatch(r'#[0-9a-fA-F]{6}', value) else fallback

    if request.method == 'POST':
        try:
            profile.display_name = _sanitize_task_text((request.POST.get('display_name') or '').strip(), multiline=False, max_len=140)
            profile.phone = _sanitize_task_text((request.POST.get('phone') or '').strip(), multiline=False, max_len=40)
            profile.license_name = _sanitize_task_text((request.POST.get('license_name') or '').strip(), multiline=False, max_len=120)
            profile.club_name = _sanitize_task_text((request.POST.get('club_name') or '').strip(), multiline=False, max_len=140)
            profile.category_label = _sanitize_task_text((request.POST.get('category_label') or '').strip(), multiline=False, max_len=120)
            profile.city = _sanitize_task_text((request.POST.get('city') or '').strip(), multiline=False, max_len=120)
            profile.document_name = _sanitize_task_text((request.POST.get('document_name') or '').strip(), multiline=False, max_len=140)
            profile.document_footer = _sanitize_task_text((request.POST.get('document_footer') or '').strip(), multiline=False, max_len=180)
            profile.signature = _sanitize_task_text((request.POST.get('signature') or '').strip(), multiline=False, max_len=140)
            profile.primary_color = _clean_color(request.POST.get('primary_color'), profile.primary_color or '#0f7a35')
            profile.secondary_color = _clean_color(request.POST.get('secondary_color'), profile.secondary_color or '#f8fafc')
            profile.accent_color = _clean_color(request.POST.get('accent_color'), profile.accent_color or '#102734')
            uploaded_crest = request.FILES.get('crest_image')
            if uploaded_crest:
                profile.crest_image = uploaded_crest
            profile.save()
            feedback = 'Perfil e identidad guardados.'
        except Exception:
            error = 'No se pudo guardar la configuración.'

    return render(
        request,
        'football/task_studio_profile.html',
        {
            'target_user': target_user,
            'profile': profile,
            'feedback': feedback,
            'error': error,
            'query_suffix': _task_studio_query_suffix(target_user, request.user),
        },
    )


@login_required
def task_studio_roster_page(request):
    forbidden = _forbid_if_no_task_studio_access(request.user)
    if forbidden:
        return forbidden
    target_user = _task_studio_target_user(request)
    forbidden = _forbid_if_task_studio_module_disabled(request, target_user, 'task_studio_roster', label='plantilla Task Studio')
    if forbidden:
        return forbidden
    feedback = ''
    error = ''
    if request.method == 'POST':
        form_action = (request.POST.get('studio_action') or 'add').strip().lower()
        player_id = _parse_int(request.POST.get('player_id'))
        roster_player = TaskStudioRosterPlayer.objects.filter(owner=target_user, id=player_id).first() if player_id else None
        try:
            if form_action == 'delete':
                if not roster_player:
                    raise ValueError('Jugador no encontrado.')
                roster_player.delete()
                feedback = 'Jugador eliminado de la plantilla.'
            else:
                name = _sanitize_task_text((request.POST.get('name') or '').strip(), multiline=False, max_len=120)
                if not name:
                    raise ValueError('Indica un nombre para el jugador.')
                defaults = {
                    'name': name,
                    'number': _parse_int(request.POST.get('number')) or None,
                    'position': _sanitize_task_text((request.POST.get('position') or '').strip(), multiline=False, max_len=60),
                    'dominant_foot': _sanitize_task_text((request.POST.get('dominant_foot') or '').strip(), multiline=False, max_len=24),
                    'birth_year': _parse_int(request.POST.get('birth_year')) or None,
                    'notes': _sanitize_task_text((request.POST.get('notes') or '').strip(), multiline=True, max_len=500),
                    'is_active': True,
                }
                if roster_player:
                    for field_name, field_value in defaults.items():
                        setattr(roster_player, field_name, field_value)
                    uploaded_photo = request.FILES.get('photo')
                    if uploaded_photo:
                        roster_player.photo = uploaded_photo
                    roster_player.save()
                    feedback = 'Jugador actualizado.'
                else:
                    roster_player = TaskStudioRosterPlayer(owner=target_user, **defaults)
                    uploaded_photo = request.FILES.get('photo')
                    if uploaded_photo:
                        roster_player.photo = uploaded_photo
                    roster_player.save()
                    feedback = 'Jugador añadido a la plantilla.'
        except ValueError as exc:
            error = str(exc)
        except Exception:
            error = 'No se pudo guardar la plantilla.'
    players = list(TaskStudioRosterPlayer.objects.filter(owner=target_user).order_by('number', 'name', 'id'))
    return render(
        request,
        'football/task_studio_roster.html',
        {
            'target_user': target_user,
            'players': players,
            'feedback': feedback,
            'error': error,
            'query_suffix': _task_studio_query_suffix(target_user, request.user),
        },
    )


@csrf_exempt
@login_required
def task_studio_task_builder_page(request, task_id=None):
    forbidden = _forbid_if_no_task_studio_access(request.user)
    if forbidden:
        return forbidden
    target_user = _task_studio_target_user(request)
    forbidden = _forbid_if_task_studio_module_disabled(request, target_user, 'task_studio_tasks', label='tareas Task Studio')
    if forbidden:
        return forbidden
    task = None
    if task_id:
        task = _task_studio_task_for_request(request, task_id)
        if not task:
            raise Http404('Tarea no encontrada')
        target_user = task.owner
    feedback = ''
    error = ''
    if request.method == 'POST':
        try:
            task = _save_task_studio_entry(request, target_user, existing_task=task)
            feedback = 'Tarea guardada correctamente.'
        except ValueError as exc:
            error = str(exc)
        except Exception:
            error = 'No se pudo guardar la tarea.'
    initial = _task_builder_initial_values(task)
    player_catalog = _build_task_studio_player_catalog(request, target_user)
    can_export_to_club = bool(task and _is_admin_user(request.user))
    club_sessions = []
    if can_export_to_club:
        sessions = list(
            TrainingSession.objects
            .select_related('microcycle__team')
            .order_by('-session_date', '-id')[:220]
        )
        team_ids = [int(getattr(getattr(session, 'microcycle', None), 'team_id', 0) or 0) for session in sessions]
        workspace_map = {
            int(item.primary_team_id): item
            for item in Workspace.objects.filter(kind=Workspace.KIND_CLUB, primary_team_id__in=[tid for tid in team_ids if tid])
        }
        for session in sessions:
            team = getattr(getattr(session, 'microcycle', None), 'team', None)
            if not team:
                continue
            ws = workspace_map.get(int(team.id))
            if not ws:
                continue
            date_label = session.session_date.strftime('%d/%m/%Y') if session.session_date else ''
            focus = str(session.focus or '').strip()
            club_sessions.append(
                {
                    'id': session.id,
                    'label': f'{ws.name} · {date_label} · {focus}',
                }
            )
    available_players = list(
        TaskStudioRosterPlayer.objects
        .filter(owner=target_user, is_active=True)
        .order_by('number', 'name')[:60]
    )
    query_suffix = _task_studio_query_suffix(target_user, request.user)
    return render(
        request,
        'football/task_builder.html',
        {
            'scope_key': 'task_studio',
            'scope_title': 'Task Studio',
            'scope_route_name': 'task-studio-home',
            'task': task,
            'feedback': feedback,
            'error': error,
            'task_blocks': SessionTask.BLOCK_CHOICES,
            'all_sessions': [],
            'can_export_to_club': can_export_to_club,
            'club_sessions': club_sessions,
            'task_surface_choices': TASK_SURFACE_CHOICES,
            'task_pitch_choices': TASK_PITCH_FORMAT_CHOICES,
            'task_complexity_choices': TASK_COMPLEXITY_CHOICES,
            'task_strategy_choices': TASK_STRATEGY_CHOICES,
            'task_coordination_skills_choices': TASK_COORDINATION_SKILLS_CHOICES,
            'task_tactical_intent_choices': TASK_TACTICAL_INTENT_CHOICES,
            'task_dynamics_choices': TASK_DYNAMICS_CHOICES,
            'task_structure_choices': TASK_STRUCTURE_CHOICES,
            'task_coordination_choices': TASK_COORDINATION_CHOICES,
            'tactical_player_catalog': player_catalog,
            'available_players': available_players,
            'initial': initial,
            'back_url': reverse('task-studio-home') + query_suffix,
            'back_label': 'Volver al estudio',
            'pdf_preview_url': reverse('task-studio-task-pdf-preview') + query_suffix,
            'task_preview_url': (reverse('task-studio-task-preview-file', args=[task.id]) if task and task.task_preview_image else ''),
            'show_session_selector': False,
            'show_dragon_nav': False,
        },
    )


@login_required
@require_POST
def task_studio_task_delete_page(request, task_id):
    forbidden = _forbid_if_no_task_studio_access(request.user)
    if forbidden:
        return forbidden
    task = _task_studio_task_for_request(request, task_id)
    if not task:
        raise Http404('Tarea no encontrada')
    forbidden = _forbid_if_task_studio_module_disabled(request, task.owner, 'task_studio_tasks', label='tareas Task Studio')
    if forbidden:
        return forbidden
    owner = task.owner
    task_title = task.title
    try:
        write_task_backup(
            task,
            kind='task_studio_task',
            reason='delete',
            actor_username=(request.user.username if getattr(request, 'user', None) and request.user.is_authenticated else ''),
        )
    except Exception:
        pass
    task.deleted_at = timezone.localtime()
    task.deleted_by = request.user if getattr(request, 'user', None) and request.user.is_authenticated else None
    task.save(update_fields=['deleted_at', 'deleted_by'])
    request.session['task_studio_feedback'] = f'Tarea enviada a papelera: {task_title}.'
    return redirect(reverse('task-studio-home') + _task_studio_query_suffix(owner, request.user))


@login_required
@require_POST
def task_studio_task_restore_page(request, task_id):
    forbidden = _forbid_if_no_task_studio_access(request.user)
    if forbidden:
        return forbidden
    task = TaskStudioTask.objects.select_related('owner').filter(id=task_id, deleted_at__isnull=False).first()
    if not task:
        raise Http404('Tarea no encontrada')
    forbidden = _forbid_if_task_studio_module_disabled(request, task.owner, 'task_studio_tasks', label='tareas Task Studio')
    if forbidden:
        return forbidden
    if not _is_admin_user(request.user) and int(task.owner_id) != int(request.user.id):
        return HttpResponse('No tienes permisos para restaurar esta tarea.', status=403)
    owner = task.owner
    task_title = task.title
    task.deleted_at = None
    task.deleted_by = None
    task.save(update_fields=['deleted_at', 'deleted_by'])
    request.session['task_studio_feedback'] = f'Tarea restaurada: {task_title}.'
    return redirect(reverse('task-studio-home') + _task_studio_query_suffix(owner, request.user))


@login_required
@require_POST
def task_studio_task_duplicate_page(request, task_id):
    forbidden = _forbid_if_no_task_studio_access(request.user)
    if forbidden:
        return forbidden
    task = _task_studio_task_for_request(request, task_id)
    if not task:
        raise Http404('Tarea no encontrada')
    forbidden = _forbid_if_task_studio_module_disabled(request, task.owner, 'task_studio_tasks', label='tareas Task Studio')
    if forbidden:
        return forbidden
    owner = task.owner
    clone = _clone_task_studio_task(task)
    request.session['task_studio_feedback'] = f'Tarea duplicada: {clone.title}.'
    return redirect(reverse('task-studio-task-edit', args=[clone.id]) + _task_studio_query_suffix(owner, request.user))


@login_required
@require_POST
def task_studio_task_export_to_session(request, task_id):
    forbidden = _forbid_if_no_task_studio_access(request.user)
    if forbidden:
        return forbidden
    if not _is_admin_user(request.user):
        return HttpResponse('No tienes permisos para exportar tareas a un cliente.', status=403)
    task = _task_studio_task_for_request(request, task_id)
    if not task:
        raise Http404('Tarea no encontrada')
    session_id = _parse_int(request.POST.get('target_session_id'))
    target_session = (
        TrainingSession.objects
        .select_related('microcycle__team')
        .filter(id=session_id)
        .first()
    )
    if not target_session:
        return HttpResponse('Sesión no encontrada.', status=404)
    team = getattr(getattr(target_session, 'microcycle', None), 'team', None)
    workspace = Workspace.objects.filter(kind=Workspace.KIND_CLUB, primary_team=team, is_active=True).first() if team else None
    if not workspace:
        return HttpResponse('No se encontró el cliente club asociado a esa sesión.', status=400)
    order = (SessionTask.objects.filter(session=target_session).aggregate(Max('order')).get('order__max') or 0) + 1
    new_task = SessionTask.objects.create(
        session=target_session,
        title=str(task.title or 'Tarea')[:160],
        block=task.block,
        duration_minutes=int(task.duration_minutes or 15),
        objective=str(task.objective or '')[:180],
        coaching_points=str(task.coaching_points or ''),
        confrontation_rules=str(task.confrontation_rules or ''),
        tactical_layout=copy.deepcopy(task.tactical_layout) if isinstance(task.tactical_layout, dict) else {},
        status=SessionTask.STATUS_PLANNED,
        order=order,
        notes=f'Importada desde Task Studio ({task.owner.username})',
    )
    request.session['active_workspace_id'] = workspace.id
    return redirect(reverse('sessions-task-edit', args=[new_task.id]))


@csrf_exempt
@login_required
@require_POST
def task_studio_task_pdf_preview(request):
    forbidden = _forbid_if_no_task_studio_access(request.user)
    if forbidden:
        return forbidden
    owner = _task_studio_target_user(request)
    forbidden = _forbid_if_task_studio_module_disabled(request, owner, 'task_studio_pdfs', label='PDF Task Studio')
    if forbidden:
        return forbidden
    pdf_style = (request.GET.get('style') or 'uefa').strip().lower()
    if pdf_style not in {'uefa', 'club'}:
        pdf_style = 'uefa'
    context = _build_task_studio_draft_pdf_context(request, owner, pdf_style=pdf_style)
    html = render_to_string('football/session_task_pdf.html', context)
    filename = slugify(f"task-studio-{context['task'].title}") or 'task-studio-tarea'
    return _build_pdf_response_or_html_fallback(request, html, filename)


@login_required
def task_studio_task_preview_file(request, task_id):
    forbidden = _forbid_if_no_task_studio_access(request.user)
    if forbidden:
        return forbidden
    task = _task_studio_task_for_request(request, task_id)
    if not task or not task.task_preview_image:
        raise Http404('Imagen de tarea no disponible')
    forbidden = _forbid_if_task_studio_module_disabled(request, task.owner, 'task_studio_tasks', label='tareas Task Studio')
    if forbidden:
        return forbidden
    file_field = task.task_preview_image
    try:
        file_field.open('rb')
    except Exception:
        return HttpResponse('No se pudo abrir la imagen de la tarea.', status=500)
    extension = Path(file_field.name).suffix.lower()
    content_type = {
        '.png': 'image/png',
        '.jpg': 'image/jpeg',
        '.jpeg': 'image/jpeg',
        '.webp': 'image/webp',
        '.gif': 'image/gif',
    }.get(extension, 'application/octet-stream')
    response = FileResponse(file_field, content_type=content_type)
    response['Content-Disposition'] = f'inline; filename="{Path(file_field.name).name}"'
    return response


@login_required
def task_studio_task_pdf(request, task_id):
    forbidden = _forbid_if_no_task_studio_access(request.user)
    if forbidden:
        return forbidden
    task = _task_studio_task_for_request(request, task_id)
    if not task:
        raise Http404('Tarea no encontrada')
    forbidden = _forbid_if_task_studio_module_disabled(request, task.owner, 'task_studio_pdfs', label='PDF Task Studio')
    if forbidden:
        return forbidden
    pdf_style = (request.GET.get('style') or 'uefa').strip().lower()
    if pdf_style not in {'uefa', 'club'}:
        pdf_style = 'uefa'
    context = _build_task_studio_pdf_context(
        request,
        owner=task.owner,
        task=task,
        tactical_layout=task.tactical_layout if isinstance(task.tactical_layout, dict) else {},
        pdf_style=pdf_style,
        preview_url=request.build_absolute_uri(reverse('task-studio-task-preview-file', args=[task.id])) if task.task_preview_image else '',
    )
    html = render_to_string('football/session_task_pdf.html', context)
    filename = slugify(f'task-studio-{task.title}') or f'task-studio-{task.id}'
    return _build_pdf_response_or_html_fallback(request, html, filename)


@login_required
@ensure_csrf_cookie
def sessions_page(request):
    return _sessions_workspace_page(request, scope_key='coach', scope_title='Sesiones · Entrenador')


@csrf_exempt
@login_required
def sessions_task_create_page(request):
    return session_task_builder_page(request, scope_key='coach', scope_title='Sesiones · Entrenador')


@csrf_exempt
@login_required
def sessions_task_edit_page(request, task_id):
    return session_task_builder_page(request, scope_key='coach', scope_title='Sesiones · Entrenador', task_id=task_id)


@login_required
@ensure_csrf_cookie
def sessions_goalkeeper_page(request):
    return _sessions_workspace_page(request, scope_key='goalkeeper', scope_title='Sesiones · Porteros')


@csrf_exempt
@login_required
def sessions_goalkeeper_task_create_page(request):
    return session_task_builder_page(request, scope_key='goalkeeper', scope_title='Sesiones · Porteros')


@csrf_exempt
@login_required
def sessions_goalkeeper_task_edit_page(request, task_id):
    return session_task_builder_page(request, scope_key='goalkeeper', scope_title='Sesiones · Porteros', task_id=task_id)


@login_required
@ensure_csrf_cookie
def sessions_fitness_page(request):
    return _sessions_workspace_page(request, scope_key='fitness', scope_title='Sesiones · Preparacion fisica')


@csrf_exempt
@login_required
def sessions_fitness_task_create_page(request):
    return session_task_builder_page(request, scope_key='fitness', scope_title='Sesiones · Preparacion fisica')


@csrf_exempt
@login_required
def sessions_fitness_task_edit_page(request, task_id):
    return session_task_builder_page(request, scope_key='fitness', scope_title='Sesiones · Preparacion fisica', task_id=task_id)


@login_required
def session_task_detail_page(request, task_id):
    if not _can_access_sessions_workspace(request.user):
        return HttpResponse('No tienes permisos para acceder a sesiones.', status=403)
    task = (
        SessionTask.objects
        .select_related('session__microcycle__team')
        .filter(id=task_id)
        .first()
    )
    if not task:
        raise Http404('Tarea no encontrada')

    scope_key = _task_scope_for_item(task)
    scope_route_name = {
        'coach': 'sessions',
        'goalkeeper': 'sessions-goalkeeper',
        'fitness': 'sessions-fitness',
    }.get(scope_key, 'sessions')
    scope_title = {
        'coach': 'Sesiones · Entrenador',
        'goalkeeper': 'Sesiones · Porteros',
        'fitness': 'Sesiones · Preparacion fisica',
    }.get(scope_key, 'Sesiones')

    feedback = ''
    error = ''
    is_editable_task = _is_task_editable(task)
    is_imported_task = _is_imported_task(task)

    # Evita confusión: para tareas editables, "detalle" abre el mismo editor visual que se usa al crear.
    # La ficha legacy queda disponible con ?legacy=1 (útil sobre todo para tareas importadas).
    if request.method == 'GET' and is_editable_task and not (request.GET.get('legacy') or '').strip():
        return redirect(reverse(_task_builder_edit_route_name(scope_key), args=[task.id]))

    if request.method == 'POST':
        detail_action = (request.POST.get('detail_action') or '').strip()
        try:
            if detail_action == 'update_task_detail':
                if not is_editable_task:
                    raise ValueError('Las tareas importadas son de solo lectura.')
                _update_library_task_from_post(task, request.POST, scope_key=scope_key)
                feedback = 'Tarea actualizada correctamente.'
                task.refresh_from_db()
            elif detail_action == 'restore_original_version':
                if not is_editable_task:
                    raise ValueError('Las tareas importadas son de solo lectura.')
                _restore_task_from_original_snapshot(task, scope_key=scope_key)
                feedback = 'Se restauró la versión original de la tarea.'
                task.refresh_from_db()
            else:
                error = 'Acción no reconocida.'
        except ValueError as exc:
            error = str(exc)
        except Exception:
            error = 'No se pudo guardar la tarea.'

    layout = task.tactical_layout if isinstance(task.tactical_layout, dict) else {}
    meta = layout.get('meta') if isinstance(layout.get('meta'), dict) else {}
    analysis_meta = meta.get('analysis') if isinstance(meta.get('analysis'), dict) else {}
    task_sheet = analysis_meta.get('task_sheet') if isinstance(analysis_meta.get('task_sheet'), dict) else {}
    original_version = meta.get('original_version') if isinstance(meta.get('original_version'), dict) else {}
    original_task_sheet = original_version.get('task_sheet') if isinstance(original_version.get('task_sheet'), dict) else {}
    original_preview_url = _storage_url_or_empty(original_version.get('task_preview_image') if isinstance(original_version, dict) else '')
    pdf_excerpt = str(meta.get('pdf_segment_excerpt') or meta.get('extracted_text_excerpt') or '').strip()
    detected_materials = analysis_meta.get('detected_materials') if isinstance(analysis_meta.get('detected_materials'), list) else []
    animation_frames = _normalize_animation_timeline(layout.get('timeline') if isinstance(layout, dict) else [])
    graphic_editor_state = meta.get('graphic_editor', {}) if isinstance(meta.get('graphic_editor'), dict) else {}
    if isinstance(graphic_editor_state, dict) and animation_frames:
        graphic_editor_state = dict(graphic_editor_state)
        graphic_editor_state['timeline'] = animation_frames
    if task.task_pdf and not task.task_preview_image:
        _ensure_task_preview_image(task)
    return render(
        request,
        'football/session_task_detail.html',
        {
            'task': task,
            'scope_key': scope_key,
            'scope_title': scope_title,
            'scope_route_name': scope_route_name,
            'analysis_meta': analysis_meta,
            'task_sheet': task_sheet,
            'pdf_excerpt': pdf_excerpt,
            'detected_materials': detected_materials,
            'feedback': feedback,
            'error': error,
            'task_blocks': SessionTask.BLOCK_CHOICES,
            'graphic_editor_state_json': json.dumps(graphic_editor_state, ensure_ascii=False),
            'animation_frames': animation_frames,
            'animation_frames_json': json.dumps(animation_frames, ensure_ascii=False),
            'original_version': original_version,
            'original_task_sheet': original_task_sheet,
            'original_preview_url': original_preview_url,
            'is_editable_task': is_editable_task,
            'is_imported_task': is_imported_task,
        },
    )


@authenticated_write
@require_POST
def save_session_task_graphic(request, task_id):
    if not _can_access_sessions_workspace(request.user):
        return JsonResponse({'error': 'No tienes permisos para acceder a sesiones.'}, status=403)
    task = (
        SessionTask.objects
        .select_related('session__microcycle__team')
        .filter(id=task_id)
        .first()
    )
    if not task:
        return JsonResponse({'error': 'Tarea no encontrada.'}, status=404)
    if not _is_task_editable(task):
        return JsonResponse({'error': 'Las tareas importadas son de solo lectura.'}, status=403)
    payload = {}
    try:
        payload = json.loads(request.body.decode('utf-8')) if request.body else {}
    except Exception:
        return JsonResponse({'error': 'Payload inválido.'}, status=400)

    canvas_state = payload.get('canvas_state')
    if not isinstance(canvas_state, dict):
        return JsonResponse({'error': 'Estado gráfico inválido.'}, status=400)
    canvas_width = _parse_int(payload.get('canvas_width'))
    canvas_height = _parse_int(payload.get('canvas_height'))

    _ensure_original_task_snapshot(task)

    preview_data = payload.get('preview_data')
    layout = task.tactical_layout if isinstance(task.tactical_layout, dict) else {}
    layout = dict(layout)
    meta = layout.get('meta') if isinstance(layout.get('meta'), dict) else {}
    meta = dict(meta)
    graphic_editor = meta.get('graphic_editor') if isinstance(meta.get('graphic_editor'), dict) else {}
    graphic_editor = dict(graphic_editor)
    graphic_editor.update(
        {
            'canvas_state': canvas_state,
            'canvas_width': canvas_width if canvas_width and canvas_width > 0 else None,
            'canvas_height': canvas_height if canvas_height and canvas_height > 0 else None,
            'updated_at': timezone.now().isoformat(),
            'updated_by': request.user.get_username() if request.user.is_authenticated else '',
        }
    )
    meta['graphic_editor'] = graphic_editor
    layout['meta'] = meta
    task.tactical_layout = layout
    update_fields = ['tactical_layout']

    if preview_data:
        raw_bytes, extension = _decode_canvas_data_url(preview_data)
        if raw_bytes and extension:
            filename = f'task-{task.id}-graphic-{uuid.uuid4().hex[:10]}{extension}'
            task.task_preview_image.save(filename, ContentFile(raw_bytes), save=False)
            update_fields.append('task_preview_image')

    task.save(update_fields=update_fields)
    return JsonResponse({'saved': True, 'task_id': task.id})


@login_required
def session_task_preview_file(request, task_id):
    if not _can_access_sessions_workspace(request.user):
        return HttpResponse('No tienes permisos para acceder a sesiones.', status=403)
    task = (
        SessionTask.objects
        .select_related('session__microcycle__team')
        .filter(id=task_id)
        .first()
    )
    if not task:
        raise Http404('Imagen de tarea no disponible')
    if not task.task_preview_image:
        if not _ensure_library_task_preview(task):
            raise Http404('Imagen de tarea no disponible')
    file_field = task.task_preview_image
    try:
        file_field.open('rb')
    except Exception:
        # If stored preview is broken/missing, try to regenerate from PDF on the fly.
        if _ensure_library_task_preview(task):
            file_field = task.task_preview_image
            try:
                file_field.open('rb')
            except Exception:
                return HttpResponse('No se pudo abrir la imagen de la tarea.', status=500)
        else:
            return HttpResponse('No se pudo abrir la imagen de la tarea.', status=500)
    extension = Path(file_field.name).suffix.lower()
    content_type = {
        '.png': 'image/png',
        '.jpg': 'image/jpeg',
        '.jpeg': 'image/jpeg',
        '.webp': 'image/webp',
        '.gif': 'image/gif',
    }.get(extension, 'application/octet-stream')
    response = FileResponse(file_field, content_type=content_type)
    response['Content-Disposition'] = f'inline; filename="{Path(file_field.name).name}"'
    return response


@login_required
def session_task_file(request, task_id):
    if not _can_access_sessions_workspace(request.user):
        return HttpResponse('No tienes permisos para acceder a sesiones.', status=403)
    task = (
        SessionTask.objects
        .select_related('session__microcycle__team')
        .filter(id=task_id)
        .first()
    )
    if not task or not task.task_pdf:
        raise Http404('Archivo de tarea no disponible')
    file_field = task.task_pdf
    try:
        file_field.open('rb')
    except Exception:
        return HttpResponse('No se pudo abrir el archivo PDF.', status=500)
    filename = (Path(file_field.name).name or f'tarea-{task.id}.pdf').replace('"', '')
    response = FileResponse(file_field, content_type='application/pdf')
    response['Content-Disposition'] = f'inline; filename="{filename}"'
    return response


@login_required
def fines_page(request):
    primary_team = _get_primary_team_for_request(request)
    if not primary_team:
        return JsonResponse({'error': 'No hay equipo principal configurado'}, status=400)
    error = ''
    message = ''
    can_manage_fines = _is_admin_user(request.user)
    if request.method == 'POST':
        if not can_manage_fines:
            error = 'Solo administradores pueden registrar o eliminar multas.'
        else:
            form_action = (request.POST.get('form_action') or 'add').strip().lower()
            if form_action == 'delete':
                fine_id = _parse_int(request.POST.get('fine_id'))
                fine = PlayerFine.objects.filter(id=fine_id, player__team=primary_team).first() if fine_id else None
                if not fine:
                    error = 'Multa no encontrada.'
                else:
                    fine.delete()
                    message = 'Multa eliminada.'
            else:
                player_id = _parse_int(request.POST.get('player_id'))
                reason = (request.POST.get('reason') or '').strip()
                amount = _parse_int(request.POST.get('amount')) or 0
                note = (request.POST.get('note') or '').strip()
                player = Player.objects.filter(id=player_id, team=primary_team).first() if player_id else None
                valid_reasons = {item[0] for item in PlayerFine.REASON_CHOICES}
                if not player:
                    error = 'Selecciona un jugador válido.'
                elif reason not in valid_reasons:
                    error = 'Selecciona un motivo válido.'
                elif amount <= 0 or amount % 5 != 0:
                    error = 'La cantidad debe ser un múltiplo de 5.'
                else:
                    PlayerFine.objects.create(
                        player=player,
                        reason=reason,
                        amount=amount,
                        note=note,
                        created_by=(request.user.get_username() if request.user.is_authenticated else ''),
                    )
                    message = 'Multa registrada.'
    players = list(Player.objects.filter(team=primary_team, is_active=True).order_by('number', 'name'))
    fines = list(PlayerFine.objects.filter(player__team=primary_team).select_related('player')[:120])
    reason_labels = dict(PlayerFine.REASON_CHOICES)
    summary_total = sum((item.amount or 0) for item in fines)
    return render(
        request,
        'football/fines.html',
        {
            'team_name': primary_team.display_name,
            'players': players,
            'fines': fines,
            'reason_choices': PlayerFine.REASON_CHOICES,
            'reason_labels': reason_labels,
            'summary_count': len(fines),
            'summary_total': summary_total,
            'message': message,
            'error': error,
            'can_manage_fines': can_manage_fines,
        },
    )


@login_required
def analysis_page(request):
    forbidden = _forbid_if_no_coach_access(request.user)
    if forbidden:
        return forbidden
    forbidden = _forbid_if_workspace_module_disabled(request, 'analysis', label='análisis')
    if forbidden:
        return forbidden
    primary_team = _get_primary_team_for_request(request)
    team_url = (request.GET.get('team_url') or '').strip()
    team_id = (request.GET.get('team_id') or '').strip()
    raw_text = ''
    roster = []
    probable_eleven = []
    lineup = []
    insights = {}
    formation = 'Auto'
    error = ''
    auto_loaded = False
    auto_team_name = ''
    video_error = ''
    video_message = ''
    folder_error = ''
    folder_message = ''
    manual_report_error = ''
    manual_report_message = ''
    extracted = {}
    preferred_next = load_preferred_next_match_payload(primary_team=primary_team) or {}
    preferred_opponent = preferred_next.get('opponent') if isinstance(preferred_next, dict) else {}
    home_rival_name = _payload_opponent_name(preferred_next)
    if home_rival_name:
        home_rival_key = normalize_label(home_rival_name)
        guessed_team = Team.objects.filter(is_primary=False).order_by('name').filter(name__icontains=home_rival_name[:12]).first()
        if guessed_team and not team_id:
            team_id = str(guessed_team.id)
    if request.method == 'POST':
        form_action = (request.POST.get('form_action') or 'analyze').strip()
        if form_action == 'create_video_folder':
            rival_team_id = _parse_int(request.POST.get('video_team_id'))
            rival_team = Team.objects.filter(id=rival_team_id).first() if rival_team_id else None
            folder_name = (request.POST.get('folder_name') or '').strip()
            if not primary_team:
                folder_error = 'No hay equipo principal configurado.'
            elif not folder_name:
                folder_error = 'Indica un nombre para la carpeta.'
            else:
                folder, created = AnalystVideoFolder.objects.get_or_create(
                    team=primary_team,
                    rival_team=rival_team,
                    name=folder_name[:140],
                    defaults={'created_by': request.user.get_username() if request.user.is_authenticated else ''},
                )
                folder_message = 'Carpeta creada correctamente.' if created else 'La carpeta ya existía y queda disponible.'
        elif form_action == 'upload_video':
            video_title = (request.POST.get('video_title') or '').strip() or 'Vídeo rival'
            video_source = (request.POST.get('video_source') or RivalVideo.SOURCE_MANUAL).strip()
            rival_team_id = _parse_int(request.POST.get('video_team_id'))
            folder_id = _parse_int(request.POST.get('video_folder_id'))
            video_file = request.FILES.get('video_file')
            rival_team = Team.objects.filter(id=rival_team_id).first() if rival_team_id else None
            folder = AnalystVideoFolder.objects.filter(id=folder_id, team=primary_team).first() if folder_id and primary_team else None
            if not video_file:
                video_error = 'Selecciona un vídeo para subir.'
            else:
                entry = RivalVideo.objects.create(
                    rival_team=rival_team,
                    folder=folder,
                    title=video_title,
                    video=video_file,
                    source=video_source if video_source in {c[0] for c in RivalVideo.SOURCE_CHOICES} else RivalVideo.SOURCE_MANUAL,
                    notes=(request.POST.get('video_notes') or '').strip(),
                )
                assigned_player_ids = [
                    player_id
                    for player_id in (_parse_int(value) for value in request.POST.getlist('assigned_player_ids'))
                    if player_id
                ]
                if assigned_player_ids:
                    entry.assigned_players.set(Player.objects.filter(team=primary_team, id__in=assigned_player_ids))
                video_message = 'Vídeo subido correctamente.'
        elif form_action == 'delete_video':
            video_id = _parse_int(request.POST.get('video_id'))
            entry = RivalVideo.objects.filter(id=video_id).first()
            if entry:
                try:
                    if entry.video:
                        entry.video.delete(save=False)
                except Exception:
                    pass
                entry.delete()
                video_message = 'Vídeo eliminado.'
        elif form_action == 'assign_video_players':
            video_id = _parse_int(request.POST.get('video_id'))
            folder_id = _parse_int(request.POST.get('video_folder_id'))
            entry = RivalVideo.objects.filter(id=video_id).first()
            if entry:
                folder = AnalystVideoFolder.objects.filter(id=folder_id, team=primary_team).first() if folder_id and primary_team else None
                entry.folder = folder
                entry.save(update_fields=['folder'])
                assigned_player_ids = [
                    player_id
                    for player_id in (_parse_int(value) for value in request.POST.getlist('assigned_player_ids'))
                    if player_id
                ]
                if primary_team:
                    entry.assigned_players.set(Player.objects.filter(team=primary_team, id__in=assigned_player_ids))
                video_message = 'Asignación de vídeo actualizada.'
        elif form_action == 'save_manual_report':
            selected_team_id = _parse_int(request.POST.get('team_id'))
            selected_team = Team.objects.filter(id=selected_team_id).first() if selected_team_id else None
            rival_name_input = (
                (request.POST.get('manual_rival_name') or '').strip()
                or (selected_team.name if selected_team else '')
                or (home_rival_name or '').strip()
            )
            if not primary_team:
                manual_report_error = 'No hay equipo principal configurado.'
            elif not rival_name_input:
                manual_report_error = 'Indica el rival para guardar el informe.'
            else:
                status_value = (request.POST.get('manual_status') or RivalAnalysisReport.STATUS_DRAFT).strip()
                if status_value not in {choice[0] for choice in RivalAnalysisReport.STATUS_CHOICES}:
                    status_value = RivalAnalysisReport.STATUS_DRAFT
                confidence = _parse_int(request.POST.get('manual_confidence')) or 3
                confidence = max(1, min(confidence, 5))
                RivalAnalysisReport.objects.create(
                    team=primary_team,
                    rival_team=selected_team,
                    rival_name=rival_name_input,
                    report_title=(request.POST.get('manual_report_title') or '').strip(),
                    match_round=(request.POST.get('manual_match_round') or '').strip(),
                    match_date=(request.POST.get('manual_match_date') or '').strip(),
                    match_location=(request.POST.get('manual_match_location') or '').strip(),
                    tactical_system=(request.POST.get('manual_tactical_system') or '').strip(),
                    attacking_patterns=(request.POST.get('manual_attacking_patterns') or '').strip(),
                    defensive_patterns=(request.POST.get('manual_defensive_patterns') or '').strip(),
                    transitions=(request.POST.get('manual_transitions') or '').strip(),
                    set_pieces_for=(request.POST.get('manual_set_pieces_for') or '').strip(),
                    set_pieces_against=(request.POST.get('manual_set_pieces_against') or '').strip(),
                    key_players=(request.POST.get('manual_key_players') or '').strip(),
                    weaknesses=(request.POST.get('manual_weaknesses') or '').strip(),
                    opportunities=(request.POST.get('manual_opportunities') or '').strip(),
                    match_plan=(request.POST.get('manual_match_plan') or '').strip(),
                    individual_tasks=(request.POST.get('manual_individual_tasks') or '').strip(),
                    alert_notes=(request.POST.get('manual_alert_notes') or '').strip(),
                    confidence_level=confidence,
                    status=status_value,
                    created_by=(request.user.get_username() if request.user.is_authenticated else ''),
                )
                manual_report_message = 'Informe manual guardado correctamente.'
        team_url = (request.POST.get('team_url') or '').strip()
        team_id = (request.POST.get('team_id') or '').strip()
        raw_text = (request.POST.get('raw_text') or '').strip()
        team = None
        if team_id:
            team = Team.objects.filter(id=team_id).first()
            if team and not team_url:
                team_url = team.preferente_url or ''
        if form_action == 'analyze':
            try:
                if raw_text:
                    roster = parse_preferente_roster(raw_text)
                elif team_url:
                    roster = fetch_preferente_team_roster(team_url)

                probable_eleven = compute_probable_eleven(roster)
                insights = build_rival_insights(roster)
                formation = compute_formation(probable_eleven)
                try:
                    lineup = assign_lineup_slots(probable_eleven, formation)
                except Exception:
                    lineup = []
                    error = 'Plantilla cargada, pero no se ha podido dibujar el 11 probable.'

                if not roster:
                    error = 'No se han encontrado jugadores en la plantilla.'
            except Exception:
                error = 'No se ha podido procesar la plantilla del rival. Revisa URL o el contenido pegado.'
        if team and team_url and team.preferente_url != team_url:
            team.preferente_url = team_url
            team.save(update_fields=['preferente_url'])
    else:
        active_match = get_active_match(primary_team) if primary_team else None
        auto_team = None
        if active_match and primary_team:
            auto_team = (
                active_match.away_team if active_match.home_team == primary_team else active_match.home_team
            )
        if auto_team:
            team_id = str(auto_team.id)
            auto_team_name = auto_team.name
            team_url = (auto_team.preferente_url or '').strip()
            if team_url:
                try:
                    roster = fetch_preferente_team_roster(team_url)
                    probable_eleven = compute_probable_eleven(roster)
                    insights = build_rival_insights(roster)
                    formation = compute_formation(probable_eleven)
                    lineup = assign_lineup_slots(probable_eleven, formation)
                    auto_loaded = True
                    if not roster:
                        error = 'No se han encontrado jugadores en la plantilla.'
                except Exception:
                    error = (
                        f'Rival detectado automáticamente ({auto_team.name}), '
                        'pero no se ha podido cargar su plantilla.'
                    )
            else:
                error = (
                    f'Rival detectado automáticamente ({auto_team.name}), '
                    'pero no tiene URL de La Preferente guardada.'
                )
    selected_team = Team.objects.filter(id=_parse_int(team_id)).first() if team_id else None
    selected_folder_id = _parse_int(request.GET.get('folder')) or _parse_int(request.POST.get('selected_folder_id'))
    universe = load_universo_snapshot() or {}
    standings = universe.get('standings') or []
    team_lookup = _build_universo_standings_lookup(universe)
    rival_name = selected_team.name if selected_team else (home_rival_name or auto_team_name)
    rival_full_name, rival_crest_url = _resolve_rival_identity(rival_name, preferred_opponent=preferred_opponent)
    rival_meta = team_lookup.get(_normalize_team_lookup_key(rival_name), {})
    extracted = {
        'source_priority': 'Universo RFAF > RFAF > La Preferente',
        'rival_name': rival_full_name,
        'rival_crest_url': _absolute_universo_url(rival_crest_url or rival_meta.get('crest_url')),
        'standings_count': len(standings),
        'next_match_round': preferred_next.get('round') if isinstance(preferred_next, dict) else '',
        'next_match_date': preferred_next.get('date') if isinstance(preferred_next, dict) else '',
        'next_match_time': preferred_next.get('time') if isinstance(preferred_next, dict) else '',
        'next_match_location': preferred_next.get('location') if isinstance(preferred_next, dict) else '',
        'preferente_url': selected_team.preferente_url if selected_team else '',
    }
    manual_report_latest = None
    manual_reports = []
    if primary_team:
        manual_reports_qs = RivalAnalysisReport.objects.filter(team=primary_team)
        if selected_team:
            manual_reports_qs = manual_reports_qs.filter(rival_team=selected_team)
        elif rival_name:
            manual_reports_qs = manual_reports_qs.filter(rival_name__icontains=rival_name[:18])
        manual_reports = list(manual_reports_qs.order_by('-updated_at', '-id')[:12])
        manual_report_latest = manual_reports[0] if manual_reports else None

    manual_initial = {
        'rival_name': rival_full_name or (selected_team.name if selected_team else ''),
        'report_title': '',
        'match_round': extracted.get('next_match_round') or '',
        'match_date': extracted.get('next_match_date') or '',
        'match_location': extracted.get('next_match_location') or '',
        'tactical_system': '',
        'attacking_patterns': '',
        'defensive_patterns': '',
        'transitions': '',
        'set_pieces_for': '',
        'set_pieces_against': '',
        'key_players': '',
        'weaknesses': '',
        'opportunities': '',
        'match_plan': '',
        'individual_tasks': '',
        'alert_notes': '',
        'confidence': 3,
        'status': RivalAnalysisReport.STATUS_DRAFT,
    }
    if manual_report_latest:
        manual_initial.update(
            {
                'rival_name': manual_report_latest.rival_name,
                'report_title': manual_report_latest.report_title,
                'match_round': manual_report_latest.match_round,
                'match_date': manual_report_latest.match_date,
                'match_location': manual_report_latest.match_location,
                'tactical_system': manual_report_latest.tactical_system,
                'attacking_patterns': manual_report_latest.attacking_patterns,
                'defensive_patterns': manual_report_latest.defensive_patterns,
                'transitions': manual_report_latest.transitions,
                'set_pieces_for': manual_report_latest.set_pieces_for,
                'set_pieces_against': manual_report_latest.set_pieces_against,
                'key_players': manual_report_latest.key_players,
                'weaknesses': manual_report_latest.weaknesses,
                'opportunities': manual_report_latest.opportunities,
                'match_plan': manual_report_latest.match_plan,
                'individual_tasks': manual_report_latest.individual_tasks,
                'alert_notes': manual_report_latest.alert_notes,
                'confidence': manual_report_latest.confidence_level,
                'status': manual_report_latest.status,
            }
        )
    video_folders = []
    if primary_team:
        folders_qs = AnalystVideoFolder.objects.filter(team=primary_team).select_related('rival_team')
        if selected_team:
            folders_qs = folders_qs.filter(Q(rival_team=selected_team) | Q(rival_team__isnull=True))
        video_folders = list(folders_qs.order_by('name', '-created_at'))

    rival_videos_qs = RivalVideo.objects.select_related('rival_team', 'folder').prefetch_related('assigned_players').order_by('-created_at')
    if selected_team:
        rival_videos_qs = rival_videos_qs.filter(rival_team=selected_team)
    else:
        rival_videos_qs = rival_videos_qs.filter(rival_team__isnull=True)
    if selected_folder_id:
        rival_videos_qs = rival_videos_qs.filter(folder_id=selected_folder_id)
    rival_videos = list(rival_videos_qs[:40])
    analyst_players = list(Player.objects.filter(team=primary_team, is_active=True).order_by('number', 'name')) if primary_team else []

    return render(
        request,
        'football/coach_analysis.html',
        {
            'section_title': 'Análisis rival',
            'description': 'Indicadores y notas tácticas para el próximo rival.',
            'team_url': team_url,
            'team_id': team_id,
            'teams': Team.objects.order_by('name'),
            'raw_text': raw_text,
            'roster': roster,
            'probable_eleven': probable_eleven,
            'lineup': lineup,
            'insights': insights,
            'formation': formation,
            'error': error,
            'auto_loaded': auto_loaded,
            'auto_team_name': auto_team_name,
            'rival_videos': rival_videos,
            'video_error': video_error,
            'video_message': video_message,
            'video_sources': RivalVideo.SOURCE_CHOICES,
            'video_folders': video_folders,
            'selected_folder_id': selected_folder_id,
            'folder_message': folder_message,
            'folder_error': folder_error,
            'analyst_players': analyst_players,
            'home_rival_name': home_rival_name,
            'extracted': extracted,
            'manual_initial': manual_initial,
            'manual_reports': manual_reports,
            'manual_report_message': manual_report_message,
            'manual_report_error': manual_report_error,
            'manual_report_status_choices': RivalAnalysisReport.STATUS_CHOICES,
        },
    )


@login_required
def manual_player_stats_page(request):
    if not _is_admin_user(request.user):
        return HttpResponse('Solo administradores pueden editar estadísticas manuales.', status=403)
    forbidden = _forbid_if_workspace_module_disabled(request, 'manual_stats', label='estadísticas manuales')
    if forbidden:
        return forbidden
    try:
        primary_team = _get_primary_team_for_request(request)
        if not primary_team:
            return JsonResponse({'error': 'No hay equipo principal configurado'}, status=400)

        season = resolve_stats_season(primary_team)
        if season is None:
            return JsonResponse({'error': 'No hay temporada activa para guardar estadísticas'}, status=400)

        players = list(Player.objects.filter(team=primary_team).order_by('name'))
        current_overrides = get_manual_player_base_overrides(primary_team, season)

        if request.method == 'POST':
            with transaction.atomic():
                for player in players:
                    overrides = {
                        'manual_pj': _parse_int(request.POST.get(f'pj_{player.id}')) or 0,
                        'manual_pt': _parse_int(request.POST.get(f'pt_{player.id}')) or 0,
                        'manual_minutes': _parse_int(request.POST.get(f'minutes_{player.id}')) or 0,
                        'manual_goals': _parse_int(request.POST.get(f'goals_{player.id}')) or 0,
                        'manual_yellow_cards': _parse_int(request.POST.get(f'yellow_{player.id}')) or 0,
                        'manual_red_cards': _parse_int(request.POST.get(f'red_{player.id}')) or 0,
                    }
                    save_manual_player_base_overrides(
                        player=player,
                        season=season,
                        values=overrides,
                    )
            current_overrides = get_manual_player_base_overrides(primary_team, season)
            _invalidate_team_dashboard_caches(primary_team)
            message = 'Estadísticas manuales guardadas.'
        else:
            message = ''

        rows = []
        try:
            roster_cache = get_roster_stats_cache()
        except Exception:
            logger.exception('No se pudo cargar la cache de plantilla para estadísticas manuales')
            roster_cache = {}
        for player in players:
            roster_entry = find_roster_entry(player.name, roster_cache) or {}
            manual = current_overrides.get(player.id, {})
            rows.append(
                {
                    'player': player,
                    'pj': manual.get('pj', roster_entry.get('pj', 0)),
                    'pt': manual.get('pt', roster_entry.get('pt', 0)),
                    'minutes': manual.get('minutes', roster_entry.get('minutes', 0)),
                    'goals': manual.get('goals', roster_entry.get('goals', 0)),
                    'yellow_cards': manual.get('yellow_cards', roster_entry.get('yellow_cards', 0)),
                    'red_cards': manual.get('red_cards', roster_entry.get('red_cards', 0)),
                }
            )

        return render(
            request,
            'football/manual_player_stats.html',
            {
                'team_name': primary_team.display_name,
                'season_name': season_display_name(season),
                'rows': rows,
                'message': message,
                'error': '',
            },
        )
    except Exception:
        logger.exception('Error en manual_player_stats_page')
        return render(
            request,
            'football/manual_player_stats.html',
            {
                'team_name': 'Equipo principal',
                'season_name': 'Temporada actual',
                'rows': [],
                'message': '',
                'error': 'No se pudieron cargar las estadísticas manuales. Revisa los datos e inténtalo de nuevo.',
            },
            status=200,
        )

@login_required
def player_detail_page(request, player_id):
    try:
        forbidden = _forbid_if_workspace_module_disabled(request, 'players', label='módulo de jugadores')
        if forbidden:
            return forbidden
        primary_team = _get_player_team_for_request(request)
        if not primary_team:
            return JsonResponse({'error': 'No hay equipo principal configurado'}, status=400)
        player = Player.objects.filter(id=player_id, team=primary_team).first()
        if not player:
            return JsonResponse({'error': 'Jugador no encontrado'}, status=404)
        forbidden = _forbid_if_no_player_access(request.user, player, primary_team=primary_team)
        if forbidden:
            return forbidden
        current_role = _get_user_role(request.user)
        can_preview_player_view = (
            request.user.is_authenticated
            and (current_role != AppUserRole.ROLE_PLAYER or _is_admin_user(request.user))
        )
        preview_mode = (request.GET.get('preview') or '').strip().lower()
        player_view_preview = can_preview_player_view and preview_mode == 'player'
        is_player_readonly = player_view_preview or (current_role == AppUserRole.ROLE_PLAYER and not _is_admin_user(request.user))
        active_match = get_active_match(primary_team)
        current_convocation = get_current_convocation_record(primary_team, match=active_match)
        is_called_up = bool(
            current_convocation
            and current_convocation.players.filter(id=player.id).exists()
        )

        def _parse_date_value(raw_value):
            value = str(raw_value or '').strip()
            if not value:
                return None
            for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%d-%m-%Y'):
                try:
                    return datetime.strptime(value, fmt).date()
                except ValueError:
                    continue
            return None

        def _parse_datetime_value(raw_value):
            value = str(raw_value or '').strip()
            if not value:
                return None
            for fmt in ('%Y-%m-%dT%H:%M', '%Y-%m-%d %H:%M', '%d/%m/%Y %H:%M'):
                try:
                    return datetime.strptime(value, fmt)
                except ValueError:
                    continue
            return None

        def _parse_decimal_value(raw_value):
            value = str(raw_value or '').strip().replace(',', '.')
            if not value:
                return None
            try:
                return round(float(value), 2)
            except Exception:
                return None

        def _resolve_season():
            if primary_team.group and primary_team.group.season:
                return primary_team.group.season
            return Season.objects.filter(is_current=True).order_by('-start_date', '-id').first()

        if request.method == 'POST' and not is_player_readonly:
            form_action = (request.POST.get('form_action') or 'profile').strip().lower()

            if form_action == 'profile':
                uploaded_photo = request.FILES.get('player_photo')
                number = request.POST.get('number', '').strip()
                injury_name = request.POST.get('injury', '').strip()
                injury_type = request.POST.get('injury_type', '').strip()
                injury_zone = request.POST.get('injury_zone', '').strip()
                injury_side = request.POST.get('injury_side', '').strip()
                injury_notes = request.POST.get('injury_notes', '').strip()
                injury_date = _parse_date_value(request.POST.get('injury_date'))
                injury_return_date = _parse_date_value(request.POST.get('injury_return_date'))
                manual_sanction_active = str(request.POST.get('manual_sanction_active') or '').lower() in {'1', 'true', 'on', 'yes'}
                manual_sanction_reason = request.POST.get('manual_sanction_reason', '').strip()
                manual_sanction_until = _parse_date_value(request.POST.get('manual_sanction_until'))
                injury_record_mode = (request.POST.get('injury_record_mode') or '').strip().lower()
                force_new_injury_record = injury_record_mode in {'new', 'add', 'create'}
                player.number = _parse_int(number) if number else None
                player.position = request.POST.get('position', '').strip()
                player.full_name = request.POST.get('full_name', '').strip()
                player.nickname = request.POST.get('nickname', '').strip()
                player.birth_date = _parse_date_value(request.POST.get('birth_date'))
                player.height_cm = _parse_int(request.POST.get('height_cm'))
                player.weight_kg = _parse_decimal_value(request.POST.get('weight_kg_base'))
                player.injury = injury_name
                player.injury_type = injury_type
                player.injury_zone = injury_zone
                player.injury_side = injury_side
                player.injury_date = injury_date
                player.manual_sanction_active = manual_sanction_active
                player.manual_sanction_reason = manual_sanction_reason
                player.manual_sanction_until = manual_sanction_until
                player.save()
                if uploaded_photo:
                    save_player_photo(player, uploaded_photo)
                _invalidate_team_dashboard_caches(primary_team)

                active_injury = (
                    PlayerInjuryRecord.objects
                    .filter(player=player, is_active=True)
                    .order_by('-injury_date', '-id')
                    .first()
                )
                if injury_name and injury_date:
                    same_record = None
                    if not force_new_injury_record:
                        same_record = (
                            PlayerInjuryRecord.objects
                            .filter(
                                player=player,
                                injury__iexact=injury_name,
                                injury_date=injury_date,
                            )
                            .order_by('-id')
                            .first()
                        )
                    if same_record:
                        updated_fields = []
                        if injury_type != same_record.injury_type:
                            same_record.injury_type = injury_type
                            updated_fields.append('injury_type')
                        if injury_zone != same_record.injury_zone:
                            same_record.injury_zone = injury_zone
                            updated_fields.append('injury_zone')
                        if injury_side != same_record.injury_side:
                            same_record.injury_side = injury_side
                            updated_fields.append('injury_side')
                        if injury_notes != same_record.notes:
                            same_record.notes = injury_notes
                            updated_fields.append('notes')
                        if injury_return_date != same_record.return_date:
                            same_record.return_date = injury_return_date
                            updated_fields.append('return_date')
                        should_be_active = not injury_return_date or injury_return_date > timezone.localdate()
                        if same_record.is_active != should_be_active:
                            same_record.is_active = should_be_active
                            updated_fields.append('is_active')
                        if updated_fields:
                            same_record.save(update_fields=updated_fields + ['updated_at'])
                    else:
                        PlayerInjuryRecord.objects.create(
                            player=player,
                            injury=injury_name,
                            injury_type=injury_type,
                            injury_zone=injury_zone,
                            injury_side=injury_side,
                            injury_date=injury_date,
                            return_date=injury_return_date,
                            notes=injury_notes,
                            is_active=(not injury_return_date or injury_return_date > timezone.localdate()),
                        )
                    if active_injury and active_injury.injury.lower() != injury_name.lower():
                        active_injury.is_active = False
                        if not active_injury.return_date:
                            active_injury.return_date = injury_return_date or timezone.localdate()
                            active_injury.save(update_fields=['is_active', 'return_date', 'updated_at'])
                        else:
                            active_injury.save(update_fields=['is_active', 'updated_at'])
                elif active_injury and injury_return_date:
                    active_injury.return_date = injury_return_date
                    active_injury.is_active = False
                    active_injury.save(update_fields=['return_date', 'is_active', 'updated_at'])

                return redirect(f"{reverse('player-detail', args=[player.id])}?tab=general")

            if form_action == 'manual_stats':
                season = _resolve_season()
                if season:
                    manual_values = {
                        'manual_pj': _parse_int(request.POST.get('manual_pj')) or 0,
                        'manual_pt': _parse_int(request.POST.get('manual_pt')) or 0,
                        'manual_minutes': _parse_int(request.POST.get('manual_minutes')) or 0,
                        'manual_goals': _parse_int(request.POST.get('manual_goals')) or 0,
                        'manual_yellow_cards': _parse_int(request.POST.get('manual_yellow_cards')) or 0,
                        'manual_red_cards': _parse_int(request.POST.get('manual_red_cards')) or 0,
                    }
                    with transaction.atomic():
                        save_manual_player_base_overrides(
                            player=player,
                            season=season,
                            values=manual_values,
                        )
                return redirect(f"{reverse('player-detail', args=[player.id])}?tab=general")

            if form_action == 'physical':
                PlayerPhysicalMetric.objects.create(
                    player=player,
                    recorded_on=_parse_date_value(request.POST.get('recorded_on')) or timezone.localdate(),
                    workload=request.POST.get('workload', '').strip(),
                    rpe=_parse_int(request.POST.get('rpe')),
                    wellness=_parse_int(request.POST.get('wellness')),
                    weight_kg=request.POST.get('weight_kg', '').strip() or None,
                    notes=request.POST.get('physical_notes', '').strip(),
                )
                return redirect(f"{reverse('player-detail', args=[player.id])}?tab=physical")

            if form_action == 'communication':
                message = request.POST.get('message', '').strip()
                if message:
                    PlayerCommunication.objects.create(
                        player=player,
                        match=active_match,
                        category=request.POST.get('category') or PlayerCommunication.CATEGORY_INTERNAL,
                        message=message,
                        scheduled_for=_parse_datetime_value(request.POST.get('scheduled_for')),
                        created_by=(request.user.get_username() if request.user.is_authenticated else ''),
                    )
                return redirect(f"{reverse('player-detail', args=[player.id])}?tab=communication")

        try:
            matches = compute_player_dashboard(primary_team)
            stats_error = ''
        except Exception:
            logger.exception('No se pudo recomponer el dashboard del jugador %s', player_id)
            matches = []
            stats_error = 'Las estadísticas consolidadas no están disponibles temporalmente.'
        detail = next((p for p in matches if p.get('player_id') == player_id), None)
        safe_stats = detail or {
            'position': player.position or '',
            'pj': 0,
            'pt': 0,
            'minutes': 0,
            'goals': 0,
            'yellow_cards': 0,
            'red_cards': 0,
            'second_yellow_cards': 0,
            'success_rate': 0,
            'duel_rate': 0,
            'total_actions': 0,
            'assists': 0,
            'goalkeeper_saves': 0,
            'field_zones': [],
            'matches': [],
            'duel_summary': {'won': 0, 'total': 0},
            'passes': {'completed': 0, 'attempts': 0, 'key_completed': 0, 'accuracy': 0},
            'shots': {'on_target': 0, 'attempts': 0, 'accuracy': 0, 'per_goal': None},
            'importance_score': 0,
            'influence_score': 0,
            'successes_per90': 0,
            'decisive_actions_per90': 0,
        }
        active_tab = (request.GET.get('tab') or 'general').strip().lower()
        physical_metrics = player.physical_metrics.all()[:20]
        latest_physical_metric = physical_metrics[0] if physical_metrics else None
        communications = player.communications.select_related('match').all()[:20]
        assigned_analysis_videos = list(
            player.assigned_analysis_videos.select_related('rival_team', 'folder').order_by('-created_at')[:20]
        )
        injury_records = player.injury_records.all()[:20]
        latest_injury_record = injury_records[0] if injury_records else None
        has_active_injury = player.id in get_active_injury_player_ids([player.id])
        if not has_active_injury and latest_injury_record:
            has_active_injury = is_injury_record_active(latest_injury_record)
        has_manual_sanction = is_manual_sanction_active(player)
        player_photo_url = resolve_player_photo_url(request, player)
        convocation_pending = bool(current_convocation and not current_convocation.players.exists())
        fines_summary = {
            'registered_fines': 0,
            'registered_total': 0,
            'manual_sanctions': 1 if has_manual_sanction else 0,
            'total': 0,
        }
        fines_records = []
        player_fines = list(player.fines.all()[:80])
        reason_labels = dict(PlayerFine.REASON_CHOICES)
        for fine in player_fines:
            fines_summary['registered_fines'] += 1
            fines_summary['registered_total'] += int(fine.amount or 0)
            fines_records.append(
                {
                    'type': reason_labels.get(fine.reason, fine.reason),
                    'amount': int(fine.amount or 0),
                    'date': fine.created_at.strftime('%d/%m/%Y'),
                    'detail': fine.note or '-',
                }
            )
        if has_manual_sanction:
            until_label = (
                player.manual_sanction_until.strftime('%d/%m/%Y')
                if player.manual_sanction_until
                else 'Sin fecha fin'
            )
            fines_records.insert(
                0,
                {
                    'type': 'Sanción manual',
                    'amount': 0,
                    'date': until_label,
                    'detail': player.manual_sanction_reason or 'Sanción configurada en ficha',
                }
            )
        fines_summary['total'] = (
            fines_summary['registered_fines']
            + fines_summary['manual_sanctions']
        )

        def _to_int_value(value):
            return _parse_int(value) or 0

        stats_source = detail or {}
        pj = _to_int_value(stats_source.get('pj'))
        pt = _to_int_value(stats_source.get('pt'))
        minutes = _to_int_value(stats_source.get('minutes'))
        goals = _to_int_value(stats_source.get('goals'))
        yellow_cards = _to_int_value(stats_source.get('yellow_cards'))
        red_cards = _to_int_value(stats_source.get('red_cards'))
        second_yellow_cards = _to_int_value(stats_source.get('second_yellow_cards'))
        competition_total_rounds = _to_int_value(stats_source.get('competition_total_rounds')) or get_competition_total_rounds(primary_team)
        participation_pct = (
            round(min((pj / competition_total_rounds) * 100, 100), 1)
            if competition_total_rounds > 0
            else 0
        )
        suplente = max(pj - pt, 0)
        goals_per_match = round((goals / pj), 2) if pj else 0
        max_minutes = pj * 90
        minute_ratio = round((minutes / max_minutes) * 100, 1) if max_minutes else 0
        minute_ratio = max(0, min(minute_ratio, 100))
        importance_score = float(stats_source.get('importance_score') or 0)
        influence_score = float(stats_source.get('influence_score') or 0)

        standings_rows = _resolve_standings_for_team(primary_team, snapshot=load_universo_snapshot())
        team_points = 0
        team_rank = 0
        team_key = _normalize_team_lookup_key(primary_team.name)
        for row in standings_rows:
            candidate_key = _normalize_team_lookup_key(row.get('full_name') or row.get('team'))
            if team_key and candidate_key == team_key:
                team_points = _to_int_value(row.get('points'))
                team_rank = _to_int_value(row.get('rank'))
                break
        if not team_points and not team_rank and standings_rows:
            preferred_aliases = ('benagalbon', 'c.d. benagalbon', 'cd benagalbon')
            for row in standings_rows:
                full_name = str(row.get('full_name') or row.get('team') or '').lower()
                if any(alias in full_name for alias in preferred_aliases):
                    team_points = _to_int_value(row.get('points'))
                    team_rank = _to_int_value(row.get('rank'))
                    break

        season_label = ''
        division_label = ''
        if primary_team.group:
            division_label = primary_team.group.name or ''
            season = primary_team.group.season
            if season:
                if season.start_date and season.end_date:
                    season_label = f'{season.start_date.year}-{season.end_date.year}'
                else:
                    season_label = season.name or ''
        if not season_label:
            season_label = 'Temporada actual'

        general_kpis = [
            {'label': 'Partidos', 'value': pj, 'pct': 100 if pj else 0},
            {'label': 'Titular', 'value': pt, 'pct': round((pt / pj) * 100, 1) if pj else 0},
            {'label': 'Suplente', 'value': suplente, 'pct': round((suplente / pj) * 100, 1) if pj else 0},
            {'label': 'Minutos', 'value': minutes, 'pct': minute_ratio},
            {'label': 'Total goles', 'value': goals, 'pct': round((goals / pj) * 100, 1) if pj else 0},
            {'label': 'Media goles/partido', 'value': goals_per_match, 'pct': round(min(goals_per_match * 100, 100), 1)},
            {'label': '% participación', 'value': participation_pct, 'pct': participation_pct},
            {'label': 'Importancia', 'value': importance_score, 'pct': importance_score},
            {'label': 'Influencia', 'value': influence_score, 'pct': influence_score},
            {'label': 'Amarillas', 'value': yellow_cards, 'pct': round(min(yellow_cards * 15, 100), 1)},
            {'label': 'Rojas', 'value': red_cards, 'pct': round(min(red_cards * 30, 100), 1)},
            {'label': 'Doble amarilla', 'value': second_yellow_cards, 'pct': round(min(second_yellow_cards * 30, 100), 1)},
        ]
        active_workspace = _get_active_workspace(request)
        home_url = _workspace_entry_url(active_workspace, user=request.user) if active_workspace else reverse('dashboard-home')

        return render(
            request,
            'football/player_detail.html',
            {
                'player': player,
                'stats': safe_stats,
                'active_tab': active_tab,
                'physical_metrics': physical_metrics,
                'latest_physical_metric': latest_physical_metric,
                'communications': communications,
                'assigned_analysis_videos': assigned_analysis_videos,
                'injury_records': injury_records,
                'latest_injury_record': latest_injury_record,
                'has_active_injury': has_active_injury,
                'has_manual_sanction': has_manual_sanction,
                'is_called_up': is_called_up,
                'current_convocation': current_convocation,
                'convocation_pending': convocation_pending,
                'active_match': active_match,
                'player_photo_url': player_photo_url,
                'general_kpis': general_kpis,
                'team_points': team_points,
                'team_rank': team_rank,
                'season_label': season_label,
                'division_label': division_label,
                'fines_summary': fines_summary,
                'fines_records': fines_records,
                'stats_error': stats_error,
                'is_player_readonly': is_player_readonly,
                'player_view_preview': player_view_preview,
                'can_preview_player_view': can_preview_player_view,
                'workspace_entry_url': home_url,
                'home_url': home_url,
            },
        )
    except Exception:
        logger.exception("Error en player_detail_page para player_id=%s", player_id)
        return HttpResponse('Error interno al cargar la ficha del jugador.', status=500)




@login_required
def player_pdf(request, player_id):
    forbidden = _forbid_if_workspace_module_disabled(request, 'players', label='módulo de jugadores')
    if forbidden:
        return forbidden
    primary_team = _get_player_team_for_request(request)
    if not primary_team:
        raise Http404('Equipo principal no configurado')
    player = Player.objects.filter(id=player_id, team=primary_team).first()
    if not player:
        raise Http404('Jugador no encontrado')
    forbidden = _forbid_if_no_player_access(request.user, player, primary_team=primary_team)
    if forbidden:
        return forbidden
    matches = compute_player_dashboard(primary_team)
    detail = next((p for p in matches if p.get('player_id') == player_id), None)
    if not detail:
        raise Http404('Sin datos para generar el PDF')
    html = render_to_string(
        'football/player_pdf.html',
        {
            'player': player,
            'stats': detail,
            'club_logo_url': resolve_team_crest_url(request, primary_team, sync=True),
            'brand_mark_url': request.build_absolute_uri(static('football/images/2j-mark.svg')),
        },
        request=request,
    )
    filename = slugify(player.name or 'jugador')
    return _build_pdf_response_or_html_fallback(request, html, filename)


@login_required
def player_presentation(request, player_id):
    forbidden = _forbid_if_workspace_module_disabled(request, 'players', label='módulo de jugadores')
    if forbidden:
        return forbidden
    primary_team = _get_player_team_for_request(request)
    if not primary_team:
        raise Http404('Equipo principal no configurado')
    player = Player.objects.filter(id=player_id, team=primary_team).first()
    if not player:
        raise Http404('Jugador no encontrado')
    forbidden = _forbid_if_no_player_access(request.user, player, primary_team=primary_team)
    if forbidden:
        return forbidden
    matches = compute_player_dashboard(primary_team)
    detail = next((p for p in matches if p.get('player_id') == player_id), None)
    if not detail:
        raise Http404('Sin datos para generar la presentación')
    return render(
        request,
        'football/player_pdf.html',
        {'player': player, 'stats': detail},
    )

@login_required
def match_stats_page(request, match_id):
    forbidden = _forbid_if_no_coach_access(request.user)
    if forbidden:
        return forbidden
    forbidden = _forbid_if_workspace_module_disabled(request, 'players', label='estadísticas de partido')
    if forbidden:
        return forbidden
    primary_team = _get_primary_team_for_request(request)
    if not primary_team:
        return JsonResponse({'error': 'No hay equipo principal configurado'}, status=400)
    match = _team_match_queryset(primary_team).filter(id=match_id).first()
    if not match:
        raise Http404('Partido no encontrado')
    opponent = match.away_team if match.home_team == primary_team else match.home_team
    payload = {
        'round': match.round or 'Partido sin jornada',
        'date': match.date.strftime('%d/%m/%Y') if match.date else 'Fecha por definir',
        'location': match.location or 'Campo por confirmar',
        'opponent': opponent.name if opponent else 'Rival desconocido',
        'home': match.home_team == primary_team,
    }
    team_metrics = compute_team_metrics_for_match(match, primary_team=primary_team)
    player_cards = compute_player_cards_for_match(match, primary_team)
    return render(
        request,
        'football/match_stats.html',
        {
            'match': payload,
            'team_metrics': team_metrics,
            'player_cards': player_cards,
        },
    )


def _build_player_match_stats_payload(primary_team, player, match):
    opponent = match.away_team if match.home_team == primary_team else match.home_team
    preferred_sources = preferred_event_source_by_match(primary_team)
    events = _filter_stats_events(
        confirmed_events_queryset()
        .filter(match=match, player=player)
        .select_related('match')
        .order_by('minute', 'id'),
        preferred_sources=preferred_sources,
    )
    match_zone_profiles, player_zone_profiles = _build_zone_inference_profiles(events)
    stats = {
        'player_id': player.id,
        'name': player.name,
        'number': player.number,
        'position': player.position,
        'total_actions': 0,
        'successes': 0,
        'goals': 0,
        'assists': 0,
        'yellow_cards': 0,
        'red_cards': 0,
        'zone_counts': {key: 0 for key in FIELD_ZONE_KEYS},
        'tercio_counts': {label: 0 for label in STANDARD_TERCIO_LABELS},
        'tercio_totals': {label: 0 for label in STANDARD_TERCIO_LABELS},
        'duels_total': 0,
        'duels_won': 0,
        'shot_attempts': 0,
        'shots_on_target': 0,
        'pass_attempts': 0,
        'passes_completed': 0,
        'key_passes_completed': 0,
        'goalkeeper_saves': 0,
    }
    for event in events:
        stats['total_actions'] += 1
        if result_is_success(event.result):
            stats['successes'] += 1
        if is_goal_event(event.event_type, event.result, event.observation):
            stats['goals'] += 1
        if is_assist_event(event.event_type, event.result, event.observation):
            stats['assists'] += 1
        if is_yellow_card_event(event.event_type, event.result, event.zone):
            stats['yellow_cards'] += 1
        if is_red_card_event(event.event_type, event.result, event.zone):
            stats['red_cards'] += 1
        duel_event = classify_duel_event(event.event_type, event.result, event.observation, event.zone)
        if duel_event.get('is_duel'):
            stats['duels_total'] += 1
            if duel_event.get('won'):
                stats['duels_won'] += 1
        zone_label = _resolve_zone_label(event, match_zone_profiles, player_zone_profiles)
        if zone_label:
            stats['zone_counts'][zone_label] += 1
        tercio_raw = (event.tercio or '').strip()
        if tercio_raw:
            mapped = map_tercio(tercio_raw)
            if mapped:
                stats['tercio_counts'][mapped] += 1
                stats['tercio_totals'][mapped] += 1
        shot_event = is_shot_attempt_event(event.event_type, event.result, event.observation)
        if shot_event:
            stats['shot_attempts'] += 1
            if is_shot_on_target_event(event.event_type, event.result, event.observation):
                stats['shots_on_target'] += 1
        if is_goalkeeper_save_event(event.event_type, event.result, event.observation):
            stats['goalkeeper_saves'] += 1
        is_pass_event = (
            contains_keyword(event.event_type, PASS_KEYWORDS)
            or contains_keyword(event.observation, PASS_KEYWORDS)
            or is_assist_event(event.event_type, event.result, event.observation)
        )
        if is_pass_event:
            stats['pass_attempts'] += 1
            is_completed_pass = (
                result_is_success(event.result)
                or is_assist_event(event.event_type, event.result, event.observation)
            )
            if is_completed_pass:
                stats['passes_completed'] += 1
                if is_key_pass_event(event.event_type, event.result, event.observation):
                    stats['key_passes_completed'] += 1
    total_tercios = sum(stats['tercio_totals'].values())
    stats['success_rate'] = round(
        (stats['successes'] / stats['total_actions']) * 100, 1
    ) if stats['total_actions'] else 0
    stats['duel_rate'] = round(
        (stats['duels_won'] / stats['duels_total']) * 100, 1
    ) if stats['duels_total'] else 0
    stats['zone_heatmap'] = sorted(
        [
            {'zone': zone, 'count': count}
            for zone, count in stats['zone_counts'].items()
            if count > 0
        ],
        key=lambda entry: entry['count'],
        reverse=True,
    )[:5]
    stats['tercio_heatmap'] = sorted(
        [{'tercio': tercio, 'count': count} for tercio, count in stats['tercio_counts'].items()],
        key=lambda entry: entry['count'],
        reverse=True,
    )[:5]
    stats['tercio_summary'] = [
        {
            'label': tercio_label,
            'count': stats['tercio_totals'].get(tercio_label, 0),
            'pct': round(
                (stats['tercio_totals'].get(tercio_label, 0) / total_tercios) * 100, 1
            )
            if total_tercios
            else 0,
        }
        for tercio_label in ('Ataque', 'Construcción', 'Defensa')
    ]
    stats['shots'] = {
        'attempts': stats['shot_attempts'],
        'on_target': stats['shots_on_target'],
        'accuracy': round((stats['shots_on_target'] / stats['shot_attempts']) * 100, 1)
        if stats['shot_attempts']
        else 0,
        'per_goal': shots_needed_per_goal(stats['shot_attempts'], stats['goals']),
    }
    stats['passes'] = {
        'attempts': stats['pass_attempts'],
        'completed': stats['passes_completed'],
        'key_completed': stats['key_passes_completed'],
        'accuracy': round((stats['passes_completed'] / stats['pass_attempts']) * 100, 1)
        if stats['pass_attempts']
        else 0,
    }
    stats['kpi_summary'] = [
        {'label': 'Acciones', 'value': stats['total_actions']},
        {'label': 'Éxitos', 'value': stats['successes']},
        {'label': 'Tasa de éxito', 'value': f"{stats['success_rate']:.1f}%"},
        {'label': 'Duelos', 'value': f"{stats['duels_won']}/{stats['duels_total']}"},
        {'label': 'Duelos %', 'value': f"{stats['duel_rate']:.1f}%"},
        {'label': 'Pases', 'value': f"{stats['passes_completed']}/{stats['pass_attempts']}"},
        {'label': 'Pases clave', 'value': stats['key_passes_completed']},
        {'label': 'Pase %', 'value': f"{stats['passes']['accuracy']:.1f}%"},
        {'label': 'Disparos', 'value': f"{stats['shots_on_target']}/{stats['shot_attempts']}"},
        {'label': 'Tiro a puerta', 'value': f"{stats['shots']['accuracy']:.1f}%"},
        {'label': 'Disparos/Gol', 'value': '-' if stats['shots']['per_goal'] is None else stats['shots']['per_goal']},
        {'label': 'Goles', 'value': stats['goals']},
        {'label': 'Asistencias', 'value': stats['assists']},
        {'label': 'Paradas', 'value': stats['goalkeeper_saves']},
    ]
    total_zone_actions = sum(int(count or 0) for count in stats['zone_counts'].values())
    stats['field_zones'] = [
        {
            **zone,
            'count': stats['zone_counts'].get(zone['key'], 0),
            'pct': round((stats['zone_counts'].get(zone['key'], 0) / total_zone_actions) * 100, 1)
            if total_zone_actions
            else 0,
        }
        for zone in FIELD_ZONES
    ]
    match_payload = {
        'round': match.round or 'Partido sin jornada',
        'date': match.date.strftime('%d/%m/%Y') if match.date else 'Fecha por definir',
        'location': match.location or 'Campo por confirmar',
        'opponent': opponent.display_name if opponent else 'Rival desconocido',
        'home': match.home_team == primary_team,
    }
    return stats, match_payload


@login_required
def player_match_stats_page(request, player_id, match_id):
    forbidden = _forbid_if_workspace_module_disabled(request, 'players', label='estadísticas de jugador')
    if forbidden:
        return forbidden
    primary_team = _get_player_team_for_request(request)
    if not primary_team:
        return JsonResponse({'error': 'No hay equipo principal configurado'}, status=400)
    player = Player.objects.filter(id=player_id, team=primary_team).first()
    if not player:
        raise Http404('Jugador no encontrado')
    forbidden = _forbid_if_no_player_access(request.user, player, primary_team=primary_team)
    if forbidden:
        return forbidden
    match = _team_match_queryset(primary_team).filter(id=match_id).first()
    if not match:
        raise Http404('Partido no encontrado')
    stats, match_payload = _build_player_match_stats_payload(primary_team, player, match)
    return render(
        request,
        'football/player_match_stats.html',
        {
            'player': player,
            'stats': stats,
            'match': match_payload,
        },
    )


@csrf_exempt
@authenticated_write
@require_POST
def refresh_scraping(request):
    # Evitar CSRF 403 en Render (cookies/hosts) y limitar acceso a administradores.
    if not (_is_admin_user(request.user) or _can_access_platform(request.user)):
        return JsonResponse({'status': 'error', 'message': 'Solo administradores pueden actualizar la clasificación.'}, status=403)
    if not cache.add(SCRAPE_LOCK_KEY, "1", timeout=SCRAPE_LOCK_TIMEOUT_SECONDS):
        return JsonResponse(
            {'status': 'error', 'message': 'Ya hay una actualización en curso. Inténtalo en unos minutos.'},
            status=429,
        )
    forbidden = _forbid_if_workspace_module_disabled(request, 'dashboard', label='dashboard')
    if forbidden:
        return JsonResponse({'status': 'error', 'message': 'El dashboard no está activo en el workspace actual.'}, status=403)
    primary_team = _get_primary_team_for_request(request)
    refresh_message = ''
    next_match_payload = None
    try:
        ok, refresh_message, next_match_payload = _refresh_rfaf_standings_inline(allow_fallback=False)
        if not ok:
            raise RuntimeError(refresh_message or 'No se pudo actualizar la clasificación.')
    except Exception as exc:
        return JsonResponse({'status': 'error', 'message': str(exc)}, status=500)
    finally:
        cache.delete(SCRAPE_LOCK_KEY)
    # Si trabajamos con Universo RFAF, evitamos depender de La Preferente en este flujo.
    preferente_refresh_enabled = str(
        os.getenv('PREFERENTE_ROSTER_REFRESH_ENABLED', '0')
    ).strip().lower() in {'1', 'true', 'yes', 'on'}
    if preferente_refresh_enabled:
        roster_ok, roster_message = refresh_primary_roster_cache(primary_team, force=True)
        roster_status = 'y plantilla actualizada' if roster_ok else f'plantilla no actualizada ({roster_message})'
    else:
        roster_status = 'plantilla gestionada por Universo RFAF'
    if primary_team:
        _invalidate_team_dashboard_caches(primary_team)
        try:
            workspace = _get_active_workspace(request)
            if workspace and workspace.kind == Workspace.KIND_CLUB and int(getattr(workspace, 'primary_team_id', 0) or 0) == int(primary_team.id):
                _sync_workspace_competition_context(workspace)
            # Persistimos el próximo partido en BD para que no dependa del filesystem (Render / múltiples instancias).
            if isinstance(next_match_payload, dict) and _next_match_payload_is_reliable(next_match_payload):
                try:
                    _upsert_match_from_next_match_payload(primary_team, next_match_payload)
                except Exception:
                    pass
                club_ws = workspace
                if not club_ws or getattr(club_ws, 'kind', None) != Workspace.KIND_CLUB:
                    club_ws = (
                        Workspace.objects
                        .filter(kind=Workspace.KIND_CLUB, is_active=True, primary_team=primary_team)
                        .select_related('competition_context')
                        .first()
                    )
                if club_ws:
                    snapshot, _ = WorkspaceCompetitionSnapshot.objects.get_or_create(workspace=club_ws)
                    snapshot.context = getattr(club_ws, 'competition_context', None)
                    snapshot.next_match_payload = normalize_next_match_payload(dict(next_match_payload))
                    snapshot.save(update_fields=['context', 'next_match_payload', 'updated_at'])
        except Exception:
            pass
    latest_updated = _team_standings_last_updated(primary_team.group) if primary_team and getattr(primary_team, 'group', None) else None
    response = JsonResponse(
        {
            'status': 'success',
            'message': f'{refresh_message} {roster_status}.',
            'standings_last_updated': latest_updated.isoformat() if latest_updated else '',
        }
    )
    response['Cache-Control'] = 'no-store'
    return response


def serialize_standings(group):
    standings = TeamStanding.objects.filter(group=group)
    current_meta = standings.aggregate(total=Count('id'), latest=Max('last_updated'))

    # Si hay grupos duplicados para misma temporada/nombre, priorizar el más reciente.
    sibling_group = (
        TeamStanding.objects.filter(
            group__season=group.season,
            group__name__iexact=group.name,
        )
        .values('group_id')
        .annotate(total=Count('id'), latest=Max('last_updated'))
        .order_by('-latest', '-total')
        .first()
    )
    if sibling_group:
        sibling_is_better = (
            current_meta['total'] == 0
            or (
                sibling_group['group_id'] != group.id
                and sibling_group['latest']
                and (
                    not current_meta['latest']
                    or sibling_group['latest'] > current_meta['latest']
                )
            )
        )
        if sibling_is_better:
            standings = TeamStanding.objects.filter(group_id=sibling_group['group_id'])

    standings = standings.order_by('position')
    if not standings.exists():
        fallback = TeamStanding.objects.filter(group__slug__icontains='grupo-2').order_by('position')
        if fallback.exists():
            standings = fallback
    crest_lookup = _build_team_crest_lookup()
    return [
        {
            'rank': standing.position,
            'team': standing.team.name.strip().upper(),
            'full_name': standing.team.name.strip(),
            'crest_url': resolve_team_crest_url(
                None,
                standing.team,
                fallback_static='',
                sync=False,
            )
            or _sanitize_universo_external_image(
                _absolute_universo_url(
                    getattr(standing.team, 'crest_url', '') or crest_lookup.get(_normalize_team_lookup_key(standing.team.name)) or ''
                )
            ),
            'played': standing.played,
            'wins': standing.wins,
            'draws': standing.draws,
            'losses': standing.losses,
            'goals_for': standing.goals_for,
            'goals_against': standing.goals_against,
            'goal_difference': standing.goal_difference,
            'points': standing.points,
        }
        for standing in standings
    ]


def get_next_match(primary_team, group):
    def _pick_undated_next(queryset):
        candidates = list(queryset.filter(date__isnull=True))
        if not candidates:
            return None
        candidates.sort(
            key=lambda match: (
                extract_round_number(match.round or '') or -1,
                match.id or 0,
            ),
            reverse=True,
        )
        return candidates[0]

    def _fetch_next_from_rfaf():
        try:
            from scripts.import_from_rfef import (
                extract_next_match_from_classification,
                fetch_next_match_from_classification,
                fetch_html,
            )
            html = fetch_html()
            payload = extract_next_match_from_classification(html)
            if not payload:
                payload = fetch_next_match_from_classification(html)
            if not isinstance(payload, dict):
                return None
            if (payload.get('status') or '').lower() != 'next':
                return None
            date_raw = payload.get('date')
            if date_raw:
                try:
                    if datetime.strptime(str(date_raw), '%Y-%m-%d').date() < timezone.localdate():
                        return None
                except ValueError:
                    return None
            # Refresh file cache so subsequent requests don't depend on network.
            try:
                NEXT_MATCH_CACHE.parent.mkdir(parents=True, exist_ok=True)
                with NEXT_MATCH_CACHE.open('w', encoding='utf-8') as handle:
                    json.dump(payload, handle)
            except Exception:
                pass
            return normalize_next_match_payload(payload)
        except Exception:
            return None

    today = timezone.localdate()
    all_team_matches_qs = (
        Match.objects.filter(Q(home_team=primary_team) | Q(away_team=primary_team))
        .select_related('home_team', 'away_team')
    )
    scoped_qs = all_team_matches_qs.filter(group=group) if group else all_team_matches_qs

    snapshot = load_universo_snapshot()
    can_use_external = _universo_snapshot_supports_team(snapshot, primary_team) if primary_team else True
    cached_next = load_cached_next_match() if can_use_external else None
    if isinstance(cached_next, dict):
        normalized_cached_next = normalize_next_match_payload(cached_next)
        if _next_match_payload_is_reliable(normalized_cached_next):
            return normalized_cached_next

    upcoming = scoped_qs.filter(date__gte=today).order_by('date').first()
    if not upcoming:
        upcoming = all_team_matches_qs.filter(date__gte=today).order_by('date').first()
    if upcoming:
        return build_match_payload(upcoming, primary_team, status='next')

    undated_next = _pick_undated_next(scoped_qs)
    if not undated_next:
        undated_next = _pick_undated_next(all_team_matches_qs)
    if undated_next:
        payload = build_match_payload(undated_next, primary_team, status='next')
        if _next_match_payload_is_reliable(payload):
            return payload

    # Live scraping on request is expensive; keep it opt-in and only as fallback.
    if RFAF_LIVE_FETCH_ON_REQUEST:
        rfaf_next = _fetch_next_from_rfaf()
        if rfaf_next:
            return rfaf_next

    latest = scoped_qs.exclude(date__isnull=True).order_by('-date').first()
    if not latest:
        latest = all_team_matches_qs.exclude(date__isnull=True).order_by('-date').first()
    if not latest:
        latest = scoped_qs.order_by('-id').first()
    if not latest:
        latest = all_team_matches_qs.order_by('-id').first()
    if not latest:
        return None

    return build_match_payload(latest, primary_team, status='latest')


def _unique_team_slug(base_name):
    base_slug = slugify(base_name) or 'rival'
    slug = base_slug
    suffix = 2
    while Team.objects.filter(slug=slug).exists():
        slug = f'{base_slug}-{suffix}'
        suffix += 1
    return slug


def _ensure_platform_team(team_name, *, region=''):
    clean_name = _sanitize_task_text(str(team_name or '').strip(), multiline=False, max_len=150)
    if not clean_name:
        return None
    existing = Team.objects.filter(name__iexact=clean_name).first()
    if existing:
        has_season = bool(getattr(getattr(existing.group, 'season', None), 'id', None)) if getattr(existing, 'group_id', None) else False
        if has_season:
            return existing
    demo_competition, _ = Competition.objects.get_or_create(
        name='Liga Demo',
        region=str(region or '').strip(),
        defaults={'slug': slugify('Liga Demo') or 'liga-demo'},
    )
    demo_season, _ = Season.objects.get_or_create(
        competition=demo_competition,
        name='2025/2026',
        defaults={'is_current': True},
    )
    group_slug = slugify(f'demo-{clean_name}')[:80] or 'demo'
    demo_group, _ = Group.objects.get_or_create(
        season=demo_season,
        slug=group_slug,
        defaults={'name': 'Grupo Demo'},
    )
    if existing:
        existing.group = demo_group
        existing.save(update_fields=['group'])
        return existing
    return Team.objects.create(
        name=clean_name,
        slug=_unique_team_slug(clean_name),
        short_name=clean_name[:60],
        group=demo_group,
    )


def _bootstrap_demo_club_workspace(workspace):
    if not workspace or workspace.kind != Workspace.KIND_CLUB or not workspace.primary_team_id:
        return {'players': 0, 'matches': 0, 'events': 0, 'sessions': 0, 'tasks': 0, 'stats': 0}
    team = workspace.primary_team
    if not team:
        return {'players': 0, 'matches': 0, 'events': 0, 'sessions': 0, 'tasks': 0, 'stats': 0}
    created = {'players': 0, 'matches': 0, 'events': 0, 'sessions': 0, 'tasks': 0, 'stats': 0}
    today = timezone.localdate()
    monday = today - timedelta(days=today.weekday())
    sunday = monday + timedelta(days=6)

    # Plantilla (players)
    if Player.objects.filter(team=team).count() < 8:
        for number in range(1, 19):
            player_name = f'Jugador {number}'
            obj, was_created = Player.objects.get_or_create(
                team=team,
                name=player_name,
                defaults={
                    'number': number,
                    'position': 'DEF' if number in {2, 3, 4, 5} else ('MED' if number in {6, 7, 8, 10} else ('DEL' if number in {9, 11, 12} else '')),
                    'is_active': True,
                },
            )
            if was_created:
                created['players'] += 1

    # Contexto (season/group) para partido y estadísticas
    group = team.group
    season = getattr(group, 'season', None) if group else None
    if not season:
        demo_team = _ensure_platform_team(team.name)
        team = demo_team or team
        workspace.primary_team = team
        workspace.save(update_fields=['primary_team'])
        group = team.group
        season = getattr(group, 'season', None) if group else None

    # Partido demo + eventos
    opponent = Team.objects.filter(name__iexact='Rival Demo').first()
    if not opponent:
        opponent = _ensure_platform_team('Rival Demo')
    match = None
    if season:
        match = Match.objects.filter(season=season, home_team=team).order_by('-id').first()
        if not match:
            match = Match.objects.create(
                season=season,
                group=group,
                round='Jornada Demo',
                date=saturday_date(monday),
                location='Campo Demo',
                home_team=team,
                away_team=opponent,
                home_score=2,
                away_score=1,
                result='2-1',
                notes='Partido de ejemplo para onboarding.',
            )
            created['matches'] += 1
    if match and MatchEvent.objects.filter(match=match).count() < 3:
        sample_players = list(Player.objects.filter(team=team).order_by('number', 'id')[:6])
        base_events = [
            (1, 8, 'Inicio', 'Kick-off'),
            (1, 12, 'Pase', 'Completo'),
            (1, 24, 'Robo', 'Recuperación alta'),
            (2, 52, 'Disparo', 'A puerta'),
            (2, 68, 'Gol', 'Finalización'),
        ]
        for idx, (period, minute, event_type, result) in enumerate(base_events):
            player = sample_players[idx % len(sample_players)] if sample_players else None
            obj, was_created = MatchEvent.objects.get_or_create(
                match=match,
                period=period,
                minute=minute,
                event_type=event_type,
                defaults={
                    'player': player,
                    'result': result,
                    'zone': 'Zona 2',
                    'tercio': 'Medio',
                    'observation': 'Evento demo',
                    'system': '4-3-3',
                    'source_file': 'demo',
                    'raw_data': {'demo': True},
                },
            )
            if was_created:
                created['events'] += 1

    # Microciclo + sesiones + tareas
    microcycle, mc_created = TrainingMicrocycle.objects.get_or_create(
        team=team,
        week_start=monday,
        defaults={
            'week_end': sunday,
            'title': 'Microciclo demo',
            'objective': 'Onboarding de sesiones y tareas.',
            'reference_match': match if match else None,
        },
    )
    if mc_created:
        created['sessions'] += 0
    session_dates = [monday + timedelta(days=1), monday + timedelta(days=3), monday + timedelta(days=5)]
    for index, session_date in enumerate(session_dates):
        session, was_created = TrainingSession.objects.get_or_create(
            microcycle=microcycle,
            session_date=session_date,
            defaults={
                'start_time': None,
                'duration_minutes': 90,
                'intensity': TrainingSession.INTENSITY_MEDIUM,
                'focus': ['Construcción', 'Presión', 'Finalización'][index],
                'content': 'Sesión de ejemplo para un usuario de prueba.',
                'status': TrainingSession.STATUS_PLANNED,
                'order': index,
            },
        )
        if was_created:
            created['sessions'] += 1
        if session.tasks.filter(deleted_at__isnull=True).count() < 2:
            base_tasks = [
                ('Rondo 6v3', SessionTask.BLOCK_ACTIVATION, 12, 'Activar + orientación corporal'),
                ('Juego de posición', SessionTask.BLOCK_MAIN_1, 20, 'Progresar y fijar'),
            ]
            for order_idx, (title, block, minutes, objective) in enumerate(base_tasks):
                task, task_created = SessionTask.objects.get_or_create(
                    session=session,
                    title=title,
                    defaults={
                        'block': block,
                        'duration_minutes': minutes,
                        'objective': objective,
                        'coaching_points': '- Perfil corporal\n- Pase tenso\n- Apoyos y tercer hombre',
                        'confrontation_rules': '- 1 punto por 6 pases\n- 2 puntos por robo y salida',
                        'tactical_layout': {
                            'tokens': [],
                            'timeline': [],
                            'meta': {
                                'strategy': 'passing_wheel',
                                'complexity': 'low',
                                'dynamics': 'extensive',
                                'structure': 'complete',
                                'coordination': 'team',
                                'tactical_intent': 'direct',
                            },
                        },
                        'status': SessionTask.STATUS_PLANNED,
                        'order': order_idx,
                    },
                )
                if task_created:
                    created['tasks'] += 1

    # Estadísticas rápidas
    if season:
        ds, _ = DataSource.objects.get_or_create(name='Demo', defaults={'base_url': '', 'notes': 'Datos de demostración.'})
        for name, value, context in [
            ('Posesión', 56.0, 'Global'),
            ('Disparos', 12.0, 'Último partido'),
            ('Recuperaciones', 38.0, 'Último partido'),
        ]:
            obj, was_created = TeamStatistic.objects.get_or_create(
                team=team,
                season=season,
                name=name,
                context=context,
                defaults={'value': float(value), 'source': ds},
            )
            if was_created:
                created['stats'] += 1
        first_player = Player.objects.filter(team=team).order_by('number', 'id').first()
        if first_player and match:
            obj, was_created = PlayerStatistic.objects.get_or_create(
                player=first_player,
                season=season,
                match=match,
                name='Minutos',
                context='Jornada Demo',
                defaults={'value': 90.0, 'source': ds},
            )
            if was_created:
                created['stats'] += 1
        for name, value in [('RPE medio', 7.2), ('Distancia total', 102.4)]:
            existing_metric = (
                CustomMetric.objects
                .filter(team=team, season=season, name=name, recorded_at__date=today)
                .order_by('-recorded_at')
                .first()
            )
            if not existing_metric:
                CustomMetric.objects.create(
                    team=team,
                    season=season,
                    name=name,
                    value=float(value),
                    recorded_at=timezone.now().replace(minute=0, second=0, microsecond=0),
                    source_notes='Demo',
                )
                created['stats'] += 1
    return created


def saturday_date(monday_date):
    if not monday_date:
        return None
    return monday_date + timedelta(days=5)


def _apply_match_info_overrides(match, primary_team, match_info_payload):
    if not match or not isinstance(match_info_payload, dict):
        return
    changed_fields = []
    round_value = (match_info_payload.get('round') or '').strip()
    location_value = (match_info_payload.get('location') or '').strip()
    datetime_value = (match_info_payload.get('datetime') or '').strip()
    opponent_name = (match_info_payload.get('opponent') or '').strip()

    if round_value != (match.round or ''):
        match.round = round_value
        changed_fields.append('round')
    if location_value != (match.location or ''):
        match.location = location_value
        changed_fields.append('location')

    parsed_date = parse_match_date_from_ui(datetime_value)
    if parsed_date and parsed_date != match.date:
        match.date = parsed_date
        changed_fields.append('date')

    if opponent_name and normalize_label(opponent_name) != normalize_label(primary_team.name):
        rival_team = Team.objects.filter(name__iexact=opponent_name).first()
        if not rival_team:
            rival_team = Team.objects.create(
                name=opponent_name,
                slug=_unique_team_slug(opponent_name),
                short_name=opponent_name[:60],
                group=match.group or primary_team.group,
            )
        if match.home_team_id == primary_team.id:
            if match.away_team_id != rival_team.id:
                match.away_team = rival_team
                changed_fields.append('away_team')
        elif match.away_team_id == primary_team.id:
            if match.home_team_id != rival_team.id:
                match.home_team = rival_team
                changed_fields.append('home_team')
        else:
            match.home_team = primary_team
            match.away_team = rival_team
            changed_fields.extend(['home_team', 'away_team'])

    if changed_fields:
        match.save(update_fields=list(dict.fromkeys(changed_fields)))


def gather_team_fields_for_group(group):
    if not group:
        return []
    seen = set()
    fields = []
    matches = (
        Match.objects.filter(group=group, home_team__isnull=False)
        .exclude(location__isnull=True)
        .exclude(location__exact='')
        .select_related('home_team')
        .order_by('home_team__name', '-date')
    )
    for match in matches:
        team = match.home_team
        if not team or team.id in seen:
            continue
        seen.add(team.id)
        fields.append(
            {
                'team_slug': team.slug,
                'team_name': team.name,
                'location': match.location.strip(),
            }
        )
    return fields


def build_match_payload(match, primary_team, status):
    opponent = match.away_team if match.home_team == primary_team else match.home_team
    return normalize_next_match_payload({
        'round': match.round,
        'date': match.date.isoformat() if match.date else None,
        'location': match.location,
        'opponent': {
            'name': opponent.name if opponent else 'Rival desconocido',
            'full_name': opponent.name if opponent else 'Rival desconocido',
            'crest_url': '',
            'team_code': '',
        },
        'home': match.home_team == primary_team,
        'status': status,
        'source': 'local-match',
    })


def preferred_event_source_by_match(primary_team):
    """
    Choose one authoritative source per match to avoid cross-source double counting.
    Priority:
    1) Any `registro-acciones` events for that match.
    2) Otherwise, most frequent non-empty source_file.
    """
    if not primary_team:
        return {}
    team_events = (
        MatchEvent.objects
        .filter(player__team=primary_team)
        .filter(
            Q(source_file='registro-acciones')
            | ~Q(system='touch-field')
        )
    )
    preferred = {}
    registro_match_ids = set(
        team_events.filter(source_file='registro-acciones')
        .values_list('match_id', flat=True)
        .distinct()
    )
    for match_id in registro_match_ids:
        preferred[match_id] = 'registro-acciones'

    fallback_rows = (
        team_events.exclude(source_file__isnull=True)
        .exclude(source_file__exact='')
        .values('match_id', 'source_file')
        .annotate(c=Count('id'))
        .order_by('match_id', '-c', 'source_file')
    )
    seen = set(preferred.keys())
    for row in fallback_rows:
        match_id = row['match_id']
        if match_id in seen:
            continue
        preferred[match_id] = row['source_file']
        seen.add(match_id)
    return preferred


def _is_manual_event_source(source_file):
    normalized = (source_file or '').strip().lower()
    return normalized == 'admin-manual' or 'manual' in normalized


def _event_matches_stats_source(event, preferred_sources=None):
    if not preferred_sources:
        return True
    preferred_source = preferred_sources.get(getattr(event, 'match_id', None))
    if not preferred_source:
        return True
    current_source = (getattr(event, 'source_file', '') or '').strip()
    return current_source == preferred_source or _is_manual_event_source(current_source)


def _filter_stats_events(rows, preferred_sources=None):
    seen_signatures = set()
    filtered = []
    for event in rows:
        if not _event_matches_stats_source(event, preferred_sources):
            continue
        signature = _event_signature(event)
        if signature in seen_signatures:
            continue
        seen_signatures.add(signature)
        filtered.append(event)
    return filtered


def _build_zone_inference_profiles(events):
    match_profiles = defaultdict(Counter)
    player_profiles = defaultdict(Counter)
    for event in events:
        player_id = getattr(event, 'player_id', None)
        zone_label = map_zone_label((getattr(event, 'zone', '') or '').strip())
        if not player_id or not zone_label:
            continue
        player_profiles[player_id][zone_label] += 1
        match_id = getattr(event, 'match_id', None)
        if match_id:
            match_profiles[(player_id, match_id)][zone_label] += 1
    return match_profiles, player_profiles


def _resolve_zone_label(event, match_profiles=None, player_profiles=None):
    direct_zone = map_zone_label((getattr(event, 'zone', '') or '').strip())
    if direct_zone:
        return direct_zone
    player_id = getattr(event, 'player_id', None)
    if not player_id:
        return None
    match_id = getattr(event, 'match_id', None)
    if match_profiles and match_id:
        match_counter = match_profiles.get((player_id, match_id))
        if match_counter:
            top_zone, top_count = match_counter.most_common(1)[0]
            if top_count > 0:
                return top_zone
    if player_profiles:
        player_counter = player_profiles.get(player_id)
        if player_counter:
            top_zone, top_count = player_counter.most_common(1)[0]
            if top_count > 0:
                return top_zone
    return None


def _normalize_excel_header(value):
    if not value:
        return ''
    return ''.join(ch.lower() for ch in str(value).strip() if ch.isalnum())


def append_events_to_bd_eventos(match, primary_team, events):
    if not events:
        return 0
    if os.getenv('WRITE_BD_EVENTOS_EXCEL', 'false').lower() != 'true':
        return 0
    path = Path(settings.BASE_DIR) / 'data' / 'excel' / 'BDT PARTIDOS BENABALBON.xlsm'
    if not path.exists():
        raise FileNotFoundError('No se encuentra el Excel de BD_EVENTOS.')
    try:
        from openpyxl import load_workbook
    except Exception as exc:
        raise RuntimeError('No se pudo cargar openpyxl para escribir BD_EVENTOS.') from exc
    workbook = load_workbook(path, keep_vba=True)
    if 'BD_EVENTOS' not in workbook.sheetnames:
        raise RuntimeError('La hoja BD_EVENTOS no existe en el Excel.')
    sheet = workbook['BD_EVENTOS']
    header_row = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not header_row:
        raise RuntimeError('La hoja BD_EVENTOS no tiene cabecera.')
    headers = [_normalize_excel_header(cell) for cell in header_row]
    header_index = {key: idx for idx, key in enumerate(headers) if key}
    opponent = match.away_team if match.home_team == primary_team else match.home_team
    match_date = match.date or timezone.localdate()
    match_round = match.round or f'Partido {match.id}'
    match_location = match.location or ''
    rival_name = opponent.name if opponent else 'Rival desconocido'
    rows_written = 0
    for event in events:
        row_payload = {
            'partidoid': match.id,
            'fecha': match_date,
            'rival': rival_name,
            'jornada': match_round,
            'campo': match_location,
            'sistema': event.system or 'touch-field-final',
            'minuto': event.minute,
            'jugador': event.player.name if event.player else '',
            'evento': event.event_type,
            'resultadoaccion': event.result,
            'zona': event.zone,
            'tercio': event.tercio,
            'observacion': event.observation,
        }
        row_values = [''] * len(headers)
        for key, value in row_payload.items():
            index = header_index.get(key)
            if index is not None:
                row_values[index] = value
        sheet.append(row_values)
        rows_written += 1
    workbook.save(path)
    return rows_written


def compute_team_metrics(primary_team):
    if not primary_team:
        return {'total_events': 0, 'top_event_types': [], 'top_results': []}
    cache_key = _team_metrics_cache_key(primary_team.id)
    cached = cache.get(cache_key)
    if isinstance(cached, dict) and cached:
        return cached
    preferred_sources = preferred_event_source_by_match(primary_team)
    events = _filter_stats_events(
        confirmed_events_queryset()
        .filter(player__team=primary_team)
        .select_related('match')
        .order_by('match_id', 'minute', 'id'),
        preferred_sources=preferred_sources,
    )
    total_events = len(events)
    event_counter = Counter(event.event_type for event in events)
    result_counter = Counter(event.result for event in events)

    top_events = [{'event': etype, 'count': count} for etype, count in event_counter.most_common(5)]
    top_results = [{'result': result, 'count': count} for result, count in result_counter.most_common(5)]

    payload = {
        'total_events': total_events,
        'top_event_types': top_events,
        'top_results': top_results,
    }
    cache.set(cache_key, payload, TEAM_METRICS_CACHE_SECONDS)
    return payload


def compute_team_metrics_for_match(match, primary_team=None):
    events_qs = confirmed_events_queryset().filter(match=match)
    preferred_sources = None
    if primary_team:
        events_qs = events_qs.filter(player__team=primary_team)
        preferred_sources = preferred_event_source_by_match(primary_team)
    events = _filter_stats_events(
        events_qs.select_related('player', 'match').order_by('minute', 'id'),
        preferred_sources=preferred_sources,
    )
    total_events = len(events)
    event_counter = Counter(event.event_type for event in events)
    result_counter = Counter(event.result for event in events)
    top_events = [{'event': etype, 'count': count} for etype, count in event_counter.most_common(6)]
    top_results = [{'result': result, 'count': count} for result, count in result_counter.most_common(6)]
    return {
        'total_events': total_events,
        'top_event_types': top_events,
        'top_results': top_results,
    }


def _event_signature(event):
    def canon(value):
        text = ' '.join(str(value or '').split())
        text = unicodedata.normalize('NFKD', text)
        text = ''.join(ch for ch in text if not unicodedata.combining(ch))
        return text.lower().strip()

    if _is_manual_event_source(getattr(event, 'source_file', '')):
        return ('manual-event', getattr(event, 'id', None))

    # "Acción realizada" for dashboard cards:
    # keep one action per player/match/minute/type.
    # If minute is missing, retain more fields to avoid over-collapsing.
    minute = event.minute
    action_type = canon(event.event_type)
    if minute is not None:
        return (
            event.match_id,
            event.player_id,
            minute,
            action_type,
        )
    return (
        event.match_id,
        event.player_id,
        minute,
        action_type,
        canon(event.result),
        canon(event.zone),
        canon(event.tercio),
        canon(event.observation),
    )


def _match_action_dedupe_signature(event):
    return (
        getattr(event, 'match_id', None),
        getattr(event, 'player_id', None),
        getattr(event, 'minute', None),
        _canonical_action_value(getattr(event, 'event_type', '')),
        _canonical_action_value(getattr(event, 'result', '')),
        _canonical_action_value(getattr(event, 'zone', '')),
        _canonical_action_value(getattr(event, 'tercio', '')),
        _canonical_action_value(getattr(event, 'observation', '')),
    )


def compute_player_cards_for_match(match, primary_team, source_file=None):
    events = confirmed_events_queryset().filter(match=match, player__team=primary_team)
    if source_file:
        events = events.filter(source_file=source_file)
        preferred_sources = None
    else:
        preferred_sources = preferred_event_source_by_match(primary_team)
    rows = _filter_stats_events(
        events.select_related('player', 'match').order_by('minute', 'id'),
        preferred_sources=preferred_sources,
    )
    per_player = {}
    for event in rows:
        player = event.player
        if not player:
            continue
        photo_path = resolve_player_photo_static_path(player)
        data = per_player.setdefault(
            player.id,
            {
                'player_id': player.id,
                'name': player.name,
                'number': player.number or '--',
                'photo_url': static(photo_path) if photo_path else '',
                'actions': 0,
                'successes': 0,
            },
        )
        data['actions'] += 1
        if result_is_success(event.result):
            data['successes'] += 1
    cards = list(per_player.values())
    for item in cards:
        total_actions = item['actions']
        success = item['successes']
        item['success_rate'] = round((success / total_actions) * 100, 1) if total_actions else 0
    return sorted(cards, key=lambda item: item['actions'], reverse=True)

def compute_player_metrics(primary_team):
    if not primary_team:
        return []
    cache_key = _player_metrics_cache_key(primary_team.id)
    cached = cache.get(cache_key)
    if isinstance(cached, list) and cached:
        return cached
    preferred_sources = preferred_event_source_by_match(primary_team)
    events = _filter_stats_events(
        confirmed_events_queryset()
        .filter(player__team=primary_team)
        .select_related('player', 'match')
        .order_by('match_id', 'minute', 'id'),
        preferred_sources=preferred_sources,
    )
    per_player = {}
    for event in events:
        player = event.player
        if not player:
            continue
        item = per_player.setdefault(
            player.id,
            {
                'player_id': player.id,
                'player': player.name,
                'actions': 0,
                'successes': 0,
            },
        )
        item['actions'] += 1
        if result_is_success(event.result):
            item['successes'] += 1
    result = sorted(per_player.values(), key=lambda item: (-item['actions'], item['player']))
    cache.set(cache_key, result, PLAYER_METRICS_CACHE_SECONDS)
    return result


def compute_player_cards(primary_team):
    if not primary_team:
        return []
    dashboard_rows = compute_player_dashboard(primary_team)
    cards = []
    for row in dashboard_rows:
        cards.append(
            {
                'player_id': row.get('player_id'),
                'name': row.get('name'),
                'photo_url': row.get('photo_url', ''),
                'pj': int(row.get('pj', 0) or 0),
                'minutes': int(row.get('minutes', 0) or 0),
                'goals': int(row.get('goals', 0) or 0),
                'yellow_cards': int(row.get('yellow_cards', 0) or 0),
                'red_cards': int(row.get('red_cards', 0) or 0),
                'total_actions': int(row.get('total_actions', 0) or 0),
                'successes': int(row.get('successes', 0) or 0),
                'shot_attempts': int(row.get('shot_attempts', 0) or 0),
                'shots_on_target': int(row.get('shots_on_target', 0) or 0),
                'duels_total': int(row.get('duels_total', 0) or 0),
                'duels_won': int(row.get('duels_won', 0) or 0),
                'success_rate': float(row.get('success_rate', 0) or 0),
                'has_active_injury': bool(row.get('has_active_injury')),
                'is_sanctioned': bool(row.get('is_sanctioned')),
                'is_apercibido': bool(row.get('is_apercibido')),
                'position': row.get('position') or '',
            }
        )
    return sorted(cards, key=lambda entry: (-entry['goals'], -entry['pj'], entry['name']))


def compute_player_dashboard(primary_team, force_refresh=False):
    if not primary_team:
        return []
    cache_key = _player_dashboard_cache_key(primary_team.id)
    if not force_refresh:
        cached_rows = cache.get(cache_key)
        if isinstance(cached_rows, list):
            return cached_rows
    player_stats = {}
    competition_total_rounds = get_competition_total_rounds(primary_team)
    roster_players = list(Player.objects.filter(team=primary_team))
    player_by_id = {player.id: player for player in roster_players}
    active_injury_ids = get_active_injury_player_ids([player.id for player in roster_players])
    active_match = get_active_match(primary_team)
    sanctioned_player_ids = get_sanctioned_player_ids_from_previous_round(
        primary_team,
        reference_match=active_match,
    )
    roster_cache = get_roster_stats_cache()
    manual_overrides = get_manual_player_base_overrides(primary_team)
    universo_snapshot = load_universo_snapshot() or {}
    universo_players = universo_snapshot.get('players') if isinstance(universo_snapshot, dict) else []
    universo_map = {}
    universo_by_number = {}
    if isinstance(universo_players, list):
        for item in universo_players:
            if not isinstance(item, dict):
                continue
            team_name = str(item.get('team') or '').strip().lower()
            if team_name and 'benagalbon' not in team_name:
                continue
            name = str(item.get('name') or '').strip()
            if not name:
                continue
            key = normalize_player_name(name)
            universo_map[key] = item
            dorsal_raw = str(item.get('dorsal') or '').strip()
            if dorsal_raw.isdigit():
                universo_by_number[int(dorsal_raw)] = item

    def _find_universo_entry(player_obj):
        key = normalize_player_name(player_obj.name)
        direct = universo_map.get(key)
        if direct:
            return direct
        if player_obj.number is not None and player_obj.number in universo_by_number:
            return universo_by_number[player_obj.number]
        compact = key.replace('-', '')
        for ukey, entry in universo_map.items():
            ucompact = ukey.replace('-', '')
            if compact in ucompact or ucompact in compact:
                return entry
        tokens = [token for token in compact.split('-') if token]
        for ukey, entry in universo_map.items():
            u_tokens = [token for token in ukey.replace('-', ' ').split() if token]
            overlap = sum(1 for token in tokens if token in u_tokens)
            if overlap >= 2:
                return entry
        return {}

    roster_entry_by_player_id = {player.id: (find_roster_entry(player.name, roster_cache) or {}) for player in roster_players}
    manual_entry_by_player_id = manual_overrides if isinstance(manual_overrides, dict) else {}
    universo_entry_by_player_id = {player.id: (_find_universo_entry(player) or {}) for player in roster_players}
    player_photo_url_by_id = {}
    for player in roster_players:
        photo_path = resolve_player_photo_static_path(player)
        player_photo_url_by_id[player.id] = resolve_player_photo_url(None, player) or (static(photo_path) if photo_path else '')
    preferred_sources = preferred_event_source_by_match(primary_team)
    lineup_by_match = {}
    convocation_qs = (
        ConvocationRecord.objects.filter(team=primary_team, match__isnull=False)
        .exclude(lineup_data={})
        .prefetch_related('players')
        .order_by('match_id', '-created_at')
    )
    for record in convocation_qs:
        if record.match_id in lineup_by_match:
            continue
        allowed = list(record.players.all())
        normalized = _normalize_lineup_payload(record.lineup_data if isinstance(record.lineup_data, dict) else {}, allowed)
        lineup_by_match[record.match_id] = normalized
    match_end_minutes = {}
    player_match_timeline = {}
    stats_events = (
        MatchEvent.objects
        .filter(player__team=primary_team)
        .filter(
            Q(system='touch-field', source_file='registro-acciones')
            | ~Q(system='touch-field')
        )
    )
    events = (
        stats_events
        .select_related('player', 'match', 'match__home_team', 'match__away_team')
        .order_by('player__name', 'match__date')
    )
    inferred_zone_events = _filter_stats_events(events, preferred_sources=preferred_sources)
    match_zone_profiles, player_zone_profiles = _build_zone_inference_profiles(inferred_zone_events)
    live_events = (
        MatchEvent.objects.filter(
            Q(player__team=primary_team) | Q(player__isnull=True),
            Q(match__home_team=primary_team) | Q(match__away_team=primary_team),
        ).filter(
            Q(system='touch-field-final')
            | Q(system='touch-field', source_file='registro-acciones')
        )
        .select_related('player', 'match', 'match__home_team', 'match__away_team')
        .order_by('player__name', 'match__date')
    )
    seen_signatures = set()
    for event in events:
        player = event.player
        if not player:
            continue
        resolved_photo_url = player_photo_url_by_id.get(player.id, '')
        match = event.match
        preferred_source = preferred_sources.get(match.id if match else None)
        event_source = (event.source_file or '').strip().lower()
        # Keep admin/manual edits visible in player stats even when another source
        # is preferred for that match.
        is_manual_source = event_source == 'admin-manual' or 'manual' in event_source
        if preferred_source and (event.source_file or '') != preferred_source and not is_manual_source:
            continue
        signature = _event_signature(event)
        if signature in seen_signatures:
            continue
        seen_signatures.add(signature)
        roster_entry = roster_entry_by_player_id.get(player.id, {})
        manual_entry = manual_entry_by_player_id.get(player.id, {})
        universo_entry = universo_entry_by_player_id.get(player.id, {})
        base_pj = (
            manual_entry.get('pj')
            if manual_entry.get('pj') is not None
            else _parse_int(universo_entry.get('pj'))
            if universo_entry.get('pj') not in (None, '')
            else roster_entry.get('pj', 0)
        )
        base_pt = (
            manual_entry.get('pt')
            if manual_entry.get('pt') is not None
            else _parse_int(universo_entry.get('pt'))
            if universo_entry.get('pt') not in (None, '')
            else roster_entry.get('pt', 0)
        )
        base_minutes = (
            manual_entry.get('minutes')
            if manual_entry.get('minutes') is not None
            else _parse_int(universo_entry.get('minutes'))
            if universo_entry.get('minutes') not in (None, '')
            else roster_entry.get('minutes', 0)
        )
        base_pc = max(roster_entry.get('pc', 0), base_pj)
        base_goals = (
            manual_entry.get('goals')
            if manual_entry.get('goals') is not None
            else _parse_int(universo_entry.get('goals'))
            if universo_entry.get('goals') not in (None, '')
            else roster_entry.get('goals', 0)
        )
        base_yellow = (
            manual_entry.get('yellow_cards')
            if manual_entry.get('yellow_cards') is not None
            else _parse_int(universo_entry.get('yellow_cards'))
            if universo_entry.get('yellow_cards') not in (None, '')
            else roster_entry.get('yellow_cards', 0)
        )
        base_red = (
            manual_entry.get('red_cards')
            if manual_entry.get('red_cards') is not None
            else _parse_int(universo_entry.get('red_cards'))
            if universo_entry.get('red_cards') not in (None, '')
            else roster_entry.get('red_cards', 0)
        )
        base_assists = roster_entry.get('assists', 0)
        stats = player_stats.setdefault(
            player.id,
            {
                'player_id': player.id,
                'name': player.name,
                'number': player.number,
                'photo_url': resolved_photo_url,
                'position': player.position or universo_entry.get('position') or roster_entry.get('position', ''),
                'total_actions': 0,
                'successes': 0,
                'pc': base_pc,
                'pj': base_pj,
                'pt': base_pt,
                'minutes': base_minutes,
                'goals': base_goals,
                'yellow_cards': base_yellow,
                'red_cards': base_red,
                'assists': base_assists,
                # Only manual-base overrides should lock auto aggregation.
                'totals_locked': any(
                    key in manual_entry
                    for key in ('pj', 'pt', 'minutes', 'goals', 'yellow_cards', 'red_cards')
                ),
                'matches': {},
                'zone_counts': {key: 0 for key in FIELD_ZONE_KEYS},
                'position_counts': {key: 0 for key in FIELD_ZONE_KEYS},
                'tercio_counts': {label: 0 for label in STANDARD_TERCIO_LABELS},
                'tercio_totals': {label: 0 for label in STANDARD_TERCIO_LABELS},
                'duels_total': 0,
                'duels_won': 0,
                'shot_attempts': 0,
                'shots_on_target': 0,
                'pass_attempts': 0,
                'passes_completed': 0,
                'key_passes_completed': 0,
                'goalkeeper_saves': 0,
                'dribbles_attempted': 0,
                'dribbles_completed': 0,
                'age': _parse_int(universo_entry.get('age')) or roster_entry.get('age'),
                'has_events': False,
            },
        )
        if base_pj > 0:
            stats['has_events'] = True
        stats['total_actions'] += 1
        if result_is_success(event.result):
            stats['successes'] += 1
        if (not stats['totals_locked']) and is_goal_event(event.event_type, event.result, event.observation):
            stats['goals'] += 1
        if (not stats['totals_locked']) and is_assist_event(event.event_type, event.result, event.observation):
            stats['assists'] += 1
        if (not stats['totals_locked']) and is_yellow_card_event(event.event_type, event.result, event.zone):
            stats['yellow_cards'] += 1
        if (not stats['totals_locked']) and is_red_card_event(event.event_type, event.result, event.zone):
            stats['red_cards'] += 1
        duel_event = classify_duel_event(event.event_type, event.result, event.observation, event.zone)
        if duel_event.get('is_duel'):
            stats['duels_total'] += 1
            if duel_event.get('won'):
                stats['duels_won'] += 1
        zone_label = _resolve_zone_label(event, match_zone_profiles, player_zone_profiles)
        if zone_label:
            stats['zone_counts'][zone_label] += 1
        tercio = (event.tercio or '').strip()
        if tercio:
            mapped = map_tercio(tercio)
            if mapped:
                stats['tercio_counts'][mapped] += 1
                stats['tercio_totals'][mapped] += 1
        position_label = categorize_position(player.position, event.zone)
        if position_label:
            stats['position_counts'][position_label] += 1
        shot_event = is_shot_attempt_event(event.event_type, event.result, event.observation)
        if shot_event:
            stats['shot_attempts'] += 1
            if is_shot_on_target_event(event.event_type, event.result, event.observation):
                stats['shots_on_target'] += 1
        if is_goalkeeper_save_event(event.event_type, event.result, event.observation):
            stats['goalkeeper_saves'] += 1
        is_pass_event = (
            contains_keyword(event.event_type, PASS_KEYWORDS)
            or contains_keyword(event.observation, PASS_KEYWORDS)
            or is_assist_event(event.event_type, event.result, event.observation)
        )
        if is_pass_event:
            stats['pass_attempts'] += 1
            is_completed_pass = (
                result_is_success(event.result)
                or is_assist_event(event.event_type, event.result, event.observation)
            )
            if is_completed_pass:
                stats['passes_completed'] += 1
                if is_key_pass_event(event.event_type, event.result, event.observation):
                    stats['key_passes_completed'] += 1
        if contains_keyword(event.event_type, DRIBBLE_KEYWORDS) or contains_keyword(event.observation, DRIBBLE_KEYWORDS):
            stats['dribbles_attempted'] += 1
            if result_is_success(event.result):
                stats['dribbles_completed'] += 1
        if not match:
            continue
        match_key = match.id
        match_entry = stats['matches'].setdefault(
            match_key,
            {
                'match_id': match.id,
                'round': match.round or 'Partido sin jornada',
                'date': match.date.isoformat() if match.date else None,
                'home': match.home_team == primary_team,
                'opponent': (
                    match.away_team.display_name
                    if match.home_team == primary_team and match.away_team
                    else match.home_team.display_name
                    if match.away_team == primary_team and match.home_team
                    else 'Rival desconocido'
                ),
                'actions': 0,
                'successes': 0,
            },
        )
        match_entry['actions'] += 1
        if result_is_success(event.result):
            match_entry['successes'] += 1
        match_entry['success_rate'] = round(
            (match_entry['successes'] / match_entry['actions']) * 100
        ) if match_entry['actions'] else 0
    for event in live_events:
        match = event.match
        if match and event.minute is not None:
            match_end_minutes[match.id] = max(match_end_minutes.get(match.id, 0), event.minute)
        player = event.player
        if not player:
            continue
        if match:
            timeline = player_match_timeline.setdefault(player.id, {}).setdefault(
                match.id,
                {'entry': None, 'exit': None, 'has_event': False},
            )
            timeline['has_event'] = True
            if is_substitution_entry(event.event_type, event.result, event.zone):
                timeline['entry'] = min_or_none(timeline['entry'], event.minute or 0)
            if is_substitution_exit(event.event_type, event.result, event.zone):
                timeline['exit'] = min_or_none(timeline['exit'], event.minute or 0)
    processed_lineup_matches = defaultdict(set)
    for player_id, matches in player_match_timeline.items():
        stats = player_stats.get(player_id)
        if not stats:
            continue
        for match_id, timeline in matches.items():
            processed_lineup_matches[player_id].add(match_id)
            match_end = match_end_minutes.get(match_id, 0)
            lineup_seed = lineup_by_match.get(match_id) or {}
            lineup_starters = {
                str(item.get('id'))
                for item in (lineup_seed.get('starters') or [])
                if isinstance(item, dict)
            }
            player_key = str(player_id)
            entry_minute = timeline.get('entry')
            exit_minute = timeline.get('exit')
            if entry_minute is None and player_key in lineup_starters:
                entry_minute = 0
            if entry_minute is None:
                entry_minute = 0
            if match_end <= 0 and player_key in lineup_starters:
                match_end = 90
            if exit_minute is None:
                exit_minute = match_end
            if exit_minute is None:
                exit_minute = entry_minute
            if exit_minute < entry_minute:
                exit_minute = entry_minute
            if not stats.get('totals_locked'):
                played_match = bool(timeline.get('has_event')) or (player_key in lineup_starters)
                stats['minutes'] += max(0, exit_minute - entry_minute)
                stats['pj'] += 1 if played_match else 0
                if player_key in lineup_starters:
                    stats['pt'] += 1
        if not stats.get('totals_locked'):
            stats['pc'] = max(stats.get('pc', 0), stats['pj'])
    # ensure roster players appear even without events
    for player in roster_players:
        if player.id not in player_stats:
            photo_path = resolve_player_photo_static_path(player)
            normalized = normalize_player_name(player.name)
            roster_entry = roster_cache.get(normalized, {})
            manual_entry = manual_overrides.get(player.id, {})
            universo_entry = _find_universo_entry(player) or {}
            base_pj = (
                manual_entry.get('pj')
                if manual_entry.get('pj') is not None
                else _parse_int(universo_entry.get('pj'))
                if universo_entry.get('pj') not in (None, '')
                else roster_entry.get('pj', 0)
            )
            base_pt = (
                manual_entry.get('pt')
                if manual_entry.get('pt') is not None
                else _parse_int(universo_entry.get('pt'))
                if universo_entry.get('pt') not in (None, '')
                else roster_entry.get('pt', 0)
            )
            player_stats[player.id] = {
                'player_id': player.id,
                'name': player.name,
                'number': player.number,
                'photo_url': static(photo_path) if photo_path else '',
                'position': player.position or universo_entry.get('position') or roster_entry.get('position'),
                'total_actions': 0,
                'successes': 0,
                'pc': max(roster_entry.get('pc', 0), base_pj),
                'pj': base_pj,
                'pt': base_pt,
                'minutes': (
                    manual_entry.get('minutes')
                    if manual_entry.get('minutes') is not None
                    else _parse_int(universo_entry.get('minutes'))
                    if universo_entry.get('minutes') not in (None, '')
                    else roster_entry.get('minutes', 0)
                ),
                'goals': (
                    manual_entry.get('goals')
                    if manual_entry.get('goals') is not None
                    else _parse_int(universo_entry.get('goals'))
                    if universo_entry.get('goals') not in (None, '')
                    else roster_entry.get('goals', 0)
                ),
                'yellow_cards': (
                    manual_entry.get('yellow_cards')
                    if manual_entry.get('yellow_cards') is not None
                    else _parse_int(universo_entry.get('yellow_cards'))
                    if universo_entry.get('yellow_cards') not in (None, '')
                    else roster_entry.get('yellow_cards', 0)
                ),
                'red_cards': (
                    manual_entry.get('red_cards')
                    if manual_entry.get('red_cards') is not None
                    else _parse_int(universo_entry.get('red_cards'))
                    if universo_entry.get('red_cards') not in (None, '')
                    else roster_entry.get('red_cards', 0)
                ),
                'assists': roster_entry.get('assists', 0),
                'totals_locked': any(
                    key in manual_entry
                    for key in ('pj', 'pt', 'minutes', 'goals', 'yellow_cards', 'red_cards')
                ),
                'matches': {},
                'zone_counts': {key: 0 for key in FIELD_ZONE_KEYS},
                'position_counts': {key: 0 for key in FIELD_ZONE_KEYS},
                'tercio_counts': {label: 0 for label in STANDARD_TERCIO_LABELS},
                'tercio_totals': {label: 0 for label in STANDARD_TERCIO_LABELS},
                'duels_total': 0,
                'duels_won': 0,
                'shot_attempts': 0,
                'shots_on_target': 0,
                'pass_attempts': 0,
                'passes_completed': 0,
                'key_passes_completed': 0,
                'goalkeeper_saves': 0,
                'dribbles_attempted': 0,
                'dribbles_completed': 0,
                'age': _parse_int(universo_entry.get('age')) or roster_entry.get('age'),
                'has_events': base_pj > 0,
            }

    # Ensure imported matches tied to PlayerStatistic are visible in player panels
    # even when no MatchEvent was captured for those fixtures.
    player_stat_matches = (
        PlayerStatistic.objects
        .filter(player__team=primary_team, match__isnull=False)
        .select_related('player', 'match', 'match__home_team', 'match__away_team')
        .values('player_id', 'match_id', 'match__round', 'match__date', 'match__home_team_id', 'match__away_team_id', 'match__home_team__name', 'match__away_team__name')
        .distinct()
    )
    for row in player_stat_matches:
        player_id = _parse_int(row.get('player_id'))
        match_id = _parse_int(row.get('match_id'))
        if not player_id or not match_id:
            continue
        stats = player_stats.get(player_id)
        if not stats:
            continue
        home_team_id = _parse_int(row.get('match__home_team_id'))
        away_team_id = _parse_int(row.get('match__away_team_id'))
        if home_team_id == int(primary_team.id):
            opponent = str(row.get('match__away_team__name') or '').strip() or 'Rival desconocido'
            is_home = True
        elif away_team_id == int(primary_team.id):
            opponent = str(row.get('match__home_team__name') or '').strip() or 'Rival desconocido'
            is_home = False
        else:
            # Imported fixtures may be detached from primary team FK.
            opponent = (
                str(row.get('match__away_team__name') or '').strip()
                or str(row.get('match__home_team__name') or '').strip()
                or 'Rival desconocido'
            )
            is_home = False
        stats['matches'].setdefault(
            match_id,
            {
                'match_id': match_id,
                'round': str(row.get('match__round') or '').strip() or 'Partido sin jornada',
                'date': row.get('match__date').isoformat() if row.get('match__date') else None,
                'home': is_home,
                'opponent': opponent,
                'actions': 0,
                'successes': 0,
                'success_rate': 0,
            },
        )

    lineup_matches = {
        match.id: match
        for match in Match.objects.filter(id__in=list(lineup_by_match.keys())).select_related('home_team', 'away_team')
    }
    for match_id, lineup_seed in lineup_by_match.items():
        starters = {
            int(item.get('id'))
            for item in (lineup_seed.get('starters') or [])
            if isinstance(item, dict) and str(item.get('id') or '').isdigit()
        }
        if not starters:
            continue
        match = lineup_matches.get(match_id)
        match_end = match_end_minutes.get(match_id, 90)
        if match_end <= 0:
            match_end = 90
        for player_id in starters:
            if match_id in processed_lineup_matches[player_id]:
                continue
            stats = player_stats.get(player_id)
            if not stats:
                continue
            if stats.get('totals_locked'):
                processed_lineup_matches[player_id].add(match_id)
                continue
            stats['pt'] += 1
            stats['pj'] += 1
            stats['minutes'] += match_end
            stats['pc'] = max(stats.get('pc', 0), stats['pj'])
            stats['has_events'] = True
            if match:
                stats['matches'].setdefault(
                    match_id,
                    {
                        'match_id': match.id,
                        'round': match.round or 'Partido sin jornada',
                        'date': match.date.isoformat() if match.date else None,
                        'home': match.home_team == primary_team,
                        'opponent': (
                            match.away_team.display_name
                            if match.home_team == primary_team and match.away_team
                            else match.home_team.display_name
                            if match.away_team == primary_team and match.home_team
                            else 'Rival desconocido'
                        ),
                        'actions': 0,
                        'successes': 0,
                        'success_rate': 0,
                    },
                )
            processed_lineup_matches[player_id].add(match_id)

    result = []
    today = timezone.localdate()
    total_possible_minutes = max(0, competition_total_rounds) * 90
    max_successes = max((int(stats.get('successes', 0) or 0) for stats in player_stats.values()), default=0)
    max_decisive_actions_per90 = 0.0
    for stats in player_stats.values():
        minutes_value = int(stats.get('minutes', 0) or 0)
        successes_value = int(stats.get('successes', 0) or 0)
        if minutes_value <= 0:
            continue
        decisive_actions = (
            successes_value
            + (int(stats.get('goals', 0) or 0) * 6)
            + (int(stats.get('assists', 0) or 0) * 4)
            + (int(stats.get('key_passes_completed', 0) or 0) * 2)
        )
        max_decisive_actions_per90 = max(
            max_decisive_actions_per90,
            round((decisive_actions / minutes_value) * 90, 2),
        )
    for stats in player_stats.values():
        matches = sorted(
            stats['matches'].values(),
            key=lambda entry: (
                extract_round_number(entry['round']) is None,
                extract_round_number(entry['round']) or 9999,
                entry['date'] or '',
            ),
        )
        roster_entry = find_roster_entry(stats['name'], roster_cache)
        total_tercios = sum(stats['tercio_totals'].values())
        position_list = sorted(
            stats['position_counts'].items(),
            key=lambda item: item[1],
            reverse=True,
        )
        field_zones = []
        total_zone_actions = sum(int(count or 0) for count in stats['zone_counts'].values())
        for zone in FIELD_ZONES:
            zone_count = stats['zone_counts'].get(zone['key'], 0)
            field_zones.append(
                {
                    **zone,
                    'count': zone_count,
                    'pct': round((zone_count / total_zone_actions) * 100, 1) if total_zone_actions else 0,
                }
            )
        merged = {
            **stats,
            'matches': matches,
            'match_count': len(matches),
            'competition_total_rounds': competition_total_rounds,
            'age': stats.get('age') or (roster_entry.get('age') if roster_entry else None),
            'success_rate': round(
                (stats['successes'] / stats['total_actions']) * 100, 1
            )
            if stats['total_actions']
            else 0,
            'duel_summary': {
                'won': stats['duels_won'],
                'total': stats['duels_total'],
            },
            'duel_rate': round(
                (stats['duels_won'] / stats['duels_total']) * 100, 1
            )
            if stats['duels_total']
            else 0,
            'zone_heatmap': sorted(
                [
                    {'zone': zone, 'count': count}
                    for zone, count in stats['zone_counts'].items()
                    if count > 0
                ],
                key=lambda entry: entry['count'],
                reverse=True,
            )[:5],
            'tercio_summary': [
                {
                    'label': tercio_label,
                    'count': stats['tercio_totals'].get(tercio_label, 0),
                    'pct': round(
                        (stats['tercio_totals'].get(tercio_label, 0) / total_tercios) * 100, 1
                    )
                    if total_tercios
                    else 0,
                }
                for tercio_label in ('Ataque', 'Construcción', 'Defensa')
            ],
            'tercio_heatmap': sorted(
                [{'tercio': tercio, 'count': count} for tercio, count in stats['tercio_counts'].items()],
                key=lambda entry: entry['count'],
                reverse=True,
            )[:5],
            'position_breakdown': [
                {'label': label, 'count': count}
                for label, count in position_list
            ],
            'dominant_position': position_list[0][0] if position_list else (stats.get('position') or 'Sin definir'),
            'field_zones': field_zones,
            'shots': {
                'attempts': stats['shot_attempts'],
                'on_target': stats['shots_on_target'],
                'accuracy': round((stats['shots_on_target'] / stats['shot_attempts']) * 100, 1)
                if stats['shot_attempts']
                else 0,
                'per_goal': shots_needed_per_goal(stats['shot_attempts'], stats['goals']),
            },
            'passes': {
                'attempts': stats['pass_attempts'],
                'completed': stats['passes_completed'],
                'key_completed': stats['key_passes_completed'],
                'accuracy': round((stats['passes_completed'] / stats['pass_attempts']) * 100, 1)
                if stats['pass_attempts']
                else 0,
            },
        }
        player_obj = player_by_id.get(stats.get('player_id'))
        red_cards_value = int(stats.get('red_cards') or 0)
        yellow_cards_value = int(stats.get('yellow_cards') or 0)
        merged['has_active_injury'] = bool(player_obj and player_obj.id in active_injury_ids)
        merged['is_sanctioned'] = bool(
            player_obj
            and (
                player_obj.id in sanctioned_player_ids
                or is_manual_sanction_active(player_obj, today=today)
            )
        )
        merged['is_apercibido'] = yellow_cards_value in {4, 9, 14}
        if competition_total_rounds > 0:
            merged['participation_pct'] = round(
                min((int(stats.get('pj') or 0) / competition_total_rounds) * 100, 100),
                1,
            )
        else:
            merged['participation_pct'] = 0
        importance = calculate_importance_score(
            minutes=merged.get('minutes', 0),
            total_possible_minutes=total_possible_minutes,
            successes=merged.get('successes', 0),
            max_successes=max_successes,
        )
        merged['availability_pct'] = importance['availability_pct']
        merged['success_volume_pct'] = importance['success_volume_pct']
        merged['importance_score'] = importance['importance_score']
        influence = calculate_influence_score(
            minutes=merged.get('minutes', 0),
            successes=merged.get('successes', 0),
            goals=merged.get('goals', 0),
            assists=merged.get('assists', 0),
            key_passes_completed=merged.get('key_passes_completed', 0),
            max_decisive_actions_per90=max_decisive_actions_per90,
        )
        merged['successes_per90'] = influence['successes_per90']
        merged['decisive_actions_per90'] = influence['decisive_actions_per90']
        merged['influence_score'] = influence['influence_score']
        profile, profile_label, smart_kpis = build_smart_kpis(stats)
        merged['profile'] = profile
        merged['profile_label'] = profile_label
        merged['smart_kpis'] = smart_kpis
        result.append(merged)
    result = sorted(
        result,
        key=lambda entry: (-entry.get('total_actions', 0), -entry.get('pj', 0), entry.get('name', '')),
    )
    cache.set(cache_key, result, PLAYER_DASHBOARD_CACHE_SECONDS)
    return result
