import csv
import re
import time
import unicodedata
from io import BytesIO, StringIO
from pathlib import Path
from typing import Optional
from urllib.parse import quote_plus, urljoin, urlparse

from PIL import Image
import pytesseract
import requests
from bs4 import BeautifulSoup
from django.conf import settings
from django.template.defaultfilters import slugify
from openpyxl import load_workbook
from django.utils import timezone

from football.models import Competition, Group, Season, Team, TeamStanding


USER_AGENT = 'webstats-crm/1.0'
DOWNLOAD_TEXT_PATTERN = re.compile(r'descarg', re.IGNORECASE)
DOWNLOAD_EXTENSIONS = ('.csv', '.xls', '.xlsx', '.png')
PLAYER_ROSTER_PATH = Path(settings.BASE_DIR) / 'data' / 'input' / 'player-roster.html'
MATCH_LISTS_PATH = Path(settings.BASE_DIR) / 'data' / 'excel' / 'FICHA_PARTIDO.xlsx'
PREFERENTE_USER_AGENT = (
    'Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) '
    'AppleWebKit/537.36 (KHTML, like Gecko) '
    'Chrome/123.0.0.0 Safari/537.36'
)
PREFERENTE_BASE_URL = 'https://www.lapreferente.com/'
ROSTER_REFRESH_SECONDS = int(getattr(settings, 'PREFERENTE_ROSTER_REFRESH_SECONDS', 6 * 3600))


def normalize_header(value):
    if value is None:
        return ''
    text = str(value).strip().lower()
    return text.replace(' ', '_')


def _preferente_headers(referer: str = PREFERENTE_BASE_URL) -> dict:
    return {
        'User-Agent': PREFERENTE_USER_AGENT,
        'Accept': 'text/html,application/xhtml+xml,application/xml;q=0.9,*/*;q=0.8',
        'Accept-Language': 'es-ES,es;q=0.9,en;q=0.8',
        'Cache-Control': 'no-cache',
        'Pragma': 'no-cache',
        'Referer': referer,
    }


def _fetch_preferente_response(team_url: str, timeout: int = 25) -> requests.Response:
    session = requests.Session()
    try:
        session.get(PREFERENTE_BASE_URL, headers=_preferente_headers(PREFERENTE_BASE_URL), timeout=timeout)
    except requests.RequestException:
        pass

    response = session.get(team_url, headers=_preferente_headers(PREFERENTE_BASE_URL), timeout=timeout)
    if response.status_code == 403:
        time.sleep(1.2)
        response = session.get(team_url, headers=_preferente_headers(PREFERENTE_BASE_URL), timeout=timeout)
    return response


def ensure_league_structure(competition_name, season_name, group_name):
    competition, _ = Competition.objects.get_or_create(
        name=competition_name,
        defaults={'slug': slugify(competition_name), 'region': 'Andalucía', 'level': 5},
    )
    season, _ = Season.objects.get_or_create(
        competition=competition,
        name=season_name,
        defaults={'is_current': True},
    )
    group_slug = slugify(group_name)
    group = Group.objects.filter(season=season, slug=group_slug).first()
    if not group:
        group = (
            Group.objects.filter(season=season, name__iexact=group_name)
            .order_by('id')
            .first()
        )
    if not group:
        group = Group.objects.create(season=season, slug=group_slug, name=group_name)
    return competition, season, group


def update_team_standings(rows, source_label, source_url, competition_name='División de Honor Andaluza', season_name='2025/2026', group_name='Grupo 2'):
    _, season, group = ensure_league_structure(competition_name, season_name, group_name)
    updated_slugs = set()
    for idx, row in enumerate(rows, start=1):
        team_name = row.get('team') or row.get('equipo')
        if not team_name:
            continue
        team = _resolve_team_for_standings(team_name, group)
        updated_slugs.add(team.slug)
        update_fields = []
        if team.name != team_name:
            team.name = team_name
            update_fields.append('name')
        if team.group != group:
            team.group = group
            update_fields.append('group')
        if 'benagalbon' in _normalize_team_key(team.name):
            if team.is_primary:
                Team.objects.exclude(id=team.id).filter(is_primary=True).update(is_primary=False)
            if not team.is_primary:
                team.is_primary = True
                update_fields.append('is_primary')
        if update_fields:
            team.save(update_fields=update_fields)

        position_value = _int_or(row.get('position'), default=idx)
        played_value = _int_or(row.get('played') or row.get('pj'))
        wins_value = _int_or(row.get('wins') or row.get('pg'))
        draws_value = _int_or(row.get('draws') or row.get('pe'))
        losses_value = _int_or(row.get('losses') or row.get('pp'))
        goals_for_value = _int_or(row.get('goals_for') or row.get('gf'))
        goals_against_value = _int_or(row.get('goals_against') or row.get('gc'))
        goal_difference_value = _int_or(row.get('goal_difference') or row.get('dg'))
        points_value = _int_or(row.get('points') or row.get('pt') or row.get('pts'))
        standing, _ = TeamStanding.objects.update_or_create(
            season=season,
            group=group,
            team=team,
            defaults={
                'position': position_value,
                'played': played_value,
                'wins': wins_value,
                'draws': draws_value,
                'losses': losses_value,
                'goals_for': goals_for_value,
                'goals_against': goals_against_value,
                'goal_difference': goal_difference_value,
                'points': points_value,
                'last_updated': timezone.now(),
            },
        )
        if standing.points is None:
            wins = standing.wins or 0
            draws = standing.draws or 0
            standing.points = wins * 3 + draws
            standing.save(update_fields=['points'])
        if standing.position is None:
            standing.position = TeamStanding.objects.filter(group=group).count()
            standing.save(update_fields=['position'])
    if updated_slugs:
        TeamStanding.objects.filter(group=group).exclude(team__slug__in=updated_slugs).delete()


def _resolve_team_for_standings(team_name: str, group: Group) -> Team:
    team_slug = slugify(team_name)
    normalized_name = _normalize_team_key(team_name)
    if 'benagalbon' in normalized_name:
        primary_team = Team.objects.filter(is_primary=True).order_by('id').first()
        if primary_team:
            return primary_team
    by_slug = Team.objects.filter(slug=team_slug).first()
    if by_slug:
        return by_slug
    for candidate in Team.objects.filter(group=group):
        if _normalize_team_key(candidate.name) == normalized_name:
            return candidate
    return Team.objects.create(slug=team_slug, name=team_name, group=group)


def _normalize_team_key(value: str) -> str:
    normalized = unicodedata.normalize('NFD', value or '')
    without_accents = ''.join(ch for ch in normalized if unicodedata.category(ch) != 'Mn')
    return re.sub(r'[^a-z0-9]+', '', without_accents.lower())


def _parse_csv_rows(content):
    text = content.decode('utf-8', errors='ignore')
    reader = csv.DictReader(StringIO(text))
    return [dict(row) for row in reader]


def _parse_excel_rows(content):
    workbook = load_workbook(filename=BytesIO(content), data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    if len(rows) <= 1:
        return []
    header = [normalize_header(cell) for cell in rows[0]]
    parsed = []
    for row in rows[1:]:
        if all(cell is None for cell in row):
            continue
        record = {}
        for idx, value in enumerate(row):
            if idx >= len(header):
                break
            key = header[idx]
            if key:
                record[key] = value
        parsed.append(record)
    return parsed


def _normalize_header_text(value):
    if value is None:
        return ''
    return re.sub(r'[^a-z0-9]+', '_', str(value).lower()).strip('_')


def _parse_html_table(soup):
    tables = soup.find_all('table')
    for table in tables:
        header_row = table.find('tr')
        if not header_row:
            continue
        headers = [
            _normalize_header_text(cell.get_text())
            for cell in header_row.find_all(['th', 'td'])
        ]
        if not headers or not any('equipo' in h or 'team' in h for h in headers):
            continue
        rows = []
        for row in table.find_all('tr')[1:]:
            cells = row.find_all(['td', 'th'])
            if not cells:
                continue
            record = {}
            for idx, cell in enumerate(cells):
                if idx >= len(headers):
                    break
                key = headers[idx]
                if not key:
                    continue
                record[key] = cell.get_text(strip=True)
            if record:
                rows.append(record)
        if rows:
            return rows
    return []


def _find_download_link(soup):
    for link in soup.find_all('a', href=True):
        text = link.get_text(' ', strip=True)
        href = link['href']
        if DOWNLOAD_TEXT_PATTERN.search(text) or any(href.lower().endswith(ext) for ext in DOWNLOAD_EXTENSIONS):
            return href
    return None


def _parse_png_rows(content):
    text = pytesseract.image_to_string(Image.open(BytesIO(content)), lang='spa')
    rows = []
    pattern = re.compile(
        r'^\s*(\d+)\s+(.+?)\s+(\d+)\s+(\d+)\s+(\d+)\s+(\d+)\s+(\d+)\s+(\d+)\s+(\d+)\s+([+-]?\d+)',
        re.UNICODE,
    )
    for line in text.splitlines():
        line = line.strip()
        if not line:
            continue
        match = pattern.match(line)
        if not match:
            continue
        position, team, points, played, wins, draws, losses, gf, gc, dg = match.groups()
        rows.append(
            {
                'position': position,
                'team': team,
                'points': points,
                'played': played,
                'wins': wins,
                'draws': draws,
                'losses': losses,
                'goals_for': gf,
                'goals_against': gc,
                'goal_difference': dg,
            }
        )
    return rows


def fetch_official_rows(url):
    response = requests.get(url, headers={'User-Agent': USER_AGENT}, timeout=15)
    response.raise_for_status()
    soup = BeautifulSoup(response.text, 'html.parser')
    download_href = _find_download_link(soup)
    if download_href:
        file_url = urljoin(url, download_href)
        file_response = requests.get(file_url, headers={'User-Agent': USER_AGENT}, timeout=15)
        file_response.raise_for_status()
        ext = urlparse(file_url).path.lower()
        if ext.endswith('.csv'):
            rows = _parse_csv_rows(file_response.content)
            if rows:
                return rows, f'Descarga CSV desde {file_url}'
        elif ext.endswith(('.xls', '.xlsx')):
            rows = _parse_excel_rows(file_response.content)
            if rows:
                return rows, f'Descarga Excel desde {file_url}'
        elif ext.endswith('.png'):
            rows = _parse_png_rows(file_response.content)
            if rows:
                return rows, f'Descarga PNG desde {file_url}'
    # fallback: parse classification table directly if download link missing
    html_rows = _parse_html_table(soup)
    if html_rows:
        return html_rows, 'Clasificación extraída desde la tabla HTML'
    return None, None


def _int_or(value, default=0):
    parsed = _parse_int(value)
    if parsed is None:
        return default
    return parsed


def _parse_int(value):
    if value is None:
        return None
    try:
        return int(float(str(value).strip()))
    except (TypeError, ValueError):
        return None


def normalize_player_name(value: str) -> str:
    return slugify(value or '')


def _normalize_table_header(value: str) -> str:
    if not value:
        return ''
    return re.sub(r'[^a-z0-9]+', '', value.lower())


def _parse_int_cell(value):
    if value is None:
        return 0
    text = str(value).strip().replace('.', '').replace(',', '')
    if not text or text == '-':
        return 0
    try:
        return int(float(text))
    except (TypeError, ValueError):
        return 0


def _extract_name_cell(cell):
    spans = cell.find_all('span')
    if spans:
        text = spans[-1].get_text(' ', strip=True) or spans[0].get_text(' ', strip=True)
        if text:
            return text
    return cell.get_text(' ', strip=True)


def _safe_cell(cells, idx):
    if idx is None:
        return None
    try:
        idx = int(idx)
    except (TypeError, ValueError):
        return None
    if idx < 0 or idx >= len(cells):
        return None
    return cells[idx]


def parse_preferente_roster(html: str) -> list[dict]:
    if not html:
        return []
    soup = BeautifulSoup(html, 'html.parser')
    tables = []
    for candidate in soup.find_all('table'):
        header = candidate.find('tr')
        if not header:
            continue
        header_text = ' '.join(
            cell.get_text(' ', strip=True).lower() for cell in header.find_all(['th', 'td'])
        )
        if 'jugador' in header_text and 'min' in header_text:
            tables.append(candidate)
    if not tables:
        return parse_preferente_roster_text(html)
    roster = []
    for table in tables:
        header_row = table.find('tr')
        if not header_row:
            continue
        headers = [_normalize_table_header(cell.get_text(' ', strip=True)) for cell in header_row.find_all(['th', 'td'])]
        index = {key: idx for idx, key in enumerate(headers) if key}
        for row in table.find_all('tr')[1:]:
            cells = row.find_all('td')
            if len(cells) < 6:
                continue
            name_idx = index.get('jugador', 0)
            pos_idx = index.get('demarcacion', 1)
            name_cell = _safe_cell(cells, name_idx) or (cells[0] if cells else None)
            position_cell = _safe_cell(cells, pos_idx) or (cells[1] if len(cells) > 1 else None)
            if not name_cell:
                continue
            name = _extract_name_cell(name_cell)
            if not name:
                continue
            position = position_cell.get_text(' ', strip=True) if position_cell else ''
            age_cell = _safe_cell(cells, index.get('edad'))
            pc_cell = _safe_cell(cells, index.get('pc'))
            pj_cell = _safe_cell(cells, index.get('pj'))
            pt_cell = _safe_cell(cells, index.get('pt'))
            min_cell = _safe_cell(cells, index.get('min'))
            goals_cell = _safe_cell(cells, index.get('goles'))
            ta_cell = _safe_cell(cells, index.get('ta'))
            tr_cell = _safe_cell(cells, index.get('tr'))
            roster.append(
                {
                    'name': name,
                    'position': position,
                    'age': _parse_int_cell(age_cell.get_text(' ', strip=True) if age_cell else None),
                    'pc': _parse_int_cell(pc_cell.get_text(' ', strip=True) if pc_cell else None),
                    'pj': _parse_int_cell(pj_cell.get_text(' ', strip=True) if pj_cell else None),
                    'pt': _parse_int_cell(pt_cell.get_text(' ', strip=True) if pt_cell else None),
                    'minutes': _parse_int_cell(min_cell.get_text(' ', strip=True) if min_cell else None),
                    'goals': _parse_int_cell(goals_cell.get_text(' ', strip=True) if goals_cell else None),
                    'yellow_cards': _parse_int_cell(ta_cell.get_text(' ', strip=True) if ta_cell else None),
                    'red_cards': _parse_int_cell(tr_cell.get_text(' ', strip=True) if tr_cell else None),
                }
            )
    return roster


def parse_preferente_roster_text(raw: str) -> list[dict]:
    if not raw:
        return []
    lines = [line.strip() for line in raw.splitlines()]
    lines = [line for line in lines if line]
    status_markers = (
        'Renovado',
        'Nuevo Fichaje',
        'Jugador',
        'Cuerpo Técnico',
        'COMPETICIONES',
        'Ex-Jugadores',
        'Total de Jugadores',
    )
    position_keywords = (
        'Portero',
        'Lateral',
        'Central',
        'Medio',
        'Interior',
        'Media',
        'Extremo',
        'Delantero',
        'Pivote',
    )
    roster = []
    last_name = ''
    for line in lines:
        if any(marker in line for marker in status_markers):
            continue
        if line.isdigit():
            continue
        tokens = line.split()
        if not tokens:
            continue
        has_position = any(keyword in line for keyword in position_keywords)
        has_numbers = any(token.replace('(', '').replace(')', '').isdigit() for token in tokens)
        if has_position and has_numbers:
            position_parts = []
            stat_tokens = []
            for token in tokens:
                cleaned = token.replace('(', '').replace(')', '')
                if cleaned.replace('-', '').isdigit() or cleaned == '-':
                    stat_tokens.append(cleaned)
                else:
                    position_parts.append(token)
            position = ' '.join(position_parts).strip()
            numbers = [int(t) for t in stat_tokens if t.isdigit()]
            while len(numbers) < 8:
                numbers.append(0)
            age, pc, pj, pt, minutes, goals, yellow, red = numbers[:8]
            if last_name:
                roster.append(
                    {
                        'name': last_name,
                        'position': position,
                        'age': age,
                        'pc': pc,
                        'pj': pj,
                        'pt': pt,
                        'minutes': minutes,
                        'goals': goals,
                        'yellow_cards': yellow,
                        'red_cards': red,
                    }
                )
            continue
        last_name = line
    return roster


def fetch_preferente_team_roster(team_url: str) -> list[dict]:
    if not team_url:
        return []
    try:
        response = _fetch_preferente_response(team_url, timeout=20)
        response.raise_for_status()
    except requests.RequestException as exc:
        raise ValueError(f'Error al consultar LaPreferente: {exc}') from exc
    return parse_preferente_roster(response.text)


def find_preferente_team_url(team_name: str) -> str:
    """
    Intenta localizar la URL del equipo en LaPreferente a partir del nombre (búsqueda).
    - Evita depender de un endpoint único: prueba varios patrones habituales.
    - Valida candidatos abriendo la página y comprobando que existe `tablePlantilla`.
    """
    if not team_name or requests is None:
        return ''
    query = str(team_name or '').strip()
    if not query:
        return ''
    q = quote_plus(query)
    search_urls = [
        f'{PREFERENTE_BASE_URL}buscar.php?buscar={q}',
        f'{PREFERENTE_BASE_URL}buscador.php?buscar={q}',
        f'{PREFERENTE_BASE_URL}index.php?buscar={q}',
        f'{PREFERENTE_BASE_URL}search.php?q={q}',
    ]

    def _norm(text: str) -> str:
        return re.sub(r'[^a-z0-9]+', '', normalize_player_name(text or ''))

    target = _norm(query)
    if not target:
        return ''

    def _score_candidate(label: str, href: str) -> int:
        label_key = _norm(label)
        href_key = _norm(href)
        score = 0
        if label_key == target:
            score += 80
        if target and (target in label_key or label_key in target):
            score += 35
        if target and (target in href_key or href_key in target):
            score += 25
        # Penaliza candidatos vacíos
        if not label_key:
            score -= 10
        return score

    def _extract_candidates(html: str) -> list[str]:
        soup = BeautifulSoup(html or '', 'html.parser')
        candidates = []
        for a in soup.find_all('a', href=True):
            href = str(a.get('href') or '').strip()
            if not href:
                continue
            lowered = href.lower()
            if 'equipo' not in lowered:
                continue
            # Filtra enlaces irrelevantes (assets, anchors, etc.)
            if lowered.startswith('javascript:') or lowered.startswith('#'):
                continue
            label = a.get_text(' ', strip=True)
            score = _score_candidate(label, href)
            if score < 20:
                continue
            full = urljoin(PREFERENTE_BASE_URL, href)
            candidates.append((score, full))
        candidates.sort(key=lambda item: item[0], reverse=True)
        return [url for _, url in candidates[:10]]

    # 1) Busca en páginas de búsqueda.
    for url in search_urls:
        try:
            response = _fetch_preferente_response(url, timeout=18)
        except Exception:
            continue
        if not getattr(response, 'text', ''):
            continue
        for candidate in _extract_candidates(response.text):
            try:
                page = _fetch_preferente_response(candidate, timeout=18)
            except Exception:
                continue
            html = page.text or ''
            if 'tablePlantilla' in html:
                return candidate

    # 2) Fallback: si el equipo ya viene con enlace incrustado en el nombre (p.ej. pegado).
    maybe_url = str(team_name or '').strip()
    if maybe_url.startswith('http://') or maybe_url.startswith('https://'):
        try:
            page = _fetch_preferente_response(maybe_url, timeout=18)
            if 'tablePlantilla' in (page.text or ''):
                return maybe_url
        except Exception:
            pass
    return ''


def infer_roster_role(position: str) -> str:
    pos = (position or '').lower().replace('.', ' ').replace('-', ' ')
    compact = re.sub(r'\s+', ' ', pos).strip()
    if not compact:
        return 'MID'
    if any(token in compact for token in ('portero', 'por', 'gk')):
        return 'GK'
    if any(token in compact for token in ('defensa', 'lateral', 'central', 'carrilero', 'li', 'ld', 'ci', 'cd')):
        return 'DEF'
    if any(token in compact for token in ('delantero', 'punta', 'extremo', 'dc', 'ei', 'ed', '9')):
        return 'ATT'
    return 'MID'


def compute_probable_eleven(players: list[dict]) -> list[dict]:
    if not players:
        return []
    eligible = [p for p in players if p.get('minutes', 0) > 0]
    eligible.sort(key=lambda p: (p.get('minutes', 0), p.get('pt', 0), p.get('pj', 0)), reverse=True)
    gks = [p for p in eligible if infer_roster_role(p.get('position') or '') == 'GK']
    lineup = []
    if gks:
        lineup.append(gks[0])
    for player in eligible:
        if player in lineup:
            continue
        lineup.append(player)
        if len(lineup) >= 11:
            break
    return lineup[:11]


def build_rival_insights(players: list[dict]) -> dict:
    if not players:
        return {'top_scorers': [], 'most_minutes': [], 'most_cards': []}

    normalized_players = []
    for player in players:
        item = dict(player)
        item['goals'] = max(0, int(item.get('goals', 0) or 0))
        item['minutes'] = max(0, int(item.get('minutes', 0) or 0))
        item['pj'] = max(0, int(item.get('pj', 0) or 0))
        item['yellow_cards'] = max(0, int(item.get('yellow_cards', 0) or 0))
        item['red_cards'] = max(0, int(item.get('red_cards', 0) or 0))
        item['_role'] = infer_roster_role(item.get('position') or '')
        # Guardrails: un portero como máximo goleador suele indicar parseo roto.
        if item['_role'] == 'GK' and item['goals'] > 3:
            item['goals'] = 0
        normalized_players.append(item)

    scorer_pool = [p for p in normalized_players if p['_role'] != 'GK']
    if not scorer_pool:
        scorer_pool = normalized_players

    top_scorers = sorted(
        scorer_pool,
        key=lambda p: (
            p.get('goals', 0),
            (p.get('goals', 0) / max(1, p.get('pj', 0))),
            p.get('minutes', 0),
        ),
        reverse=True,
    )[:3]
    most_minutes = sorted(normalized_players, key=lambda p: p.get('minutes', 0), reverse=True)[:3]
    most_cards = sorted(
        normalized_players,
        key=lambda p: (p.get('red_cards', 0) * 2 + p.get('yellow_cards', 0)),
        reverse=True,
    )[:3]
    for row in top_scorers + most_minutes + most_cards:
        row.pop('_role', None)
    return {
        'top_scorers': top_scorers,
        'most_minutes': most_minutes,
        'most_cards': most_cards,
    }


def assign_lineup_slots(players: list[dict], formation: Optional[str] = None) -> list[dict]:
    assigned = []

    def role_from_position(position: str) -> str:
        pos = (position or '').lower()
        compact = pos.replace('.', '').replace('-', ' ').strip()
        if any(token in compact for token in ('portero', 'por', 'gk')):
            return 'GK'
        if 'lateral izquierdo' in compact or compact in {'li', 'dfi'}:
            return 'LI'
        if 'lateral derecho' in compact or compact in {'ld', 'dfd'}:
            return 'LD'
        if 'central izquierdo' in compact or compact in {'ci'}:
            return 'CI'
        if 'central derecho' in compact or compact in {'cd'}:
            return 'CD'
        if 'central' in compact or compact in {'c', 'dfc'}:
            return 'C'
        if 'interior izquierdo' in compact or compact in {'mi'}:
            return 'MI'
        if 'interior derecho' in compact or compact in {'md'}:
            return 'MD'
        if 'media punta' in compact or compact in {'mp', 'mco'}:
            return 'MP'
        if 'extremo izquierdo' in compact or compact in {'ei'}:
            return 'EI'
        if 'extremo derecho' in compact or compact in {'ed'}:
            return 'ED'
        if 'pivote' in compact or 'medio centro' in compact or compact in {'mc', 'mcd', 'mci'}:
            return 'MC'
        if any(token in compact for token in ('delantero', 'punta', '9', 'dc')):
            return 'DC'
        if any(token in compact for token in ('medio', 'interior', 'volante', 'media')):
            return 'MC'
        if any(token in compact for token in ('defensa', 'lateral', 'central')):
            return 'C'
        return '?'

    def classify(position: str) -> str:
        role = role_from_position(position)
        if role == 'GK':
            return 'gk'
        if role in {'LI', 'LD', 'CI', 'CD', 'C'}:
            return 'def'
        if role in {'MI', 'MD', 'MC', 'MP'}:
            return 'mid'
        if role in {'EI', 'ED', 'DC'}:
            return 'att'
        return 'mid'

    groups = {'gk': [], 'def': [], 'mid': [], 'att': []}
    for player in players:
        groups[classify(player.get('position') or '')].append(player)

    def parse_counts(value: Optional[str]) -> tuple[int, int, int]:
        if not value:
            return (4, 4, 2)
        parts = [p for p in str(value).split('-') if p.isdigit()]
        if len(parts) != 3:
            return (4, 4, 2)
        return (int(parts[0]), int(parts[1]), int(parts[2]))

    def_cap, mid_cap, att_cap = parse_counts(formation)
    gk_cap = 1

    def line_slots(line_key: str, count: int) -> list[str]:
        if line_key == 'def':
            presets = {
                1: ['C'],
                2: ['LI', 'LD'],
                3: ['LI', 'C', 'LD'],
                4: ['LI', 'CI', 'CD', 'LD'],
                5: ['LI', 'CI', 'C', 'CD', 'LD'],
            }
        elif line_key == 'mid':
            presets = {
                1: ['MC'],
                2: ['MC', 'MC'],
                3: ['MI', 'MC', 'MD'],
                4: ['MI', 'MC', 'MC', 'MD'],
                5: ['EI', 'MC', 'MC', 'MD', 'ED'],
            }
        else:  # att
            presets = {
                1: ['DC'],
                2: ['DC', 'DC'],
                3: ['EI', 'DC', 'ED'],
                4: ['EI', 'DC', 'DC', 'ED'],
                5: ['EI', 'DC', 'DC', 'DC', 'ED'],
            }
        return presets.get(count, ['MC'] * count)

    def score_role_to_slot(role: str, slot: str) -> int:
        if role == slot:
            return 10
        compatible = {
            'C': {'CI', 'CD', 'C'},
            'CI': {'CI', 'C'},
            'CD': {'CD', 'C'},
            'LI': {'LI', 'MI'},
            'LD': {'LD', 'MD'},
            'MC': {'MC', 'MP', 'MI', 'MD'},
            'MP': {'MP', 'MC', 'DC'},
            'MI': {'MI', 'EI', 'MC'},
            'MD': {'MD', 'ED', 'MC'},
            'EI': {'EI', 'MI', 'DC'},
            'ED': {'ED', 'MD', 'DC'},
            'DC': {'DC', 'MP', 'EI', 'ED'},
            'GK': {'GK'},
            '?': {'MC', 'C', 'DC'},
        }
        if slot in compatible.get(role, set()):
            return 6
        return 1

    def best_assign(line_players: list[dict], slots: list[str]) -> list[tuple[dict, str]]:
        if not line_players or not slots:
            return []
        players_with_role = [(player, role_from_position(player.get('position') or '')) for player in line_players]
        best_score = -1
        best_pick = []

        def walk(slot_idx: int, remaining: list[tuple[dict, str]], acc: list[tuple[dict, str]], score: int):
            nonlocal best_score, best_pick
            if slot_idx >= len(slots) or not remaining:
                if score > best_score:
                    best_score = score
                    best_pick = acc.copy()
                return
            slot = slots[slot_idx]
            for idx, pair in enumerate(remaining):
                player, role = pair
                pair_score = score_role_to_slot(role, slot)
                next_remaining = remaining[:idx] + remaining[idx + 1:]
                acc.append((player, slot))
                walk(slot_idx + 1, next_remaining, acc, score + pair_score)
                acc.pop()
            walk(slot_idx + 1, remaining, acc, score)

        walk(0, players_with_role, [], 0)
        return best_pick

    def take_from(group_key: str, count: int) -> list[dict]:
        picked = groups[group_key][:count]
        groups[group_key] = groups[group_key][count:]
        return picked

    remaining = []
    for key in ('def', 'mid', 'att'):
        remaining.extend(groups[key])

    lineup = []
    lineup.extend(take_from('gk', gk_cap))
    def_line = take_from('def', def_cap)
    mid_line = take_from('mid', mid_cap)
    att_line = take_from('att', att_cap)

    for line, cap in ((def_line, def_cap), (mid_line, mid_cap), (att_line, att_cap)):
        while len(line) < cap and remaining:
            line.append(remaining.pop(0))

    for (line_key, line, top) in (('def', def_line, 70), ('mid', mid_line, 52), ('att', att_line, 32)):
        slots = line_slots(line_key, len(line))
        assignment = best_assign(line, slots)
        ordered = [player for player, _ in assignment]
        slot_labels = [slot for _, slot in assignment]
        xs = {
            1: [50],
            2: [42, 58],
            3: [30, 50, 70],
            4: [22, 40, 60, 78],
            5: [16, 33, 50, 67, 84],
        }.get(len(ordered), [50])
        for idx, player in enumerate(ordered):
            enriched = dict(player)
            enriched['left'] = xs[idx]
            enriched['top'] = top
            enriched['badge'] = slot_labels[idx] if idx < len(slot_labels) else role_from_position(player.get('position') or '')
            assigned.append(enriched)

    gk = (groups['gk'][0] if groups['gk'] else (lineup[0] if lineup else None))
    if gk:
        assigned.append(
            {
                **gk,
                'left': 50,
                'top': 88,
                'badge': 'GK',
            }
        )

    return assigned


def compute_formation(players: list[dict]) -> str:
    if not players:
        return 'Auto'
    def_count = 0
    mid_count = 0
    att_count = 0
    for player in players:
        pos = (player.get('position') or '').lower()
        if 'portero' in pos:
            continue
        if 'lateral' in pos or 'central' in pos or 'defensa' in pos:
            def_count += 1
            continue
        if (
            'medio' in pos
            or 'interior' in pos
            or 'pivote' in pos
            or 'media punta' in pos
            or 'mediapunta' in pos
        ):
            mid_count += 1
            continue
        if 'delantero' in pos or 'extremo' in pos:
            att_count += 1
            continue
        mid_count += 1
    total = def_count + mid_count + att_count
    if total == 0:
        return 'Auto'
    if def_count == 0:
        def_count = 4
    if att_count == 0:
        att_count = 2
    mid_count = max(0, total - def_count - att_count)
    return f'{def_count}-{mid_count}-{att_count}'


def load_player_roster_stats() -> dict:
    if not PLAYER_ROSTER_PATH.exists():
        return {}
    try:
        html = PLAYER_ROSTER_PATH.read_text(encoding='utf-8')
    except Exception:
        try:
            html = PLAYER_ROSTER_PATH.read_text(encoding='latin-1', errors='ignore')
        except Exception:
            return {}
    soup = BeautifulSoup(html, 'html.parser')
    table = soup.find('table', id='tablePlantilla')
    if not table:
        return {}
    roster = {}
    for row in table.find_all('tr'):
        cells = row.find_all('td')
        if len(cells) < 11:
            continue
        name_cell = cells[2]
        spans = name_cell.find_all('span')
        if not spans:
            continue
        full_name = spans[-1].get_text(' ', strip=True) or spans[0].get_text(' ', strip=True)
        if not full_name:
            continue
        position = cells[3].get_text(' ', strip=True)
        normalized_name = normalize_player_name(full_name)
        roster[normalized_name] = {
            'name': full_name,
            'position': position,
            'age': _parse_int(cells[5].get_text(' ', strip=True)) or 0,
            'pc': _parse_int(cells[6].get_text(' ', strip=True)) or 0,
            'pj': _parse_int(cells[7].get_text(' ', strip=True)) or 0,
            'pt': _parse_int(cells[8].get_text(' ', strip=True)) or 0,
            'minutes': _parse_int(cells[9].get_text(' ', strip=True)) or 0,
            'goals': _parse_int(cells[10].get_text(' ', strip=True)) or 0,
            'yellow_cards': _parse_int(cells[11].get_text(' ', strip=True)) or 0,
            'red_cards': _parse_int(cells[12].get_text(' ', strip=True)) or 0,
            'assists': 0,
        }
    return roster


_ROSTER_CACHE = None
_ROSTER_CACHE_MTIME = None


def refresh_primary_roster_cache(primary_team, force: bool = False):
    if not primary_team:
        return False, 'Equipo principal no configurado'
    team_url = (primary_team.preferente_url or '').strip()
    if not team_url:
        return False, 'Sin URL de La Preferente en equipo principal'
    if not force and PLAYER_ROSTER_PATH.exists():
        try:
            age_seconds = time.time() - PLAYER_ROSTER_PATH.stat().st_mtime
            if age_seconds < ROSTER_REFRESH_SECONDS:
                return False, f'Cache reciente ({int(age_seconds // 60)} min)'
        except OSError:
            pass
    try:
        response = _fetch_preferente_response(team_url, timeout=25)
        response.raise_for_status()
    except requests.RequestException as exc:
        if PLAYER_ROSTER_PATH.exists():
            return False, f'La Preferente no respondió ({exc}); se mantiene la última plantilla en caché.'
        return False, f'Error consultando La Preferente: {exc}'
    html = response.text or ''
    if 'tablePlantilla' not in html:
        if PLAYER_ROSTER_PATH.exists():
            return False, 'La respuesta no incluyó la tabla de plantilla; se mantiene la última caché.'
        return False, 'HTML sin tabla de plantilla (tablePlantilla)'
    PLAYER_ROSTER_PATH.parent.mkdir(parents=True, exist_ok=True)
    PLAYER_ROSTER_PATH.write_text(html, encoding='utf-8')
    global _ROSTER_CACHE, _ROSTER_CACHE_MTIME
    _ROSTER_CACHE = None
    _ROSTER_CACHE_MTIME = None
    return True, f'Plantilla actualizada desde {team_url}'


def get_roster_stats_cache() -> dict:
    global _ROSTER_CACHE, _ROSTER_CACHE_MTIME
    try:
        current_mtime = PLAYER_ROSTER_PATH.stat().st_mtime if PLAYER_ROSTER_PATH.exists() else None
    except OSError:
        current_mtime = None
    if _ROSTER_CACHE is None or _ROSTER_CACHE_MTIME != current_mtime:
        _ROSTER_CACHE = load_player_roster_stats()
        _ROSTER_CACHE_MTIME = current_mtime
    return _ROSTER_CACHE


ALIAS_MAP = {
    'antonio': 'antonio-gamez-paniagua',
    'andrew': 'andrew-brayce-gonzales-ticona',
    'andrews': 'andrew-brayce-gonzales-ticona',
    'andrew-brayce-gonzales-ticona': 'andrew-brayce-gonzales-ticona',
    'manu': 'manuel-torres-palenzuela',
    'lolo': 'manuel-fernandez-canete',
    'jaime': 'javier-gutierrez-palma',
    'javi': 'javier-gutierrez-palma',
    'martinez': 'antonio-martinez-campens',
    'nico': 'nicolas-villalba-alcaide',
    'nicolas': 'nicolas-villalba-alcaide',
    'nacho': 'ignacio-dorado-morales',
    'ivan': 'ivan-fernandez-reina',
    'yaco': 'yaco-uriel-campoamor',
    'acosta': 'jose-garcia-acosta',
    'francis': 'francisco-javier-ruiz-perez',
    'juanmi': 'juan-miguel-anaya-bustamante',
    'victor': 'victor-ruiz-postigo',
    'antonio-ruiz': 'antonio-vilches',
}


def canonical_roster_key(player_name: str) -> str:
    normalized = normalize_player_name(player_name)
    return ALIAS_MAP.get(normalized, normalized)


def find_roster_entry(player_name: str, roster: dict) -> Optional[dict]:
    if not isinstance(roster, dict):
        return None
    key = canonical_roster_key(player_name)
    if key:
        entry = roster.get(key)
        if isinstance(entry, dict):
            return entry
    target = player_name.lower().strip()
    if not target:
        return None
    for entry in roster.values():
        if not isinstance(entry, dict):
            continue
        entry_name = str(entry.get('name') or '').lower()
        if not entry_name:
            continue
        if target in entry_name or entry_name in target:
            return entry
    return None


_MATCH_LIST_CACHE = None
_MATCH_RESULT_CACHE = None


def _read_match_list_sheet():
    global _MATCH_LIST_CACHE, _MATCH_RESULT_CACHE
    if _MATCH_LIST_CACHE is not None and _MATCH_RESULT_CACHE is not None:
        return _MATCH_LIST_CACHE, _MATCH_RESULT_CACHE
    actions = []
    results = []
    seen_actions = set()
    seen_results = set()
    if not MATCH_LISTS_PATH.exists():
        _MATCH_LIST_CACHE = actions
        _MATCH_RESULT_CACHE = results
        return actions, results
    try:
        workbook = load_workbook(filename=MATCH_LISTS_PATH, read_only=True, data_only=True)
        if 'LISTAS' in workbook.sheetnames:
            sheet = workbook['LISTAS']
            for row in sheet.iter_rows(values_only=True):
                if not row:
                    continue
                action_label = (row[0] or '').strip()
                result_label = (row[1] or '').strip()
                if action_label:
                    key = action_label.upper()
                    if key not in seen_actions:
                        seen_actions.add(key)
                        actions.append(action_label)
                if result_label:
                    key = result_label.upper()
                    if key not in seen_results:
                        seen_results.add(key)
                        results.append(result_label)
    except Exception:
        pass
    _MATCH_LIST_CACHE = actions
    _MATCH_RESULT_CACHE = results
    return actions, results


DEFAULT_QUICK_ACTIONS = [
    'Disparo',
    'Pase clave',
    'Robo',
    'Falta',
    'Cambio',
    'Duelo aéreo',
    'Regate',
    'Parada',
    'Saque de esquina a favor',
    'Saque de esquina en contra',
]

def load_match_actions():
    actions, _ = _read_match_list_sheet()
    if not actions:
        return DEFAULT_QUICK_ACTIONS.copy()
    ordered = []
    seen = set()
    for action in actions:
        normalized = action.strip().lower()
        if not normalized or normalized in seen:
            continue
        seen.add(normalized)
        ordered.append(action.strip())
    for extra in DEFAULT_QUICK_ACTIONS:
        normalized = extra.strip().lower()
        if normalized not in seen:
            seen.add(normalized)
            ordered.append(extra)
    return ordered


def load_match_results():
    _, results = _read_match_list_sheet()
    return results or ['Ganado', 'Perdido', 'Neutral']
